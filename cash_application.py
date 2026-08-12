"""
cash_application.py — Cash Application (O2C, the final stage).

The shape here is deliberately: record cash first, categorize it second.
A payment arrives in the bank before anyone knows exactly which
invoice(s) it settles — record_payment() just logs that money showed up.
Matching it to specific invoices (or recognizing it as an advance, if it
doesn't map to a real invoice yet) is a separate, explicit action, same
as everywhere else in this app: nothing gets categorized speculatively.

Both matching actions reuse ONE mechanism — Payment_Applications rows,
with the sentinel Invoice_ID "ADVANCE" standing in for "recognized as a
customer advance" rather than a real invoice. That keeps a payment's
"how much is still unclassified" number correct without a third table.

Accounting posting happens at match time, not at receipt time, because
that's the first point where the actual account (AR vs. a liability) is
known:
  apply_payment()      -> Dr Cash and Bank / Cr Accounts Receivable
  record_as_advance()  -> Dr Cash and Bank / Cr Customer Advances

Short payments (a customer sends less than the invoice total — routine
in India due to TDS deduction, and also just discounts or disputes) are
recorded with a reason, but this module does NOT attempt to auto-close
the invoice or fabricate a TDS Receivable account for the difference.
The invoice honestly stays Partially Paid with the real shortfall
visible; reconciling a TDS certificate against Form 26AS is real
accounting work this PoC doesn't attempt to automate.

Building block left for later, not wired in now: get_customer_open_ar()
computes a customer's real outstanding AR from actual applied/unpaid
invoices — exactly what customer_onboarding.check_credit_available()'s
docstring flagged as missing. Wiring it into the credit check is a
follow-up (needs sales_order.py to accept an injected credit_check_fn,
same DI shape as elsewhere) — not done here to keep this module's
first pass focused and to avoid touching already-tested modules
without being asked.

New sheets:
  Payments              Payment_ID | Customer_ID | Customer_Name |
                         Payment_Date | Amount | Payment_Method |
                         Reference_No | Unapplied_Amount | Notes
  Payment_Applications   Payment_ID | Line_Item | Invoice_ID |
                         Applied_Amount | Short_Payment_Reason |
                         Application_Date

SQLite pilot: Payments and Payment_Applications now live in
erp_pilot.db (tables `payments`, `payment_applications`) — this module
is the exclusive owner of both. This is the final table set of the
whole application to move — every business table in the app is now
SQLite-backed.
"""

import os, io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import billing as bl
import customer_onboarding as co
import accounting as acct
import rma

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

PAY_SHEET = "Payments"
PAY_APP_SHEET = "Payment_Applications"

PAY_COLS = ["Payment_ID", "Customer_ID", "Customer_Name", "Payment_Date", "Amount",
           "Payment_Method", "Reference_No", "Unapplied_Amount", "Notes"]
PAY_APP_COLS = ["Payment_ID", "Line_Item", "Invoice_ID", "Applied_Amount",
               "Short_Payment_Reason", "Application_Date"]

PAYMENT_METHODS = ["Bank Transfer", "Cheque", "UPI", "Cash", "Other"]
SHORT_PAYMENT_REASONS = ["", "TDS Deducted", "Discount Allowed", "Dispute", "Bank Charges", "Other"]
ADVANCE_SENTINEL = "ADVANCE"


def ensure_sheets(wb=None):
    """Kept for signature compatibility — Payments/Payment_Applications
    no longer live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


# ── Read: payments ──────────────────────────────────────────────────────────────
def get_payments(customer_id=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM payments ORDER BY payment_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"payment_id": r["payment_id"], "customer_id": r["customer_id"],
               "customer_name": r["customer_name"], "payment_date": r["payment_date"],
               "amount": r["amount"], "payment_method": r["payment_method"],
               "reference_no": r["reference_no"], "unapplied_amount": r["unapplied_amount"],
               "notes": r["notes"]}
        if customer_id and row["customer_id"] != customer_id:
            continue
        out.append(row)
    return out


def get_payment(payment_id, data_file=None):
    for p in get_payments(data_file=data_file):
        if p["payment_id"] == payment_id:
            return p
    return None


def get_payment_applications(payment_id=None, invoice_id=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM payment_applications ORDER BY payment_id, line_item").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"payment_id": r["payment_id"], "line_item": r["line_item"],
               "invoice_id": r["invoice_id"], "applied_amount": r["applied_amount"],
               "short_payment_reason": r["short_payment_reason"],
               "application_date": r["application_date"]}
        if payment_id and row["payment_id"] != payment_id:
            continue
        if invoice_id and row["invoice_id"] != invoice_id:
            continue
        out.append(row)
    return out


# ── Invoice-side view — computed, not stored on the Invoices sheet ─────────────
def get_invoice_payment_info(invoice_id, data_file=None):
    fpath = data_file or DATA_FILE
    inv = bl.get_invoice(invoice_id, fpath)
    if inv is None:
        raise ValueError(f"{invoice_id} not found.")
    applications = [a for a in get_payment_applications(invoice_id=invoice_id, data_file=fpath)
                    if a["invoice_id"] != ADVANCE_SENTINEL]
    paid = round(sum(a["applied_amount"] for a in applications), 2)
    # RMA-US-04's own credit memos reduce what's genuinely still owed on this
    # invoice, the same way a payment does — read live from credit_memos
    # (rma.py's own authoritative source) rather than a second, separately-
    # maintained AR figure that could drift from it.
    credited = round(sum(cm["credit_total"] for cm in rma.get_credit_memos(
        invoice_id=invoice_id, data_file=fpath)), 2)
    balance_due = round(inv["grand_total"] - paid - credited, 2)
    if inv["status"] == "Cancelled":
        payment_status = "N/A (Cancelled)"
    elif balance_due <= 0:
        payment_status = "Paid" if credited == 0 or paid > 0 else "Credited"
    elif paid > 0 or credited > 0:
        payment_status = "Partially Paid"
    else:
        payment_status = "Unpaid"
    return {"grand_total": inv["grand_total"], "paid_amount": paid, "credited_amount": credited,
            "balance_due": balance_due, "payment_status": payment_status}


def get_customer_open_invoices(customer_id, data_file=None):
    """Issued invoices for this customer with a real remaining balance —
    the natural candidate list when applying a payment."""
    fpath = data_file or DATA_FILE
    out = []
    for inv in bl.get_invoices(status="Issued", data_file=fpath):
        if inv["customer_id"] != customer_id:
            continue
        info = get_invoice_payment_info(inv["invoice_id"], fpath)
        if info["balance_due"] > 0:
            out.append({**inv, **info})
    return out


def get_customer_open_ar(customer_id, data_file=None):
    """Real outstanding AR for a customer — sum of balance_due across their
    Issued invoices. Not yet wired into credit checking; see module docstring."""
    return round(sum(inv["balance_due"] for inv in get_customer_open_invoices(customer_id, data_file)), 2)


def get_overdue_invoices(data_file=None):
    """Issued invoices with a real balance still due, past their due date —
    a straightforward collections worklist."""
    fpath = data_file or DATA_FILE
    today = date.today().isoformat()
    out = []
    for inv in bl.get_invoices(status="Issued", data_file=fpath):
        info = get_invoice_payment_info(inv["invoice_id"], fpath)
        if info["balance_due"] > 0 and str(inv["due_date"]) < today:
            out.append({**inv, **info})
    return out


# ── Record a payment ────────────────────────────────────────────────────────────
def _next_payment_id(conn):
    rows = conn.execute("SELECT payment_id FROM payments WHERE payment_id LIKE 'PMT-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["payment_id"].split("-")[1]))
        except Exception: pass
    return f"PMT-{mx+1:05d}"


def record_payment(customer_id, amount, payment_date=None, payment_method="Bank Transfer",
                   reference_no="", notes="", data_file=None):
    fpath = data_file or DATA_FILE
    customer = co.get_customer(customer_id, fpath)
    if customer is None:
        raise ValueError(f"Customer {customer_id} not found.")
    if amount <= 0:
        raise ValueError("Payment amount must be positive.")

    pdate = payment_date or date.today()
    pdate_str = pdate.strftime("%Y-%m-%d") if hasattr(pdate, "strftime") else str(pdate)

    db.init_schema()
    conn = db.get_connection()
    try:
        payment_id = _next_payment_id(conn)
        conn.execute(
            "INSERT INTO payments (payment_id, customer_id, customer_name, payment_date, "
            "amount, payment_method, reference_no, unapplied_amount, notes) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (payment_id, customer_id, customer.get("Customer_Name", customer_id), pdate_str,
             round(float(amount), 2), payment_method, reference_no, round(float(amount), 2), notes),
        )
        conn.commit()
    finally:
        conn.close()
    return payment_id


# ── Match a payment ────────────────────────────────────────────────────────────
def _next_app_line(conn):
    """Global max across ALL payments' applications, not per-payment —
    faithfully replicates the pre-migration behavior (the Excel version
    scanned the whole Line_Item column unfiltered too)."""
    row = conn.execute("SELECT MAX(line_item) AS mx FROM payment_applications").fetchone()
    return (row["mx"] or 0) + 1


def _record_application(payment_id, invoice_id, amount, reason, data_file):
    db.init_schema()
    conn = db.get_connection()
    try:
        line = _next_app_line(conn)
        conn.execute(
            "INSERT INTO payment_applications (payment_id, line_item, invoice_id, "
            "applied_amount, short_payment_reason, application_date) VALUES (?,?,?,?,?,?)",
            (payment_id, line, invoice_id, round(amount, 2), reason,
             date.today().strftime("%Y-%m-%d")),
        )
        conn.execute(
            "UPDATE payments SET unapplied_amount = ROUND(COALESCE(unapplied_amount, 0) - ?, 2) "
            "WHERE payment_id = ?",
            (amount, payment_id),
        )
        conn.commit()
    finally:
        conn.close()


def apply_payment(payment_id, invoice_id, amount, short_payment_reason="", data_file=None):
    fpath = data_file or DATA_FILE
    payment = get_payment(payment_id, fpath)
    if payment is None:
        raise ValueError(f"{payment_id} not found.")
    if amount <= 0:
        raise ValueError("Applied amount must be positive.")
    if amount > payment["unapplied_amount"] + 0.005:
        raise ValueError(f"{payment_id} only has \u20b9{payment['unapplied_amount']:,.2f} "
                          "unapplied — can't apply more than that.")

    inv = bl.get_invoice(invoice_id, fpath)
    if inv is None:
        raise ValueError(f"{invoice_id} not found.")
    if inv["status"] != "Issued":
        raise ValueError(f"{invoice_id} is '{inv['status']}' — only an Issued invoice can take a payment.")
    info = get_invoice_payment_info(invoice_id, fpath)
    if amount > info["balance_due"] + 0.005:
        raise ValueError(f"{invoice_id} only has \u20b9{info['balance_due']:,.2f} outstanding — "
                          "can't apply more than that to a single invoice.")

    _record_application(payment_id, invoice_id, amount, short_payment_reason, fpath)
    je_id = acct.post_journal_entry(
        "Payment", payment_id,
        f"Payment {payment_id} applied to {invoice_id}" + (f" ({short_payment_reason})" if short_payment_reason else ""),
        [{"account_code": "1000", "debit": amount, "credit": 0, "description": f"Cash received, {payment_id}"},
         {"account_code": "1100", "debit": 0, "credit": amount, "description": f"AR settled, {invoice_id}"}],
        fpath)
    return je_id


def record_as_advance(payment_id, amount, data_file=None):
    fpath = data_file or DATA_FILE
    payment = get_payment(payment_id, fpath)
    if payment is None:
        raise ValueError(f"{payment_id} not found.")
    if amount <= 0:
        raise ValueError("Advance amount must be positive.")
    if amount > payment["unapplied_amount"] + 0.005:
        raise ValueError(f"{payment_id} only has \u20b9{payment['unapplied_amount']:,.2f} "
                          "unapplied — can't recognize more than that as an advance.")

    _record_application(payment_id, ADVANCE_SENTINEL, amount, "", fpath)
    je_id = acct.post_journal_entry(
        "Payment", payment_id, f"Payment {payment_id} recognized as customer advance",
        [{"account_code": "1000", "debit": amount, "credit": 0, "description": f"Cash received, {payment_id}"},
         {"account_code": "2300", "debit": 0, "credit": amount, "description": f"Advance, {payment_id}"}],
        fpath)
    return je_id


# ── Document generation ───────────────────────────────────────────────────────
def generate_payment_receipt(payment_id, data_file=None):
    fpath = data_file or DATA_FILE
    payment = get_payment(payment_id, fpath)
    if payment is None:
        raise ValueError(f"{payment_id} not found.")
    applications = get_payment_applications(payment_id=payment_id, data_file=fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = "0C4A6E"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Receipt"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "PAYMENT RECEIPT")
    ws.merge_cells("A1:E1")
    label("A3", "Payment ID:");  value("B3", payment["payment_id"])
    label("A4", "Date:");        value("B4", payment["payment_date"])
    label("A5", "From:");        value("B5", payment["customer_name"])
    label("D3", "Amount:");      value("E3", payment["amount"])
    label("D4", "Method:");      value("E4", payment["payment_method"])
    label("D5", "Reference:");   value("E5", payment["reference_no"])

    hdr_row = 8
    hdrs = ["#", "Applied To", "Amount", "Reason", "Date"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=navy)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for i, a in enumerate(applications, 1):
        target = "Customer Advance" if a["invoice_id"] == ADVANCE_SENTINEL else a["invoice_id"]
        vals = [i, target, a["applied_amount"], a["short_payment_reason"] or "", a["application_date"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="1A1A2E")
            c.border = bdr
        r += 1
    if payment["unapplied_amount"] > 0:
        ws.cell(r, 2, "Unapplied (not yet matched)").font = Font(name="Arial", size=9, italic=True, color="94A3B8")
        ws.cell(r, 3, payment["unapplied_amount"]).font = Font(name="Arial", size=9, italic=True, color="94A3B8")

    widths = [4, 22, 14, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{payment_id}_{payment['customer_id']}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    payments = get_payments(data_file=data_file)
    total_received = round(sum(p["amount"] for p in payments), 2)
    total_unapplied = round(sum(p["unapplied_amount"] for p in payments), 2)
    overdue = get_overdue_invoices(data_file)
    return {"total_payments": len(payments), "total_received": total_received,
            "total_unapplied": total_unapplied,
            "overdue_count": len(overdue),
            "overdue_amount": round(sum(o["balance_due"] for o in overdue), 2)}


if __name__ == "__main__":
    print("Cash application stats:", stats())
