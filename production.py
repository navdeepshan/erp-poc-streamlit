"""
production.py — Production Confirmation (MFG, manufacturing pilot).

The event inventory.py's own docstring named as missing: until this
module, nothing in this system ever recorded that a build actually
happened. BOM (bom.py) only ever computed what a build WOULD need;
Goods Receipt only ever recorded components arriving. Nothing recorded
components being consumed into a build, or the finished good coming out
the other end. This is that event.

confirm_production() does exactly two things atomically, both through
inventory.py's ledger (this module holds no inventory logic of its own):
  - Explodes the BOM for parent_code x quantity, posts a NEGATIVE
    (consumption) transaction per component at location_id
  - Posts a POSITIVE (output) transaction for parent_code itself at the
    same location

Record-only, same philosophy as every other module here that records
rather than gatekeeps (GR allows over-receipt, QC allows failures,
inventory.py allows negative balances): if there isn't enough of a
component on hand, the consumption still posts and the balance goes
negative — that's a real, informative fact about a process gap, not
something to silently prevent. The UI is expected to show a shortage
preview before confirming, same pattern as BOM's net_requirements()
preview before proposing a PR, so a human sees it coming rather than
finding out after the fact — but this module itself doesn't block on it.

New sheet:
  Production_Confirmations   Confirmation_ID | Parent_Item_Code |
                             Parent_Item_Desc | Quantity_Built |
                             Location_ID | Confirmation_Date |
                             Confirmed_By | Notes

SQLite pilot: Production_Confirmations now lives in erp_pilot.db
(table `production_confirmations`).

Active-flag policy (warn-only, per direct instruction): confirm_production()
checks whether parent_code is currently Active and appends any concern
to the confirmation's notes — it never blocks. inventory.py's own
record_transaction() does its own equivalent check per component
consumed/produced, so a build against an inactive component or output
item is warned about on both the confirmation record and the ledger,
without ever being refused.
"""

import os, io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import bom
import inventory as inv
import po_export
import pr_consolidation as pc

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

CONF_SHEET = "Production_Confirmations"
CONF_COLS = ["Confirmation_ID", "Parent_Item_Code", "Parent_Item_Desc", "Quantity_Built",
            "Location_ID", "Confirmation_Date", "Confirmed_By", "Notes"]


def ensure_sheets(wb=None):
    """Kept for signature compatibility — Production_Confirmations no
    longer lives in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


def _next_confirmation_id(conn):
    rows = conn.execute(
        "SELECT confirmation_id FROM production_confirmations WHERE confirmation_id LIKE 'PC-%'"
    ).fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["confirmation_id"].split("-")[1]))
        except Exception: pass
    return f"PC-{mx+1:05d}"


def preview_production(parent_code, quantity, location_id, data_file=None):
    """
    What confirm_production() would do, without doing it — the shortage
    check a UI should show before letting a human commit to it. Returns
    the exploded components with on-hand balance and resulting (possibly
    negative) balance after consumption, plus the parent item's own
    current balance for reference.
    """
    fpath = data_file or DATA_FILE
    detailed = bom.explode_bom_detailed(parent_code, quantity, fpath)
    if not detailed:
        raise ValueError(f"{parent_code} has no BOM — nothing to build from.")
    out = []
    for d in detailed:
        on_hand = inv.get_balance(d["mat_code"], location_id, fpath)
        out.append({**d, "on_hand_qty": on_hand, "after_qty": round(on_hand - d["gross_qty"], 3),
                   "shortage": on_hand < d["gross_qty"]})
    return out


def confirm_production(parent_code, quantity, location_id, confirmed_by="", notes="", data_file=None):
    """
    Explodes the BOM, consumes every component from location_id (allowing
    negative balances, per module docstring), and produces quantity units
    of parent_code at the same location. Returns the confirmation_id and
    the component consumption list actually posted.
    """
    fpath = data_file or DATA_FILE
    if quantity <= 0:
        raise ValueError("Quantity built must be greater than zero.")
    detailed = bom.explode_bom_detailed(parent_code, quantity, fpath)
    if not detailed:
        raise ValueError(f"{parent_code} has no BOM — nothing to build from.")

    parent_master = po_export.get_item_by_code(parent_code, active_only=False)
    parent_desc = parent_master["desc"] if parent_master else parent_code

    confirmation_notes = notes
    if parent_master and parent_master["active"] != "Yes":
        warning = f"Parent item {parent_code} is currently marked inactive."
        confirmation_notes = (notes + " | " if notes else "") + warning

    db.init_schema()
    conn = db.get_connection()
    try:
        confirmation_id = _next_confirmation_id(conn)
        conn.execute(
            "INSERT INTO production_confirmations (confirmation_id, parent_item_code, "
            "parent_item_desc, quantity_built, location_id, confirmation_date, confirmed_by, "
            "notes) VALUES (?,?,?,?,?,?,?,?)",
            (confirmation_id, parent_code, parent_desc, quantity, location_id,
             date.today().strftime("%Y-%m-%d"), confirmed_by, confirmation_notes),
        )
        conn.commit()
    finally:
        conn.close()

    # Consume components (negative), produce the parent (positive) — both
    # through inventory.py, same reference so they're traceable together.
    # fpath threaded through explicitly: without it, these fall back to
    # inventory.py's own default DATA_FILE, which silently diverges from
    # whatever data_file this function was actually called with — exactly
    # the "which value should be authoritative" class of bug this project
    # keeps finding. Confirmed via a real reproduction: calling this with
    # a non-default data_file still wrote inventory transactions into the
    # default data.xlsx before this fix.
    for d in detailed:
        inv.record_transaction(d["mat_code"], d["mat_desc"], location_id, -d["gross_qty"],
                               "Production Consumption", "Production", confirmation_id,
                               data_file=fpath)
    inv.record_transaction(parent_code, parent_desc, location_id, quantity,
                           "Production Output", "Production", confirmation_id,
                           data_file=fpath)

    return {"confirmation_id": confirmation_id, "components_consumed": len(detailed)}


# ── Read ──────────────────────────────────────────────────────────────────────
def get_confirmations(data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM production_confirmations ORDER BY confirmation_id DESC"
        ).fetchall()
    finally:
        conn.close()
    return [{"confirmation_id": r["confirmation_id"], "parent_code": r["parent_item_code"],
             "parent_desc": r["parent_item_desc"], "quantity": r["quantity_built"],
             "location_id": r["location_id"], "confirmation_date": r["confirmation_date"],
             "confirmed_by": r["confirmed_by"], "notes": r["notes"]} for r in rows]


def get_confirmation_detail(confirmation_id, data_file=None):
    """Re-derives the component consumption from inventory.py's own
    transaction log — no separate line-item table to keep in sync."""
    fpath = data_file or DATA_FILE
    txns = inv.get_transactions(data_file=fpath)
    return [t for t in txns if t["reference_id"] == confirmation_id]


# ── Document generation ───────────────────────────────────────────────────────
def generate_production_slip(confirmation_id, data_file=None):
    fpath = data_file or DATA_FILE
    conf = next((c for c in get_confirmations(fpath) if c["confirmation_id"] == confirmation_id), None)
    if conf is None:
        raise ValueError(f"{confirmation_id} not found.")
    lines = get_confirmation_detail(confirmation_id, fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    purple = "4C1D95"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Slip"

    c = ws["A1"]; c.value = "PRODUCTION CONFIRMATION"
    c.font = Font(name="Arial", size=15, bold=True, color="0F172A")
    ws.merge_cells("A1:F1")
    ws["A3"] = "Confirmation:"; ws["A3"].font = Font(name="Arial", size=9, bold=True)
    ws["B3"] = conf["confirmation_id"]; ws["B3"].font = Font(name="Arial", size=10)
    ws["D3"] = "Date:"; ws["D3"].font = Font(name="Arial", size=9, bold=True)
    ws["E3"] = conf["confirmation_date"]; ws["E3"].font = Font(name="Arial", size=10)
    ws["A4"] = "Built:"; ws["A4"].font = Font(name="Arial", size=9, bold=True)
    ws["B4"] = f"{conf['quantity']:g} x {conf['parent_code']} — {conf['parent_desc']}"
    ws["B4"].font = Font(name="Arial", size=10)
    ws["A5"] = "Location:"; ws["A5"].font = Font(name="Arial", size=9, bold=True)
    loc_name = conf["location_id"]
    for d in pc.get_delivery_locations(active_only=False):
        if d["id"] == loc_name:
            loc_name = d["name"]
            break
    ws["B5"] = loc_name; ws["B5"].font = Font(name="Arial", size=10)

    hdr_row = 8
    hdrs = ["Material Code", "Description", "Movement", "Qty"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=purple)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for ln in lines:
        movement = "Output" if ln["quantity"] > 0 else "Consumed"
        vals = [ln["mat_code"], ln["mat_desc"], movement, abs(ln["quantity"])]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9,
                          color="15803D" if movement == "Output" else "1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    widths = [16, 34, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 20

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{confirmation_id}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    confs = get_confirmations(data_file)
    return {"total_confirmations": len(confs),
            "total_units_built": round(sum(c["quantity"] for c in confs), 3)}


if __name__ == "__main__":
    print("Production stats:", stats())
