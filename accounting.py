"""
accounting.py — Chart of Accounts + Journal Entries (O2C).

Two posting events, triggered by the UI at the exact business moment they
represent (not automatically embedded inside billing.py/fulfillment.py —
this module imports them, not the other way around, keeping the same
one-directional dependency shape used throughout this app):

  Invoice marked Issued  -> post_invoice_entry()
    Dr Accounts Receivable         grand_total
    Cr Sales Revenue               subtotal
    Cr GST Output - CGST/SGST/IGST tax amounts (whichever apply)

  Fulfillment marked Delivered -> post_fulfillment_entry()
    Dr Cost of Goods Sold          qty_shipped x Item Master cost
    Cr Inventory Clearing          same amount

Every entry is validated to balance (total debits == total credits)
before it's written — an unbalanced entry is refused outright, not
written with a warning, since a ledger that doesn't balance isn't
a ledger.

On "Inventory Clearing" rather than "Inventory Asset": a Goods Receipt
process exists in this organization's actual production system, but it
is explicitly out of scope for this PoC (confirmed directly, not
assumed) — so there is no corresponding entry anywhere that ever debits
Inventory. Crediting a real "Inventory Asset" account with nothing ever
debiting it would make that account run permanently negative and look
like a bug to anyone reading a balance sheet. "Inventory Clearing" keeps
COGS correct on the P&L (genuinely useful now) while being honest that
the balance-sheet side isn't complete without Goods Receipt feeding it
— that's a real gap in this PoC's scope, not a bug in this module.

New sheets:
  Chart_of_Accounts     Account_Code | Account_Name | Account_Type |
                         Description
  Journal_Entries        JE_ID | Entry_Date | Source_Type | Source_ID |
                         Description | Total_Debit | Total_Credit
  Journal_Entry_Lines    JE_ID | Line_Item | Account_Code | Account_Name |
                         Debit | Credit | Description

SQLite pilot: Chart_of_Accounts, Journal_Entries, and Journal_Entry_Lines
now live in erp_pilot.db (tables `chart_of_accounts`, `journal_entries`,
`journal_entry_lines`) — this module is the exclusive owner of all
three. SEED_ACCOUNTS is inserted by migrate_o2c_accounting.py on first
migration (mirroring the old "seed on first get_accounts() call"
behavior) rather than lazily here, since a schema init is now a
one-time, explicit step (db.init_schema()) rather than something that
happens implicitly on every read.
"""

import os
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

import db
import po_export
import billing as bl
import fulfillment as ful
import goods_receipt as gr

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

COA_SHEET = "Chart_of_Accounts"
JE_SHEET = "Journal_Entries"
JE_LINES_SHEET = "Journal_Entry_Lines"

COA_COLS = ["Account_Code", "Account_Name", "Account_Type", "Description"]
JE_COLS = ["JE_ID", "Entry_Date", "Source_Type", "Source_ID", "Description",
          "Total_Debit", "Total_Credit"]
JE_LINE_COLS = ["JE_ID", "Line_Item", "Account_Code", "Account_Name",
               "Debit", "Credit", "Description"]

# Seeded on first use — exactly the accounts this PoC's postings need.
# add_account() can extend this later (e.g. Cash/Bank for Cash Application).
SEED_ACCOUNTS = [
    ("1000", "Cash and Bank", "Asset", "Cash and bank balances"),
    ("1100", "Accounts Receivable", "Asset", "Amounts owed by customers"),
    ("1200", "Inventory Clearing", "Asset",
     "Clearing account for COGS postings — see module docstring: Goods "
     "Receipt exists in production but isn't modeled in this PoC, so "
     "nothing ever debits Inventory here"),
    ("2300", "Customer Advances", "Liability",
     "Payments received but not yet applied to a specific invoice"),
    ("4000", "Sales Revenue", "Revenue", "Revenue from sale of goods"),
    ("5000", "Cost of Goods Sold", "Expense", "Cost of goods sold, from Item Master cost basis"),
    ("2200", "GST Output - CGST", "Liability", "Central GST collected on sales, payable to government"),
    ("2210", "GST Output - SGST", "Liability", "State GST collected on sales, payable to government"),
    ("2220", "GST Output - IGST", "Liability", "Integrated GST collected on inter-state sales, payable to government"),
]


# ── Sheet bootstrap ────────────────────────────────────────────────────────────
def ensure_sheets(wb=None):
    """Kept for signature compatibility — Chart_of_Accounts/Journal_Entries/
    Journal_Entry_Lines no longer live in the Excel workbook. `wb` is
    accepted and ignored. Seed accounts are inserted by
    migrate_o2c_accounting.py, not here — see module docstring."""
    db.init_schema()


# ── Chart of Accounts ─────────────────────────────────────────────────────────
def get_accounts(data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM chart_of_accounts ORDER BY account_code").fetchall()
    finally:
        conn.close()
    return [{"account_code": r["account_code"], "account_name": r["account_name"],
             "account_type": r["account_type"], "description": r["description"]} for r in rows]


def get_account(account_code, data_file=None):
    for a in get_accounts(data_file):
        if a["account_code"] == account_code:
            return a
    return None


def add_account(account_code, account_name, account_type, description="", data_file=None):
    db.init_schema()
    conn = db.get_connection()
    try:
        existing = conn.execute(
            "SELECT 1 FROM chart_of_accounts WHERE account_code = ?", (account_code,)
        ).fetchone()
        if existing is not None:
            raise ValueError(f"Account {account_code} already exists.")
        conn.execute(
            "INSERT INTO chart_of_accounts (account_code, account_name, account_type, "
            "description) VALUES (?,?,?,?)",
            (account_code, account_name, account_type, description),
        )
        conn.commit()
    finally:
        conn.close()


# ── Journal Entries — generic posting primitive ─────────────────────────────────
def _next_je_id(conn):
    rows = conn.execute("SELECT je_id FROM journal_entries WHERE je_id LIKE 'JE-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["je_id"].split("-")[1]))
        except Exception: pass
    return f"JE-{mx+1:05d}"


def post_journal_entry(source_type, source_id, description, lines, data_file=None):
    """
    lines: list of {account_code, debit, credit, description}. Refuses to
    post if total debits != total credits — an unbalanced entry is a bug,
    not a warning.
    """
    total_debit = round(sum(l.get("debit", 0) or 0 for l in lines), 2)
    total_credit = round(sum(l.get("credit", 0) or 0 for l in lines), 2)
    if total_debit != total_credit:
        raise ValueError(f"Journal entry doesn't balance: debits {total_debit} != "
                          f"credits {total_credit}. Refusing to post.")
    if total_debit == 0:
        raise ValueError("Journal entry has zero value on both sides — nothing to post.")

    db.init_schema()
    conn = db.get_connection()
    try:
        je_id = _next_je_id(conn)
        conn.execute(
            "INSERT INTO journal_entries (je_id, entry_date, source_type, source_id, "
            "description, total_debit, total_credit) VALUES (?,?,?,?,?,?,?)",
            (je_id, date.today().strftime("%Y-%m-%d"), source_type, source_id,
             description, total_debit, total_credit),
        )
        accounts = {r["account_code"]: r["account_name"] for r in
                    conn.execute("SELECT account_code, account_name FROM chart_of_accounts").fetchall()}
        for seq, line in enumerate(lines, 1):
            acct_name = accounts.get(line["account_code"], line["account_code"])
            conn.execute(
                "INSERT INTO journal_entry_lines (je_id, line_item, account_code, account_name, "
                "debit, credit, description) VALUES (?,?,?,?,?,?,?)",
                (je_id, seq, line["account_code"], acct_name,
                 line.get("debit") or 0, line.get("credit") or 0, line.get("description", "")),
            )
        conn.commit()
    finally:
        conn.close()
    return je_id


def get_journal_entries(source_type=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM journal_entries ORDER BY je_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"je_id": r["je_id"], "entry_date": r["entry_date"], "source_type": r["source_type"],
               "source_id": r["source_id"], "description": r["description"],
               "total_debit": r["total_debit"], "total_credit": r["total_credit"]}
        if source_type and row["source_type"] != source_type:
            continue
        out.append(row)
    return out


def get_journal_entry(je_id, data_file=None):
    for je in get_journal_entries(data_file=data_file):
        if je["je_id"] == je_id:
            return je
    return None


def get_journal_entry_lines(je_id, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM journal_entry_lines WHERE je_id = ? ORDER BY line_item", (je_id,)
        ).fetchall()
    finally:
        conn.close()
    return [{"je_id": r["je_id"], "line_item": r["line_item"], "account_code": r["account_code"],
             "account_name": r["account_name"], "debit": r["debit"], "credit": r["credit"],
             "description": r["description"]} for r in rows]


def already_posted(source_type, source_id, data_file=None):
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT 1 FROM journal_entries WHERE source_type = ? AND source_id = ? LIMIT 1",
            (source_type, source_id),
        ).fetchone()
    finally:
        conn.close()
    return row is not None


# ── Specific postings ────────────────────────────────────────────────────────
def post_invoice_entry(invoice_id, data_file=None):
    fpath = data_file or DATA_FILE
    if already_posted("Invoice", invoice_id, fpath):
        raise ValueError(f"{invoice_id} has already been posted to the ledger.")
    inv = bl.get_invoice(invoice_id, fpath)
    if inv is None:
        raise ValueError(f"{invoice_id} not found.")
    if inv["status"] != "Issued":
        raise ValueError(f"{invoice_id} is '{inv['status']}' — only an Issued "
                          "invoice gets posted to the ledger.")

    lines = [{"account_code": "1100", "debit": inv["grand_total"], "credit": 0,
             "description": f"AR for {invoice_id}"},
             {"account_code": "4000", "debit": 0, "credit": inv["subtotal"],
             "description": f"Revenue for {invoice_id}"}]
    if inv["cgst_total"]:
        lines.append({"account_code": "2200", "debit": 0, "credit": inv["cgst_total"],
                      "description": f"CGST on {invoice_id}"})
    if inv["sgst_total"]:
        lines.append({"account_code": "2210", "debit": 0, "credit": inv["sgst_total"],
                      "description": f"SGST on {invoice_id}"})
    if inv["igst_total"]:
        lines.append({"account_code": "2220", "debit": 0, "credit": inv["igst_total"],
                      "description": f"IGST on {invoice_id}"})

    return post_journal_entry("Invoice", invoice_id,
                              f"Invoice {invoice_id} — {inv['customer_name']}", lines, fpath)


def post_fulfillment_entry(fulfillment_id, data_file=None):
    fpath = data_file or DATA_FILE
    if already_posted("Fulfillment", fulfillment_id, fpath):
        raise ValueError(f"{fulfillment_id} has already been posted to the ledger.")
    f = ful.get_fulfillment(fulfillment_id, fpath)
    if f is None:
        raise ValueError(f"{fulfillment_id} not found.")
    if f["status"] != "Delivered":
        raise ValueError(f"{fulfillment_id} is '{f['status']}' — only a Delivered "
                          "fulfillment gets posted to the ledger.")

    items = [i for i in ful.get_fulfillment_items(fulfillment_id, fpath) if (i["qty_shipped"] or 0) > 0]
    if not items:
        raise ValueError(f"{fulfillment_id} has no shipped quantity — nothing to post.")

    cost_by_code = {i["code"]: i["price"] for i in po_export.load_item_master(fpath)}
    cogs_total = 0.0
    for item in items:
        cost = cost_by_code.get(item["mat_code"])
        if cost is None:
            raise ValueError(f"No cost found for {item['mat_code']} in Item Master — can't post COGS.")
        cogs_total += float(item["qty_shipped"]) * float(cost)
    cogs_total = round(cogs_total, 2)

    lines = [{"account_code": "5000", "debit": cogs_total, "credit": 0,
             "description": f"COGS for {fulfillment_id}"},
             {"account_code": "1200", "debit": 0, "credit": cogs_total,
             "description": f"Inventory clearing for {fulfillment_id}"}]

    return post_journal_entry("Fulfillment", fulfillment_id,
                              f"Fulfillment {fulfillment_id} — {f['customer_name']}", lines, fpath)


def post_gr_entry(gr_id, data_file=None):
    """
    Dr Inventory Clearing / Cr GR/IR Clearing — ex-GST, both sides. This
    is stage 1 of the proper 3-way match (GR / Invoice Receipt /
    Payment), not the final liability — see vendor_invoices.create_invoice()
    for stage 2 (where GR/IR Clearing gets debited back out, GST Input
    gets claimed, and Accounts Payable — the real, final vendor
    liability — gets credited for the first time).

    GST is deliberately NOT determined here, even though it was in an
    earlier version of this function. Real practice (SAP's GR/IR
    process, for instance) determines tax at invoice verification, not
    goods receipt — a GR is a warehouse/quantity event, often happening
    before the vendor's tax invoice with its GST breakdown even arrives;
    forcing a GST split at GR time was getting ahead of information
    that, in the real process this models, doesn't exist yet at this
    stage. Moving it to the invoice step isn't just more textbook-
    correct, it's also what gives that step a real accounting entry —
    previously "Simulate Invoice" posted nothing at all, since this
    function had already claimed the entire liability (including GST)
    the moment the GR was created.
    """
    fpath = data_file or DATA_FILE
    if already_posted("GR", gr_id, fpath):
        raise ValueError(f"{gr_id} has already been posted to the ledger.")
    g = gr.get_gr(gr_id, fpath)
    if g is None:
        raise ValueError(f"{gr_id} not found.")
    items = [i for i in gr.get_gr_items(gr_id, fpath) if (i["qty_received"] or 0) > 0]
    if not items:
        raise ValueError(f"{gr_id} has no received quantity on any line — nothing to post.")

    inventory_total = 0.0
    unpriced = []
    for item in items:
        unit_price = float(item["unit_price"] or 0)
        if unit_price <= 0:
            unpriced.append(item["mat_code"])
        inventory_total += round(float(item["qty_received"]) * unit_price, 2)
    inventory_total = round(inventory_total, 2)

    if inventory_total <= 0:
        # A real, identifiable gap, not a generic bug: every line on this GR
        # has no unit price on file (PO_Items.Unit_Price is blank/zero — most
        # commonly a Direct PO Entry line for an item with no price in Item
        # Master, or a manually-typed PO). The physical receipt is still
        # recorded either way (create_gr() already succeeded by the time
        # this runs) — this only blocks the accounting entry, which
        # genuinely has nothing to post: Dr 0 / Cr 0 isn't a real journal
        # entry, it's a no-op post_journal_entry() correctly refuses.
        names = ", ".join(unpriced)
        raise ValueError(
            f"Can't post {gr_id} — every received line has no unit price on file "
            f"({names}), so there's nothing of value to record. Set a price on the "
            f"PO line (or the item in Item Master) and try posting again — the goods "
            f"receipt itself is already saved, only the accounting entry is pending.")

    lines = [
        {"account_code": "1200", "debit": inventory_total, "credit": 0,
         "description": f"Inventory received for {gr_id}"},
        {"account_code": "2100", "debit": 0, "credit": inventory_total,
         "description": f"GR/IR clearing for {gr_id} — {g['vendor_name']}, pending invoice"},
    ]
    return post_journal_entry("GR", gr_id, f"Goods Receipt {gr_id} — {g['vendor_name']}", lines, fpath)


# ── Reporting ────────────────────────────────────────────────────────────────
def get_trial_balance(data_file=None):
    accounts = get_accounts(data_file)
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT account_code, SUM(debit) AS total_debit, SUM(credit) AS total_credit "
            "FROM journal_entry_lines GROUP BY account_code"
        ).fetchall()
    finally:
        conn.close()
    totals = {r["account_code"]: {"debit": r["total_debit"] or 0.0, "credit": r["total_credit"] or 0.0}
              for r in rows}

    out = []
    for a in accounts:
        t = totals.get(a["account_code"], {"debit": 0.0, "credit": 0.0})
        debit_normal = a["account_type"] in ("Asset", "Expense")
        balance = round((t["debit"] - t["credit"]) if debit_normal else (t["credit"] - t["debit"]), 2)
        out.append({**a, "total_debit": round(t["debit"], 2), "total_credit": round(t["credit"], 2),
                   "balance": balance})
    return out


# ── Document generation ───────────────────────────────────────────────────────
def generate_journal_voucher(je_id, data_file=None):
    fpath = data_file or DATA_FILE
    je = get_journal_entry(je_id, fpath)
    if je is None:
        raise ValueError(f"{je_id} not found.")
    lines = get_journal_entry_lines(je_id, fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    slate = "1E293B"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Voucher"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "JOURNAL VOUCHER")
    ws.merge_cells("A1:E1")
    label("A3", "JE No:");   value("B3", je["je_id"])
    label("A4", "Date:");    value("B4", je["entry_date"])
    label("D3", "Source:");  value("E3", f"{je['source_type']} {je['source_id']}")
    label("A5", "Narration:"); value("B5", je["description"])

    hdr_row = 8
    hdrs = ["#", "Account Code", "Account Name", "Debit", "Credit", "Description"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=slate)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for i, ln in enumerate(lines, 1):
        vals = [i, ln["account_code"], ln["account_name"], ln["debit"] or "",
                ln["credit"] or "", ln["description"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 1
    ws.cell(r, 3, "Total").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 4, je["total_debit"]).font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 5, je["total_credit"]).font = Font(name="Arial", size=10, bold=True)

    widths = [4, 14, 26, 12, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{je_id}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    entries = get_journal_entries(data_file=data_file)
    by_source = {}
    for je in entries:
        by_source[je["source_type"]] = by_source.get(je["source_type"], 0) + 1
    return {"total": len(entries), "by_source": by_source}


if __name__ == "__main__":
    print("Chart of Accounts:")
    for a in get_accounts():
        print(" ", a["account_code"], a["account_name"], f"({a['account_type']})")
    print()
    print("Accounting stats:", stats())
