"""
migrate_s2c_remainder.py — one-time port of the remaining S2C tables
(Contracts, Contract_Items, VRQ_Requests, VRQ_Responses, RFx_Quotes,
RFx_Invitations) from data.xlsx into erp_pilot.db.

Idempotent: clears each table before inserting, so re-running against a
refreshed data.xlsx doesn't duplicate rows.

All six sheets use the "no title row" convention (header row 1, data
row 2+) — same as every other pilot-created sheet (Inventory_Transactions,
GR_Header, RFx sheets already migrated, etc.), unlike the original S2C
core sheets (PR_Header/PO_Header/RFP) which have a title row.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_contracts(wb, conn):
    if "Contracts" not in wb.sheetnames:
        return 0, 0
    ws = wb["Contracts"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[hdrs.get("Contract_ID", 0)] or "").strip()
        if not cid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((
            cid, g("Vendor_ID"), g("Vendor_Name"), g("Status"), g("Start_Date"),
            g("End_Date"), g("Payment_Terms"), g("Delivery_SLA_Days"), g("Currency"),
            g("Created_Date"), g("Source_PO"), g("Auto_Renew"), g("Notes"),
        ))
    conn.execute("DELETE FROM contracts")
    conn.executemany(
        "INSERT INTO contracts (contract_id, vendor_id, vendor_name, status, start_date, "
        "end_date, payment_terms, delivery_sla_days, currency, created_date, source_po, "
        "auto_renew, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    n_contracts = len(rows)

    n_items = 0
    if "Contract_Items" in wb.sheetnames:
        ws2 = wb["Contract_Items"]
        hdrs2 = _headers(ws2)
        item_rows = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            cid = str(row[hdrs2.get("Contract_ID", 0)] or "").strip()
            if not cid:
                continue
            def g2(col):
                i = hdrs2.get(col)
                return row[i] if i is not None else None
            item_rows.append((
                cid, g2("Line_Item"), g2("Material_Code"), g2("Material_Desc"),
                g2("UOM"), g2("Contracted_Unit_Price"), g2("Min_Order_Qty"),
                g2("Lead_Time_Days"),
            ))
        conn.execute("DELETE FROM contract_items")
        conn.executemany(
            "INSERT INTO contract_items (contract_id, line_item, material_code, "
            "material_desc, uom, contracted_unit_price, min_order_qty, lead_time_days) "
            "VALUES (?,?,?,?,?,?,?,?)",
            item_rows,
        )
        n_items = len(item_rows)
    return n_contracts, n_items


def migrate_vrq(wb, conn):
    n_reqs = 0
    if "VRQ_Requests" in wb.sheetnames:
        ws = wb["VRQ_Requests"]
        hdrs = _headers(ws)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vid = str(row[hdrs.get("VRQ_ID", 0)] or "").strip()
            if not vid:
                continue
            def g(col):
                i = hdrs.get(col)
                return row[i] if i is not None else None
            rows.append((
                vid, g("Vendor_Name"), g("Contact_Email"), g("Sent_Date"),
                g("Status"), g("Filename"), g("Vendor_ID"),
            ))
        conn.execute("DELETE FROM vrq_requests")
        conn.executemany(
            "INSERT INTO vrq_requests (vrq_id, vendor_name, contact_email, sent_date, "
            "status, filename, vendor_id) VALUES (?,?,?,?,?,?,?)",
            rows,
        )
        n_reqs = len(rows)

    n_resp = 0
    if "VRQ_Responses" in wb.sheetnames:
        ws2 = wb["VRQ_Responses"]
        hdrs2 = _headers(ws2)
        rows2 = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            vid = str(row[hdrs2.get("VRQ_ID", 0)] or "").strip()
            key = str(row[hdrs2.get("Question_Key", 1)] or "").strip()
            if not vid or not key:
                continue
            def g2(col):
                i = hdrs2.get(col)
                return row[i] if i is not None else None
            rows2.append((vid, key, g2("Section"), g2("Question"), g2("Answer")))
        conn.execute("DELETE FROM vrq_responses")
        conn.executemany(
            "INSERT INTO vrq_responses (vrq_id, question_key, section, question, answer) "
            "VALUES (?,?,?,?,?)",
            rows2,
        )
        n_resp = len(rows2)
    return n_reqs, n_resp


def migrate_rfx(wb, conn):
    n_quotes = 0
    if "RFx_Quotes" in wb.sheetnames:
        ws = wb["RFx_Quotes"]
        hdrs = _headers(ws)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            qid = str(row[hdrs.get("Quote_ID", 0)] or "").strip()
            if not qid:
                continue
            def g(col):
                i = hdrs.get(col)
                return row[i] if i is not None else None
            rows.append((
                qid, g("RFP_Number"), g("Vendor_ID"), g("Vendor_Name"), g("Quoted_Price"),
                g("Lead_Time_Days"), g("MOQ"), g("Quote_Date"), g("Status"), g("Notes"),
            ))
        conn.execute("DELETE FROM rfx_quotes")
        conn.executemany(
            "INSERT INTO rfx_quotes (quote_id, rfp_number, vendor_id, vendor_name, "
            "quoted_price, lead_time_days, moq, quote_date, status, notes) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        n_quotes = len(rows)

    n_invites = 0
    if "RFx_Invitations" in wb.sheetnames:
        ws2 = wb["RFx_Invitations"]
        hdrs2 = _headers(ws2)
        rows2 = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            iid = str(row[hdrs2.get("Invitation_ID", 0)] or "").strip()
            if not iid:
                continue
            def g2(col):
                i = hdrs2.get(col)
                return row[i] if i is not None else None
            rows2.append((
                iid, g2("RFP_Number"), g2("Vendor_ID"), g2("Vendor_Name"),
                g2("Invited_Date"), g2("Filename"),
            ))
        conn.execute("DELETE FROM rfx_invitations")
        conn.executemany(
            "INSERT INTO rfx_invitations (invitation_id, rfp_number, vendor_id, vendor_name, "
            "invited_date, filename) VALUES (?,?,?,?,?,?)",
            rows2,
        )
        n_invites = len(rows2)
    return n_quotes, n_invites


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_contracts, n_citems = migrate_contracts(wb, conn)
        n_vrq_req, n_vrq_resp = migrate_vrq(wb, conn)
        n_quotes, n_invites = migrate_rfx(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"contracts": n_contracts, "contract_items": n_citems,
            "vrq_requests": n_vrq_req, "vrq_responses": n_vrq_resp,
            "rfx_quotes": n_quotes, "rfx_invitations": n_invites}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate S2C remainder tables from Excel to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to source data.xlsx")
    parser.add_argument("--db", default=None, help="Path to target SQLite db (default: erp_pilot.db)")
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
