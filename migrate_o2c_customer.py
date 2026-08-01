"""
migrate_o2c_customer.py — one-time port of Customer_Master,
Customer_Documents, and Org_Profile from data.xlsx into erp_pilot.db.

First O2C slice — foundational, dependency-free tables (everything else
in O2C depends on Customer_Master downstream, nothing depends on it
upstream). Idempotent: clears each table before inserting.

All three sheets use the "no title row" convention (header row 1, data
row 2+), matching every other pilot-created sheet.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_customers(wb, conn):
    if "Customer_Master" not in wb.sheetnames:
        return 0
    ws = wb["Customer_Master"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[hdrs.get("Customer_ID", 0)] or "").strip()
        if not cid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((
            cid, g("Customer_Name"), g("Customer_Type"), g("Geolocation"), g("City"),
            g("Country"), g("Address"), g("Contact_Name"), g("Contact_Email"),
            g("Contact_Phone"), g("GSTIN"), g("PAN"), g("Credit_Limit"), g("Credit_Status"),
            g("Payment_Terms"), g("Onboarding_Status"), g("KYC_Flag"), g("Onboarded_Date"),
            g("Active"),
        ))
    conn.execute("DELETE FROM customer_master")
    conn.executemany(
        "INSERT INTO customer_master (customer_id, customer_name, customer_type, geolocation, "
        "city, country, address, contact_name, contact_email, contact_phone, gstin, pan, "
        "credit_limit, credit_status, payment_terms, onboarding_status, kyc_flag, "
        "onboarded_date, active) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_customer_documents(wb, conn):
    if "Customer_Documents" not in wb.sheetnames:
        return 0
    ws = wb["Customer_Documents"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        did = str(row[hdrs.get("Document_ID", 0)] or "").strip()
        if not did:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((did, g("Customer_ID"), g("Doc_Type"), g("Filename"),
                    g("Uploaded_Date"), g("Status"), g("Notes")))
    conn.execute("DELETE FROM customer_documents")
    conn.executemany(
        "INSERT INTO customer_documents (document_id, customer_id, doc_type, filename, "
        "uploaded_date, status, notes) VALUES (?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_org_profile(wb, conn):
    if "Org_Profile" not in wb.sheetnames:
        return 0
    ws = wb["Org_Profile"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        oid = str(row[hdrs.get("Org_ID", 0)] or "").strip()
        if not oid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((oid, g("Legal_Name"), g("GSTIN"), g("PAN"), g("Address"), g("City"),
                    g("State"), g("Country"), g("Bank_Account_No"), g("IFSC"), g("Bank_Name"),
                    g("Contact_Email"), g("Contact_Phone")))
    conn.execute("DELETE FROM org_profile")
    conn.executemany(
        "INSERT INTO org_profile (org_id, legal_name, gstin, pan, address, city, state, "
        "country, bank_account_no, ifsc, bank_name, contact_email, contact_phone) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_cust = migrate_customers(wb, conn)
        n_docs = migrate_customer_documents(wb, conn)
        n_org = migrate_org_profile(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"customer_master": n_cust, "customer_documents": n_docs, "org_profile": n_org}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Customer_Master/Customer_Documents/Org_Profile to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--db", default=None)
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
