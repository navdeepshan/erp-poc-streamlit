"""
fulfillment.py — Fulfillment (O2C): pick, ship, deliver.

Scope note, stated up front rather than discovered the hard way: Item
Master has an "In Stock" column, but like Category/Sub-Category before
categorization.py ran, it's never actually been populated — checked
before building this, not assumed. Building a real available-to-promise
check would mean fabricating stock data that doesn't exist, so this
module doesn't pretend to. What it does instead, and what's genuinely
useful without inventory data: real partial-shipment tracking. A human
records what was actually picked and shipped per line — Qty_Shipped is
independent of Qty_Ordered — which is how a warehouse actually discovers
a shortage (at the shelf, not from a database that was never fed).

A proper inventory ledger needs both halves — receiving (procurement side
incrementing stock) and fulfillment (this side decrementing it) — and
neither exists yet. Building only the decrement half would show a stock
number that only ever falls and was never seeded correctly; better to
wait and build both together than ship something quietly misleading.

Lifecycle: Pending -> Picking -> Shipped -> Delivered, or -> Cancelled.
is_ready_for_billing() is the hook a future Billing module calls — same
dependency-injection shape used throughout (contracts.py on
pr_consolidation.py, sales_order.py on customer_onboarding.py).

New sheets:
  Fulfillments       Fulfillment_ID | SO_ID | Customer_ID | Customer_Name |
                      Status | Created_Date | Shipped_Date | Delivered_Date |
                      Carrier | Tracking_Ref | Delivery_Location |
                      POD_Reference | Notes
  Fulfillment_Items  Fulfillment_ID | Line_Item | Material_Code |
                      Material_Desc | UOM | Qty_Ordered | Qty_Shipped

SQLite pilot: Fulfillments and Fulfillment_Items now live in
erp_pilot.db (tables `fulfillments`, `fulfillment_items`) — this
module is the exclusive owner of both.
"""

import os, io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import pr_consolidation as pc
import sales_order as so

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

FUL_SHEET = "Fulfillments"
FUL_ITEMS_SHEET = "Fulfillment_Items"

FUL_COLS = ["Fulfillment_ID", "SO_ID", "Customer_ID", "Customer_Name", "Status",
           "Created_Date", "Shipped_Date", "Delivered_Date", "Carrier",
           "Tracking_Ref", "Delivery_Location", "POD_Reference", "Notes"]
FUL_ITEM_COLS = ["Fulfillment_ID", "Line_Item", "Material_Code", "Material_Desc",
                "UOM", "Qty_Ordered", "Qty_Shipped"]


# ── Sheet bootstrap ────────────────────────────────────────────────────────────
def ensure_sheets(wb=None):
    """Kept for signature compatibility — Fulfillments/Fulfillment_Items
    no longer live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


# ── Read ──────────────────────────────────────────────────────────────────────
def get_fulfillments(status=None, so_id=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM fulfillments ORDER BY fulfillment_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"fulfillment_id": r["fulfillment_id"], "so_id": r["so_id"],
               "customer_id": r["customer_id"], "customer_name": r["customer_name"],
               "status": r["status"], "created_date": r["created_date"],
               "shipped_date": r["shipped_date"], "delivered_date": r["delivered_date"],
               "carrier": r["carrier"], "tracking_ref": r["tracking_ref"],
               "delivery_location": r["delivery_location"], "pod_reference": r["pod_reference"],
               "notes": r["notes"]}
        if status and row["status"] != status:
            continue
        if so_id and row["so_id"] != so_id:
            continue
        out.append(row)
    return out


def get_fulfillment(fulfillment_id, data_file=None):
    for f in get_fulfillments(data_file=data_file):
        if f["fulfillment_id"] == fulfillment_id:
            return f
    return None


def get_fulfillment_items(fulfillment_id, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM fulfillment_items WHERE fulfillment_id = ? ORDER BY line_item",
            (fulfillment_id,),
        ).fetchall()
    finally:
        conn.close()
    return [{"fulfillment_id": r["fulfillment_id"], "line_item": r["line_item"],
             "mat_code": r["material_code"], "mat_desc": r["material_desc"], "uom": r["uom"],
             "qty_ordered": r["qty_ordered"], "qty_shipped": r["qty_shipped"]} for r in rows]


def so_already_fulfilled(so_id, data_file=None):
    """True if an active (non-cancelled) fulfillment already exists for this SO."""
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT 1 FROM fulfillments WHERE so_id = ? AND status != 'Cancelled' LIMIT 1",
            (so_id,),
        ).fetchone()
    finally:
        conn.close()
    return row is not None


# ── Create ────────────────────────────────────────────────────────────────────
def _next_fulfillment_id(conn):
    rows = conn.execute(
        "SELECT fulfillment_id FROM fulfillments WHERE fulfillment_id LIKE 'FUL-%'"
    ).fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["fulfillment_id"].split("-")[1]))
        except Exception: pass
    return f"FUL-{mx+1:05d}"


def create_fulfillment(so_id, data_file=None):
    fpath = data_file or DATA_FILE
    if not so.is_ready_for_fulfillment(so_id, fpath):
        raise ValueError(f"{so_id} is not Confirmed — only a confirmed order can be fulfilled.")
    if so_already_fulfilled(so_id, fpath):
        raise ValueError(f"{so_id} already has an active fulfillment record.")
    order = so.get_order(so_id, fpath)
    items = so.get_order_items(so_id, fpath)

    db.init_schema()
    conn = db.get_connection()
    try:
        fid = _next_fulfillment_id(conn)
        conn.execute(
            "INSERT INTO fulfillments (fulfillment_id, so_id, customer_id, customer_name, "
            "status, created_date, shipped_date, delivered_date, carrier, tracking_ref, "
            "delivery_location, pod_reference, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (fid, so_id, order["customer_id"], order["customer_name"], "Pending",
             date.today().strftime("%Y-%m-%d"), "", "", "", "",
             order["delivery_location"] or "", "", ""),
        )
        for item in items:
            conn.execute(
                "INSERT INTO fulfillment_items (fulfillment_id, line_item, material_code, "
                "material_desc, uom, qty_ordered, qty_shipped) VALUES (?,?,?,?,?,?,?)",
                (fid, item["line_item"], item["mat_code"], item["mat_desc"],
                 item["uom"], item["qty"], 0),
            )
        conn.commit()
    finally:
        conn.close()
    return fid


# ── Lifecycle ─────────────────────────────────────────────────────────────────
_FIELD_TO_COL = {  # matches the Capitalized keys _set_field's callers already use
    "Status": "status", "Shipped_Date": "shipped_date", "Carrier": "carrier",
    "Tracking_Ref": "tracking_ref", "Delivered_Date": "delivered_date",
    "POD_Reference": "pod_reference", "Notes": "notes",
}


def start_picking(fulfillment_id, data_file=None):
    f = get_fulfillment(fulfillment_id, data_file)
    if f is None:
        raise ValueError(f"{fulfillment_id} not found.")
    if f["status"] != "Pending":
        raise ValueError(f"{fulfillment_id} is '{f['status']}' — only Pending can start picking.")
    _set_field(fulfillment_id, {"Status": "Picking"}, data_file)


def record_shipment(fulfillment_id, line_qty_shipped, carrier="", tracking_ref="", data_file=None):
    """
    line_qty_shipped: {mat_code: qty_shipped}. Any line not in this dict
    keeps its previous Qty_Shipped (defaults to 0). Allows genuine partial
    shipment — a line shipped less than ordered just shows that plainly,
    no inventory system required to know it happened.

    Posts a real Goods Issue against inventory.py's own ledger for each
    shipped line, at this fulfillment's own delivery_location (a real,
    tracked location in both pilots — the Plant or regional office stock
    genuinely ships from, not a raw customer address) — a real, found
    gap until now: outbound fulfillment shipments never touched the
    inventory ledger at all, so stock that had genuinely left the
    building and been delivered to a real customer still showed as
    on-hand. Same "record what happened, don't gatekeep it" principle
    already used throughout inventory.py (GR allows over-receipt, this
    allows a resulting negative balance too) — the shipment already
    physically happened by the time this runs; refusing to record it
    wouldn't undo that, it would just make the ledger wrong in a
    different way. A real warning is raised (not silently absorbed)
    when a line's own post-shipment balance goes negative, the same
    honest-signal spirit as every other module here.
    """
    fpath = data_file or DATA_FILE
    f = get_fulfillment(fulfillment_id, fpath)
    if f is None:
        raise ValueError(f"{fulfillment_id} not found.")
    if f["status"] not in ("Pending", "Picking"):
        raise ValueError(f"{fulfillment_id} is '{f['status']}' — can't ship from this state.")

    import inventory as inv
    items_by_mat = {i["mat_code"]: i for i in get_fulfillment_items(fulfillment_id, fpath)}
    goods_issue_warnings = []

    conn = db.get_connection()
    try:
        for mat, qty in line_qty_shipped.items():
            conn.execute(
                "UPDATE fulfillment_items SET qty_shipped = ? "
                "WHERE fulfillment_id = ? AND material_code = ?",
                (qty, fulfillment_id, mat),
            )
        conn.execute(
            "UPDATE fulfillments SET status='Shipped', shipped_date=?, carrier=?, "
            "tracking_ref=? WHERE fulfillment_id=?",
            (date.today().strftime("%Y-%m-%d"), carrier, tracking_ref, fulfillment_id),
        )
        conn.commit()
    finally:
        conn.close()

    for mat, qty in line_qty_shipped.items():
        if qty <= 0:
            continue
        item = items_by_mat.get(mat)
        mat_desc = item["mat_desc"] if item else mat
        inv.record_transaction(mat, mat_desc, f["delivery_location"], -qty, "Goods Issue",
                               reference_type="Fulfillment", reference_id=fulfillment_id,
                               notes=f"Shipped against {f['so_id']}", data_file=data_file)
        new_balance = inv.get_balance(mat, f["delivery_location"], data_file)
        if new_balance < -0.005:
            goods_issue_warnings.append(
                f"{mat} at {f['delivery_location']} is now {new_balance:g} — this shipment "
                f"posted against stock the ledger didn't actually show as available.")

    return {"warnings": goods_issue_warnings}


def record_delivery(fulfillment_id, pod_reference="", data_file=None):
    f = get_fulfillment(fulfillment_id, data_file)
    if f is None:
        raise ValueError(f"{fulfillment_id} not found.")
    if f["status"] != "Shipped":
        raise ValueError(f"{fulfillment_id} is '{f['status']}' — only Shipped can be marked Delivered.")
    _set_field(fulfillment_id, {"Status": "Delivered",
        "Delivered_Date": date.today().strftime("%Y-%m-%d"), "POD_Reference": pod_reference}, data_file)


def cancel_fulfillment(fulfillment_id, reason="", data_file=None):
    """
    A Shipped fulfillment being cancelled needs a real reversal, not
    just a status flip — its own Goods Issue already reduced a real
    location's balance in inventory.py's own ledger; leaving that
    unreversed would mean cancelling a shipment permanently loses that
    stock from the ledger, which never physically happened. A Pending
    or Picking fulfillment never posted a Goods Issue in the first
    place (nothing shipped yet), so nothing needs reversing there.
    """
    f = get_fulfillment(fulfillment_id, data_file)
    if f and f["status"] == "Delivered":
        raise ValueError(f"{fulfillment_id} is already Delivered — can't cancel a completed fulfillment.")

    if f and f["status"] == "Shipped":
        import inventory as inv
        for item in get_fulfillment_items(fulfillment_id, data_file):
            qty = item["qty_shipped"] or 0
            if qty > 0:
                inv.record_transaction(item["mat_code"], item["mat_desc"], f["delivery_location"],
                                       qty, "Goods Issue Reversal", reference_type="Fulfillment",
                                       reference_id=fulfillment_id,
                                       notes=f"Cancelled: {reason}" if reason else "Cancelled",
                                       data_file=data_file)

    _set_field(fulfillment_id, {"Status": "Cancelled",
        "Notes": f"{(f or {}).get('notes','') or ''} | Cancelled: {reason}".strip(" |")}, data_file)


def _set_field(fulfillment_id, field_values, data_file=None):
    conn = db.get_connection()
    try:
        for k, v in field_values.items():
            col = _FIELD_TO_COL[k]
            conn.execute(
                f"UPDATE fulfillments SET {col} = ? WHERE fulfillment_id = ?",
                (v, fulfillment_id),
            )
        conn.commit()
    finally:
        conn.close()


def is_ready_for_billing(fulfillment_id, data_file=None):
    """The hook a future Billing module calls."""
    f = get_fulfillment(fulfillment_id, data_file)
    return f is not None and f["status"] == "Delivered"


# ── Document generation ───────────────────────────────────────────────────────
def generate_delivery_note(fulfillment_id, data_file=None):
    fpath = data_file or DATA_FILE
    f = get_fulfillment(fulfillment_id, fpath)
    if f is None:
        raise ValueError(f"{fulfillment_id} not found.")
    items = get_fulfillment_items(fulfillment_id, fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = "14532D"; amber = "FFF8E8"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delivery Note"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "DELIVERY NOTE")
    ws.merge_cells("A1:G1")

    label("A3", "Fulfillment ID:"); value("B3", f["fulfillment_id"])
    label("A4", "Order:");          value("B4", f["so_id"])
    label("A5", "Status:");         value("B5", f["status"])
    label("D3", "Customer:");       value("E3", f["customer_name"])
    deliv_name = f["delivery_location"] or ""
    for d in pc.get_delivery_locations(active_only=False):
        if d["id"] == deliv_name:
            deliv_name = d["name"]
            break
    label("D4", "Deliver To:");     value("E4", deliv_name)
    label("D5", "Carrier:");        value("E5", f"{f['carrier'] or 'TBD'} {f['tracking_ref'] or ''}".strip())

    hdr_row = 8
    hdrs = ["#", "Material Code", "Description", "UOM", "Qty Ordered", "Qty Shipped", "Backorder"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=green)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for i, item in enumerate(items, 1):
        backorder = max(0, (item["qty_ordered"] or 0) - (item["qty_shipped"] or 0))
        vals = [i, item["mat_code"], item["mat_desc"], item["uom"],
                item["qty_ordered"], item["qty_shipped"], backorder or ""]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color=("B91C1C" if ci == 7 and backorder else "1A1A2E"))
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if ci == 7 and backorder:
                c.fill = PatternFill("solid", fgColor=amber)
        r += 1

    r += 3
    ws.cell(r, 1, "Received in good condition by:").font = Font(name="Arial", size=10, bold=True)
    r += 3
    ws.cell(r, 1, "Signature: ____________________").font = Font(name="Arial", size=9)
    r += 1
    ws.cell(r, 1, "Name / Date: ____________________").font = Font(name="Arial", size=9)

    widths = [4, 15, 32, 8, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 24

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{f['fulfillment_id']}_{f['so_id']}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    fulfillments = get_fulfillments(data_file=data_file)
    by_status = {}
    for f in fulfillments:
        by_status[f["status"]] = by_status.get(f["status"], 0) + 1
    return {"total": len(fulfillments), "by_status": by_status}


if __name__ == "__main__":
    print("Fulfillment stats:", stats())
