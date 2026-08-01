"""
migrate_bom_to_sqlite.py — one-time port of BOM_Items from data.xlsx
into erp_pilot.db.

Idempotent: clears bom_items before inserting, so re-running against a
refreshed data.xlsx doesn't duplicate rows or require manual cleanup.

This is a data port only — it does not touch bom.py's logic. Run this
once before bom.py's functions (now backed by SQLite) can return
anything.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")
BOM_SHEET = "BOM_Items"


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    if BOM_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{BOM_SHEET} sheet not found in {xlsx_path}")

    ws = wb[BOM_SHEET]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        parent_code, parent_desc, component_code, component_desc, qty_per, uom, notes = row
        rows.append((parent_code, parent_desc, component_code, component_desc,
                      qty_per, uom, notes))
    wb.close()

    conn = db.get_connection(db_path)
    try:
        conn.execute("DELETE FROM bom_items")
        conn.executemany(
            "INSERT INTO bom_items "
            "(parent_code, parent_desc, component_code, component_desc, qty_per, uom, notes) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            rows,
        )
        conn.commit()
        count = conn.execute("SELECT COUNT(*) FROM bom_items").fetchone()[0]
    finally:
        conn.close()

    return count


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate BOM_Items from Excel to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to source data.xlsx")
    parser.add_argument("--db", default=None, help="Path to target SQLite db (default: erp_pilot.db)")
    args = parser.parse_args()

    n = migrate(args.xlsx, args.db)
    print(f"Migrated {n} BOM_Items row(s) from {args.xlsx} into SQLite.")
