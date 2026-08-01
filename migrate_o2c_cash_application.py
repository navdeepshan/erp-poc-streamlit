"""
migrate_o2c_cash_application.py — one-time port of Payments and
Payment_Applications from data.xlsx into erp_pilot.db.

Final O2C slice (and the final slice of the entire pilot) — depends on
Invoices, Customer_Master, Chart_of_Accounts already being migrated.
Idempotent.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_payments(wb, conn):
    if "Payments" not in wb.sheetnames:
        return 0
    ws = wb["Payments"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = str(row[hdrs.get("Payment_ID", 0)] or "").strip()
        if not pid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((pid, g("Customer_ID"), g("Customer_Name"), g("Payment_Date"), g("Amount"),
                    g("Payment_Method"), g("Reference_No"), g("Unapplied_Amount"), g("Notes")))
    conn.execute("DELETE FROM payments")
    conn.executemany(
        "INSERT INTO payments (payment_id, customer_id, customer_name, payment_date, amount, "
        "payment_method, reference_no, unapplied_amount, notes) VALUES (?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_payment_applications(wb, conn):
    if "Payment_Applications" not in wb.sheetnames:
        return 0
    ws = wb["Payment_Applications"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = str(row[hdrs.get("Payment_ID", 0)] or "").strip()
        if not pid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((pid, g("Line_Item"), g("Invoice_ID"), g("Applied_Amount"),
                    g("Short_Payment_Reason"), g("Application_Date")))
    conn.execute("DELETE FROM payment_applications")
    conn.executemany(
        "INSERT INTO payment_applications (payment_id, line_item, invoice_id, applied_amount, "
        "short_payment_reason, application_date) VALUES (?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_p = migrate_payments(wb, conn)
        n_pa = migrate_payment_applications(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"payments": n_p, "payment_applications": n_pa}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Payments/Payment_Applications to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--db", default=None)
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
