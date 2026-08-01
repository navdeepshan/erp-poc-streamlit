"""
po_export.py
============
Core export logic used by erp_ui.py.
No UI dependencies — pure Python + openpyxl.
Place in the SAME FOLDER as erp_ui.py and data.xlsx.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
import os, re, unicodedata, io
from datetime import datetime
from typing import List, Dict, Tuple
from copy import copy

import db

# ── AV export — uses BytesIO, no filesystem write ─────────────────────────────
def make_av_bytes(po_number, supplier_id, po_type, legal_entity,
                  purch_entity, purch_group, currency, plant_code, po_lines):
    """Returns (filename, bytes). po_lines need mat_code/code, mat_desc/desc,
    uom, qty. Moved here from erp_ui.py — this is a pure document generator
    with no Streamlit dependency, and needed to be importable from non-UI
    code (demo_scenario.py, any future script) without triggering an
    entire Streamlit page's top-level code to execute as an import
    side effect, which is what importing it out of erp_ui.py directly
    used to do."""
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{po_number}_AV_{ts}.xlsx"
    wb       = openpyxl.Workbook()
    ws       = wb.active
    ws.title = "PO_Upload"
    hdrs = ["PO_Number","PO_Type","Legal_Entity","Purchase_Entity","Purchasing_Group",
            "Currency","Plant_code","Supplier_ID","Supplier_Name","Material_Code",
            "Material_Desc","UOM","Quantity","Delivery_Date","Delivery_Location"]
    hf = Font(name="Arial", bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor="1E3A5F")
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(1, ci, h); c.font = hf; c.fill = hb
    for ri, ln in enumerate(po_lines, 2):
        ws.cell(ri,  1, po_number);    ws.cell(ri,  2, po_type)
        ws.cell(ri,  3, legal_entity); ws.cell(ri,  4, purch_entity)
        ws.cell(ri,  5, purch_group);  ws.cell(ri,  6, currency)
        ws.cell(ri,  7, plant_code);   ws.cell(ri,  8, supplier_id)
        ws.cell(ri,  9, "")
        ws.cell(ri, 10, ln.get("mat_code") or ln.get("code") or "")
        ws.cell(ri, 11, ln.get("mat_desc") or ln.get("desc") or "")
        ws.cell(ri, 12, ln.get("uom",""))
        ws.cell(ri, 13, ln.get("qty", 0))
        ws.cell(ri, 14, ln.get("deliv_date",""))
        ws.cell(ri, 15, ln.get("deliv_loc",""))
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 50)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return filename, buf.read()

# ── Organisation constants — edit these ──────────────────────────────────────
PO_TYPE           = "NB"
LEGAL_ENTITY      = "LE-001"
PURCHASE_ENTITY   = "PE-001"
PURCHASING_GROUP  = "PG-001"
CURRENCY          = "INR"
PLANT_CODE        = "PLANT-01"
DELIVERY_LOCATION = "WH-MAIN"

VENDOR_SURG_0074 = "VND-SURG-0074"
VENDOR_DENT_0023 = "VND-DENT-0023"

ITEM_MASTER_SHEET = "Item Master"


def _normalize(text: str) -> str:
    text = str(text).replace("\u00ae", "").replace("\u2122", "")
    text = "".join(c for c in text if unicodedata.category(c) != "So")
    return re.sub(r"\s+", " ", text).strip().upper()


# ══════════════════════════════════════════════════════════════════════════════
# Item Master
# ══════════════════════════════════════════════════════════════════════════════
def load_item_master(data_path=None, active_only=True):
    """
    Item Master now lives in SQLite (erp_pilot.db, table `item_master`)
    — this is the canonical reader every other module should delegate
    to rather than reading the table directly, closing a real
    inconsistency that existed before this migration: some readers
    filtered on Active (this function did), others didn't
    (categorization.py, item_tax.py, rfx.py's own Item Master reads).

    active_only=True (the default, matching every existing caller's
    prior behavior — item pickers, BOM explosion, cost lookups) returns
    only Active='Yes' items. active_only=False returns everything,
    for callers whose job is completeness/coverage over the WHOLE
    catalog regardless of active status (categorization.py, item_tax.py)
    — Active gates what's newly selectable, not whether an item's
    master data is allowed to be complete.

    `data_path` is accepted for signature compatibility (every existing
    caller passes an Excel path here) and is unused — see db.py: one
    live SQLite database, same as one live data.xlsx.
    """
    conn = db.get_connection()
    try:
        if active_only:
            rows = conn.execute(
                "SELECT * FROM item_master WHERE active = 'Yes' ORDER BY item_code"
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM item_master ORDER BY item_code").fetchall()
    finally:
        conn.close()
    return [{
        "code": r["item_code"], "desc": r["item_desc"] or "",
        "category": r["category"] or "", "subcategory": r["subcategory"] or "",
        "uom": r["uom"] or "", "price": float(r["unit_price"]) if r["unit_price"] else 0.0,
        "stock": r["in_stock"] or 0, "tags": r["tags"] or "",
        "lead_time": r["lead_time_days"], "active": r["active"] or "",
        "hsn_code": r["hsn_code"], "gst_rate": r["gst_rate"],
        "weight_kg": float(r["weight_kg"]) if r["weight_kg"] is not None else None,
    } for r in rows]


def get_item_by_code(material_code, active_only=False):
    """Single-item point lookup by code — for callers that already know
    the exact code (e.g. an RFx line's material, a PO line) and need
    its master data regardless of current Active status. active_only
    defaults to False here (unlike load_item_master's listing default)
    precisely because a point lookup for an already-referenced code
    should still resolve even if the item was deactivated after the
    reference was created."""
    conn = db.get_connection()
    try:
        r = conn.execute(
            "SELECT * FROM item_master WHERE item_code = ?" +
            (" AND active = 'Yes'" if active_only else ""),
            (material_code,),
        ).fetchone()
    finally:
        conn.close()
    if not r:
        return None
    return {
        "code": r["item_code"], "desc": r["item_desc"] or "",
        "category": r["category"] or "", "subcategory": r["subcategory"] or "",
        "uom": r["uom"] or "", "price": float(r["unit_price"]) if r["unit_price"] else 0.0,
        "stock": r["in_stock"] or 0, "tags": r["tags"] or "",
        "lead_time": r["lead_time_days"], "active": r["active"] or "",
        "hsn_code": r["hsn_code"], "gst_rate": r["gst_rate"],
        "weight_kg": float(r["weight_kg"]) if r["weight_kg"] is not None else None,
    }


def fuzzy_search(query: str, items: List[Dict], max_results: int = 15) -> List[Dict]:
    """Multi-token AND search across desc + category + subcategory + tags + code."""
    if not query or len(query.strip()) < 2:
        return []
    tokens = query.lower().strip().split()
    results = []
    for item in items:
        hay = " ".join([
            item["desc"].lower(), item["category"].lower(),
            item["subcategory"].lower(), item["tags"].lower(),
            item["code"].lower(),
        ])
        if all(t in hay for t in tokens):
            results.append(item)
        if len(results) >= max_results:
            break
    return results


# ══════════════════════════════════════════════════════════════════════════════
# AV Upload export (Direct PO flow)
# ══════════════════════════════════════════════════════════════════════════════
def send_to_av(po_number, supplier_id, deliv_date, po_lines,
               data_path, output_folder, prefix=""):
    """
    Create the AV Upload file and (if applicable) the vendor-specific file.
    Returns: (av_filename, vendor_filename_or_empty, list_of_messages)
    """
    if not po_lines:
        return "", "", ["No items to export."]

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_prefix = prefix.strip() or po_number or "PO"

    # ── AV Upload file ────────────────────────────────────────────────────────
    av_filename = f"{file_prefix}_Upload_{timestamp}.xlsx"
    av_path     = os.path.join(output_folder, av_filename)

    wb_av = openpyxl.Workbook()
    ws_av = wb_av.active
    ws_av.title = "PO_Upload"

    headers = [
        "PO_Number", "PO_Type", "Legal_Entity", "Purchase_Entity",
        "Purchasing_Group", "Currency", "Plant_code", "Supplier_ID",
        "Supplier_Name", "Material_Code", "Material_Desc",
        "Quantity", "Delivery_Date", "Delivery_Location"
    ]
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2E4057")
    for ci, h in enumerate(headers, 1):
        c = ws_av.cell(1, ci, h)
        c.font = hdr_font; c.fill = hdr_fill

    for ri, line in enumerate(po_lines, 2):
        ws_av.cell(ri,  1, po_number);      ws_av.cell(ri,  2, PO_TYPE)
        ws_av.cell(ri,  3, LEGAL_ENTITY);   ws_av.cell(ri,  4, PURCHASE_ENTITY)
        ws_av.cell(ri,  5, PURCHASING_GROUP);ws_av.cell(ri,  6, CURRENCY)
        ws_av.cell(ri,  7, PLANT_CODE);     ws_av.cell(ri,  8, supplier_id)
        ws_av.cell(ri,  9, "");             ws_av.cell(ri, 10, line["code"])
        ws_av.cell(ri, 11, "");             ws_av.cell(ri, 12, line["qty"])
        ws_av.cell(ri, 13, deliv_date);     ws_av.cell(ri, 14, DELIVERY_LOCATION)

    for col in ws_av.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws_av.column_dimensions[col[0].column_letter].width = min(w + 4, 50)
    wb_av.save(av_path)

    messages = [f"AV file created: {av_filename}"]

    # ── Vendor-specific export ────────────────────────────────────────────────
    vendor_filename, vendor_msgs = _export_vendor_template(
        supplier_id, po_lines, data_path, output_folder, timestamp)
    messages.extend(vendor_msgs)

    return av_filename, vendor_filename, messages


# ══════════════════════════════════════════════════════════════════════════════
# Vendor dispatcher
# ══════════════════════════════════════════════════════════════════════════════
def _export_vendor_template(supplier_id, po_lines, data_path,
                             output_folder, timestamp):
    wb = openpyxl.load_workbook(data_path, data_only=True)
    if supplier_id not in wb.sheetnames:
        wb.close(); return "", []

    if supplier_id == VENDOR_SURG_0074:
        result = _export_surg_0074(wb, po_lines, supplier_id, output_folder, timestamp)
    elif supplier_id == VENDOR_DENT_0023:
        result = _export_dent_0023(wb, po_lines, supplier_id, output_folder, timestamp)
    else:
        wb.close(); return "", []

    wb.close()
    return result


# ══════════════════════════════════════════════════════════════════════════════
# VND-SURG-0074
# ══════════════════════════════════════════════════════════════════════════════
def _export_surg_0074(wb, po_lines, supplier_id, output_folder, timestamp):
    ws    = wb[supplier_id]
    TQCOL = 7; TROW1 = 8

    lookup = {}
    for r in range(TROW1, ws.max_row + 1):
        vcode = str(ws.cell(r, 2).value or "").strip()
        if not vcode: continue
        d4 = str(ws.cell(r, 4).value or "").strip()
        d5 = str(ws.cell(r, 5).value or "").strip()
        lookup[vcode] = {"row": r, "desc": d4 + d5}

    errors, matched = [], {}
    for line in po_lines:
        desc, code = line["desc"], line["code"]
        if "~" not in desc:
            errors.append(f"{code} — no '~' separator"); continue
        sep   = desc.index("~")
        vcode = desc[:sep].strip(); vdesc = desc[sep+1:].strip()
        if vcode not in lookup:
            errors.append(f"{code} (VCode={vcode}) — not in template"); continue
        e = lookup[vcode]
        if _normalize(e["desc"]) != _normalize(vdesc):
            errors.append(f"{code} — desc mismatch"); continue
        matched[e["row"]] = line["qty"]

    if not matched:
        msgs = ["Vendor SURG-0074: no items matched — vendor file not created."]
        if errors: msgs.append("Unmatched:\n" + "\n".join(f"  • {e}" for e in errors))
        return "", msgs
    return _write_vendor_file(wb, supplier_id, matched, TQCOL,
                               output_folder, timestamp, errors)


# ══════════════════════════════════════════════════════════════════════════════
# VND-DENT-0023
# ══════════════════════════════════════════════════════════════════════════════
def _export_dent_0023(wb, po_lines, supplier_id, output_folder, timestamp):
    ws    = wb[supplier_id]
    TQCOL = 6; TROW1 = 6

    # Find grand total row (SUBTOTAL in col F)
    t_last = ws.max_row
    for r in range(TROW1, min(TROW1 + 500, ws.max_row + 1)):
        cell = ws.cell(r, TQCOL)
        if cell.data_type == 'f' and cell.value and \
                "SUBTOTAL" in str(cell.value).upper():
            t_last = r - 1; break

    lookup = {}
    for r in range(TROW1, t_last + 1):
        vcode = str(ws.cell(r, 2).value or "").strip()
        vdesc = str(ws.cell(r, 3).value or "")
        if vcode:
            lookup[vcode] = {"row": r, "desc": _normalize(vdesc)}

    errors, matched = [], {}
    for line in po_lines:
        desc, code = line["desc"].strip(), line["code"]
        if not desc:
            errors.append(f"{code} — empty description"); continue
        pd_ = desc.find("-"); ps = desc.find(" ")
        if pd_ == -1 and ps == -1:
            errors.append(f"{code} — cannot split"); continue
        elif pd_ == -1: sp = ps
        elif ps == -1:  sp = pd_
        else:           sp = min(pd_, ps)
        vcode = desc[:sp].strip(); vdesc = desc[sp+1:].strip()
        if vcode not in lookup:
            errors.append(f"{code} (VCode={vcode}) — not in template"); continue
        e = lookup[vcode]
        if e["desc"] != _normalize(vdesc):
            errors.append(f"{code} — desc mismatch"); continue
        matched[e["row"]] = line["qty"]

    if not matched:
        msgs = ["Vendor DENT-0023: no items matched — vendor file not created."]
        if errors: msgs.append("Unmatched:\n" + "\n".join(f"  • {e}" for e in errors))
        return "", msgs
    return _write_vendor_file(wb, supplier_id, matched, TQCOL,
                               output_folder, timestamp, errors)


# ══════════════════════════════════════════════════════════════════════════════
# Write vendor output file
# ══════════════════════════════════════════════════════════════════════════════
def _write_vendor_file(wb_src, supplier_id, matched, qty_col,
                        output_folder, timestamp, errors):
    ws_src = wb_src[supplier_id]
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = supplier_id

    for col_letter, dim in ws_src.column_dimensions.items():
        ws_out.column_dimensions[col_letter].width = dim.width

    for row in ws_src.iter_rows():
        ws_out.row_dimensions[row[0].row].height = (
            ws_src.row_dimensions[row[0].row].height or 15)
        for cell in row:
            dst = ws_out.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                try:
                    dst.font      = copy(cell.font)
                    dst.fill      = copy(cell.fill)
                    dst.alignment = copy(cell.alignment)
                    dst.number_format = cell.number_format
                except: pass

    for mc in ws_src.merged_cells.ranges:
        try: ws_out.merge_cells(str(mc))
        except: pass

    if matched:
        for r in range(min(matched), max(matched) + 1):
            ws_out.cell(r, qty_col).value = None
    for row_num, qty in matched.items():
        ws_out.cell(row_num, qty_col).value = qty

    vendor_filename = f"{supplier_id}_Upload_{timestamp}.xlsx"
    wb_out.save(os.path.join(output_folder, vendor_filename))

    msgs = [f"Vendor file created: {vendor_filename} ({len(matched)} item(s) matched)"]
    if errors:
        msgs.append("Unmatched:\n" + "\n".join(f"  • {e}" for e in errors))
    return vendor_filename, msgs
