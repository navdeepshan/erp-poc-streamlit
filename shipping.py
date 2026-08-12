"""
shipping.py — real shipping-detail collection and Excel export for
third-party courier handoff.

Built 2026-07-31, deliberately scoped down at direct request: just
BlueDart (not Delhivery yet), just the Excel export (not "Ship All",
not a live API connector) — all explicitly deferred to a later pass,
not overlooked. COURIERS below is a list of one specifically so
extending it later is a one-line change, not a redesign.

Generalized 2026-08-08 (LOG-US-01/FUL-US-05 connection): this module's
booking/tracking/export logic was, until now, reachable only from the
internal stock-transfer screen (INV-US-05) — every function took a real
`inventory.stock_transfers` row shape specifically. The design note
below (left as-is, historical) already anticipated exactly this need:
O2C shipping to an external customer, where the consignee is a real
customer address, not one of this org's own Delivery Locations, and
there's no `stock_transfers` row underneath it at all.

The fix landed as a `kind` registry (_KINDS below), not a rewrite of the
proven internal-transfer path: `build_shipment_details()` (stock
transfer) and the new `build_customer_shipment_details()` (O2C
fulfillment) are now two adapters feeding the SAME
`generate_shipping_excel()` / `submit_batch_to_courier()` /
`get_tracking_status()` / `skip_ahead_tracking()` functions, selected by
a `kind` parameter that defaults to `"stock_transfer"` — every existing
call site (mfg_ui.py's Ship / Position & Transfers screen) keeps working
completely unchanged. Route consolidation, least-cost carrier selection,
weight-slab pricing, and the rate cards themselves are shared as-is;
only the "which table do I read/write, and how do I build a shipment
dict from it" indirection is new.

DESIGN NOTE for real (credentialed) implementation, flagged directly
per request, not yet acted on — a real courier integration needs the
same reuse discipline, but against a real network call: swapping in a
real BlueDart/Delhivery connection later means replacing the single
`requests.post(...)` line noted in submit_to_courier()'s own docstring,
not redesigning the data flow around it. Not built now — noted so it
isn't designed away by accident later.
"""

import re
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import po_export
import pr_consolidation as pc
import org_profile as op

COURIERS = ["BlueDart", "Delhivery", "Cargo Service"]

# Real, simulated rate cards -- not verified against either courier's actual
# current published tariff (no real access to either), but structured the
# way real Indian B2B Surface tariffs actually work: a per-shipment minimum
# charge, and per-kg rates that step down at real weight breakpoints (a
# genuine bulk discount shape, not a flat linear rate that would misrepresent
# a real slab-boundary saving). Deliberately not identical between the two
# couriers -- if one were simply always cheaper, "least-cost" comparison
# would have nothing real to demonstrate.
RATE_CARDS = {
    "BlueDart": {
        "min_charge": 150, "max_weight_kg": 50, "transit_days": 4,
        "slabs": [(5, 80), (20, 65), (50, 50)],  # (upper bound kg, rate per kg)
    },
    "Delhivery": {
        "min_charge": 140, "max_weight_kg": 40, "transit_days": 3,
        "slabs": [(5, 75), (20, 68), (40, 48)],
    },
    # A real, distinct service tier, not a third "courier" competing on the
    # same terms -- a genuine gap found from direct testing: a 65kg autoclave
    # or a 105kg box of pouches would never actually move via a standard
    # Surface parcel service in real life; it goes via Cargo/Freight, which
    # this system didn't model at all until this was hit. Deliberately a
    # higher minimum charge (real freight bookings carry fixed overhead a
    # small parcel doesn't) and a slower transit (freight is genuinely
    # slower than parcel Surface), so this never quietly wins the least-cost
    # comparison for a shipment light enough for BlueDart/Delhivery to
    # already handle -- it only becomes competitive, and structurally
    # necessary, once a shipment is too heavy for either of them.
    "Cargo Service": {
        "min_charge": 800, "max_weight_kg": 1000, "transit_days": 6,
        "slabs": [(50, 45), (200, 35), (1000, 25)],
    },
}


def estimate_cost(courier, weight_kg):
    """
    Simulated weight-slab cost for a single courier and a total chargeable
    weight -- the rate applied is the slab the *total* weight falls into,
    not summed across slabs, matching how real slab-based freight tariffs
    are typically quoted (a shipment in the 20-50kg slab is charged at that
    slab's own rate for its whole weight, not 5kg at the first rate plus
    15kg at the second). Returns None if the weight exceeds this courier's
    own maximum per-shipment limit -- it is not this function's job to
    silently allow an over-limit booking.
    """
    card = RATE_CARDS[courier]
    if weight_kg > card["max_weight_kg"]:
        return None
    for upper, rate in card["slabs"]:
        if weight_kg <= upper:
            return round(max(card["min_charge"], weight_kg * rate), 2)
    return None



def _extract_pincode(address, city):
    """A real 6-digit Indian PIN, pulled from the location's own real
    address/city text, not fabricated. Returns "" if genuinely none is
    present, rather than guessing one — every real Delivery Location's
    address in this system's seed data does end with one, checked
    directly before relying on this, but a location added later
    without one should show an honest blank, not a wrong number."""
    m = re.search(r"\b(\d{6})\b", f"{address or ''} {city or ''}")
    return m.group(1) if m else ""


def _location_lookup(data_file=None):
    return {l["id"]: l for l in pc.get_delivery_locations(active_only=False)}


def build_shipment_details(transfer, data_file=None):
    """
    transfer: a real row from inventory.get_stock_transfers() — an
    actual, already-shipped movement with its own real transfer_id,
    quantities, and carrier, not a transfer OPPORTUNITY (a suggestion
    that may never be acted on). Every field below is either read
    directly from real data (item master, delivery locations, org
    profile, the transfer record itself) or computed from it — nothing
    here is a placeholder value.

    Consignor and consignee contact both use the org's own profile
    contact (Contact_Phone/Contact_Email), not a fabricated separate
    number per location — honest, since these are the same company's
    own two facilities, and no per-location contact is captured
    anywhere in this system to draw on instead.

    Weight and declared value are both real computations (per-unit
    weight/price from the item master times the real shipped
    quantity), not flat estimates re-applied per shipment.
    """
    locs = _location_lookup(data_file)
    from_loc = locs.get(transfer["from_location"], {})
    to_loc = locs.get(transfer["to_location"], {})
    org = op.get_org_profile(data_file) or {}
    item = po_export.get_item_by_code(transfer["material_code"], active_only=False) or {}

    qty = float(transfer["quantity"])
    weight_per_unit = item.get("weight_kg")
    total_weight = round(weight_per_unit * qty, 2) if weight_per_unit is not None else None
    unit_price = item.get("price") or 0
    declared_value = round(unit_price * qty, 2)

    return {
        "Shipment Reference": transfer["transfer_id"],
        "Courier": transfer.get("carrier") or "BlueDart",
        "Ship Date": transfer.get("shipped_date") or date.today().strftime("%Y-%m-%d"),
        "Mode": "Surface",
        "Consignor Name": org.get("Legal_Name", ""),
        "Consignor Address": from_loc.get("address", ""),
        "Consignor City": from_loc.get("city", ""),
        "Consignor State": from_loc.get("state", ""),
        "Consignor Pincode": _extract_pincode(from_loc.get("address", ""), from_loc.get("city", "")),
        "Consignor Contact": org.get("Contact_Phone", ""),
        "Consignee Name": org.get("Legal_Name", ""),
        "Consignee Address": to_loc.get("address", ""),
        "Consignee City": to_loc.get("city", ""),
        "Consignee State": to_loc.get("state", ""),
        "Consignee Pincode": _extract_pincode(to_loc.get("address", ""), to_loc.get("city", "")),
        "Consignee Contact": org.get("Contact_Phone", ""),
        "Material Code": transfer["material_code"],
        "Product Description": transfer["material_desc"],
        "HSN Code": item.get("hsn_code", ""),
        "Quantity": qty,
        "UOM": transfer.get("uom", ""),
        "Pieces": int(qty) if qty == int(qty) else qty,
        "Weight per Unit (kg)": weight_per_unit,
        "Total Weight (kg)": total_weight,
        "Declared Value (INR)": declared_value,
    }


def build_customer_shipment_details(fulfillment, data_file=None):
    """
    fulfillment: a real row from fulfillment.get_fulfillment() — the
    external-consignee counterpart to build_shipment_details() above,
    per LOG-US-01's own documented generalization (S2S V1.15 / O2C
    V1.9): the consignee is resolved from the real customer master
    record (CUST-US-01), never from this org's own Delivery_Locations
    network — that network is this org's own supply-side Plants, not a
    genuine customer address (see CLAUDE.md's own documented
    conflation note on Sales Order delivery_location; this function
    deliberately does not reuse it for the consignee side, only for
    the consignor/fulfilling-Plant side, where it's genuinely correct).

    Consignee state is resolved from the customer's own GSTIN via the
    same vendor_onboarding.validate_gstin() state-code lookup
    billing.py already relies on for GST determination — one real
    state-resolution mechanism, not a second, free-text-parsed one.

    A fulfillment can carry more than one material (unlike a stock
    transfer, which is always single-material) — cargo fields that
    only make sense per-line (Material Code, Product Description, HSN
    Code) are honestly summarized across lines rather than picking
    just the first one; HSN Code is left blank if lines don't share a
    single one, the same "honest blank over a wrong number" principle
    _extract_pincode() already uses. Weight and value are real
    aggregate sums across every shipped line — if ANY line's weight is
    unknown (Item Master's weight_kg unset), the aggregate is honestly
    None rather than a partial, understated total.
    """
    import customer_onboarding as co
    import vendor_onboarding as vo
    import fulfillment as ful

    org = op.get_org_profile(data_file) or {}
    locs = _location_lookup(data_file)
    from_loc = locs.get(fulfillment["delivery_location"], {})
    customer = co.get_customer(fulfillment["customer_id"], data_file=data_file) or {}

    consignee_state = ""
    if customer.get("GSTIN"):
        ok, _, details = vo.validate_gstin(customer["GSTIN"])
        if ok:
            consignee_state = details["state_name"]

    items = ful.get_fulfillment_items(fulfillment["fulfillment_id"], data_file=data_file)
    shipped_items = [it for it in items if (it["qty_shipped"] or 0) > 0]

    total_weight = 0.0
    weight_known = True
    total_value = 0.0
    mat_codes, descriptions, hsn_codes, uoms = [], [], [], set()
    total_qty = 0.0
    for it in shipped_items:
        qty = float(it["qty_shipped"] or 0)
        total_qty += qty
        item = po_export.get_item_by_code(it["mat_code"], active_only=False) or {}
        w = item.get("weight_kg")
        if w is None:
            weight_known = False
        else:
            total_weight += w * qty
        total_value += (item.get("price") or 0) * qty
        mat_codes.append(it["mat_code"])
        descriptions.append(f"{it['mat_desc']} x {qty:g}")
        if item.get("hsn_code"):
            hsn_codes.append(item["hsn_code"])
        uoms.add(it["uom"])

    single_line = len(shipped_items) == 1
    hsn_common = hsn_codes[0] if hsn_codes and len(set(hsn_codes)) == 1 else ""
    uom_common = next(iter(uoms)) if len(uoms) == 1 else "Mixed"

    return {
        "Shipment Reference": fulfillment["fulfillment_id"],
        "Courier": fulfillment.get("carrier") or "BlueDart",
        "Ship Date": fulfillment.get("shipped_date") or date.today().strftime("%Y-%m-%d"),
        "Mode": "Surface",
        "Consignor Name": org.get("Legal_Name", ""),
        "Consignor Address": from_loc.get("address", ""),
        "Consignor City": from_loc.get("city", ""),
        "Consignor State": from_loc.get("state", ""),
        "Consignor Pincode": _extract_pincode(from_loc.get("address", ""), from_loc.get("city", "")),
        "Consignor Contact": org.get("Contact_Phone", ""),
        "Consignee Name": customer.get("Customer_Name", ""),
        "Consignee Address": customer.get("Address", ""),
        "Consignee City": customer.get("City", ""),
        "Consignee State": consignee_state,
        "Consignee Pincode": _extract_pincode(customer.get("Address", ""), customer.get("City", "")),
        "Consignee Contact": customer.get("Contact_Phone", ""),
        "Material Code": mat_codes[0] if single_line else ", ".join(mat_codes),
        "Product Description": descriptions[0].rsplit(" x ", 1)[0] if single_line else "; ".join(descriptions),
        "HSN Code": hsn_common,
        "Quantity": total_qty,
        "UOM": uom_common,
        "Pieces": int(total_qty) if total_qty == int(total_qty) else total_qty,
        "Weight per Unit (kg)": (total_weight / total_qty) if (single_line and weight_known and total_qty) else None,
        "Total Weight (kg)": round(total_weight, 2) if weight_known else None,
        "Declared Value (INR)": round(total_value, 2),
    }


def generate_shipping_excel(shipment):
    """
    One real shipment -> one formatted .xlsx, laid out as a real
    courier booking form would be (labeled sections, not a raw data
    dump) — ready to hand-fill into BlueDart's own portal, or attach as
    a booking reference alongside the physical consignment. No
    formulas (every value here is already a real, final number — there
    is nothing for a formula to compute from something else in this
    sheet), so no recalculation step applies. Returns (filename, bytes).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipping Details"

    title_font = Font(name="Arial", bold=True, size=14)
    section_font = Font(name="Arial", bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="1E1E2E", end_color="1E1E2E", fill_type="solid")
    label_font = Font(name="Arial", bold=True)
    value_font = Font(name="Arial")

    ws["A1"] = f"{shipment['Courier']} — Shipment Booking Details"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    ws["A2"] = f"Reference: {shipment['Shipment Reference']}  ·  Generated: " \
              f"{date.today().strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.merge_cells("A2:B2")

    sections = [
        ("Shipment", ["Shipment Reference", "Courier", "Ship Date", "Mode"]),
        ("Consignor (From)", ["Consignor Name", "Consignor Address", "Consignor City",
                              "Consignor State", "Consignor Pincode", "Consignor Contact"]),
        ("Consignee (To)", ["Consignee Name", "Consignee Address", "Consignee City",
                            "Consignee State", "Consignee Pincode", "Consignee Contact"]),
        ("Cargo", ["Material Code", "Product Description", "HSN Code", "Quantity", "UOM",
                  "Pieces", "Weight per Unit (kg)", "Total Weight (kg)", "Declared Value (INR)"]),
    ]

    row = 4
    for section_title, fields in sections:
        ws.cell(row=row, column=1, value=section_title).font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=2).fill = section_fill
        row += 1
        for f in fields:
            c1 = ws.cell(row=row, column=1, value=f)
            c1.font = label_font
            c2 = ws.cell(row=row, column=2, value=shipment.get(f, ""))
            c2.font = value_font
            if "Value" in f:
                c2.number_format = "₹#,##0.00"
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 52

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_ref = shipment["Shipment Reference"].replace("/", "-")
    filename = f"{shipment['Courier']}_Shipment_{safe_ref}.xlsx"
    return filename, buf.getvalue()


# ── Mock courier API connector ───────────────────────────────────────────────
# No real BlueDart credentials or API documentation are available here, and
# guessing at an integration contract for a real courier and presenting it as
# genuine would be actively misleading, not merely incomplete. This is a
# clearly-labeled simulation, not a real connection — no network call happens
# anywhere below. Structured deliberately so that swapping in a real
# integration later means replacing the single `requests.post(...)` line
# noted in submit_to_courier()'s own docstring, not redesigning the data
# flow around it — build_bluedart_api_payload() already produces the real
# request shape a genuine integration would send.

TRANSIT_DAYS_SURFACE = 4  # a reasonable default for domestic Surface mode,
                          # not zone-calculated — there's no real distance/
                          # zone-rating data in this system to compute one from


def build_bluedart_api_payload(shipment):
    """
    The real request payload a BlueDart waybill-generation call would
    send — structured the way this class of Indian B2B courier API
    typically shapes a booking request (Shipper/Consignee blocks,
    package/weight detail, service type), not verified against
    BlueDart's own current API documentation (not available here).
    Real integration work would need to confirm this against their
    actual current contract before going live — this is a realistic
    starting shape, not a guarantee of exact field names.
    """
    return {
        "Shipper": {
            "Name": shipment["Consignor Name"],
            "Address": shipment["Consignor Address"],
            "City": shipment["Consignor City"],
            "State": shipment["Consignor State"],
            "Pincode": shipment["Consignor Pincode"],
            "Phone": shipment["Consignor Contact"],
        },
        "Consignee": {
            "Name": shipment["Consignee Name"],
            "Address": shipment["Consignee Address"],
            "City": shipment["Consignee City"],
            "State": shipment["Consignee State"],
            "Pincode": shipment["Consignee Pincode"],
            "Phone": shipment["Consignee Contact"],
        },
        "Shipment": {
            "ReferenceNumber": shipment["Shipment Reference"],
            "ProductDescription": shipment["Product Description"],
            "HSNCode": shipment["HSN Code"],
            "Pieces": shipment["Pieces"],
            "WeightKg": shipment["Total Weight (kg)"],
            "DeclaredValueINR": shipment["Declared Value (INR)"],
            "ServiceType": "Surface",
            "PickupDate": shipment["Ship Date"],
        },
    }


def build_delhivery_api_payload(shipment):
    """
    Delhivery's own request shape, deliberately structured differently
    from build_bluedart_api_payload() above -- flatter, snake_case,
    pickup/delivery terminology rather than Shipper/Consignee blocks.
    Real courier APIs genuinely differ this way; if both payloads were
    shaped identically there would be nothing for a normalization layer
    to actually normalize. Same caveat as BlueDart's: a realistic
    starting shape, not verified against Delhivery's own current API
    documentation (not available here).
    """
    return {
        "pickup_details": {
            "name": shipment["Consignor Name"],
            "address_line1": shipment["Consignor Address"],
            "city": shipment["Consignor City"],
            "state": shipment["Consignor State"],
            "pin": shipment["Consignor Pincode"],
            "phone": shipment["Consignor Contact"],
        },
        "delivery_details": {
            "name": shipment["Consignee Name"],
            "address_line1": shipment["Consignee Address"],
            "city": shipment["Consignee City"],
            "state": shipment["Consignee State"],
            "pin": shipment["Consignee Pincode"],
            "phone": shipment["Consignee Contact"],
        },
        "package_details": {
            "client_reference": shipment["Shipment Reference"],
            "product_desc": shipment["Product Description"],
            "hsn": shipment["HSN Code"],
            "quantity": shipment["Pieces"],
            "weight_kg": shipment["Total Weight (kg)"],
            "declared_value": shipment["Declared Value (INR)"],
            "mode": "surface",
            "pickup_date": shipment["Ship Date"],
        },
    }


PAYLOAD_BUILDERS = {"BlueDart": build_bluedart_api_payload, "Delhivery": build_delhivery_api_payload,
                    # No real distinct API shape for a Cargo/Freight booking to model here --
                    # reuses BlueDart's own shape explicitly rather than relying on a silent
                    # fallback, since the same consignor/consignee/cargo fields genuinely apply.
                    "Cargo Service": build_bluedart_api_payload}


def _group_legs_by_route(legs):
    """Groups real shipment legs by (from_location, to_location) only --
    same consignor AND same consignee, never route similarity alone (e.g.
    same destination city from a different source Plant), since a real
    courier waybill requires one physical pickup point."""
    groups = {}
    for leg in legs:
        key = (leg["from_location"], leg["to_location"])
        groups.setdefault(key, []).append(leg)
    return groups


def _split_into_bins(legs, max_weight_kg):
    """
    Real bin-packing, not a naive one-booking-per-leg fallback: sorts
    legs largest-first (first-fit-decreasing, a well-known, simple
    approximation -- not a claim of true optimal packing, which is
    NP-hard, but it reliably minimizes bin count for realistic shipment
    sizes) and greedily fills bins that each stay within max_weight_kg,
    consolidating as many legs together as the limit allows rather than
    splitting more than necessary.
    """
    sorted_legs = sorted(legs, key=lambda l: l["_weight"], reverse=True)
    bins = []  # list of (list_of_legs, total_weight)
    for leg in sorted_legs:
        placed = False
        for i, (bin_legs, bin_weight) in enumerate(bins):
            if bin_weight + leg["_weight"] <= max_weight_kg:
                bin_legs.append(leg)
                bins[i] = (bin_legs, bin_weight + leg["_weight"])
                placed = True
                break
        if not placed:
            bins.append(([leg], leg["_weight"]))
    return [b[0] for b in bins]


def _resolve_courier_and_bins(legs, pinned_courier=None):
    """
    Returns (bookings, unbookable) -- bookings is a list of (courier,
    bin_legs, total_weight, cost) for every group that can actually be
    booked; unbookable is a list of (bin_legs, reason) for any bin whose
    own weight exceeds every available courier's maximum -- a real,
    genuine situation (a single shipment too heavy for standard Surface
    service needs a Cargo/Freight service this system doesn't model),
    not an error to crash on. A single transfer heavier than every
    courier's own limit can never be split further -- it is reported,
    never silently dropped or forced into an over-limit booking.

    A pinned courier is never overridden -- its own weight limit governs
    splitting directly. With no pin, the group's total weight is first
    checked against every courier whole (no split) and the cheapest
    courier that can carry the whole group wins; only when no single
    courier's limit covers the full group does splitting happen at all,
    using the courier with the larger limit (fewest resulting bookings),
    then each resulting bin is re-priced against every courier that can
    carry it -- a bin small enough to fit the other courier can still
    win it if that's genuinely cheaper for that specific bin's weight.
    """
    total_weight = sum(l["_weight"] for l in legs)

    if pinned_courier:
        limit = RATE_CARDS[pinned_courier]["max_weight_kg"]
        if total_weight <= limit:
            bins = [legs]
        else:
            bins = _split_into_bins(legs, limit)
        bookings, unbookable = [], []
        for b in bins:
            w = sum(l["_weight"] for l in b)
            cost = estimate_cost(pinned_courier, w)
            if cost is None:
                unbookable.append((b, f"{w:g}kg exceeds {pinned_courier}'s own "
                                      f"{limit:g}kg maximum and cannot be split further "
                                      f"(a single leg alone exceeds it)."))
            else:
                bookings.append((pinned_courier, b, w, cost))
        return bookings, unbookable

    # No pin: whole-group least-cost first.
    fits_whole = [(c, estimate_cost(c, total_weight)) for c in COURIERS
                 if total_weight <= RATE_CARDS[c]["max_weight_kg"]]
    if fits_whole:
        best_courier = min(fits_whole, key=lambda x: x[1])[0]
        return [(best_courier, legs, total_weight, estimate_cost(best_courier, total_weight))], []

    # No single courier covers the whole group -- split using the larger
    # limit (fewest bookings), then re-price each resulting bin per-courier.
    split_courier = max(COURIERS, key=lambda c: RATE_CARDS[c]["max_weight_kg"])
    max_limit = RATE_CARDS[split_courier]["max_weight_kg"]
    bins = _split_into_bins(legs, max_limit)
    bookings, unbookable = [], []
    for b in bins:
        w = sum(l["_weight"] for l in b)
        candidates = [(c, estimate_cost(c, w)) for c in COURIERS
                     if estimate_cost(c, w) is not None]
        if not candidates:
            unbookable.append((b, f"{w:g}kg exceeds every available courier's maximum "
                                  f"(largest is {split_courier} at {max_limit:g}kg) and "
                                  f"cannot be split further (a single leg alone exceeds it)."))
        else:
            best_courier = min(candidates, key=lambda x: x[1])[0]
            bookings.append((best_courier, b, w, estimate_cost(best_courier, w)))
    return bookings, unbookable


# ── Kind registry: lets ONE shared booking/tracking implementation serve
# more than one real shipment source ──────────────────────────────────────
# "stock_transfer" (INV-US-05, the original) and "customer_fulfillment"
# (FUL-US-05, new) each need their own way to fetch a leg, build its
# shipment dict, and persist tracking back — everything downstream of that
# (route consolidation, least-cost selection, AWB generation, the tracking
# simulation) is identical, so it's written once, not twice. A generic leg
# dict always carries from_location/to_location/transfer_id keys regardless
# of kind (transfer_id doubles as "whatever this shipment's own real
# reference id is" — a fulfillment_id here, not a misnomer, just the one
# generic key name every downstream line already uses).

def _get_stock_transfer_leg(ref_id, data_file=None):
    import inventory as inv
    return inv.get_stock_transfer(ref_id, data_file=data_file)


def _get_customer_fulfillment_leg(ref_id, data_file=None):
    import fulfillment as ful
    f = ful.get_fulfillment(ref_id, data_file=data_file)
    if f is None:
        return None
    f = dict(f)
    f["from_location"] = f["delivery_location"]
    f["to_location"] = f["customer_id"]  # consignee identity, for route-grouping only
    f["transfer_id"] = f["fulfillment_id"]  # generic ref-id key, reused by every line below
    return f


def _update_stock_transfer_tracking(ref_id, awb, carrier, data_file=None):
    import inventory as inv
    inv.update_transfer_tracking(ref_id, awb, carrier=carrier, data_file=data_file)


def _update_customer_fulfillment_tracking(ref_id, awb, carrier, data_file=None):
    import fulfillment as ful
    ful.update_dispatch_tracking(ref_id, awb, carrier=carrier, data_file=data_file)


def _update_stock_transfer_shipdate_by_awb(awb, new_date, data_file=None):
    import db
    conn = db.get_connection()
    try:
        conn.execute(
            "UPDATE stock_transfers SET shipped_date=? WHERE tracking_ref=? AND "
            "tracking_ref IS NOT NULL AND tracking_ref != ''",
            (str(new_date), awb))
        conn.commit()
    finally:
        conn.close()


def _update_customer_fulfillment_shipdate_by_awb(awb, new_date, data_file=None):
    import db
    conn = db.get_connection()
    try:
        conn.execute(
            "UPDATE fulfillments SET shipped_date=? WHERE tracking_ref=? AND "
            "tracking_ref IS NOT NULL AND tracking_ref != ''",
            (str(new_date), awb))
        conn.commit()
    finally:
        conn.close()


_KINDS = {
    "stock_transfer": {
        "get_leg": _get_stock_transfer_leg,
        "build_details": build_shipment_details,
        "update_tracking": _update_stock_transfer_tracking,
        "update_shipdate_by_awb": _update_stock_transfer_shipdate_by_awb,
    },
    "customer_fulfillment": {
        "get_leg": _get_customer_fulfillment_leg,
        "build_details": build_customer_shipment_details,
        "update_tracking": _update_customer_fulfillment_tracking,
        "update_shipdate_by_awb": _update_customer_fulfillment_shipdate_by_awb,
    },
}


def submit_batch_to_courier(ref_ids, kind="stock_transfer", data_file=None):
    """
    The real LOG-US-01 entry point: books one or more shipment legs
    together, applying real route consolidation and weight-slab-aware
    least-cost selection across the batch, rather than booking and
    pricing each leg in isolation. submit_to_courier() (single-shipment)
    now delegates here with a one-element batch, so both paths share one
    real implementation, not two that could drift.

    kind selects which real source these ref_ids come from — internal
    stock transfers (default, unchanged behavior) or O2C customer
    fulfillments (see _KINDS above) — everything else about this
    function is identical either way.

    Each leg's own `carrier` field (set at Ship time) is treated as a
    pinned preference; a leg shipped with no carrier recorded (an
    empty string) is left for this function's own least-cost proposal.
    Legs are only ever grouped when they share both consignor and
    consignee location AND the same pinned-courier-or-none status --
    two legs pinned to different couriers on the same route can never
    share one booking, since a real waybill is with exactly one courier.

    Returns a list of per-booking-group results, each carrying its own
    real AWB, courier, the real ref ids it covers, and each leg's own
    proportional share of that booking's real cost.
    """
    from datetime import datetime, timedelta

    handlers = _KINDS[kind]
    legs = []
    for rid in ref_ids:
        t = handlers["get_leg"](rid, data_file=data_file)
        if t is None or t.get("tracking_ref"):
            continue  # already booked or doesn't exist -- caller reports these separately
        shipment = handlers["build_details"](t, data_file=data_file)
        weight = shipment.get("Total Weight (kg)")
        if weight is None:
            continue  # no resolvable weight -- caller reports this separately
        t["_weight"] = weight
        t["_shipment"] = shipment
        t["_pinned"] = t.get("carrier") or None
        legs.append(t)

    route_groups = _group_legs_by_route(legs)
    results = []
    unbookable_out = []
    for (from_loc, to_loc), route_legs in route_groups.items():
        # Sub-group by pinned courier (including "no pin") within this route --
        # a shared route with mixed pins can never all be one booking.
        by_pin = {}
        for l in route_legs:
            by_pin.setdefault(l["_pinned"], []).append(l)

        for pinned_courier, pin_legs in by_pin.items():
            bookings, unbookable = _resolve_courier_and_bins(pin_legs, pinned_courier)
            for bin_legs, reason in unbookable:
                for l in bin_legs:
                    unbookable_out.append({"transfer_id": l["transfer_id"], "reason": reason})
            for courier, bin_legs, total_weight, cost in bookings:
                ref_transfer_id = bin_legs[0]["transfer_id"]
                digits = "".join(ch for ch in ref_transfer_id if ch.isdigit()).rjust(8, "0")[-8:]
                prefix = {"BlueDart": "77", "Delhivery": "DL", "Cargo Service": "CG"}.get(courier, "XX")
                awb = f"{prefix}{digits}"
                transit_days = RATE_CARDS[courier]["transit_days"]
                ship_date = datetime.strptime(bin_legs[0]["shipped_date"], "%Y-%m-%d")
                eta = (ship_date + timedelta(days=transit_days)).strftime("%Y-%m-%d")

                leg_shares = []
                for l in bin_legs:
                    share = round(cost * (l["_weight"] / total_weight), 2) if total_weight else 0
                    handlers["update_tracking"](l["transfer_id"], awb, courier, data_file=data_file)
                    leg_shares.append({"transfer_id": l["transfer_id"], "weight_kg": l["_weight"],
                                       "freight_share": share})

                results.append({
                    "courier": courier, "awb_number": awb, "from_location": from_loc,
                    "to_location": to_loc, "total_weight_kg": total_weight,
                    "total_cost": cost, "estimated_delivery": eta,
                    "consolidated_legs": leg_shares,
                    "pinned": pinned_courier is not None,
                })
    return {"bookings": results, "unbookable": unbookable_out}


def submit_to_courier(shipment, kind="stock_transfer", data_file=None):
    """
    SIMULATED submission for a single shipment -- a thin convenience
    wrapper around submit_batch_to_courier() (the real, shared LOG-US-01
    entry point) with a one-element batch, so single- and multi-shipment
    callers share exactly one real implementation rather than two that
    could quietly drift apart. Kept for callers (and any existing tests)
    that only ever have one shipment in hand at a time.

    Idempotent by design, inherited directly from the batch function:
    a shipment that already has a real tracking_ref on file returns
    that same persisted AWB rather than generating a new one.

    Returns {"request_payload": ..., "response": ..., "already_submitted": bool} --
    the same shape this function has always returned, even though the
    real work now happens in the batch function.
    """
    handlers = _KINDS[kind]
    ref_id = shipment["Shipment Reference"]
    existing = handlers["get_leg"](ref_id, data_file=data_file)
    already_submitted = bool(existing and existing.get("tracking_ref"))
    if already_submitted:
        courier = existing["carrier"] or "BlueDart"
        payload = PAYLOAD_BUILDERS.get(courier, build_bluedart_api_payload)(shipment)
        tracking = get_tracking_status(ref_id, kind=kind, data_file=data_file)
        response = {"awb_number": existing["tracking_ref"], "status": "Already Booked",
                   "estimated_delivery": (tracking or {}).get("checkpoints", [{}])[-1].get("date", ""),
                   "service_type": "Surface", "courier": courier}
        return {"request_payload": payload, "response": response, "already_submitted": True}

    result = submit_batch_to_courier([ref_id], kind=kind, data_file=data_file)
    if result["unbookable"]:
        raise ValueError(result["unbookable"][0]["reason"])
    if not result["bookings"]:
        raise ValueError(f"Could not book {ref_id} -- no resolvable weight or courier.")
    r = result["bookings"][0]
    payload = PAYLOAD_BUILDERS.get(r["courier"], build_bluedart_api_payload)(shipment)
    response = {"awb_number": r["awb_number"], "status": "Booked",
               "estimated_delivery": r["estimated_delivery"], "service_type": "Surface",
               "courier": r["courier"]}
    return {"request_payload": payload, "response": response, "already_submitted": False}


def get_tracking_status(ref_id, kind="stock_transfer", data_file=None):
    """
    SIMULATED tracking timeline — same reasoning as submit_to_courier():
    no real courier tracking API exists to call here. Deterministic
    from the real ship date and the booked courier's own real transit_
    days (from RATE_CARDS, e.g. BlueDart 4 days vs Delhivery 3 -- not
    one flat assumption for every courier, which would have made a
    Delhivery shipment's own tracking timeline disagree with the ETA
    already shown at booking time), not randomized — refreshing or
    reopening this later on the same day shows the identical status,
    the way a real tracking page checked twice in one day would. Only
    callable for a shipment that's already been submitted (has a real
    tracking_ref/AWB on file) — there's nothing to track before a
    booking exists.

    kind selects which real source ref_id comes from (see _KINDS) —
    everything else about the simulation is identical either way.

    The five checkpoint stages are spaced proportionally across
    whichever courier's own real transit_days applies (e.g. Delhivery's
    3-day service compresses two stages onto the same day rather than
    inventing a slower timeline than the courier's own ETA promised).

    Checkpoint cities are the shipment's own real consignor/consignee
    cities (not fabricated waypoints/hub names this system has no real
    data for) — "In Transit" deliberately doesn't claim a specific
    intermediate hub it can't back up.

    Returns {"awb_number", "current_status", "checkpoints": [
    {"day", "label", "location", "date", "completed"}, ...]}.
    """
    from datetime import datetime, timedelta, date as date_cls

    handlers = _KINDS[kind]
    transfer = handlers["get_leg"](ref_id, data_file=data_file)
    if not transfer or not transfer.get("tracking_ref"):
        return None

    courier = transfer.get("carrier") or "BlueDart"
    transit_days = RATE_CARDS.get(courier, RATE_CARDS["BlueDart"])["transit_days"]

    shipment = handlers["build_details"](transfer, data_file=data_file)
    ship_date = datetime.strptime(shipment["Ship Date"], "%Y-%m-%d").date()
    days_elapsed = max(0, (date_cls.today() - ship_date).days)
    days_elapsed = min(days_elapsed, transit_days)

    stage_fractions = [
        (0.0, "Booked", shipment["Consignor City"]),
        (0.25, "Picked Up", shipment["Consignor City"]),
        (0.5, "In Transit", f"{shipment['Consignor City']} → {shipment['Consignee City']}"),
        (0.75, "Arrived at Destination Hub", shipment["Consignee City"]),
        (1.0, None, shipment["Consignee City"]),  # label resolved below
    ]
    stages = []
    for frac, label, location in stage_fractions:
        day = round(frac * transit_days)
        if label is None:
            label = "Delivered" if days_elapsed >= transit_days else "Out for Delivery"
        stages.append((day, label, location))

    checkpoints = []
    current_status = "Booked"
    for day, label, location in stages:
        completed = day <= days_elapsed
        if completed:
            current_status = label
        checkpoints.append({
            "day": day, "label": label, "location": location,
            "date": (ship_date + timedelta(days=day)).strftime("%Y-%m-%d"),
            "completed": completed,
        })

    return {
        "awb_number": transfer["tracking_ref"],
        "current_status": current_status,
        "checkpoints": checkpoints,
    }


def skip_ahead_tracking(ref_id, kind="stock_transfer", data_file=None):
    """
    A labeled demo control, not a second tracking mechanism: the real
    simulation in get_tracking_status() is deterministic from elapsed
    real time (today minus the real ship date), which is the right
    honesty standard for the simulation itself but makes a live demo
    wait real days to show progress. This advances the SAME simulation
    by moving the shipped_date further into the past by just enough to
    reach the next not-yet-completed checkpoint -- get_tracking_status()
    then computes a genuinely later status from that adjusted real
    date, not a separately faked one. A no-op, returning the status
    unchanged, once every checkpoint is already complete (Delivered) --
    there is nothing further to advance to.

    kind selects which real table (stock_transfers or fulfillments)
    actually gets the adjusted ship date written back to it.

    Real bug fixed here, found by testing through an actual browser,
    not assumed correct from the single-transfer case working:
    consolidated shipments (submit_batch_to_courier()'s whole reason
    to exist) share one AWB across more than one real ref id -- the
    same physical booking. Advancing only the one ref id whose dialog
    happened to be open left its consolidated siblings behind on the
    old date, so the same physical shipment could show two different
    statuses depending on which row's AWB link someone clicked. Every
    leg sharing this ref id's own tracking_ref now moves together,
    always.
    """
    from datetime import timedelta, date as date_cls

    handlers = _KINDS[kind]
    status = get_tracking_status(ref_id, kind=kind, data_file=data_file)
    if not status:
        return None

    next_day = None
    for cp in status["checkpoints"]:
        if not cp["completed"]:
            next_day = cp["day"]
            break
    if next_day is None:
        return status  # already fully Delivered -- nothing further to skip to

    new_ship_date = date_cls.today() - timedelta(days=next_day)
    handlers["update_shipdate_by_awb"](status["awb_number"], new_ship_date, data_file=data_file)

    return get_tracking_status(ref_id, kind=kind, data_file=data_file)
