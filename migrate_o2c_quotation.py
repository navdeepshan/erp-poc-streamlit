"""
migrate_o2c_quotation.py — one-time port of Quotes and Quote_Items from
data.xlsx into erp_pilot.db.

Second O2C slice — depends on Customer_Master already being migrated
(migrate_o2c_customer.py), since Quotes references Customer_ID (not a
foreign key in SQLite here, but the same "run migrations in dependency
order" discipline as everywhere else). Idempotent.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_quotes(wb, conn):
    if "Quotes" not in wb.sheetnames:
        return 0
    ws = wb["Quotes"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        qid = str(row[hdrs.get("Quote_ID", 0)] or "").strip()
        if not qid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((qid, g("Customer_ID"), g("Customer_Name"), g("Quote_Date"),
                    g("Valid_Until"), g("Payment_Terms"), g("Status"), g("Currency"),
                    g("Total_Value"), g("Notes"), g("Filename")))
    conn.execute("DELETE FROM quotes")
    conn.executemany(
        "INSERT INTO quotes (quote_id, customer_id, customer_name, quote_date, valid_until, "
        "payment_terms, status, currency, total_value, notes, filename) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_quote_items(wb, conn):
    if "Quote_Items" not in wb.sheetnames:
        return 0
    ws = wb["Quote_Items"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        qid = str(row[hdrs.get("Quote_ID", 0)] or "").strip()
        if not qid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((qid, g("Line_Item"), g("Material_Code"), g("Material_Desc"),
                    g("UOM"), g("Qty"), g("Unit_Price"), g("Line_Total")))
    conn.execute("DELETE FROM quote_items")
    conn.executemany(
        "INSERT INTO quote_items (quote_id, line_item, material_code, material_desc, uom, "
        "qty, unit_price, line_total) VALUES (?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_q = migrate_quotes(wb, conn)
        n_qi = migrate_quote_items(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"quotes": n_q, "quote_items": n_qi}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Quotes/Quote_Items to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--db", default=None)
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
