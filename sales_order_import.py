"""
sales_order_import.py — Bulk Sales Order import via a downloadable
Excel template plus an upload-and-validate step, so a tenant with real
historical or planned demand can seed a batch of Sales Orders without
hand-entering each one through the Sales Orders screen. Built to make
INV-US-03's "Sales Order Based" time-phased planning mode
demonstrable for a tenant (e.g. IDS Denmed) that might not have O2C
actively creating orders through the normal flow yet — a real, general
O2C capability, not something specific to any one pilot.

Reuses sales_order.create_direct_order() for every order this creates
— not a separate insert path — so every uploaded order goes through
the exact same real-time credit check a manually-entered one would.
Some uploaded orders may land on Credit Hold rather than Confirmed;
that's correct behaviour, not a bug in the importer.
"""
import io
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

import db
import customer_onboarding as co
import pr_consolidation as pc
import po_export
import sales_order as so

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REF_HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")

TEMPLATE_HEADERS = [
    "Order Reference", "Customer ID", "Delivery Location (blank = customer default)",
    "Requested Delivery Date (YYYY-MM-DD)", "Material Code", "Quantity",
    "Unit Price (blank = list price)",
]


def _style_header(ws, headers, fill, font, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30


def generate_template(data_file=None):
    """
    Builds the downloadable Sales Order import template: one data-
    entry sheet (Order Reference groups multiple rows into one order —
    several rows sharing the same reference become that order's
    separate line items) with dropdown-validated Customer/Location/
    Material columns, plus reference sheets a person can check while
    filling it out without needing to memorize any ID. Delivery
    Location is deliberately optional per row: leave it blank to use
    that customer's own stored default_delivery_location, or fill in
    a different, valid location to override it for just that row —
    the same blank-means-default / filled-means-override rule
    import_sales_orders() itself enforces at upload time.

    Returns raw .xlsx bytes, ready for a Streamlit download button —
    no temp file involved.
    """
    fpath = data_file
    customers = [c for c in co.list_customers(fpath) if c["Active"] == "Yes"]
    locations = pc.get_delivery_locations(active_only=True)
    materials = [m for m in po_export.load_item_master(fpath, active_only=True)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import"
    _style_header(ws, TEMPLATE_HEADERS, HEADER_FILL, HEADER_FONT)
    widths = [16, 18, 30, 24, 16, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    # A couple of blank data rows ready to fill, well beyond the last
    # validated row so the dropdowns are visibly there before anyone
    # starts typing.
    for r in range(2, 12):
        ws.cell(row=r, column=1)

    # Customer dropdown
    cust_ids = [c["Customer_ID"] for c in customers]
    if cust_ids:
        dv_cust = DataValidation(type="list", formula1='"' + ",".join(cust_ids) + '"',
                                 allow_blank=False, showDropDown=False)
        dv_cust.error = "Select a valid, active customer ID from the dropdown."
        ws.add_data_validation(dv_cust)
        dv_cust.add(f"B2:B1000")

    # Delivery Location dropdown (optional — blank allowed)
    loc_ids = [l["id"] for l in locations]
    if loc_ids:
        dv_loc = DataValidation(type="list", formula1='"' + ",".join(loc_ids) + '"',
                                allow_blank=True, showDropDown=False)
        dv_loc.error = "Leave blank to use the customer's default, or select a valid location."
        ws.add_data_validation(dv_loc)
        dv_loc.add(f"C2:C1000")

    # Material dropdown
    mat_codes = [m["code"] for m in materials]
    if mat_codes:
        dv_mat = DataValidation(type="list", formula1='"' + ",".join(mat_codes) + '"',
                                allow_blank=False, showDropDown=False)
        dv_mat.error = "Select a valid, active material code from the dropdown."
        ws.add_data_validation(dv_mat)
        dv_mat.add(f"E2:E1000")

    # Reference: Customers
    ws2 = wb.create_sheet("Customers")
    ref_headers = ["Customer ID", "Customer Name", "City", "Default Delivery Location", "Credit Status"]
    _style_header(ws2, ref_headers, REF_HEADER_FILL, Font(bold=True))
    for r, c in enumerate(customers, 2):
        loc_name = next((l["name"] for l in locations if l["id"] == c["Default_Delivery_Location"]), "")
        ws2.cell(row=r, column=1, value=c["Customer_ID"])
        ws2.cell(row=r, column=2, value=c["Customer_Name"])
        ws2.cell(row=r, column=3, value=c["City"])
        ws2.cell(row=r, column=4, value=f"{c['Default_Delivery_Location'] or '(none set)'} "
                                        f"{'- ' + loc_name if loc_name else ''}".strip())
        ws2.cell(row=r, column=5, value=c["Credit_Status"])
    for i, w in enumerate([16, 30, 16, 30, 14], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Reference: Materials
    ws3 = wb.create_sheet("Materials")
    _style_header(ws3, ["Material Code", "Description", "UOM", "List Price"], REF_HEADER_FILL, Font(bold=True))
    for r, m in enumerate(materials, 2):
        ws3.cell(row=r, column=1, value=m["code"])
        ws3.cell(row=r, column=2, value=m["desc"])
        ws3.cell(row=r, column=3, value=m["uom"])
        ws3.cell(row=r, column=4, value=m["price"])
    for i, w in enumerate([16, 40, 10, 14], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Reference: Delivery Locations
    ws4 = wb.create_sheet("Delivery Locations")
    _style_header(ws4, ["Location ID", "Name", "City"], REF_HEADER_FILL, Font(bold=True))
    for r, l in enumerate(locations, 2):
        ws4.cell(row=r, column=1, value=l["id"])
        ws4.cell(row=r, column=2, value=l["name"])
        ws4.cell(row=r, column=3, value=l["city"])
    for i, w in enumerate([16, 30, 16], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_sales_orders(file_bytes, data_file=None):
    """
    Reads an uploaded template (generate_template()'s own shape, though
    doesn't require it came from there), groups rows by Order
    Reference, validates each resulting order as a whole, and creates
    every valid one via sales_order.create_direct_order() — the exact
    same function, and exact same real-time credit check, a manually-
    entered order goes through. An uploaded order landing on Credit
    Hold instead of Confirmed is correct behaviour, not a failure.

    Validation granularity is per-ORDER, not per-line: if any line
    within one Order Reference group fails validation, the whole order
    is rejected with every failing line's specific reason — unlike
    INV-US-06's independent per-line consumption records,
    create_direct_order() takes one order's full line list in a single
    call, so there's no clean way to create "most of" an order and
    silently drop one bad line from it.

    Returns {accepted: [{order_ref, so_id, status}, ...],
             rejected: [{order_ref, reasons: [...]}, ...]}.
    """
    fpath = data_file
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Import" not in wb.sheetnames:
        raise ValueError("This file doesn't have an 'Import' sheet — use the downloaded template.")
    ws = wb["Import"]

    customers = {c["Customer_ID"]: c for c in co.list_customers(fpath)}
    locations = {l["id"]: l for l in pc.get_delivery_locations(active_only=True)}
    materials = {m["code"]: m for m in po_export.load_item_master(fpath, active_only=True)}

    groups = {}  # order_ref -> list of raw row dicts
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [c.value for c in row]
        if not any(vals):
            continue
        order_ref = str(vals[0]).strip() if vals[0] else ""
        if not order_ref:
            continue
        groups.setdefault(order_ref, []).append({
            "customer_id": vals[1], "delivery_location": vals[2],
            "req_date": vals[3], "mat_code": vals[4], "qty": vals[5], "unit_price": vals[6],
        })

    accepted, rejected = [], []
    for order_ref, rows in groups.items():
        reasons = []
        customer_id = str(rows[0]["customer_id"]).strip() if rows[0]["customer_id"] else ""
        customer = customers.get(customer_id)
        if not customer:
            reasons.append(f"Customer '{customer_id}' not found.")
        elif customer["Active"] != "Yes":
            reasons.append(f"Customer '{customer_id}' is not active.")

        # Delivery location: consistent across all lines in one order —
        # take the first non-blank override in the group, or fall back
        # to the customer's own default.
        override_loc = next((str(r["delivery_location"]).strip() for r in rows
                             if r["delivery_location"]), None)
        resolved_loc = override_loc or (customer.get("Default_Delivery_Location") if customer else None)
        if not resolved_loc:
            reasons.append("No delivery location given, and this customer has no default "
                          "delivery location on file — provide one explicitly.")
        elif resolved_loc not in locations:
            reasons.append(f"Delivery location '{resolved_loc}' is not a valid, active location.")

        req_date = rows[0]["req_date"]
        req_date_str = None
        if not req_date:
            reasons.append("Requested delivery date is required.")
        else:
            if isinstance(req_date, datetime):
                req_date_str = req_date.date().isoformat()
            elif isinstance(req_date, date):
                req_date_str = req_date.isoformat()
            else:
                try:
                    req_date_str = datetime.strptime(str(req_date).strip(), "%Y-%m-%d").date().isoformat()
                except ValueError:
                    reasons.append(f"Requested delivery date '{req_date}' is not a valid YYYY-MM-DD date.")
            if req_date_str and req_date_str < date.today().isoformat():
                reasons.append(f"Requested delivery date {req_date_str} is in the past.")

        line_items = []
        for i, r in enumerate(rows, 1):
            mat_code = str(r["mat_code"]).strip() if r["mat_code"] else ""
            item = materials.get(mat_code)
            if not item:
                reasons.append(f"Line {i}: material '{mat_code}' not found or not active.")
                continue
            try:
                qty = float(r["qty"])
                if qty <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                reasons.append(f"Line {i}: quantity '{r['qty']}' must be a positive number.")
                continue
            unit_price = item["price"]
            if r["unit_price"] not in (None, ""):
                try:
                    unit_price = float(r["unit_price"])
                except (TypeError, ValueError):
                    reasons.append(f"Line {i}: unit price '{r['unit_price']}' is not a valid number.")
                    continue
            line_items.append({"mat_code": mat_code, "mat_desc": item["desc"],
                               "uom": item["uom"], "qty": qty, "unit_price": unit_price})

        if reasons:
            rejected.append({"order_ref": order_ref, "reasons": reasons})
            continue

        loc_geo = locations[resolved_loc]["geo"]
        result = so.create_direct_order(customer_id, line_items, delivery_location=resolved_loc,
                                        delivery_geo=loc_geo, requested_delivery_date=req_date_str,
                                        notes=f"Bulk-imported ({order_ref})", data_file=fpath)
        accepted.append({"order_ref": order_ref, "so_id": result["so_id"], "status": result["status"]})

    return {"accepted": accepted, "rejected": rejected}
