"""
goods_receipt.py — Goods Receipt (S2C, manufacturing pilot).

Deliberately lightweight, per direct instruction: a full production-grade
receiving system (put-away, lot/serial tracking, warehouse bins) already
exists outside this PoC's scope. What this module provides is the minimum
real foundation the next two steps need: a genuine GR event, with actual
received quantities distinct from ordered quantities, that Three-Way
Match and Quality Inspection can both anchor to.

Mirrors fulfillment.py's shape on purpose (that's the O2C side's "goods
moving out" event; this is S2C's "goods moving in" event) — Qty_Ordered
vs Qty_Received tracked separately, genuine partial-receipt support (a
vendor can under-ship, or ship across several GRs against the same PO),
nothing forced to reconcile that hasn't actually happened.

Reads PO_Header/PO_Items by column NAME, not position — pr_consolidation.py's
PO_ITM_COLS has changed shape before (Unit_Price was added mid-project),
and a data.xlsx that's never had a PO consolidated against it yet won't
reflect the current schema until it does. Name-based lookup is resilient
to that; position-based lookup would silently misread columns.

New sheets:
  GR_Header   GR_ID | PO_Number | Vendor_ID | Vendor_Name | GR_Date |
              Status | Delivery_Location | Received_By | Notes
  GR_Items    GR_ID | Line_Item | Material_Code | Material_Desc | UOM |
              PO_Qty | Qty_Received | Unit_Price | Notes

SQLite pilot: GR_Header and GR_Items now live in erp_pilot.db (tables
`gr_header`, `gr_items`). goods_receipt.py is the exclusive owner —
quality_inspection.py reads GR_Items but delegates to this module's
get_gr_items_index() rather than reading the table directly (see that
module for why this matters).

Active-flag policy (warn-only, per direct instruction): create_gr()
checks whether each received material is currently Active and appends
any concerns to that GR line's notes — it never blocks the receipt. A
GR is nearly always completing a commitment made when the item WAS
active (the PO already exists), so refusing to record physical goods
that have actually arrived would make the system actively wrong, not
safer.
"""

import os, io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import pr_consolidation as pc
import inventory as inv

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

GR_HDR_SHEET = "GR_Header"
GR_ITM_SHEET = "GR_Items"

GR_HDR_COLS = ["GR_ID", "PO_Number", "Vendor_ID", "Vendor_Name", "GR_Date",
              "Status", "Delivery_Location", "Received_By", "Notes"]
GR_ITM_COLS = ["GR_ID", "Line_Item", "Material_Code", "Material_Desc", "UOM",
              "PO_Qty", "Qty_Received", "Unit_Price", "Notes"]


def ensure_sheets(wb=None):
    """Kept for signature compatibility — GR_Header/GR_Items no longer
    live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


# ── Read PO data (name-based lookup — schema-resilient) ─────────────────────────
# All data-row iteration below uses iter_rows() (sequential), not .cell(r,c)
# (random access) — measured the latter at 277x slower in read-only mode on
# this file, which is what caused BOM's explosion to hang (net_against_open_pos
# calls these once per leaf material, so the difference compounds badly).
def _po_header_row(po_number, data_file=None):
    """PO_Header now lives in SQLite (pr_consolidation.py's pilot) —
    delegates and re-shapes into the same Capitalized-key dict this
    function has always returned, so every downstream .get("Supplier_ID")
    etc. call in this file keeps working unchanged."""
    row = pc.get_po_header(po_number)
    if not row:
        return None
    return {"PO_Number": row["po_number"], "PO_Type": row["po_type"],
            "Legal_Entity": row["legal_entity"], "Purchase_Entity": row["purchase_entity"],
            "Purchasing_Group": row["purchasing_group"], "Currency": row["currency"],
            "Plant_code": row["plant_code"], "Supplier_ID": row["supplier_id"],
            "Supplier_Name": row["supplier_name"], "Supplier_Geolocation": row["supplier_geolocation"],
            "Status": row["status"]}


def _po_item_rows(po_number, data_file=None):
    """PO_Items now lives in SQLite — same re-shaping approach as
    _po_header_row() above."""
    rows = pc.get_po_items(po_number)
    return [{"PO_Number": r["po_number"], "PO_Item": r["po_item"],
             "Material_Code": r["material_code"], "Material_Desc": r["material_desc"],
             "UOM": r["uom"], "Quantity": r["quantity"], "Unit_Price": r["unit_price"],
             "Delivery_Date": r["delivery_date"], "Delivery_Location": r["delivery_location"],
             "Delivery_Geolocation": r["delivery_geolocation"],
             "Source_PR_Number": r["source_pr_number"], "Source_PR_Line_Item": r["source_pr_line_item"],
             "Requester_ID": r["requester_id"], "Requester_Dept": r["requester_dept"],
             "Project_ID": r["project_id"]}
            for r in rows]


def _all_po_numbers(data_file=None):
    """PO_Header now lives in SQLite — same re-shaping approach."""
    rows = pc.get_all_po_headers()
    return [{"PO_Number": r["po_number"], "Supplier_ID": r["supplier_id"],
             "Supplier_Name": r["supplier_name"], "Status": r["status"]} for r in rows]


# ── Read GR data ──────────────────────────────────────────────────────────────
def get_grs(status=None, po_number=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM gr_header ORDER BY gr_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"gr_id": r["gr_id"], "po_number": r["po_number"], "vendor_id": r["vendor_id"],
               "vendor_name": r["vendor_name"], "gr_date": r["gr_date"], "status": r["status"],
               "delivery_location": r["delivery_location"], "received_by": r["received_by"],
               "notes": r["notes"]}
        if status and row["status"] != status:
            continue
        if po_number and row["po_number"] != po_number:
            continue
        out.append(row)
    return out


def get_gr(gr_id, data_file=None):
    for g in get_grs(data_file=data_file):
        if g["gr_id"] == gr_id:
            return g
    return None


def get_gr_items(gr_id, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM gr_items WHERE gr_id = ? ORDER BY line_item", (gr_id,)
        ).fetchall()
    finally:
        conn.close()
    return [{"gr_id": r["gr_id"], "line_item": r["line_item"], "po_item": r["po_item"],
             "mat_code": r["material_code"],
             "mat_desc": r["material_desc"], "uom": r["uom"], "po_qty": r["po_qty"],
             "qty_received": r["qty_received"], "unit_price": r["unit_price"],
             "notes": r["notes"]} for r in rows]


def get_gr_items_index(data_file=None):
    """{gr_id: [items]} for every GR — loaded once. quality_inspection.py
    delegates to this instead of reading gr_items directly, so GR_Items
    has exactly one reader implementation (this module), the same
    "single owner" principle applied everywhere else in this migration."""
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM gr_items ORDER BY gr_id, line_item").fetchall()
    finally:
        conn.close()
    index = {}
    for r in rows:
        item = {"gr_id": r["gr_id"], "line_item": r["line_item"], "po_item": r["po_item"],
                "mat_code": r["material_code"],
                "mat_desc": r["material_desc"], "uom": r["uom"], "po_qty": r["po_qty"],
                "qty_received": r["qty_received"], "unit_price": r["unit_price"],
                "notes": r["notes"]}
        index.setdefault(item["gr_id"], []).append(item)
    return index


# ── PO receipt status — computed live, matches every other _compute_status() ──
def get_po_receipt_status(po_number, data_file=None):
    """
    One row per PO LINE, not per material — a real bug fixed here: this
    used to aggregate received quantity by material_code across the
    whole PO, which silently conflated two different PO lines that
    happen to share a material code (now possible since PR consolidation
    can legitimately produce two lines for the same item at two
    different delivery locations). Receiving against one such line used
    to make BOTH lines look received. Now keyed by po_item, the one
    thing that's always unique per PO line.
    """
    fpath = data_file or DATA_FILE
    po_items = _po_item_rows(po_number, fpath)
    if not po_items:
        return []
    grs = [g for g in get_grs(po_number=po_number, data_file=fpath) if g["status"] != "Cancelled"]
    received_by_po_item = {}
    for g in grs:
        for item in get_gr_items(g["gr_id"], fpath):
            if item["po_item"] is not None:
                received_by_po_item[item["po_item"]] = (
                    received_by_po_item.get(item["po_item"], 0) + (item["qty_received"] or 0))

    out = []
    for pi in po_items:
        po_item_no = pi.get("PO_Item")
        mat = pi.get("Material_Code")
        po_qty = float(pi.get("Quantity") or 0)
        received = float(received_by_po_item.get(po_item_no, 0))
        outstanding = round(po_qty - received, 3)
        if received <= 0:
            recv_status = "Not Received"
        elif outstanding > 0.001:
            recv_status = "Partially Received"
        else:
            recv_status = "Fully Received"
        out.append({"po_item": po_item_no, "mat_code": mat, "mat_desc": pi.get("Material_Desc"),
                    "uom": pi.get("UOM"), "unit_price": pi.get("Unit_Price"), "po_qty": po_qty,
                    "received_qty": received, "outstanding_qty": max(outstanding, 0),
                    "receipt_status": recv_status, "deliv_loc": pi.get("Delivery_Location")})
    return out


def get_receivable_pos(data_file=None):
    """POs with at least one line not yet fully received — the natural
    candidate list for creating a new GR. Only POs at status='Created'
    are eligible — a PO still at 'Proposed' (written by Consolidate but
    never actually sent to the vendor) has no business being received
    against; nothing's been sent for the vendor to fulfill yet. This is
    the actual enforcement point: create_gr() below checks the same
    thing defensively, but this is what keeps a Proposed PO from ever
    showing up as an option in the first place."""
    fpath = data_file or DATA_FILE
    out = []
    for po in _all_po_numbers(fpath):
        if po.get("Status") != "Created":
            continue
        po_number = po.get("PO_Number")
        lines = get_po_receipt_status(po_number, fpath)
        outstanding_lines = [l for l in lines if l["receipt_status"] != "Fully Received"]
        if outstanding_lines:
            out.append({"po_number": po_number, "vendor_id": po.get("Supplier_ID"),
                        "vendor_name": po.get("Supplier_Name"), "lines": len(lines),
                        "outstanding_lines": len(outstanding_lines)})
    return out


# ── Create ────────────────────────────────────────────────────────────────────
def _next_gr_id(conn):
    rows = conn.execute("SELECT gr_id FROM gr_header WHERE gr_id LIKE 'GR-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["gr_id"].split("-")[1]))
        except Exception: pass
    return f"GR-{mx+1:05d}"


def _parse_source_prs(source_pr_number, source_pr_line_item):
    """PO_Items' Source_PR_Number/Source_PR_Line_Item are comma-separated
    when consolidated mode merged multiple PR lines into one PO line
    (e.g. 'PR-001, PR-002' / '1, 1') — confirmed against real output
    before writing this, not assumed. Single-source PO lines (1:1 mode,
    or a consolidated line with only one contributor) have no comma and
    still parse correctly as a one-element list."""
    prs = [p.strip() for p in str(source_pr_number or "").split(",") if p.strip()]
    lines = [l.strip() for l in str(source_pr_line_item or "").split(",") if l.strip()]
    if len(prs) != len(lines):
        return []
    out = []
    for pr, ln in zip(prs, lines):
        try:
            out.append((pr, int(float(ln))))
        except ValueError:
            continue
    return out


def _load_pr_items_index(data_file=None):
    """PR_Items now lives in SQLite — delegates to
    pr_consolidation.get_pr_items_index(), same {(pr_number, pr_line):
    {requested_qty, accepted_qty}} shape as before."""
    return pc.get_pr_items_index()


def _read_pr_line(pr_number, pr_line, data_file):
    """Single-line convenience wrapper — still correct, but prefer
    _load_pr_items_index() directly when looking up more than one line
    (get_po_lines_source_prs() below does exactly that)."""
    return _load_pr_items_index(data_file).get((pr_number, pr_line))


def get_po_line_source_prs(po_number, po_item, data_file=None):
    """Every PR line that feeds this PO line, with requested and
    already-accepted quantities — what the UI needs to show before a
    human decides whether the default proportional split is right or
    needs overriding. For rendering an entire PO's worth of lines, use
    get_po_lines_source_prs() instead — it loads PR_Items once for all
    lines rather than once per line.

    Keyed by po_item (the PO line number), not material_code — a real
    bug fixed here: two PO lines can now legitimately share a material
    code (see _build_po_lines' consolidated-mode fix), so material_code
    alone can no longer identify a single line."""
    fpath = data_file or DATA_FILE
    po_items = _po_item_rows(po_number, fpath)
    line = next((pi for pi in po_items if pi.get("PO_Item") == po_item), None)
    if line is None:
        return []
    pairs = _parse_source_prs(line.get("Source_PR_Number"), line.get("Source_PR_Line_Item"))
    pr_index = _load_pr_items_index(fpath)
    out = []
    for pr_number, pr_line in pairs:
        detail = pr_index.get((pr_number, pr_line))
        if detail:
            out.append({"pr_number": pr_number, "pr_line": pr_line, **detail})
    return out


def get_po_lines_source_prs(po_number, data_file=None):
    """Batch version — computes source PRs for EVERY line of a PO in one
    pass (one PO_Items read, one PR_Items read), not one call per line.
    Returns {po_item: [source_pr_details]} — keyed by PO line number,
    not material_code (see get_po_line_source_prs()'s docstring for
    why). This is what the Goods Receipt UI should call once per page
    render, looping over the result in memory, instead of calling
    get_po_line_source_prs() once per line — that was the actual N+1
    pattern causing the 2.8s render time."""
    fpath = data_file or DATA_FILE
    po_items = _po_item_rows(po_number, fpath)
    pr_index = _load_pr_items_index(fpath)
    out = {}
    for pi in po_items:
        po_item_no = pi.get("PO_Item")
        pairs = _parse_source_prs(pi.get("Source_PR_Number"), pi.get("Source_PR_Line_Item"))
        lines = []
        for pr_number, pr_line in pairs:
            detail = pr_index.get((pr_number, pr_line))
            if detail:
                lines.append({"pr_number": pr_number, "pr_line": pr_line, **detail})
        out[po_item_no] = lines
    return out


def compute_default_allocation(source_prs, qty_to_allocate):
    """
    Proportional-by-requested-qty split — the default rule, not the only
    possible one. Per direct instruction, a real shortage might warrant a
    different split (e.g. by PR criticality), which this app has no data
    field for yet — this computes the honest default, and create_gr()'s
    pr_allocations parameter lets a human override it per line rather
    than being stuck with proportional-only.
    source_prs: get_po_line_source_prs()'s output. Returns
    {(pr_number, pr_line): allocated_qty}.
    """
    if not source_prs:
        return {}
    total_requested = sum(s["requested_qty"] for s in source_prs)
    if total_requested <= 0:
        # No basis to weight by — split evenly rather than divide by zero.
        share = qty_to_allocate / len(source_prs)
        return {(s["pr_number"], s["pr_line"]): share for s in source_prs}
    return {(s["pr_number"], s["pr_line"]): round(qty_to_allocate * s["requested_qty"] / total_requested, 3)
            for s in source_prs}


def _apply_pr_line_acceptances_batch(allocations, data_file=None):
    """PR_Items now lives in SQLite — delegates to
    pr_consolidation.apply_pr_line_acceptances_batch(), same one-
    transaction batching benefit as before (one SQLite transaction
    instead of one workbook open/save)."""
    pc.apply_pr_line_acceptances_batch(allocations)


def _apply_pr_line_acceptance(pr_number, pr_line, additional_qty, data_file=None):
    """Single-allocation convenience wrapper — prefer
    _apply_pr_line_acceptances_batch() when applying more than one.
    (The pre-migration version of this function called `wb.save(fpath)`
    after delegating to the batch version, referencing a `wb` that was
    never defined in this function's scope — a NameError waiting to
    happen if this wrapper were ever actually called. Not something the
    SQLite migration introduced; fixed here since the whole function
    was being rewritten anyway.)"""
    _apply_pr_line_acceptances_batch([(pr_number, pr_line, additional_qty)], data_file)


def get_pr_fulfillment_status(pr_number, data_file=None):
    """Requested vs accepted per line of a PR — the PR-side mirror of
    get_po_receipt_status(), live-computed the same way. PR_Items now
    lives in SQLite — delegates to pr_consolidation.get_pr_items_for_pr()."""
    out = []
    for row in pc.get_pr_items_for_pr(pr_number):
        requested = row["quantity"] or 0
        accepted = row["quantity_accepted"] or 0
        outstanding = max(0, round(requested - accepted, 3))
        if accepted <= 0:
            status = "Not Fulfilled"
        elif outstanding > 0.001:
            status = "Partially Fulfilled"
        else:
            status = "Fully Fulfilled"
        out.append({"pr_line": row["pr_line_item"], "mat_code": row["material_code"],
                    "mat_desc": row["material_desc"], "requested_qty": requested,
                    "accepted_qty": accepted, "outstanding_qty": outstanding, "status": status})
    return out


def create_gr(po_number, line_receipts, delivery_location="", received_by="", notes="",
              pr_allocations=None, data_file=None):
    """
    line_receipts: {po_item: qty_received} — keyed by PO line number, NOT
    material code. This changed from material-code-keying after a real
    bug: PR consolidation can legitimately produce two PO lines sharing
    a material code (same item, two different delivery locations,
    merged under the same vendor) — see _build_po_lines' fix. Only
    po_item is guaranteed unique per PO line; material_code no longer
    is. Only lines with a positive quantity are recorded — a line left
    at 0 (or omitted) simply isn't part of this GR, no need to
    explicitly zero it out.

    pr_allocations: optional {po_item: [(pr_number, pr_line, qty), ...]}
    — overrides the default proportional-by-requested-qty split for that
    line's source PR(s). Omit a line entirely to use the default for
    it. This is how a human says "80/40 isn't right, PR-002 is more
    urgent, make it 60/60 instead" — the default is a starting point,
    not the only option.
    """
    fpath = data_file or DATA_FILE
    po = _po_header_row(po_number, fpath)
    if po is None:
        raise ValueError(f"PO {po_number} not found.")
    if po.get("Status") != "Created":
        raise ValueError(f"{po_number} is still at status '{po.get('Status')}' — a PO has "
                          "to be Created (its export file actually generated and sent to "
                          "the vendor) before anything can be received against it.")
    po_items = _po_item_rows(po_number, fpath)
    if not po_items:
        raise ValueError(f"PO {po_number} has no line items.")

    # BUGFIX: when a manual allocation override is given for a line, its
    # sum is the one true received quantity for that line — not a second,
    # independently-trusted number. Previously line_receipts and
    # pr_allocations could disagree (e.g. line_receipts still held the
    # default outstanding qty while the user had edited the allocation
    # breakdown to a different total), and GR_Items recorded the wrong one
    # while the PR-side allocation recorded the right one. Reconciling here
    # makes that divergence structurally impossible.
    reconciled_receipts = dict(line_receipts)
    for po_item_no, override in (pr_allocations or {}).items():
        if override:
            reconciled_receipts[po_item_no] = round(sum(q for _, _, q in override), 3)

    lines_to_record = [(pi, reconciled_receipts.get(pi.get("PO_Item"), 0)) for pi in po_items]
    lines_to_record = [(pi, qty) for pi, qty in lines_to_record if qty and qty > 0]
    if not lines_to_record:
        raise ValueError("Enter a received quantity greater than zero on at least one line.")

    db.init_schema()
    conn = db.get_connection()
    try:
        gr_id = _next_gr_id(conn)
        conn.execute(
            "INSERT INTO gr_header (gr_id, po_number, vendor_id, vendor_name, gr_date, "
            "status, delivery_location, received_by, notes) VALUES (?,?,?,?,?,?,?,?,?)",
            (gr_id, po_number, po.get("Supplier_ID"), po.get("Supplier_Name"),
             date.today().strftime("%Y-%m-%d"), "Posted", delivery_location, received_by, notes),
        )
        for seq, (pi, qty) in enumerate(lines_to_record, 1):
            mat_code = pi.get("Material_Code")
            line_notes = ""
            master = conn.execute(
                "SELECT active FROM item_master WHERE item_code = ?", (mat_code,)
            ).fetchone()
            if master is not None and (master["active"] or "Yes") != "Yes":
                line_notes = f"Material {mat_code} is currently marked inactive."
            conn.execute(
                "INSERT INTO gr_items (gr_id, line_item, po_item, material_code, material_desc, "
                "uom, po_qty, qty_received, unit_price, notes) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (gr_id, seq, pi.get("PO_Item"), mat_code, pi.get("Material_Desc"), pi.get("UOM"),
                 pi.get("Quantity"), qty, pi.get("Unit_Price"), line_notes),
            )
        conn.commit()
    finally:
        conn.close()

    # Allocate accepted quantity back to source PR line(s) — default
    # proportional split, or the caller's explicit override per PO line.
    # Batched: one source-PR lookup for the whole PO, one PR_Items
    # write for every allocation, instead of one call each per line —
    # that was the confirmed severe bottleneck (up to ~7.8s for a 7-line,
    # 6-way-consolidated GR before this fix).
    source_prs_by_line = get_po_lines_source_prs(po_number, fpath)
    all_allocations = []
    for pi, qty in lines_to_record:
        po_item_no = pi.get("PO_Item")
        source_prs = source_prs_by_line.get(po_item_no, [])
        if not source_prs:
            continue  # no traceable source PR line(s) — nothing to allocate
        override = (pr_allocations or {}).get(po_item_no)
        if override:
            allocation = {(pr, ln): q for pr, ln, q in override}
        else:
            allocation = compute_default_allocation(source_prs, qty)
        for (pr_number, pr_line), alloc_qty in allocation.items():
            if alloc_qty:
                all_allocations.append((pr_number, pr_line, alloc_qty))
    _apply_pr_line_acceptances_batch(all_allocations, fpath)

    # Post stock-in to the inventory ledger. `delivery_location`, if given,
    # is a genuine override of wherever the PO's own line said goods would
    # land — for the real case of a shipment actually arriving somewhere
    # different from what the PO specified. Left blank, this falls back to
    # EACH line's own PO Delivery_Location independently — correct even
    # when a single GR spans multiple PO lines with genuinely different
    # locations (a multi-location consolidated PO). The override, if
    # given, is intentionally uniform across the whole GR — it models
    # "this entire delivery physically arrived somewhere else," not a
    # per-line correction; a GR that's receiving against a multi-location
    # PO and needs a DIFFERENT correction per line isn't something a
    # single override field can express, and isn't attempted here.
    # (Previously this parameter was accepted but silently ignored for
    # the actual inventory posting — it only ever reached GR_Header's own
    # delivery_location column, a cosmetic field with no effect on where
    # stock landed. Every received line still increases stock regardless
    # of whether it traces back to a PR — unlike PR allocation above, this
    # isn't conditional.) data_file=fpath passed explicitly — without it
    # this silently falls back to inventory.py's own default DATA_FILE
    # instead of whatever file this GR was actually posted against (same
    # bug class found and fixed in production.py's record_transaction calls).
    for pi, qty in lines_to_record:
        loc = delivery_location or pi.get("Delivery_Location") or "Unspecified"
        inv.record_transaction(pi.get("Material_Code"), pi.get("Material_Desc"), loc, qty,
                               "GR Receipt", "GR", gr_id, data_file=fpath)

    # ATP-US-03's own real supply-arrival trigger: a GR is one of the two
    # real events (the other being inventory.receive_transfer()) that
    # genuinely increases on-hand for a material at a Plant, so every
    # Open/Partially Fulfilled Backorder for each (material, location)
    # this GR touched gets a real, automatic re-evaluation attempt here
    # — never a batch job, never something a Planner has to remember to
    # trigger by hand.
    import backorder as bo
    for pi, qty in lines_to_record:
        loc = delivery_location or pi.get("Delivery_Location") or "Unspecified"
        bo.reevaluate_backorders(pi.get("Material_Code"), loc)

    return gr_id


def get_po_delivery_location(po_number, data_file=None):
    """The PO's own delivery location, for defaulting the GR form's
    override field to something sensible rather than blank. Returns the
    first line's Delivery_Location — in practice every line on a given
    PO shares the same one (pr_consolidation.py sets it uniformly per
    PO), so "first line" is effectively "the PO's location", not an
    arbitrary pick."""
    items = _po_item_rows(po_number, data_file)
    for pi in items:
        if pi.get("Delivery_Location"):
            return pi["Delivery_Location"]
    return None


def cancel_gr(gr_id, reason="", data_file=None):
    conn = db.get_connection()
    try:
        row = conn.execute("SELECT notes FROM gr_header WHERE gr_id = ?", (gr_id,)).fetchone()
        if row is not None:
            existing = row["notes"] or ""
            new_notes = f"{existing} | Cancelled: {reason}".strip(" |")
            conn.execute(
                "UPDATE gr_header SET status='Cancelled', notes=? WHERE gr_id=?",
                (new_notes, gr_id),
            )
            conn.commit()
    finally:
        conn.close()


def is_ready_for_quality_check(gr_id, data_file=None):
    """The hook quality_inspection.py calls."""
    g = get_gr(gr_id, data_file)
    return g is not None and g["status"] != "Cancelled"


# ── Document generation ───────────────────────────────────────────────────────
def generate_gr_note(gr_id, data_file=None):
    fpath = data_file or DATA_FILE
    g = get_gr(gr_id, fpath)
    if g is None:
        raise ValueError(f"{gr_id} not found.")
    items = get_gr_items(gr_id, fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = "14532D"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Goods Receipt Note"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "GOODS RECEIPT NOTE")
    ws.merge_cells("A1:G1")
    label("A3", "GR No:");    value("B3", g["gr_id"])
    label("A4", "Date:");     value("B4", g["gr_date"])
    label("A5", "Received By:"); value("B5", g["received_by"] or "")
    label("D3", "PO Number:"); value("E3", g["po_number"])
    label("D4", "Vendor:");    value("E4", g["vendor_name"])
    label("D5", "Deliver To:"); value("E5", g["delivery_location"] or "")

    hdr_row = 8
    hdrs = ["#", "Material Code", "Description", "UOM", "PO Qty", "Qty Received", "Variance"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=green)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for i, item in enumerate(items, 1):
        variance = round((item["qty_received"] or 0) - (item["po_qty"] or 0), 3)
        vals = [i, item["mat_code"], item["mat_desc"], item["uom"],
                item["po_qty"], item["qty_received"], variance if variance else ""]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9,
                          color="B91C1C" if (ci == 7 and variance and variance < 0) else "1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 3
    ws.cell(r, 1, "Received in good order by:").font = Font(name="Arial", size=10, bold=True)
    r += 3
    ws.cell(r, 1, "Signature: ____________________").font = Font(name="Arial", size=9)
    r += 1
    ws.cell(r, 1, "Name / Date: ____________________").font = Font(name="Arial", size=9)

    widths = [4, 15, 32, 8, 10, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{gr_id}_{g['po_number']}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    grs = get_grs(data_file=data_file)
    by_status = {}
    for g in grs:
        by_status[g["status"]] = by_status.get(g["status"], 0) + 1
    return {"total": len(grs), "by_status": by_status}


if __name__ == "__main__":
    print("Goods Receipt stats:", stats())
