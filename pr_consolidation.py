"""
pr_consolidation.py — PR -> PO consolidation.

SQLite pilot (BOM -> PR -> Consolidation -> PO): PR_Header, PR_Items,
PO_Header, PO_Items, and RFP now live in erp_pilot.db (tables
pr_header, pr_items, po_header, po_items, rfp) — not in data.xlsx.

This is the CORE ENGINE slice of the pilot, and a harder one than
bom.py: unlike BOM_Items (bom.py's exclusive domain), these 5 tables
are also written by rfx.py's generate_pos() (POs issued from RFx
awards) and read/written by goods_receipt.py (PR acceptance
quantities) and contracts.py (reading POs to convert to contracts).
Verified before writing any code: goods_receipt.py, contracts.py,
customer_onboarding.py, fulfillment.py, inventory.py, org_profile.py,
and sales_order.py all import this module, but ONLY for the generic
Excel-styling helpers below (_write_hdrs, _cell) — none of them touch
PR/PO/RFP data directly except goods_receipt.py and contracts.py
(handled explicitly) and rfx.py (handled explicitly, including its
shared PR_Items/PO writeback).

All PR/PO/RFP reads and writes for this table set now go through the
functions in this module rather than being duplicated in each
consumer — the same reasoning that made a full clear-and-rebuild of
PO_Header a real bug before: two independent writers (or readers) of
the same data is exactly how that class of bug happens. In
particular, PO numbering and the PR_Items "mark as ordered" writeback
used to be implemented independently in both this module's run() and
rfx.py's generate_pos() — next_po_number()/mark_pr_items_ordered()
below are now the one shared implementation both call.

Data still in Excel (unaffected by this migration): Vendor_Master,
Delivery_Locations, Item Master — reference data read here via
openpyxl, same as always. `data_file` parameters throughout continue
to mean "the Excel path for that reference data" — never a SQLite
path. Keeping one, unambiguous meaning per parameter is what avoids
the "which kind of path" class of bug this project keeps finding
(same design decision made for bom.py's migration).

Excel-styling helpers below (_hdr, _cell, _write_hdrs, _autofit, ...)
are NOT specific to PR/PO/RFP — they're generic cell-formatting
utilities used by other modules for their OWN, unrelated Excel sheets.
They are untouched by this migration.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, io
from datetime import datetime, date, timedelta
from collections import defaultdict

import db
import vendor_onboarding as vo
import po_export

ONE_TO_ONE_PR_PO_LINES = True

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

_thin = Side(style="thin", color="CBD5E1")
_bdr  = Border(left=_thin,right=_thin,top=_thin,bottom=_thin)
def _fill(c):  return PatternFill("solid", fgColor=c)
def _left(w=False): return Alignment(horizontal="left",vertical="center",wrap_text=w)
def _center():   return Alignment(horizontal="center",vertical="center")
def _right():    return Alignment(horizontal="right",vertical="center")

def _hdr(cell,bg="1E3A5F",fg="FFFFFF"):
    cell.font=Font(name="Arial",size=9,bold=True,color=fg)
    cell.fill=_fill(bg); cell.alignment=_center(); cell.border=_bdr

def _cell(cell,bg=None,align=None,bold=False,color="1A1A2E"):
    cell.font=Font(name="Arial",size=9,bold=bold,color=color)
    cell.alignment=align or _left(); cell.border=_bdr
    if bg: cell.fill=_fill(bg)

def _write_hdrs(ws,row,cols,bg="1E3A5F"):
    ws.row_dimensions[row].height=18
    for ci,col in enumerate(cols,1): _hdr(ws.cell(row,ci,col),bg=bg)

def _autofit(ws):
    for col in ws.columns:
        mx=max((len(str(c.value or "")) for c in col),default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(mx+4,50)

def _parse_ll(s):
    try:
        p=str(s or "").split(",")
        if len(p)==2: return float(p[0].strip()),float(p[1].strip())
    except Exception: pass
    return None


def _parse_date(s):
    """Parses a stored delivery_date string ('YYYY-MM-DD') into a real
    date, or None if it's blank/unparseable — a line with no delivery
    date set is excluded by any date-range filter rather than crashing
    it, matching how the PO Line Items list already treats this case."""
    try:
        return datetime.strptime(str(s or "").strip()[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


# ── Reference data (SQLite now — Vendor_Master/Delivery_Locations migrated) ──
def _read_vendor_master():
    """Returns {vendor_id: geo_string}, active vendors only. Delegates to
    vendor_onboarding.py — the canonical Vendor_Master owner — instead
    of reading Excel directly. This closes a real inconsistency that
    predated this migration: this function never filtered on Active at
    all (rfx.py had its own separate copy that did), so a deactivated
    vendor could still appear on the PO consolidation map. There is now
    exactly one implementation of "read Vendor_Master," and it filters
    consistently everywhere it's used."""
    return {v["Vendor_ID"]: v["Geolocation"] for v in vo.list_vendors(include_inactive=False)
            if v.get("Geolocation")}


def get_delivery_locations(active_only=True):
    """
    Canonical Delivery_Locations reader — Delivery_Locations now lives
    in SQLite. Previously read independently (and inconsistently — none
    of the three filtered on Active despite the column existing) by
    erp_ui.py, mfg_ui.py, o2c_ui.py, and this module; all four now
    delegate here. Returns a list of dicts: id, name, geo, city, state,
    country, address, active.
    """
    conn = db.get_connection()
    try:
        if active_only:
            rows = conn.execute(
                "SELECT * FROM delivery_locations WHERE active = 'Yes' ORDER BY location_id"
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM delivery_locations ORDER BY location_id").fetchall()
    finally:
        conn.close()
    return [{"id": r["location_id"], "name": r["location_name"], "geo": r["geolocation"],
             "city": r["city"], "state": r["state"], "country": r["country"],
             "address": r["address"], "active": r["active"]} for r in rows]


def _read_delivery_master():
    """Returns {location_id: geo_string}, active locations only."""
    return {l["id"]: l["geo"] for l in get_delivery_locations(active_only=True) if l["geo"]}


# ── SQLite: PR_Header / PR_Items ─────────────────────────────────────────────
def next_pr_number(data_file=None):
    """Same PR-{YYYYMMDD}-{NNN} convention used everywhere a PR gets
    created (Create PR, BOM proposal) — single source of truth now.
    Previously erp_ui.py and bom.py each scanned Excel independently for
    this; both now call this one function instead, closing the risk of
    the two diverging (the same risk class as PO numbering below)."""
    conn = db.get_connection()
    try:
        today = date.today().strftime("%Y%m%d")
        prefix = f"PR-{today}-"
        rows = conn.execute(
            "SELECT pr_number FROM pr_header WHERE pr_number LIKE ?", (prefix + "%",)
        ).fetchall()
    finally:
        conn.close()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["pr_number"][len(prefix):]))
        except (ValueError, IndexError): pass
    return f"{prefix}{mx+1:03d}"


def create_pr(pr_number, requester_id="", requester_name="", requester_dept="",
              project_id="", po_type="NB", legal_entity="LE-001",
              purchase_entity="PE-001", purchasing_group="PG-001",
              currency="INR", plant_code="PLANT-01", lines=None, pr_date=None, data_file=None):
    """
    Writes one pr_header row + N pr_items rows in a single transaction.
    lines: list of dicts with mat_code, mat_desc, uom, qty, and
    optionally vendor/req_date/deliv_loc/deliv_geo per line — the shape
    Create PR's staged items and bom.py's exploded/netted lines both
    already produce. Replaces erp_ui.py's old write_pr() (direct
    openpyxl write) and bom.py's inline PR_Header/PR_Items writing.

    pr_date defaults to today, same pattern as every other date field
    in this app — a real column now (added 2026-07-29), not parsed out
    of pr_number at read time. The PR number itself still usually
    embeds today's date too (see next_pr_number()), but that's now just
    an ID convention, not the only place the date lives; pass pr_date
    explicitly to backdate a PR the same way GR/SO/etc. already support.
    `data_file` accepted for signature compatibility; unused (see
    module docstring).
    """
    db.init_schema()
    conn = db.get_connection()
    try:
        conn.execute(
            "INSERT INTO pr_header (pr_number, requester_id, requester_name, "
            "requester_dept, project_id, po_type, legal_entity, purchase_entity, "
            "purchasing_group, currency, plant_code, pr_date) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (pr_number, requester_id, requester_name, requester_dept, project_id,
             po_type, legal_entity, purchase_entity, purchasing_group, currency, plant_code,
             pr_date or date.today().isoformat()),
        )
        for line_num, line in enumerate(lines or [], 1):
            conn.execute(
                "INSERT INTO pr_items (pr_number, pr_line_item, preferred_vendor, "
                "material_code, material_desc, uom, quantity, required_date, "
                "delivery_location, delivery_geolocation, status) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (pr_number, line_num, line.get("vendor", ""), line["mat_code"],
                 line["mat_desc"], line["uom"], line["qty"],
                 str(line.get("req_date", "") or ""), line.get("deliv_loc", ""),
                 line.get("deliv_geo", ""), "Open"),
            )
        conn.commit()
    finally:
        conn.close()
    return pr_number


def get_pr_headers(data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM pr_header").fetchall()
    finally:
        conn.close()
    out = {}
    for r in rows:
        d = dict(r)
        out[d["pr_number"]] = {
            "req_id": d["requester_id"] or "", "req_name": d["requester_name"] or "",
            "req_dept": d["requester_dept"] or "", "project_id": d["project_id"] or "",
            "po_type": d["po_type"] or "", "legal_entity": d["legal_entity"] or "",
            "purch_entity": d["purchase_entity"] or "", "purch_group": d["purchasing_group"] or "",
            "currency": d["currency"] or "", "plant_code": d["plant_code"] or "",
            "pr_date": d["pr_date"] or "",
        }
    return out


def get_pr_items(open_only=True, data_file=None):
    """open_only=True mirrors the old _read_pr_items() (skips anything
    not 'Open', blank/None status defaults to Open). open_only=False
    mirrors _read_pr_items_all() (returns everything, for reporting)."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM pr_items ORDER BY pr_number, pr_line_item"
        ).fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        d = dict(r)
        status = d["status"] or "Open"
        if open_only and status != "Open":
            continue
        out.append({
            "pr_number": d["pr_number"], "pr_line": d["pr_line_item"],
            "vendor": d["preferred_vendor"] or "", "mat_code": d["material_code"],
            "mat_desc": d["material_desc"] or "", "uom": d["uom"] or "",
            "qty": float(d["quantity"] or 0), "req_date": d["required_date"] or "",
            "deliv_loc": d["delivery_location"] or "", "deliv_geo": d["delivery_geolocation"] or "",
            "status": status, "po_number": d["po_number"], "po_item": d["po_item"],
            "qty_accepted": d["quantity_accepted"] or 0,
        })
    return out


def get_pr_items_for_pr(pr_number, data_file=None):
    """Every line of one PR, regardless of status — what
    get_pr_fulfillment_status() needs."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM pr_items WHERE pr_number = ? ORDER BY pr_line_item", (pr_number,)
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


def get_pr_items_index(data_file=None):
    """{(pr_number, pr_line): {requested_qty, accepted_qty}} for every
    PR line — what goods_receipt.py's source-PR allocation UI needs,
    loaded once rather than once per lookup."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT pr_number, pr_line_item, quantity, quantity_accepted FROM pr_items"
        ).fetchall()
    finally:
        conn.close()
    return {
        (r["pr_number"], r["pr_line_item"]): {
            "requested_qty": r["quantity"] or 0, "accepted_qty": r["quantity_accepted"] or 0
        }
        for r in rows
    }


def apply_pr_line_acceptances_batch(allocations, data_file=None):
    """allocations: iterable of (pr_number, pr_line, qty). Applies every
    allocation in one transaction, same batching benefit the Excel
    version had (one open/save instead of N) — now one SQLite
    transaction instead of one workbook save."""
    allocations = list(allocations)
    if not allocations:
        return
    conn = db.get_connection()
    try:
        today_str = date.today().strftime("%Y-%m-%d")
        for pr_number, pr_line, qty in allocations:
            conn.execute(
                "UPDATE pr_items SET quantity_accepted = ROUND(COALESCE(quantity_accepted, 0) + ?, 3), "
                "last_accepted_date = ? WHERE pr_number = ? AND pr_line_item = ?",
                (qty, today_str, pr_number, int(pr_line)),
            )
        conn.commit()
    finally:
        conn.close()


def get_pr_report(data_file=None):
    """Flat PR+PR_Items view — what erp_ui.py's Consolidate page table
    needs. Replaces its old load_pr_data(), which read PR_Header/
    PR_Items via openpyxl directly."""
    headers = get_pr_headers(data_file)
    items = get_pr_items(open_only=False, data_file=data_file)
    out = []
    for it in items:
        hdr = headers.get(it["pr_number"], {})
        out.append({
            "PR": it["pr_number"], "Line": it["pr_line"], "Vendor": it["vendor"],
            "Code": it["mat_code"], "Description": it["mat_desc"], "UOM": it["uom"],
            "Qty": it["qty"], "Required": it["req_date"], "Status": it["status"],
            "Requester": hdr.get("req_name", ""), "PR_Date": hdr.get("pr_date", ""),
            "PO_Number": it["po_number"], "PO_Item": it["po_item"],
        })
    return out


# ── SQLite: PO_Header / PO_Items ─────────────────────────────────────────────
def next_po_number(year=None, data_file=None):
    """Single source of truth for PO numbering — both run() (PR
    consolidation) and rfx.py's generate_pos() (RFx-awarded POs) call
    this, instead of each independently scanning for the highest
    existing PO-{year}-NNNN (which is exactly how the old PO-numbering
    collision bug happened: two independent scanners, one file wiped
    between them)."""
    year = year or datetime.now().year
    conn = db.get_connection()
    try:
        prefix = f"PO-{year}-"
        rows = conn.execute(
            "SELECT po_number FROM po_header WHERE po_number LIKE ?", (prefix + "%",)
        ).fetchall()
    finally:
        conn.close()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["po_number"][len(prefix):]))
        except (ValueError, IndexError): pass
    return f"{prefix}{mx+1:04d}"


def insert_po(po_number, header, items, data_file=None):
    """
    header: dict with po_type, legal_entity, purch_entity, purch_group,
    currency, plant_code, supplier_id, supplier_name (optional),
    supplier_geo. items: list of dicts in _build_po_lines()'s output
    shape (mat_code, mat_desc, uom, qty, unit_price, deliv_date,
    deliv_loc, deliv_geo, source_pr, source_pr_line, req_id, req_dept,
    project_id). One po_header row + N po_items rows, one transaction.
    Returns the items list with po_item (sequence number) assigned —
    both run() and rfx.generate_pos() need this for their PR writeback.

    Always inserts at status='Proposed' (the schema default) — a PO
    only becomes usable downstream (Goods Receipt, and transitively
    Vendor Invoices) once mark_po_created() runs, which every creation
    path calls right after generating that PO's vendor-facing export
    file: immediately, in the same click, for Direct PO Entry and
    RFx-awarded POs (neither has a separate "send" stage to model);
    after a real, separate "Create PO" click for PR-Consolidation-
    sourced POs, the one path where a genuine gap exists between a PO
    being written and being sent. One mechanism for the transition
    (mark_po_created()), not two — simpler than also branching this
    function's own insert on which path is calling it.
    """
    db.init_schema()
    conn = db.get_connection()
    try:
        conn.execute(
            "INSERT INTO po_header (po_number, po_type, legal_entity, purchase_entity, "
            "purchasing_group, currency, plant_code, supplier_id, supplier_name, "
            "supplier_geolocation) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (po_number, header.get("po_type", ""), header.get("legal_entity", ""),
             header.get("purch_entity", ""), header.get("purch_group", ""),
             header.get("currency", ""), header.get("plant_code", ""),
             header.get("supplier_id", ""), header.get("supplier_name", ""),
             header.get("supplier_geo", "")),
        )
        out_items = []
        for seq, it in enumerate(items, 1):
            it = {**it, "po_item": seq}
            conn.execute(
                "INSERT INTO po_items (po_number, po_item, material_code, material_desc, "
                "uom, quantity, unit_price, delivery_date, delivery_location, "
                "delivery_geolocation, source_pr_number, source_pr_line_item, "
                "requester_id, requester_dept, project_id) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (po_number, seq, it["mat_code"], it["mat_desc"], it["uom"], it["qty"],
                 it.get("unit_price"), it.get("deliv_date", ""), it.get("deliv_loc", ""),
                 it.get("deliv_geo", ""), str(it.get("source_pr", "")),
                 str(it.get("source_pr_line", "")), it.get("req_id", ""),
                 it.get("req_dept", ""), it.get("project_id", "")),
            )
            out_items.append(it)
        conn.commit()
    finally:
        conn.close()
    return out_items


def get_po_header(po_number, data_file=None):
    """Raw po_header row (DB column names), or None. Consumers
    (goods_receipt.py, contracts.py) each adapt this into their own
    existing key-naming convention locally, so their public behavior
    doesn't change — only the source does."""
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT * FROM po_header WHERE po_number = ?", (po_number,)
        ).fetchone()
    finally:
        conn.close()
    return dict(row) if row else None


def get_po_items(po_number, data_file=None):
    """Raw po_items rows (DB column names) for one PO, ordered by po_item."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM po_items WHERE po_number = ? ORDER BY po_item", (po_number,)
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


def get_all_po_headers(data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM po_header").fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


def get_procurement_lead_time(material_code, data_file=None):
    """
    Extracted from get_open_po_lines_with_dates()'s own inline logic
    (2026-07-29) so bom.py's Phase 4 recommendation engine can reuse
    the identical contract-vs-Item-Master precedence without
    duplicating it — same DRY reasoning as everywhere else in this
    codebase that pulls shared logic into one place rather than two
    copies quietly drifting apart.

    Returns {lead_time_days, source} — source is 'contract' or
    'item_master_default' or 'unknown' (no lead time resolvable
    anywhere, e.g. an inactive or unrecognized material).
    """
    import contracts as ct
    contract = ct.find_active_contract_for_item(material_code, data_file)
    if contract and contract.get("lead_time_days"):
        return {"lead_time_days": contract["lead_time_days"], "source": "contract"}
    item = po_export.get_item_by_code(material_code, active_only=False)
    if item and item.get("lead_time"):
        return {"lead_time_days": item["lead_time"], "source": "item_master_default"}
    return {"lead_time_days": None, "source": "unknown"}


def get_open_po_lines_with_dates(material_code=None, data_file=None):
    """
    Per-line detail for every outstanding quantity on a Created PO —
    the date-aware input bom._batch_open_po_exposure() never had (that
    function only ever returns a flat {material: total} sum, no dates,
    no per-line detail, and — worth noting since this function
    deliberately does NOT repeat it — no status filter of its own,
    meaning it currently counts Proposed PO quantity as if it were
    already outstanding-and-ordered).

    Only Created POs are considered here: a Proposed PO hasn't been
    sent to the vendor yet, so there's no real 'sent' date the
    lead-time clock could honestly start counting from. This is a
    deliberate, narrower scope than _batch_open_po_exposure()'s own,
    not an oversight — that function is left exactly as it is; this is
    a new, separate function for time-phased planning to use instead
    of it, not a drop-in replacement.

    Lead time precedence: an active Contract's own lead_time_days (a
    real, vendor-negotiated figure) takes precedence over Item
    Master's generic default — mirroring run()'s own price-precedence
    rule (contract rate beats list price) for the same underlying
    reason: a specific commercial term beats a generic fallback
    wherever one actually exists.

    Returns a list of dicts: po_number, po_item, material_code,
    material_desc, delivery_location, outstanding_qty, po_date,
    lead_time_days, lead_time_source ('contract' / 'item_master_default'
    / 'unknown'), expected_arrival_date. expected_arrival_date is
    computed live on every call — po_date + lead_time_days — never
    stored, same discipline this whole codebase already applies to
    every other derived figure (on-hand balances included).
    """
    conn = db.get_connection()
    try:
        where = "WHERE ph.status = 'Created'"
        params = []
        if material_code:
            where += " AND pi.material_code = ?"
            params.append(material_code)
        rows = conn.execute(
            f"SELECT pi.po_number, pi.po_item, pi.material_code, pi.material_desc, "
            f"pi.quantity, pi.delivery_location, ph.po_date "
            f"FROM po_items pi JOIN po_header ph ON pi.po_number = ph.po_number "
            f"{where}",
            params,
        ).fetchall()

        active_grs = {r["gr_id"]: r["po_number"] for r in conn.execute(
            "SELECT gr_id, po_number FROM gr_header WHERE status != 'Cancelled'"
        ).fetchall()}
        received_by_line = {}
        if active_grs:
            placeholders = ",".join("?" for _ in active_grs)
            gr_rows = conn.execute(
                f"SELECT gr_id, po_item, qty_received FROM gr_items WHERE gr_id IN ({placeholders})",
                list(active_grs),
            ).fetchall()
            for r in gr_rows:
                key = (active_grs[r["gr_id"]], r["po_item"])
                received_by_line[key] = received_by_line.get(key, 0) + (r["qty_received"] or 0)
    finally:
        conn.close()

    out = []
    for r in rows:
        key = (r["po_number"], r["po_item"])
        outstanding = max(0, (r["quantity"] or 0) - received_by_line.get(key, 0))
        if outstanding <= 0:
            continue

        lt = get_procurement_lead_time(r["material_code"], data_file)
        lead_time, source = lt["lead_time_days"], lt["source"]

        expected_arrival = None
        if r["po_date"] and lead_time is not None:
            expected_arrival = (datetime.strptime(r["po_date"], "%Y-%m-%d").date() +
                               timedelta(days=int(lead_time))).isoformat()

        out.append({
            "po_number": r["po_number"], "po_item": r["po_item"],
            "material_code": r["material_code"], "material_desc": r["material_desc"],
            "delivery_location": r["delivery_location"], "outstanding_qty": outstanding,
            "po_date": r["po_date"], "lead_time_days": lead_time, "lead_time_source": source,
            "expected_arrival_date": expected_arrival,
        })
    return out


def get_all_po_line_items(data_file=None):
    """Flat listing of every PO line ever created, joined with its own
    header's vendor info — one SQL JOIN, not one get_po_items() call per
    PO. Powers erp_ui.py's 'All POs Map' panel's line-items table."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT h.po_number, h.supplier_id, h.supplier_name, h.currency, "
            "i.po_item, i.material_code, i.material_desc, i.uom, i.quantity, "
            "i.unit_price, i.delivery_date, i.delivery_location, i.source_pr_number "
            "FROM po_items i JOIN po_header h ON h.po_number = i.po_number "
            "ORDER BY i.delivery_date DESC, h.po_number, i.po_item"
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


def mark_pr_items_ordered(pairs, data_file=None):
    """pairs: iterable of (pr_number, pr_line, po_number, po_item, vendor).
    Sets status='PO Proposed' + po_number/po_item/preferred_vendor on each
    matching pr_items row, one transaction. Shared by run() (PR
    consolidation) AND rfx.py's generate_pos() (RFx-awarded PO issuance)
    — the same write to the same table, now expressed once instead of
    twice (previously run()'s _writeback_pr() and rfx.py's inline loop
    were two independent implementations of the same operation —
    precisely the split-brain risk this migration exists to close).

    preferred_vendor is written here for a real reason: a PR line often
    starts with no vendor tagged (the normal case for anything that
    gets a vendor via RFx award, contract auto-match, or fully-
    consolidated grouping) — before this, that line's vendor stayed
    blank in PR_Items forever, even once the PO genuinely had one. The
    PR report's "Vendor" column was reading that stale blank, not
    reality, for every line that didn't already have a vendor before
    consolidation — which is most of them.

    Status is 'PO Proposed': by the time this runs, insert_po() has
    already written the PO row — the PO genuinely exists — but it isn't
    yet usable downstream (see po_header.status, and the gate in
    goods_receipt.py) until it's actually been sent to the vendor. See
    mark_po_created() for that next real state transition, once the
    PO's export file has actually been generated."""
    pairs = list(pairs)
    if not pairs:
        return
    conn = db.get_connection()
    try:
        for pr_number, pr_line, po_number, po_item, vendor in pairs:
            conn.execute(
                "UPDATE pr_items SET status='PO Proposed', po_number=?, po_item=?, "
                "preferred_vendor=? WHERE pr_number=? AND pr_line_item=?",
                (po_number, po_item, vendor, pr_number, int(pr_line)),
            )
        conn.commit()
    finally:
        conn.close()


def mark_po_created(po_number, data_file=None, po_date=None):
    """
    The real state transition that makes a PO usable downstream. Three
    things happen atomically:
      1. po_header.status flips from 'Proposed' to 'Created' — this is
         the actual gate goods_receipt.py checks; a PO stuck at
         'Proposed' can't be received against or invoiced.
      2. po_header.po_date is set (default today, or pass po_date to
         backdate — same pattern as create_pr()'s own pr_date). This is
         the moment the lead-time clock genuinely starts in a real
         business sense: the PO briefly exists at 'Proposed' purely as
         an internal draft, which the vendor never sees — po_date marks
         when it actually got sent, generated in the same click as the
         vendor-facing AV export file, not when it was first drafted.
      3. Every PR line that feeds this PO advances from 'PO Proposed'
         to 'PO Created' — cosmetic/reporting only, the real gate is
         (1), but keeps the PR report's status column meaningful too.

    Called from three places, each right after generating that PO's
    vendor-facing export file: erp_ui.py's "Create PO"/"Create All POs"
    buttons (PR-Consolidation-sourced POs — the only path where there's
    a genuine gap between a PO being written and being sent), and
    immediately/automatically for Direct PO Entry and RFx-awarded POs
    (both already generate that same export file in the same click
    that creates the PO, so this just runs right there instead of
    waiting on a separate user action — every PO briefly exists at
    'Proposed' from insert_po()'s own default, but for these two paths
    nothing is ever shown or usable in that state before this runs).

    Traces PR lines the way goods_receipt.get_po_lines_source_prs()
    does — parsing each PO line's comma-separated
    Source_PR_Number/Source_PR_Line_Item back to (pr_number, pr_line)
    pairs — but doesn't import that module (would be a real circular-
    import risk: goods_receipt already imports this module). A PO with
    no source PR lines at all (Direct PO Entry) just finds zero pairs
    here — harmless, the po_header update still applies.
    """
    conn = db.get_connection()
    try:
        conn.execute("UPDATE po_header SET status='Created', po_date=? WHERE po_number=?",
                     (po_date or date.today().isoformat(), po_number))
        rows = conn.execute(
            "SELECT source_pr_number, source_pr_line_item FROM po_items WHERE po_number = ?",
            (po_number,),
        ).fetchall()
        pairs = set()
        for r in rows:
            prs = str(r["source_pr_number"] or "").split(",")
            lns = str(r["source_pr_line_item"] or "").split(",")
            for pr_r, ln_r in zip(prs, lns):
                pr_r, ln_r = pr_r.strip(), ln_r.strip()
                if pr_r and ln_r:
                    try:
                        pairs.add((pr_r, int(float(ln_r))))
                    except ValueError:
                        pass
        for pr_number, pr_line in pairs:
            conn.execute(
                "UPDATE pr_items SET status='PO Created' WHERE pr_number=? AND pr_line_item=? "
                "AND status='PO Proposed'",
                (pr_number, pr_line),
            )
        conn.commit()
    finally:
        conn.close()
    return len(pairs)


def mark_pr_items_rfp(pairs, data_file=None):
    """pairs: iterable of (pr_number, pr_line). Sets status='RFP' — the
    writeback that stops a future run() from picking these lines up
    again as 'Open, no vendor' and generating a duplicate RFP entry."""
    pairs = list(pairs)
    if not pairs:
        return
    conn = db.get_connection()
    try:
        for pr_number, pr_line in pairs:
            conn.execute(
                "UPDATE pr_items SET status='RFP' WHERE pr_number=? AND pr_line_item=?",
                (pr_number, int(pr_line)),
            )
        conn.commit()
    finally:
        conn.close()


def build_all_pos_geo_data(data_file=None, date_from=None, date_to=None):
    """
    Same {vendors, deliveries, routes} shape run()'s geo_data uses (so
    erp_ui.py's existing _map() renders it with zero changes), but built
    from every PO that's ever been created — not just whatever the most
    recent consolidation round produced. Same VHEX/DHEX color scheme for
    visual consistency between the two map views.

    date_from/date_to (both optional, inclusive) filter on each line's
    own delivery_date — the same field and the same filtering the PO
    Line Items list below the map uses. Passing neither returns
    everything, unfiltered. This exists so the map and the list can be
    kept in sync: before this, the map always showed every PO ever
    created regardless of what date range was selected for the list,
    so a PO with an old delivery date would show a route on the map
    while being invisible in the list right below it — confusing, and
    not what "filtered by this date range" should mean.
    """
    fpath = data_file or DATA_FILE
    vendor_geo = _read_vendor_master()

    VHEX = {"GARMY": "#7F77DD", "HINDMED": "#1D9E75", "MANI": "#BA7517",
           "Pharmalines": "#D85A30", "dntl": "#D4537E"}
    DHEX = ["#4F8EF7", "#9B59B6", "#E24B4A", "#F39C12", "#1ABC9C"]

    po_vendor = {r["po_number"]: r["supplier_id"] for r in get_all_po_headers()}

    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT po_number, delivery_location, delivery_geolocation, delivery_date FROM po_items"
        ).fetchall()
    finally:
        conn.close()

    po_deliveries = {}   # po_number -> {deliv_loc: [deliv_geo, item_count]}
    for row in rows:
        po, dl, dg, dd = row["po_number"], row["delivery_location"], row["delivery_geolocation"], row["delivery_date"]
        if date_from or date_to:
            parsed = _parse_date(dd)
            if parsed is None:
                continue
            if date_from and parsed < date_from:
                continue
            if date_to and parsed > date_to:
                continue
        slot = po_deliveries.setdefault(po, {})
        if dl not in slot:
            slot[dl] = [dg, 0]
        slot[dl][1] += 1

    routes = []
    uniq_v, uniq_d = {}, {}
    for po, vendor in po_vendor.items():
        sup_geo = vendor_geo.get(vendor, "")
        for dl, (dg, count) in po_deliveries.get(po, {}).items():
            if not sup_geo or not dg:
                continue
            if vendor not in uniq_v:
                ll = _parse_ll(sup_geo)
                uniq_v[vendor] = {"id": vendor, "geo": sup_geo,
                                  "lat": ll[0] if ll else None, "lng": ll[1] if ll else None,
                                  "color": VHEX.get(vendor, DHEX[len(uniq_v) % len(DHEX)])}
            if dl not in uniq_d:
                ll = _parse_ll(dg)
                uniq_d[dl] = {"id": dl, "geo": dg,
                              "lat": ll[0] if ll else None, "lng": ll[1] if ll else None}
            routes.append({"vendor": vendor, "supplier_geo": sup_geo, "deliv_loc": dl,
                           "deliv_geo": dg, "po_number": po, "items": count,
                           "color": uniq_v[vendor]["color"]})

    return {"vendors": list(uniq_v.values()), "deliveries": list(uniq_d.values()), "routes": routes}


# ── SQLite: RFP ───────────────────────────────────────────────────────────────
def next_rfp_seq(base, data_file=None):
    """base like 'RFP-20260721'. Returns the next sequence number after
    the highest existing {base}-NNN."""
    conn = db.get_connection()
    try:
        prefix = f"{base}-"
        rows = conn.execute(
            "SELECT rfp_number FROM rfp WHERE rfp_number LIKE ?", (prefix + "%",)
        ).fetchall()
    finally:
        conn.close()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["rfp_number"][len(prefix):]))
        except (ValueError, IndexError): pass
    return mx + 1


def insert_rfp_lines(rfp_lines, data_file=None):
    """rfp_lines: list of dicts in _build_rfp()'s output shape."""
    rfp_lines = list(rfp_lines)
    if not rfp_lines:
        return
    conn = db.get_connection()
    try:
        for item in rfp_lines:
            conn.execute(
                "INSERT INTO rfp (rfp_number, rfp_date, material_code, material_desc, uom, "
                "total_qty, required_by_date, delivery_location, delivery_geolocation, "
                "source_pr_numbers, source_pr_lines, requester_depts, project_ids, "
                "specifications, closing_date, status) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (item["rfp_number"], item["rfp_date"], item["mat_code"], item["mat_desc"],
                 item["uom"], item["total_qty"], item["req_by_date"], item["deliv_loc"],
                 item["deliv_geo"], item["source_prs"], item["source_lines"], item["req_depts"],
                 item["project_ids"], item["specs"], item["closing_date"], item["status"]),
            )
        conn.commit()
    finally:
        conn.close()


# ── Pure computation (unchanged from before the migration) ──────────────────
def _build_po_lines(lines,po_num,vendor,one_to_one):
    if one_to_one:
        return [{"po_number":po_num,"po_type":l["po_type"],"legal_entity":l["legal_entity"],
                 "purch_entity":l["purch_entity"],"purch_group":l["purch_group"],
                 "currency":l["currency"],"plant_code":l["plant_code"],
                 "supplier_id":vendor,"mat_code":l["mat_code"],"mat_desc":l["mat_desc"],
                 "uom":l["uom"],"qty":l["qty"],"unit_price":l.get("unit_price"),
                 "deliv_date":l["req_date"],
                 "deliv_loc":l["deliv_loc"],"deliv_geo":l["deliv_geo"],
                 "source_pr":l["pr_number"],"source_pr_line":l["pr_line"],
                 "req_id":l["req_id"],"req_dept":l["req_dept"],"project_id":l["project_id"]}
                for l in lines]
    else:
        # Grouped by (material, delivery location), not material alone — a
        # real gap found and fixed here: grouping by material alone meant
        # two PRs for the same item from DIFFERENT locations, routed to the
        # same vendor, would merge into one PO line carrying only the
        # first PR's location — the combined quantity was correct, but
        # part of it would silently be posted (at Goods Receipt) to a
        # location that requester never asked for, with no error anywhere
        # to catch it. Grouping on the pair keeps the real benefit of
        # "consolidated" mode (fewer lines, by merging same-item
        # same-destination requests) without ever letting one PO line
        # represent goods bound for more than one place.
        grps=defaultdict(list)
        for l in lines: grps[(l["mat_code"], l["deliv_loc"])].append(l)
        out=[]
        for (mc, _dl),grp in grps.items():
            dates=sorted([g["req_date"] for g in grp if g["req_date"]])
            prices={g.get("unit_price") for g in grp if g.get("unit_price") is not None}
            unit_price = prices.pop() if len(prices)==1 else None  # None if mixed/unknown
            out.append({"po_number":po_num,"po_type":grp[0]["po_type"],
                "legal_entity":grp[0]["legal_entity"],"purch_entity":grp[0]["purch_entity"],
                "purch_group":grp[0]["purch_group"],"currency":grp[0]["currency"],
                "plant_code":grp[0]["plant_code"],"supplier_id":vendor,
                "mat_code":mc,"mat_desc":grp[0]["mat_desc"],"uom":grp[0]["uom"],
                "qty":sum(g["qty"] for g in grp),"unit_price":unit_price,
                "deliv_date":dates[0] if dates else "",
                "deliv_loc":grp[0]["deliv_loc"],"deliv_geo":grp[0]["deliv_geo"],
                "source_pr":", ".join(g["pr_number"] for g in grp),
                "source_pr_line":", ".join(str(g["pr_line"]) for g in grp),
                "req_id":", ".join(sorted(set(g["req_id"] for g in grp))),
                "req_dept":", ".join(sorted(set(g["req_dept"] for g in grp))),
                "project_id":", ".join(sorted(set(g["project_id"] for g in grp)))})
        return out


def _build_rfp(rfp_lines, deliv_geo_m, start_seq=1):
    """
    Grouped by (material, delivery location), not material alone — same
    bug and same fix as _build_po_lines' consolidated mode: this
    function has no toggle, it ALWAYS merges across PRs, so grouping by
    material alone meant any two PRs for the same unassigned-vendor
    item from different locations would silently collapse onto one
    RFP line, one location. Since generate_pos() carries an RFP's
    deliv_loc straight through to the PO it creates, this would have
    propagated the same "quantity right, location wrong for part of
    it" problem into the PO/GR/inventory chain even for RFP-sourced
    purchases, not just directly-consolidated ones.
    """
    today=date.today(); closing=(today+timedelta(days=14)).strftime("%Y-%m-%d")
    base=f"RFP-{today.strftime('%Y%m%d')}"
    grps=defaultdict(list)
    for l in rfp_lines: grps[(l["mat_code"], l["deliv_loc"])].append(l)
    out=[]
    for i,((mc, _dl),grp) in enumerate(grps.items()):
        seq = start_seq + i
        dates=sorted([g["req_date"] for g in grp if g["req_date"]])
        dloc=grp[0]["deliv_loc"]
        dgeo=grp[0].get("deliv_geo","") or deliv_geo_m.get(dloc,"")
        out.append({"rfp_number":f"{base}-{seq:03d}","rfp_date":today.strftime("%Y-%m-%d"),
            "mat_code":mc,"mat_desc":grp[0]["mat_desc"],"uom":grp[0]["uom"],
            "total_qty":sum(g["qty"] for g in grp),"req_by_date":dates[0] if dates else "",
            "deliv_loc":dloc,"deliv_geo":dgeo,
            "source_prs":", ".join(sorted(set(g["pr_number"] for g in grp))),
            "source_lines":", ".join(str(g["pr_line"]) for g in grp),
            "req_depts":", ".join(sorted(set(g["req_dept"] for g in grp))),
            "project_ids":", ".join(sorted(set(g["project_id"] for g in grp))),
            "specs":grp[0]["mat_desc"],"closing_date":closing,"status":"Open"})
    return out


# ── Core engine ───────────────────────────────────────────────────────────────
def run(data_file=None, one_to_one=None, contract_lookup_fn=None):
    """
    contract_lookup_fn: optional callable(material_code) -> {"vendor_id","price",
    "contract_id"} or None. Injected by the caller (erp_ui.py wires this to
    contracts.find_active_contract_for_item) rather than imported directly here,
    so this module has zero dependency on contracts.py — avoids a circular
    import (contracts.py already imports this module for its style helpers).
    When supplied, PR lines with no vendor pre-assigned are checked against it
    first; a match skips RFP entirely and goes straight to a priced PO line,
    same as a line that already had a vendor assigned in the PR itself.
    """
    fpath = data_file or DATA_FILE
    mode  = ONE_TO_ONE_PR_PO_LINES if one_to_one is None else one_to_one

    vendor_geo  = _read_vendor_master()
    deliv_geo_m = _read_delivery_master()

    pr_headers = get_pr_headers(fpath)
    pr_items   = get_pr_items(open_only=True, data_file=fpath)
    pr_items_all_count = len(get_pr_items(open_only=False, data_file=fpath))

    for item in pr_items:
        hdr = pr_headers.get(item["pr_number"],{})
        item.update({k: hdr.get(k,"") for k in
            ["req_id","req_name","req_dept","project_id","po_type",
             "legal_entity","purch_entity","purch_group","currency","plant_code"]})
        if not item.get("deliv_geo") and item.get("deliv_loc"):
            item["deliv_geo"] = deliv_geo_m.get(item["deliv_loc"],"")

    vendor_items = [i for i in pr_items if i["vendor"]]
    rfp_items_in = [i for i in pr_items if not i["vendor"]]

    contract_matched_count = 0
    if contract_lookup_fn:
        still_unmatched = []
        for it in rfp_items_in:
            match = contract_lookup_fn(it["mat_code"])
            if match:
                it["vendor"] = match["vendor_id"]
                it["unit_price"] = match["price"]
                it["contract_id"] = match["contract_id"]
                vendor_items.append(it)
                contract_matched_count += 1
            else:
                still_unmatched.append(it)
        rfp_items_in = still_unmatched

    # A PR line reaching this point without a price falls back to the
    # item's Item Master list price — a real gap fixed here, not a
    # cosmetic one. PR_Items has never had a unit_price column at all;
    # the ONLY PR line that ever carried a price before this was one
    # auto-matched to a contract just above. Every other PR-Consolidation
    # PO line — the overwhelming majority of real usage — has always
    # gone out with unit_price=None, which meant its GR could never post
    # a real accounting entry (an all-zero JE, correctly refused — see
    # accounting.post_gr_entry()). Direct PO Entry never had this
    # problem, since its item picker already pulls price from Item
    # Master directly; this brings PR Consolidation's PO lines up to
    # the same standard. Only fills in a MISSING price — a line that
    # already has one (from a PR-level tag or a contract match) keeps
    # it untouched.
    for it in vendor_items:
        if it.get("unit_price") is None:
            info = po_export.get_item_by_code(it["mat_code"], active_only=False)
            if info and info["price"]:
                it["unit_price"] = info["price"]

    by_vendor = defaultdict(list)
    for it in vendor_items: by_vendor[it["vendor"]].append(it)

    VHEX={"GARMY":"#7F77DD","HINDMED":"#1D9E75","MANI":"#BA7517",
          "Pharmalines":"#D85A30","dntl":"#D4537E"}
    DHEX=["#4F8EF7","#9B59B6","#E24B4A","#F39C12","#1ABC9C"]

    deliv_map = {}
    for it in pr_items:
        if it["deliv_loc"] and it.get("deliv_geo"):
            deliv_map[it["deliv_loc"]] = it["deliv_geo"]

    year = datetime.now().year
    po_summary=[]; all_po_items=[]; geo_routes=[]; pr_writeback_pairs=[]

    for vi,(vendor,lines) in enumerate(sorted(by_vendor.items())):
        po_num  = next_po_number(year, fpath)
        sup_geo = vendor_geo.get(vendor,"")
        po_lns  = _build_po_lines(lines, po_num, vendor, mode)
        vendor_row = vo.get_vendor(vendor)
        header = {"po_type": lines[0]["po_type"], "legal_entity": lines[0]["legal_entity"],
                  "purch_entity": lines[0]["purch_entity"], "purch_group": lines[0]["purch_group"],
                  "currency": lines[0]["currency"], "plant_code": lines[0]["plant_code"],
                  "supplier_id": vendor,
                  "supplier_name": vendor_row["Vendor_Name"] if vendor_row else "",
                  "supplier_geo": sup_geo}
        written = insert_po(po_num, header, po_lns, fpath)
        all_po_items.extend(written)

        for it in written:
            prs = str(it["source_pr"]).split(", ")
            lns = str(it["source_pr_line"]).split(", ")
            for pr_r, ln_r in zip(prs, lns):
                pr_r, ln_r = pr_r.strip(), ln_r.strip()
                if pr_r and ln_r:
                    pr_writeback_pairs.append((pr_r, int(float(ln_r)), po_num, it["po_item"], vendor))

        seen=set()
        for pol in written:
            dl=pol["deliv_loc"]
            if dl not in seen:
                seen.add(dl)
                dg=deliv_map.get(dl,"")
                geo_routes.append({"vendor":vendor,"supplier_geo":sup_geo,
                    "deliv_loc":dl,"deliv_geo":dg,"po_number":po_num,
                    "items":sum(1 for p in written if p["deliv_loc"]==dl),
                    "color":VHEX.get(vendor,DHEX[vi%len(DHEX)])})

        po_summary.append({"po_number":po_num,"vendor":vendor,
                           "lines":len(written),"supplier_geo":sup_geo})

    rfp_start_seq = next_rfp_seq(f"RFP-{date.today().strftime('%Y%m%d')}", fpath)
    rfp_out = _build_rfp(rfp_items_in, deliv_geo_m, rfp_start_seq)
    insert_rfp_lines(rfp_out, fpath)

    mark_pr_items_ordered(pr_writeback_pairs, fpath)
    mark_pr_items_rfp([(i["pr_number"], i["pr_line"]) for i in rfp_items_in], fpath)

    uniq_v={}; uniq_d={}
    for r in geo_routes:
        if r["vendor"] not in uniq_v:
            ll=_parse_ll(r["supplier_geo"])
            uniq_v[r["vendor"]]={"id":r["vendor"],"geo":r["supplier_geo"],
                "lat":ll[0] if ll else None,"lng":ll[1] if ll else None,
                "color":VHEX.get(r["vendor"],DHEX[len(uniq_v)%len(DHEX)])}
        if r["deliv_loc"] not in uniq_d:
            ll=_parse_ll(r["deliv_geo"])
            uniq_d[r["deliv_loc"]]={"id":r["deliv_loc"],"geo":r["deliv_geo"],
                "lat":ll[0] if ll else None,"lng":ll[1] if ll else None}
        r["color"]=uniq_v[r["vendor"]]["color"]

    msgs=[
        f"Read {len(pr_items)} Open PR lines from {len(pr_headers)} PRs.",
        f"Skipped non-Open lines (already PO Proposed, PO Created, or other status)." if pr_items_all_count > len(pr_items) else "",
        f"Auto-matched {contract_matched_count} line(s) to active contracts — skipped RFx." if contract_matched_count else "",
        f"Created {len(po_summary)} PO(s): "+", ".join(f"{p['po_number']} ({p['vendor']}, {p['lines']} items)" for p in po_summary),
        f"RFP: {len(rfp_out)} material(s) without identified vendor.",
        f"Mode: {'1:1 PR→PO' if mode else 'consolidated qty'}.",
        f"PR_Items updated: {len(all_po_items)} lines → 'PO Proposed'.",
    ]
    msgs=[m for m in msgs if m]

    return {"po_summary":po_summary,"rfp_count":len(rfp_out),"messages":msgs,
            "pr_items":pr_items,"pr_headers":pr_headers,
            "po_items":all_po_items,"rfp_items":rfp_out,
            "geo_data":{"vendors":list(uniq_v.values()),
                        "deliveries":list(uniq_d.values()),"routes":geo_routes}}


def clear_consolidation_data(data_file=None):
    """
    Moved here from erp_ui.py (extraction, per the pilot plan — this was
    ad-hoc UI code before, now delegated to the module that owns the
    data). Same behavior: full reset of PO_Header/PO_Items/RFP, PR_Items
    status reverted to Open for 'PO Proposed', 'PO Created', and 'RFP',
    PLUS RFx_Quotes/RFx_Invitations cleared (now SQLite too, same as RFP).

    Deliberately does NOT touch preferred_vendor — mark_pr_items_ordered()
    may have written one there that wasn't requester-specified (a
    contract auto-match, an RFx award, or fully-consolidated grouping).
    Clearing it would either destroy a real requester-entered tag (if
    there was one) or, for an auto-derived one, just force the exact
    same vendor to be rediscovered the slow way on the next Consolidate
    run — leaving it in place means a redo after Clear reuses what's
    already known instead of re-running RFx from scratch.

    Currently unused by the UI — see clear_po_proposals() below for the
    scoped version "Clear PO Proposals" actually calls. Left in place,
    still correct and tested, for if a genuine full-reset need comes up.
    """
    conn = db.get_connection()
    try:
        conn.execute("DELETE FROM po_items")
        conn.execute("DELETE FROM po_header")
        conn.execute("DELETE FROM rfp")
        conn.execute("DELETE FROM rfx_quotes")
        conn.execute("DELETE FROM rfx_invitations")
        cur = conn.execute(
            "UPDATE pr_items SET status='Open', po_number=NULL, po_item=NULL "
            "WHERE status IN ('PO Proposed','PO Created','RFP')"
        )
        reset_count = cur.rowcount
        conn.commit()
    finally:
        conn.close()
    return reset_count


def clear_po_proposals(data_file=None):
    """
    "Clear PO Proposals" — deliberately scoped to only PO_Header rows
    still at status='Proposed' (and their PO_Items), reverting those
    specific PR lines back to Open so they rejoin the next Consolidate
    run alongside anything else still Open. Anything already at
    status='Created' — genuinely sent to a vendor, or a Direct PO
    Entry/RFx PO that was Created from the start — is left completely
    untouched; those are the vendor's or the ledger's business now, not
    something to silently undo. RFP data is untouched too — this is
    scoped to PO Proposals specifically, not a general "start over"
    reset (see clear_consolidation_data() above for that, if it's ever
    actually needed).

    Same preferred_vendor reasoning as clear_consolidation_data(): left
    alone, not cleared.
    """
    conn = db.get_connection()
    try:
        proposed = [r["po_number"] for r in conn.execute(
            "SELECT po_number FROM po_header WHERE status = 'Proposed'").fetchall()]
        if proposed:
            placeholders = ",".join("?" for _ in proposed)
            conn.execute(f"DELETE FROM po_items WHERE po_number IN ({placeholders})", proposed)
            conn.execute(f"DELETE FROM po_header WHERE po_number IN ({placeholders})", proposed)
        cur = conn.execute(
            "UPDATE pr_items SET status='Open', po_number=NULL, po_item=NULL "
            "WHERE status='PO Proposed'"
        )
        reset_count = cur.rowcount
        conn.commit()
    finally:
        conn.close()
    return {"po_count": len(proposed), "pr_lines_reset": reset_count}


if __name__=="__main__":
    print(f"ONE_TO_ONE={ONE_TO_ONE_PR_PO_LINES}\n")
    result=run()
    print()
    for m in result["messages"]: print("✓",m)
