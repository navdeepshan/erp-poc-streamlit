"""
contracts.py — Contract Management (the "C" in S2C).

Closes the loop that source-to-contract implies but didn't yet exist:
recurring items shouldn't re-run RFx every time. This module lets a
completed PO (from RFx or direct entry) become a rate contract — locked
per-item pricing, a validity window, SLAs — and then plugs that contract
back into pr_consolidation.run() as a lookup: a future PR line for a
contracted material auto-matches to the contracted vendor and price,
skipping RFx entirely.

Design note on the integration: pr_consolidation.py does NOT import this
module. It accepts an optional contract_lookup_fn callable instead, which
erp_ui.py wires up (contracts.find_active_contract_for_item). That keeps
pr_consolidation.py dependency-free of contracts.py. This module does
import rfx.py (for award_rfx_to_contracts, the RFx-award-straight-to-
contract path) — safe, since rfx.py never imports this module, so the
dependency chain stays one-directional: pr_consolidation <- rfx <- contracts.

New sheets:
  Contracts        Contract_ID | Vendor_ID | Vendor_Name | Status |
                    Start_Date | End_Date | Payment_Terms |
                    Delivery_SLA_Days | Currency | Created_Date |
                    Source_PO | Auto_Renew | Notes
  Contract_Items    Contract_ID | Line_Item | Material_Code | Material_Desc |
                    UOM | Contracted_Unit_Price | Min_Order_Qty | Lead_Time_Days

Status stored on the Contracts row is a cached label ("Active"/"Expired"/
"Terminated"/"Draft"); the source of truth for "is this usable right now"
is always the live date check in find_active_contract_for_item() and
refresh_statuses(), so a stale stored label never silently misleads a
lookup — it just misleads a dashboard count until refreshed.

SQLite pilot: Contracts and Contract_Items now live in erp_pilot.db
(tables `contracts`, `contract_items`), not in data.xlsx. Contracts.py
is the exclusive owner of both — no other module reads or writes them
directly (erp_ui.py only routes to this module's page, pr_consolidation.py
only calls the injected contract_lookup_fn callable, never the tables
themselves) — so this was a contained, single-file migration, same
category of "safest slice" as bom.py's original BOM_Items move.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, io
from datetime import date, datetime, timedelta

import db
import pr_consolidation as pc
import rfx

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

CONTRACTS_SHEET = "Contracts"
CONTRACT_ITEMS_SHEET = "Contract_Items"

CONTRACT_COLS = ["Contract_ID", "Vendor_ID", "Vendor_Name", "Status", "Start_Date",
                 "End_Date", "Payment_Terms", "Delivery_SLA_Days", "Currency",
                 "Created_Date", "Source_PO", "Auto_Renew", "Notes"]
CONTRACT_ITEM_COLS = ["Contract_ID", "Line_Item", "Material_Code", "Material_Desc",
                      "UOM", "Contracted_Unit_Price", "Min_Order_Qty", "Lead_Time_Days"]


# ── Sheet bootstrap ────────────────────────────────────────────────────────────
def ensure_sheets(wb=None):
    """Kept for signature compatibility — Contracts/Contract_Items no
    longer live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


def _compute_status(start_date, end_date, stored_status):
    if stored_status == "Terminated":
        return "Terminated"
    today = date.today().isoformat()
    if start_date and today < str(start_date):
        return "Scheduled"
    if end_date and today > str(end_date):
        return "Expired"
    return "Active"


# ── Readers ─────────────────────────────────────────────────────────────────────
def get_contracts(status=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM contracts ORDER BY contract_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"contract_id": r["contract_id"], "vendor_id": r["vendor_id"],
               "vendor_name": r["vendor_name"], "status": r["status"],
               "start_date": r["start_date"], "end_date": r["end_date"],
               "payment_terms": r["payment_terms"], "delivery_sla_days": r["delivery_sla_days"],
               "currency": r["currency"], "created_date": r["created_date"],
               "source_po": r["source_po"], "auto_renew": r["auto_renew"], "notes": r["notes"]}
        row["status"] = _compute_status(row["start_date"], row["end_date"], row["status"])
        if status and row["status"] != status:
            continue
        out.append(row)
    return out


def get_contract(contract_id, data_file=None):
    for c in get_contracts(data_file=data_file):
        if c["contract_id"] == contract_id:
            return c
    return None


def get_contract_items(contract_id, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM contract_items WHERE contract_id = ? ORDER BY line_item",
            (contract_id,),
        ).fetchall()
    finally:
        conn.close()
    return [{"contract_id": r["contract_id"], "line_item": r["line_item"],
             "mat_code": r["material_code"], "mat_desc": r["material_desc"],
             "uom": r["uom"], "unit_price": r["contracted_unit_price"],
             "min_order_qty": r["min_order_qty"], "lead_time_days": r["lead_time_days"]}
            for r in rows]


def find_active_contract_for_item(material_code, data_file=None):
    """
    The loop-closer: does material_code have an active, in-date contract?
    Returns {"contract_id","vendor_id","vendor_name","price"} or None.
    If more than one active contract covers the material (shouldn't happen
    in normal use, but PoC data can be messy), the lowest price wins.
    """
    conn = db.get_connection()
    try:
        contract_rows = conn.execute(
            "SELECT contract_id, vendor_id, vendor_name, status, start_date, end_date "
            "FROM contracts"
        ).fetchall()
        active_contracts = {}
        for r in contract_rows:
            if _compute_status(r["start_date"], r["end_date"], r["status"]) == "Active":
                active_contracts[r["contract_id"]] = {"vendor_id": r["vendor_id"],
                                                       "vendor_name": r["vendor_name"]}
        if not active_contracts:
            return None
        placeholders = ",".join("?" for _ in active_contracts)
        item_rows = conn.execute(
            f"SELECT contract_id, contracted_unit_price, lead_time_days FROM contract_items "
            f"WHERE material_code = ? AND contract_id IN ({placeholders}) "
            f"AND contracted_unit_price IS NOT NULL",
            [material_code] + list(active_contracts.keys()),
        ).fetchall()
    finally:
        conn.close()

    candidates = [{"contract_id": r["contract_id"],
                   "vendor_id": active_contracts[r["contract_id"]]["vendor_id"],
                   "vendor_name": active_contracts[r["contract_id"]]["vendor_name"],
                   "price": float(r["contracted_unit_price"]),
                   "lead_time_days": r["lead_time_days"]} for r in item_rows]
    if not candidates:
        return None
    return min(candidates, key=lambda c: c["price"])


# ── Creation from a PO ──────────────────────────────────────────────────────────
def _next_contract_id(conn):
    rows = conn.execute("SELECT contract_id FROM contracts WHERE contract_id LIKE 'CTR-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["contract_id"].split("-")[1]))
        except Exception: pass
    return f"CTR-{mx+1:04d}"


def _read_po(po_number, data_file):
    """PO_Header/PO_Items now live in SQLite (pr_consolidation.py's
    pilot) — delegates instead of scanning Excel cell-by-cell (which,
    incidentally, was the exact .cell(r,c)-in-a-loop pattern measured
    elsewhere in this project at ~277x slower than iter_rows() in
    read-only mode; this migration removes that cost here too, not just
    the PR/PO writes)."""
    row = pc.get_po_header(po_number)
    if row is None:
        return None, []
    header = {"po_number": po_number, "vendor_id": row["supplier_id"],
              "vendor_name": row["supplier_name"] or row["supplier_id"],
              "currency": row["currency"]}
    items = [{"po_item": r["po_item"], "mat_code": r["material_code"],
              "mat_desc": r["material_desc"], "uom": r["uom"], "qty": r["quantity"],
              "unit_price": r["unit_price"]} for r in pc.get_po_items(po_number)]
    return header, items


def list_pos_available_for_contract(data_file=None):
    """POs that have at least one line with a known Unit_Price and aren't
    already the source of an existing contract — the realistic candidate list
    for 'Convert to Contract'. PO_Header/PO_Items and Contracts now both
    live in SQLite."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT DISTINCT source_po FROM contracts WHERE source_po IS NOT NULL"
        ).fetchall()
    finally:
        conn.close()
    already_used = {str(r["source_po"]) for r in rows}

    priced_pos = set()
    po_line_counts = {}
    all_headers = pc.get_all_po_headers()
    for header in all_headers:
        po = header["po_number"]
        items = pc.get_po_items(po)
        po_line_counts[po] = len(items)
        if any(it["unit_price"] is not None for it in items):
            priced_pos.add(po)

    out = []
    for header in all_headers:
        po = header["po_number"]
        if not po or po in already_used or po not in priced_pos:
            continue
        vid = header["supplier_id"]
        out.append({"po_number": po, "vendor_id": vid,
                    "vendor_name": header["supplier_name"] or vid,
                    "lines": po_line_counts.get(po, 0)})
    return out


def create_contract_from_po(po_number, start_date, end_date, payment_terms="Net 30",
                             delivery_sla_days=7, auto_renew=False, notes="", data_file=None):
    fpath = data_file or DATA_FILE
    header, items = _read_po(po_number, fpath)
    if header is None:
        raise ValueError(f"PO {po_number} not found.")
    priced_items = [i for i in items if i["unit_price"] is not None]
    if not priced_items:
        raise ValueError(f"PO {po_number} has no line with a known Unit_Price — "
                          "nothing to lock into a rate contract.")

    db.init_schema()
    conn = db.get_connection()
    try:
        contract_id = _next_contract_id(conn)
        status = _compute_status(str(start_date), str(end_date), "Active")
        conn.execute(
            "INSERT INTO contracts (contract_id, vendor_id, vendor_name, status, start_date, "
            "end_date, payment_terms, delivery_sla_days, currency, created_date, source_po, "
            "auto_renew, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (contract_id, header["vendor_id"], header["vendor_name"], status, str(start_date),
             str(end_date), payment_terms, delivery_sla_days, header.get("currency") or "INR",
             date.today().strftime("%Y-%m-%d"), po_number, "Yes" if auto_renew else "No", notes),
        )
        for seq, item in enumerate(priced_items, 1):
            conn.execute(
                "INSERT INTO contract_items (contract_id, line_item, material_code, "
                "material_desc, uom, contracted_unit_price, min_order_qty, lead_time_days) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (contract_id, seq, item["mat_code"], item["mat_desc"], item["uom"],
                 item["unit_price"], 1, delivery_sla_days),
            )
        conn.commit()
    finally:
        conn.close()

    return {"contract_id": contract_id, "lines": len(priced_items),
            "skipped_unpriced": len(items) - len(priced_items)}


def award_rfx_to_contracts(start_date, end_date, payment_terms="Net 30",
                            delivery_sla_days=7, auto_renew=False, notes="", data_file=None):
    """
    The direct path: issues every pending RFx award as a PO (via
    rfx.generate_pos(), unchanged) and immediately locks each resulting PO
    into a rate contract with the given terms — one action instead of two.
    This lets the *first* purchase of an item be contract-first too, not
    just repeat purchases: RFx award -> PO + Contract together, rather than
    RFx award -> PO now, remember to convert it to a contract later.
    Returns rfx.generate_pos()'s result list, each dict augmented with
    'contract_id' (or 'contract_error' if that one PO couldn't be contracted,
    e.g. no priced lines — kept per-PO so one failure doesn't sink the batch).
    """
    fpath = data_file or DATA_FILE
    po_results = rfx.generate_pos(fpath)
    combined = []
    for po in po_results:
        entry = dict(po)
        try:
            c = create_contract_from_po(po["po_number"], start_date, end_date,
                payment_terms, delivery_sla_days, auto_renew, notes, fpath)
            entry["contract_id"] = c["contract_id"]
            entry["contract_lines"] = c["lines"]
        except ValueError as e:
            entry["contract_id"] = None
            entry["contract_error"] = str(e)
        combined.append(entry)
    return combined


# ── Lifecycle ─────────────────────────────────────────────────────────────────
def terminate_contract(contract_id, reason="", data_file=None):
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT notes FROM contracts WHERE contract_id = ?", (contract_id,)
        ).fetchone()
        if row is not None:
            existing_notes = row["notes"] or ""
            new_notes = f"{existing_notes} | Terminated: {reason}".strip(" |")
            conn.execute(
                "UPDATE contracts SET status='Terminated', notes=? WHERE contract_id=?",
                (new_notes, contract_id),
            )
        conn.commit()
    finally:
        conn.close()


def renew_contract(contract_id, new_end_date, data_file=None):
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT status FROM contracts WHERE contract_id = ?", (contract_id,)
        ).fetchone()
        if row is not None and (row["status"] or "") == "Terminated":
            raise ValueError(f"{contract_id} was terminated — renewal doesn't apply to a "
                              "for-cause termination. Create a new contract instead.")
        if row is not None:
            conn.execute(
                "UPDATE contracts SET end_date=?, status='Active' WHERE contract_id=?",
                (str(new_end_date), contract_id),
            )
            conn.commit()
    finally:
        conn.close()


def renewals_due(within_days=30, data_file=None):
    """Active contracts whose End_Date falls within the next N days."""
    cutoff = (date.today() + timedelta(days=within_days)).isoformat()
    today = date.today().isoformat()
    out = []
    for c in get_contracts(data_file=data_file):
        if c["status"] == "Active" and c["end_date"] and today <= str(c["end_date"]) <= cutoff:
            out.append(c)
    return sorted(out, key=lambda c: c["end_date"])


# ── Contract document ─────────────────────────────────────────────────────────
def generate_contract_document(contract_id, data_file=None):
    """A formal rate-contract document: terms block + priced line-item
    schedule + signature blocks. Same generation pattern as the RFQ/VRQ docs."""
    fpath = data_file or DATA_FILE
    contract = get_contract(contract_id, fpath)
    if contract is None:
        raise ValueError(f"Contract {contract_id} not found.")
    items = get_contract_items(contract_id, fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = "0C4A6E"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "RATE CONTRACT / FRAMEWORK AGREEMENT")
    ws.merge_cells("A1:F1")

    label("A3", "Contract ID:");     value("B3", contract["contract_id"])
    label("A4", "Vendor:");          value("B4", f"{contract['vendor_name']} ({contract['vendor_id']})")
    label("A5", "Status:");          value("B5", contract["status"])
    label("D3", "Start Date:");      value("E3", contract["start_date"])
    label("D4", "End Date:");        value("E4", contract["end_date"])
    label("D5", "Auto-Renew:");      value("E5", contract["auto_renew"])
    label("A6", "Payment Terms:");   value("B6", contract["payment_terms"])
    label("D6", "Delivery SLA:");    value("E6", f"{contract['delivery_sla_days']} days")

    ws["A8"] = ("This agreement fixes unit pricing for the items below for the validity "
                "period stated. Purchase Orders issued against this contract during that "
                "period do not require re-quoting.")
    ws["A8"].font = Font(name="Arial", size=9, italic=True, color="64748B")
    ws.merge_cells("A8:F8")

    hdr_row = 10
    hdrs = ["#", "Material Code", "Description", "UOM", "Contracted Unit Price",
            "Min Order Qty", "Lead Time (days)"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=navy)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = hdr_row + 1
    for i, item in enumerate(items, 1):
        vals = [i, item["mat_code"], item["mat_desc"], item["uom"],
                item["unit_price"], item["min_order_qty"], item["lead_time_days"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 2
    ws.cell(r, 1, "For the Buyer").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 4, "For the Vendor").font = Font(name="Arial", size=10, bold=True)
    r += 3
    ws.cell(r, 1, "Signature: ____________________").font = Font(name="Arial", size=9)
    ws.cell(r, 4, "Signature: ____________________").font = Font(name="Arial", size=9)
    r += 1
    ws.cell(r, 1, "Name / Date: ____________________").font = Font(name="Arial", size=9)
    ws.cell(r, 4, "Name / Date: ____________________").font = Font(name="Arial", size=9)

    widths = [4, 15, 34, 8, 18, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 26

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{contract['contract_id']}_{contract['vendor_id']}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    contracts = get_contracts(data_file=data_file)
    by_status = {}
    for c in contracts:
        by_status[c["status"]] = by_status.get(c["status"], 0) + 1
    return {"total": len(contracts), "by_status": by_status,
            "renewals_due_30d": len(renewals_due(30, data_file))}


if __name__ == "__main__":
    print("Contract stats:", stats())
    candidates = list_pos_available_for_contract()
    print(f"\n{len(candidates)} PO(s) available to convert to a contract:")
    for c in candidates:
        print(f"  {c['po_number']} — {c['vendor_name']} ({c['lines']} lines)")
