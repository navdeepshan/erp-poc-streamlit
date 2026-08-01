"""
quotation.py — Sales Quotation (pre-sales, O2C).

The mirror image of RFQ, not a repeat of it. In RFQ, a vendor fills in a
blank price column — we don't know the number yet. In a sales quote, WE
state the price — the customer just accepts or declines. So the document
this module generates has no fill-in columns; it has a priced line-item
table and an acceptance signature block, structurally closer to
contracts.py's contract document than to rfx.py's RFQ workbook.

Pricing honesty note: suggest_sale_price() marks up Item Master's Unit_Price
(which is a procurement COST — what we pay vendors) by DEFAULT_MARKUP to
get a suggested sale price. This is a deliberate placeholder, not a real
pricing engine — there's no customer-tier pricing, no volume breaks, no
category-specific margins. It's editable per line on every quote. A real
next step would be a price-list/rate-card module; until then this keeps
quoting usable without silently pretending to be more sophisticated than
it is.

Approval gating note: quote creation does NOT require the customer to be
KYC-approved or credit-active — quoting a brand-new prospect before they're
fully onboarded is completely normal in sales. The credit gate belongs at
Sales Order conversion time, not here — is_accepted() is the hook a future
Sales Order Management module calls (same dependency-injection shape as
contracts.find_active_contract_for_item and
customer_onboarding.check_credit_available), so this module stays
independent of whatever's built next.

New sheets:
  Quotes        Quote_ID | Customer_ID | Customer_Name | Quote_Date |
                Valid_Until | Payment_Terms | Status | Currency |
                Total_Value | Notes | Filename
  Quote_Items   Quote_ID | Line_Item | Material_Code | Material_Desc |
                UOM | Qty | Unit_Price | Line_Total

Status values: Draft -> Sent -> Accepted / Rejected, or Sent -> Expired
(computed live off Valid_Until, same approach as contracts.py's
_compute_status — a stale stored label never lands on a live lookup,
only on a dashboard count until refreshed).

SQLite pilot: Quotes and Quote_Items now live in erp_pilot.db (tables
`quotes`, `quote_items`) — this module is the exclusive owner of both.
"""

import os, io, hashlib
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import customer_onboarding as co

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

QUOTES_SHEET = "Quotes"
QUOTE_ITEMS_SHEET = "Quote_Items"

QUOTE_COLS = ["Quote_ID", "Customer_ID", "Customer_Name", "Quote_Date", "Valid_Until",
             "Payment_Terms", "Status", "Currency", "Total_Value", "Notes", "Filename"]
QUOTE_ITEM_COLS = ["Quote_ID", "Line_Item", "Material_Code", "Material_Desc",
                   "UOM", "Qty", "Unit_Price", "Line_Total"]

DEFAULT_MARKUP = 0.30  # placeholder distributor margin — see module docstring


def suggest_sale_price(cost_price):
    """Pure helper: cost -> suggested sale price at DEFAULT_MARKUP. No I/O,
    always overridable per line — never silently applied without review."""
    if cost_price is None:
        return None
    return round(float(cost_price) * (1 + DEFAULT_MARKUP), 2)


# ── Sheet bootstrap ────────────────────────────────────────────────────────────
def ensure_sheets(wb=None):
    """Kept for signature compatibility — Quotes/Quote_Items no longer
    live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


def _compute_status(valid_until, stored_status):
    if stored_status in ("Accepted", "Rejected", "Draft"):
        return stored_status
    if stored_status == "Sent" and valid_until and date.today().isoformat() > str(valid_until):
        return "Expired"
    return stored_status or "Draft"


# ── Read ──────────────────────────────────────────────────────────────────────
def get_quotes(status=None, customer_id=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM quotes ORDER BY quote_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"quote_id": r["quote_id"], "customer_id": r["customer_id"],
               "customer_name": r["customer_name"], "quote_date": r["quote_date"],
               "valid_until": r["valid_until"], "payment_terms": r["payment_terms"],
               "status": r["status"], "currency": r["currency"],
               "total_value": r["total_value"], "notes": r["notes"], "filename": r["filename"]}
        row["status"] = _compute_status(row["valid_until"], row["status"])
        if status and row["status"] != status:
            continue
        if customer_id and row["customer_id"] != customer_id:
            continue
        out.append(row)
    return out


def get_quote(quote_id, data_file=None):
    for q in get_quotes(data_file=data_file):
        if q["quote_id"] == quote_id:
            return q
    return None


def get_quote_items(quote_id, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM quote_items WHERE quote_id = ? ORDER BY line_item", (quote_id,)
        ).fetchall()
    finally:
        conn.close()
    return [{"quote_id": r["quote_id"], "line_item": r["line_item"], "mat_code": r["material_code"],
             "mat_desc": r["material_desc"], "uom": r["uom"], "qty": r["qty"],
             "unit_price": r["unit_price"], "line_total": r["line_total"]} for r in rows]


# ── Create ────────────────────────────────────────────────────────────────────
def _next_quote_id(conn):
    rows = conn.execute("SELECT quote_id FROM quotes WHERE quote_id LIKE 'QT-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["quote_id"].split("-")[1]))
        except Exception: pass
    return f"QT-{mx+1:05d}"


def create_quote(customer_id, line_items, valid_days=14, payment_terms=None,
                  notes="", currency="INR", data_file=None):
    """
    line_items: list of {mat_code, mat_desc, uom, qty, unit_price}.
    Deliberately does not require the customer to be KYC/credit-approved —
    see module docstring. Looks up Customer_Name (and a default
    Payment_Terms, if none given) from customer_onboarding.
    """
    fpath = data_file or DATA_FILE
    if not line_items:
        raise ValueError("A quote needs at least one line item.")
    customer = co.get_customer(customer_id, fpath)
    if customer is None:
        raise ValueError(f"Customer {customer_id} not found — onboard them first "
                          "(even a draft record is fine for quoting).")
    customer_name = customer.get("Customer_Name", customer_id)
    if payment_terms is None:
        payment_terms = customer.get("Payment_Terms") or "Net 30"

    total_value = sum(float(li["qty"]) * float(li["unit_price"]) for li in line_items)

    db.init_schema()
    conn = db.get_connection()
    try:
        quote_id = _next_quote_id(conn)
        conn.execute(
            "INSERT INTO quotes (quote_id, customer_id, customer_name, quote_date, valid_until, "
            "payment_terms, status, currency, total_value, notes, filename) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (quote_id, customer_id, customer_name, date.today().strftime("%Y-%m-%d"),
             (date.today() + timedelta(days=valid_days)).strftime("%Y-%m-%d"),
             payment_terms, "Draft", currency, round(total_value, 2), notes, ""),
        )
        for seq, li in enumerate(line_items, 1):
            line_total = round(float(li["qty"]) * float(li["unit_price"]), 2)
            conn.execute(
                "INSERT INTO quote_items (quote_id, line_item, material_code, material_desc, "
                "uom, qty, unit_price, line_total) VALUES (?,?,?,?,?,?,?,?)",
                (quote_id, seq, li["mat_code"], li["mat_desc"], li["uom"],
                 li["qty"], li["unit_price"], line_total),
            )
        conn.commit()
    finally:
        conn.close()
    return quote_id


# ── Document generation ───────────────────────────────────────────────────────
def generate_quote_document(quote_id, data_file=None):
    """The actual quote document — priced line items, no blank fill-in
    columns (we're stating terms, not soliciting them), and an acceptance
    signature block so the customer can sign to accept."""
    fpath = data_file or DATA_FILE
    quote = get_quote(quote_id, fpath)
    if quote is None:
        raise ValueError(f"Quote {quote_id} not found.")
    items = get_quote_items(quote_id, fpath)
    customer = co.get_customer(quote["customer_id"], fpath) or {}

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = "9D174D"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "SALES QUOTATION")
    ws.merge_cells("A1:F1")

    label("A3", "Quote ID:");     value("B3", quote["quote_id"])
    label("A4", "Date:");         value("B4", quote["quote_date"])
    label("A5", "Valid Until:");  value("B5", quote["valid_until"])
    label("D3", "Customer:");     value("E3", quote["customer_name"])
    label("D4", "Contact:");      value("E4", customer.get("Contact_Email", ""))
    label("D5", "Payment Terms:"); value("E5", quote["payment_terms"])

    if quote.get("notes"):
        ws["A7"] = quote["notes"]
        ws["A7"].font = Font(name="Arial", size=9, italic=True, color="64748B")
        ws.merge_cells("A7:F7")

    hdr_row = 9
    hdrs = ["#", "Material Code", "Description", "UOM", "Qty", "Unit Price", "Line Total"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=navy)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for i, item in enumerate(items, 1):
        vals = [i, item["mat_code"], item["mat_desc"], item["uom"],
                item["qty"], item["unit_price"], item["line_total"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 1
    ws.cell(r, 6, "Total:").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 7, quote["total_value"]).font = Font(name="Arial", size=10, bold=True)

    r += 3
    ws.cell(r, 1, "Quoted by (seller)").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 4, "Accepted by (customer)").font = Font(name="Arial", size=10, bold=True)
    r += 3
    ws.cell(r, 1, "Signature: ____________________").font = Font(name="Arial", size=9)
    ws.cell(r, 4, "Signature: ____________________").font = Font(name="Arial", size=9)
    r += 1
    ws.cell(r, 1, "Name / Date: ____________________").font = Font(name="Arial", size=9)
    ws.cell(r, 4, "Name / Date: ____________________").font = Font(name="Arial", size=9)

    widths = [4, 15, 34, 8, 8, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 24

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{quote_id}_{quote['customer_id']}.xlsx"
    return filename, buf.read()


# ── Lifecycle ─────────────────────────────────────────────────────────────────
def _set_status(quote_id, status, data_file=None, filename=None):
    conn = db.get_connection()
    try:
        conn.execute("UPDATE quotes SET status = ? WHERE quote_id = ?", (status, quote_id))
        if filename is not None:
            conn.execute("UPDATE quotes SET filename = ? WHERE quote_id = ?", (filename, quote_id))
        conn.commit()
    finally:
        conn.close()


def mark_sent(quote_id, filename=None, data_file=None):
    _set_status(quote_id, "Sent", data_file, filename)


def record_response(quote_id, accepted, reason="", data_file=None):
    """Records the customer's decision. Refuses on an Expired quote — the
    fix for a lapsed quote is a new one at current pricing, not reviving
    the old one, same reasoning as contracts.py blocking renewal of a
    terminated contract."""
    quote = get_quote(quote_id, data_file)
    if quote is None:
        raise ValueError(f"Quote {quote_id} not found.")
    if quote["status"] == "Expired":
        raise ValueError(f"{quote_id} has expired — create a new quote instead "
                          "of recording a response against a lapsed one.")
    if quote["status"] not in ("Sent",):
        raise ValueError(f"{quote_id} is '{quote['status']}' — only a Sent quote "
                          "can be accepted or rejected.")
    _set_status(quote_id, "Accepted" if accepted else "Rejected", data_file)
    if reason:
        conn = db.get_connection()
        try:
            row = conn.execute("SELECT notes FROM quotes WHERE quote_id = ?", (quote_id,)).fetchone()
            existing = row["notes"] or "" if row else ""
            new_notes = f"{existing} | {reason}".strip(" |")
            conn.execute("UPDATE quotes SET notes = ? WHERE quote_id = ?", (new_notes, quote_id))
            conn.commit()
        finally:
            conn.close()


def simulate_response(quote_id, data_file=None):
    """Demo helper: deterministic accept/reject so re-running the same
    quote_id always gives the same demo outcome. Biased toward acceptance
    (70%) — a quote that never gets accepted makes for a dull demo, and
    real objection handling isn't what this simulation is trying to show."""
    h = int(hashlib.sha1(quote_id.encode()).hexdigest(), 16)
    accepted = (h % 10) < 7
    record_response(quote_id, accepted,
                     "Simulated response (demo)" if accepted else "Simulated decline (demo) — price too high",
                     data_file)
    return accepted


def is_accepted(quote_id, data_file=None):
    """The hook a future Sales Order Management module calls to check
    whether a quote is ready to convert. Kept as a simple boolean function
    rather than this module reaching into an order module that doesn't
    exist yet — same dependency-injection shape used elsewhere in this app."""
    q = get_quote(quote_id, data_file)
    return q is not None and q["status"] == "Accepted"


def stats(data_file=None):
    quotes = get_quotes(data_file=data_file)
    by_status = {}
    for q in quotes:
        by_status[q["status"]] = by_status.get(q["status"], 0) + 1
    total_value = sum(q["total_value"] or 0 for q in quotes)
    accepted_value = sum(q["total_value"] or 0 for q in quotes if q["status"] == "Accepted")
    return {"total": len(quotes), "by_status": by_status,
            "total_value": total_value, "accepted_value": accepted_value}


if __name__ == "__main__":
    print("Suggested price for cost 100:", suggest_sale_price(100))
    print("Quote stats:", stats())
