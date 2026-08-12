"""
sales_order.py — Sales Order Management (O2C).

The piece that finally uses customer_onboarding.check_credit_available() —
the hook that module built specifically for this: quoting a customer never
required credit approval, but committing to fulfill an order does. Every
order this module creates runs the customer through a credit check first;
a customer who fails it still gets an order record, just held (Status =
"Credit Hold") rather than confirmed — the real-world behavior is "review
and release," not "silently reject."

Two ways in, mirroring the quote-vs-standing-agreement split discussed
earlier:
  1. create_order_from_quote() — the primary path. Only converts an
     Accepted quote (checked via quotation.is_accepted()), and only once
     per quote (checked by scanning existing orders' Source_Quote, not by
     touching quotation.py's schema — this module stays a pure consumer
     of quotation.py and customer_onboarding.py, imports only, no reverse
     dependency).
  2. create_direct_order() — for repeat/routine orders with no separate
     quote step, same shape as the S2P side's Direct PO Entry. There's no
     "Customer Price Agreement" module yet (the sell-side mirror of
     contracts.py) to source a pre-negotiated price from, so this prices
     off Item Master cost with the same placeholder markup quotation.py
     uses — a real next module, not pretended to be more than it is.

New sheets:
  Sales_Orders       SO_ID | Customer_ID | Customer_Name | Order_Date |
                      Status | Source_Quote | Payment_Terms | Currency |
                      Total_Value | Delivery_Location |
                      Delivery_Geolocation | Requested_Delivery_Date | Notes
  Sales_Order_Items  SO_ID | Line_Item | Material_Code | Material_Desc |
                      UOM | Qty | Unit_Price | Line_Total

Status values: Confirmed | Credit Hold | Cancelled. is_ready_for_fulfillment()
is the hook a future Fulfillment module calls — same dependency-injection
shape used throughout this app (contracts.py on pr_consolidation.py,
customer_onboarding.py's own check_credit_available()).

SQLite pilot: Sales_Orders and Sales_Order_Items now live in erp_pilot.db
(tables `sales_orders`, `sales_order_items`) — this module is the
exclusive owner of both.
"""

import os, io
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import customer_onboarding as co
import quotation as qt
import atp

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

SO_SHEET = "Sales_Orders"
SO_ITEMS_SHEET = "Sales_Order_Items"

SO_COLS = ["SO_ID", "Customer_ID", "Customer_Name", "Order_Date", "Status",
          "Source_Quote", "Payment_Terms", "Currency", "Total_Value",
          "Delivery_Location", "Delivery_Geolocation", "Requested_Delivery_Date", "Notes"]
SO_ITEM_COLS = ["SO_ID", "Line_Item", "Material_Code", "Material_Desc",
               "UOM", "Qty", "Unit_Price", "Line_Total"]


# ── Sheet bootstrap ────────────────────────────────────────────────────────────
def ensure_sheets(wb=None):
    """Kept for signature compatibility — Sales_Orders/Sales_Order_Items
    no longer live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


# ── Read ──────────────────────────────────────────────────────────────────────
def get_orders(status=None, customer_id=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM sales_orders ORDER BY so_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"so_id": r["so_id"], "customer_id": r["customer_id"],
               "customer_name": r["customer_name"], "order_date": r["order_date"],
               "status": r["status"], "source_quote": r["source_quote"],
               "payment_terms": r["payment_terms"], "currency": r["currency"],
               "total_value": r["total_value"], "delivery_location": r["delivery_location"],
               "delivery_geo": r["delivery_geolocation"],
               "requested_delivery_date": r["requested_delivery_date"], "notes": r["notes"]}
        if status and row["status"] != status:
            continue
        if customer_id and row["customer_id"] != customer_id:
            continue
        out.append(row)
    return out


def get_order(so_id, data_file=None):
    for o in get_orders(data_file=data_file):
        if o["so_id"] == so_id:
            return o
    return None


def get_order_items(so_id, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM sales_order_items WHERE so_id = ? ORDER BY line_item", (so_id,)
        ).fetchall()
    finally:
        conn.close()
    return [{"so_id": r["so_id"], "line_item": r["line_item"], "mat_code": r["material_code"],
             "mat_desc": r["material_desc"], "uom": r["uom"], "qty": r["qty"],
             "unit_price": r["unit_price"], "line_total": r["line_total"],
             "atp_outcome": r["atp_outcome"], "promised_qty": r["promised_qty"],
             "backordered_qty": r["backordered_qty"], "reservation_id": r["reservation_id"],
             "backorder_id": r["backorder_id"]}
            for r in rows]


def quote_already_converted(quote_id, data_file=None):
    """True if some order already sources from this quote — enforces one
    order per quote without touching quotation.py's schema."""
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT 1 FROM sales_orders WHERE source_quote = ? LIMIT 1", (quote_id,)
        ).fetchone()
    finally:
        conn.close()
    return row is not None


# ── Create ────────────────────────────────────────────────────────────────────
def _next_so_id(conn):
    rows = conn.execute("SELECT so_id FROM sales_orders WHERE so_id LIKE 'SO-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["so_id"].split("-")[1]))
        except Exception: pass
    return f"SO-{mx+1:05d}"


def _run_credit_check(customer_id, total_value, data_file):
    result = co.check_credit_available(customer_id, total_value, data_file)
    return ("Confirmed", result["reason"]) if result["ok"] else ("Credit Hold", result["reason"])


def _create_order(customer_id, line_items, source_quote, delivery_location,
                   delivery_geo, requested_delivery_date, notes, payment_terms, data_file):
    fpath = data_file or DATA_FILE
    if not line_items:
        raise ValueError("An order needs at least one line item.")
    customer = co.get_customer(customer_id, fpath)
    if customer is None:
        raise ValueError(f"Customer {customer_id} not found.")

    total_value = round(sum(float(li["qty"]) * float(li["unit_price"]) for li in line_items), 2)
    status, reason = _run_credit_check(customer_id, total_value, fpath)

    db.init_schema()
    conn = db.get_connection()
    try:
        # Real, pre-existing race condition, found and fixed here: two
        # concurrent order-creation calls could both read the same
        # "highest existing SO number" before either inserted, and both
        # try to claim the same next so_id -- a real UNIQUE constraint
        # violation, not a hypothetical one (found via a real concurrent
        # stress test built to verify ATP-US-01's own atomicity claim,
        # not something this story set out to fix on its own). The same
        # real BEGIN IMMEDIATE atomic pattern reservation.py's own
        # create_reservation_up_to() already uses and has already proven
        # under real concurrent load.
        conn.execute("BEGIN IMMEDIATE")
        so_id = _next_so_id(conn)
        conn.execute(
            "INSERT INTO sales_orders (so_id, customer_id, customer_name, order_date, status, "
            "source_quote, payment_terms, currency, total_value, delivery_location, "
            "delivery_geolocation, requested_delivery_date, notes) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (so_id, customer_id, customer.get("Customer_Name", customer_id),
             date.today().strftime("%Y-%m-%d"), status, source_quote or "",
             payment_terms or customer.get("Payment_Terms") or "Net 30", "INR",
             total_value, delivery_location, delivery_geo,
             str(requested_delivery_date) if requested_delivery_date else "", notes),
        )
        for seq, li in enumerate(line_items, 1):
            line_total = round(float(li["qty"]) * float(li["unit_price"]), 2)
            conn.execute(
                "INSERT INTO sales_order_items (so_id, line_item, material_code, material_desc, "
                "uom, qty, unit_price, line_total) VALUES (?,?,?,?,?,?,?,?)",
                (so_id, seq, li["mat_code"], li["mat_desc"], li["uom"],
                 li["qty"], li["unit_price"], line_total),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    # ATP-US-01's own real check-and-promise, one line at a time, run
    # independently of the credit check above -- neither blocks the
    # other, matching the story's own Preconditions exactly. Runs even
    # for a Credit Hold order: this check's own job is only to
    # determine what's genuinely available and reserve it, not to
    # decide whether the order is allowed to ship.
    atp_outcomes = []
    for li in line_items:
        result = atp.check_and_promise_line(so_id, line_items.index(li) + 1,
            li["mat_code"], li["mat_desc"], delivery_location, float(li["qty"]))
        atp_outcomes.append(result)

    conn = db.get_connection()
    try:
        for seq, result in enumerate(atp_outcomes, 1):
            conn.execute(
                "UPDATE sales_order_items SET atp_outcome=?, promised_qty=?, "
                "backordered_qty=?, reservation_id=?, backorder_id=? WHERE so_id=? AND line_item=?",
                (result["outcome"], result["promised_qty"], result["backordered_qty"],
                 result["reservation_id"], result["backorder_id"], so_id, seq))
        conn.commit()
    finally:
        conn.close()

    return {"so_id": so_id, "status": status, "credit_reason": reason,
           "total_value": total_value, "atp_outcomes": atp_outcomes}


def create_order_from_quote(quote_id, delivery_location="", delivery_geo="",
                             requested_delivery_date=None, notes="", data_file=None):
    fpath = data_file or DATA_FILE
    if not qt.is_accepted(quote_id, fpath):
        raise ValueError(f"{quote_id} is not Accepted — only an accepted quote can become an order.")
    if quote_already_converted(quote_id, fpath):
        raise ValueError(f"{quote_id} has already been converted to a Sales Order.")
    quote = qt.get_quote(quote_id, fpath)
    items = qt.get_quote_items(quote_id, fpath)
    line_items = [{"mat_code": i["mat_code"], "mat_desc": i["mat_desc"], "uom": i["uom"],
                   "qty": i["qty"], "unit_price": i["unit_price"]} for i in items]
    return _create_order(quote["customer_id"], line_items, quote_id, delivery_location,
                         delivery_geo, requested_delivery_date, notes or quote.get("notes", ""),
                         quote["payment_terms"], fpath)


def create_direct_order(customer_id, line_items, delivery_location="", delivery_geo="",
                         requested_delivery_date=None, notes="", payment_terms=None, data_file=None):
    return _create_order(customer_id, line_items, None, delivery_location, delivery_geo,
                         requested_delivery_date, notes, payment_terms, data_file)


# ── Lifecycle ─────────────────────────────────────────────────────────────────
def release_credit_hold(so_id, data_file=None):
    """Re-runs the credit check (e.g. after the customer's limit was
    raised or a hold was manually released). Only actually flips to
    Confirmed if the check now passes — doesn't just force it through."""
    fpath = data_file or DATA_FILE
    order = get_order(so_id, fpath)
    if order is None:
        raise ValueError(f"{so_id} not found.")
    if order["status"] != "Credit Hold":
        raise ValueError(f"{so_id} is '{order['status']}', not on Credit Hold.")
    status, reason = _run_credit_check(order["customer_id"], order["total_value"], fpath)

    conn = db.get_connection()
    try:
        conn.execute("UPDATE sales_orders SET status = ? WHERE so_id = ?", (status, so_id))
        conn.commit()
    finally:
        conn.close()
    return {"status": status, "reason": reason}


def cancel_order(so_id, reason="", data_file=None):
    """
    Cancelling the order itself is only half of it — found directly
    (2026-08-11) as a real bug, not a refinement: this used to flip
    only sales_orders.status, leaving every one of this order's own
    Open reservations and Open/Partially Fulfilled backorders sitting
    there untouched. Since bom.get_inventory_position() and
    reservation.get_available_to_promise() both net real on-hand
    against every currently-Open reservation regardless of whether its
    own Sales Order still exists, a cancelled order's reservation went
    on suppressing real, physically-available stock forever — for
    Position & Transfers' own on-hand figure and for every future
    order's real-time ATP check alike, not just a stale display.
    reservation.release_reservation()'s own docstring already names
    "a Sales Order line was cancelled" as its intended trigger; this is
    that wiring, not new policy.

    Backorders are cancelled *before* reservations are released, not
    the other order (2026-08-11) — release_reservation() itself now
    re-triggers backorder.reevaluate_backorders() for that same
    (material, location), since a released reservation is a real
    increase in Available-to-Promise other open backorders deserve a
    shot at, the same as new supply arriving. If this order's own
    backorder were still Open at that moment, that re-check could hand
    it a split-second reservation of its own, immediately orphaned the
    instant this loop then cancels it. Cancelling first removes it from
    the re-check's own candidate pool entirely, closing that window.
    """
    conn = db.get_connection()
    try:
        row = conn.execute("SELECT notes FROM sales_orders WHERE so_id = ?", (so_id,)).fetchone()
        if row is not None:
            existing = row["notes"] or ""
            new_notes = f"{existing} | Cancelled: {reason}".strip(" |")
            conn.execute(
                "UPDATE sales_orders SET status='Cancelled', notes=? WHERE so_id=?",
                (new_notes, so_id),
            )
            conn.commit()
    finally:
        conn.close()

    import reservation as res
    import backorder as bo
    cancel_note = f"{so_id} cancelled" + (f": {reason}" if reason else "")
    conn = db.get_connection()
    try:
        open_reservations = conn.execute(
            "SELECT reservation_id FROM reservations WHERE so_id=? AND status='Open'",
            (so_id,)).fetchall()
        open_backorders = conn.execute(
            "SELECT backorder_id FROM backorders WHERE so_id=? AND status IN "
            "('Open', 'Partially Fulfilled')", (so_id,)).fetchall()
    finally:
        conn.close()
    for b in open_backorders:
        bo.cancel_backorder(b["backorder_id"], reason=cancel_note)
    for r in open_reservations:
        res.release_reservation(r["reservation_id"], reason=cancel_note)


def is_ready_for_fulfillment(so_id, data_file=None):
    """The hook a future Fulfillment module calls."""
    o = get_order(so_id, data_file)
    return o is not None and o["status"] == "Confirmed"


# ── Document generation ───────────────────────────────────────────────────────
def generate_order_confirmation(so_id, data_file=None):
    fpath = data_file or DATA_FILE
    order = get_order(so_id, fpath)
    if order is None:
        raise ValueError(f"{so_id} not found.")
    items = get_order_items(so_id, fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = "1E3A5F"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "SALES ORDER CONFIRMATION")
    ws.merge_cells("A1:F1")

    label("A3", "Order ID:");    value("B3", order["so_id"])
    label("A4", "Date:");        value("B4", order["order_date"])
    label("A5", "Status:");      value("B5", order["status"])
    label("D3", "Customer:");    value("E3", order["customer_name"])
    label("D4", "Delivery:");    value("E4", order["delivery_location"] or "TBD")
    label("D5", "Payment Terms:"); value("E5", order["payment_terms"])
    if order.get("source_quote"):
        label("A6", "Source Quote:"); value("B6", order["source_quote"])

    hdr_row = 9
    hdrs = ["#", "Material Code", "Description", "UOM", "Qty", "Unit Price", "Line Total"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=navy)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for i, item in enumerate(items, 1):
        vals = [i, item["mat_code"], item["mat_desc"], item["uom"],
                item["qty"], item["unit_price"], item["line_total"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 1
    ws.cell(r, 6, "Total:").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 7, order["total_value"]).font = Font(name="Arial", size=10, bold=True)

    widths = [4, 15, 34, 8, 8, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 24

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{so_id}_{order['customer_id']}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    orders = get_orders(data_file=data_file)
    by_status = {}
    for o in orders:
        by_status[o["status"]] = by_status.get(o["status"], 0) + 1
    total_value = sum(o["total_value"] or 0 for o in orders)
    confirmed_value = sum(o["total_value"] or 0 for o in orders if o["status"] == "Confirmed")
    return {"total": len(orders), "by_status": by_status,
            "total_value": total_value, "confirmed_value": confirmed_value}


if __name__ == "__main__":
    print("Sales Order stats:", stats())
