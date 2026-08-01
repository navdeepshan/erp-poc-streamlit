"""
migrate_o2c_sales_order.py — one-time port of Sales_Orders and
Sales_Order_Items from data.xlsx into erp_pilot.db.

Third O2C slice — depends on Customer_Master and Quotes already being
migrated. Idempotent.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_orders(wb, conn):
    if "Sales_Orders" not in wb.sheetnames:
        return 0
    ws = wb["Sales_Orders"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[hdrs.get("SO_ID", 0)] or "").strip()
        if not sid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((sid, g("Customer_ID"), g("Customer_Name"), g("Order_Date"), g("Status"),
                    g("Source_Quote"), g("Payment_Terms"), g("Currency"), g("Total_Value"),
                    g("Delivery_Location"), g("Delivery_Geolocation"),
                    g("Requested_Delivery_Date"), g("Notes")))
    conn.execute("DELETE FROM sales_orders")
    conn.executemany(
        "INSERT INTO sales_orders (so_id, customer_id, customer_name, order_date, status, "
        "source_quote, payment_terms, currency, total_value, delivery_location, "
        "delivery_geolocation, requested_delivery_date, notes) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_order_items(wb, conn):
    if "Sales_Order_Items" not in wb.sheetnames:
        return 0
    ws = wb["Sales_Order_Items"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[hdrs.get("SO_ID", 0)] or "").strip()
        if not sid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((sid, g("Line_Item"), g("Material_Code"), g("Material_Desc"),
                    g("UOM"), g("Qty"), g("Unit_Price"), g("Line_Total")))
    conn.execute("DELETE FROM sales_order_items")
    conn.executemany(
        "INSERT INTO sales_order_items (so_id, line_item, material_code, material_desc, uom, "
        "qty, unit_price, line_total) VALUES (?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_so = migrate_orders(wb, conn)
        n_soi = migrate_order_items(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"sales_orders": n_so, "sales_order_items": n_soi}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Sales_Orders/Sales_Order_Items to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--db", default=None)
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
