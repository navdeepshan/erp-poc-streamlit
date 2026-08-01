"""
migrate_reference_data.py — one-time port of the reference/master-data
tables (Item Master, Vendor_Master, Delivery_Locations, Vendor_Documents)
from data.xlsx into erp_pilot.db.

Idempotent: clears each table before inserting, so re-running against a
refreshed data.xlsx doesn't duplicate rows.

Item Master and Vendor_Master were both extended additively over the
project's Excel life (HSN_Code/GST_Rate on Item Master; GSTIN/PAN/Bank/
KYC columns on Vendor_Master) — this migration reads by column NAME
from the Excel header row, not position, so it's correct regardless of
which columns exist or what order they're in.

Note on `Active`: all three master tables (Item Master, Vendor_Master,
Delivery_Locations) already carry an Active column in the source Excel
— this migration carries it over as-is (values like "Yes"/"No"). It
does not invent a new column; the earlier assumption that this project
needed one added was wrong, confirmed by inspecting the actual sheets.
Vendor_Documents is a document *log*, not master data, and deliberately
has no Active column, matching its source sheet.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws, header_row):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))}


def migrate_item_master(wb, conn):
    ws = wb["Item Master"]
    hdrs = _headers(ws, 2)  # title row 1, header row 2, data row 3+
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = str(row[hdrs.get("Item Code", 0)] or "").strip()
        if not code:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((
            code, g("Item Description"), g("Category"), g("Sub-Category"),
            g("Unit of Measure"), g("Unit Price"), g("Lead Time (days)"),
            g("In Stock"), g("Tags / Keywords"), g("Active"),
            g("HSN_Code"), g("GST_Rate"),
        ))
    conn.execute("DELETE FROM item_master")
    conn.executemany(
        "INSERT INTO item_master (item_code, item_desc, category, subcategory, uom, "
        "unit_price, lead_time_days, in_stock, tags, active, hsn_code, gst_rate) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_vendor_master(wb, conn):
    ws = wb["Vendor_Master"]
    hdrs = _headers(ws, 1)  # no title row — header row 1, data row 2+
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vid = str(row[hdrs.get("Vendor_ID", 0)] or "").strip()
        if not vid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((
            vid, g("Vendor_Name"), g("Geolocation"), g("City"), g("Country"),
            g("Address"), g("Contact_Name"), g("Contact_Email"), g("Active"),
            g("GSTIN"), g("PAN"), g("Bank_Account_No"), g("IFSC"), g("Bank_Name"),
            g("Bank_Branch"), g("Onboarding_Status"), g("KYC_Flag"), g("Onboarded_Date"),
        ))
    conn.execute("DELETE FROM vendor_master")
    conn.executemany(
        "INSERT INTO vendor_master (vendor_id, vendor_name, geolocation, city, country, "
        "address, contact_name, contact_email, active, gstin, pan, bank_account_no, ifsc, "
        "bank_name, bank_branch, onboarding_status, kyc_flag, onboarded_date) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_delivery_locations(wb, conn):
    ws = wb["Delivery_Locations"]
    hdrs = _headers(ws, 1)  # no title row — header row 1, data row 2+
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lid = str(row[hdrs.get("Location_ID", 0)] or "").strip()
        if not lid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((
            lid, g("Location_Name"), g("Geolocation"), g("City"), g("State"),
            g("Country"), g("Address"), g("Active"),
        ))
    conn.execute("DELETE FROM delivery_locations")
    conn.executemany(
        "INSERT INTO delivery_locations (location_id, location_name, geolocation, city, "
        "state, country, address, active) VALUES (?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_vendor_documents(wb, conn):
    if "Vendor_Documents" not in wb.sheetnames:
        return 0
    ws = wb["Vendor_Documents"]
    hdrs = _headers(ws, 1)  # no title row — header row 1, data row 2+
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        did = str(row[hdrs.get("Document_ID", 0)] or "").strip()
        if not did:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((
            did, g("Vendor_ID"), g("Doc_Type"), g("Filename"),
            g("Uploaded_Date"), g("Status"), g("Notes"),
        ))
    conn.execute("DELETE FROM vendor_documents")
    conn.executemany(
        "INSERT INTO vendor_documents (document_id, vendor_id, doc_type, filename, "
        "uploaded_date, status, notes) VALUES (?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_items = migrate_item_master(wb, conn)
        n_vendors = migrate_vendor_master(wb, conn)
        n_locs = migrate_delivery_locations(wb, conn)
        n_docs = migrate_vendor_documents(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"item_master": n_items, "vendor_master": n_vendors,
            "delivery_locations": n_locs, "vendor_documents": n_docs}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate reference/master data from Excel to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to source data.xlsx")
    parser.add_argument("--db", default=None, help="Path to target SQLite db (default: erp_pilot.db)")
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
