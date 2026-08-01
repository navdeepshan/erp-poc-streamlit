"""
billing.py — Billing & Invoicing (O2C).

The payoff of the two prerequisite modules: org_profile.py (who's billing)
and item_tax.py (what tax applies to each line). Both are checked here,
not assumed — create_invoice() refuses to run if the seller has no valid
GSTIN on file, and refuses per-line if that line's HSN code or GST rate
isn't set. Neither failure is silent; both come back as a clear
ValueError naming exactly what's missing, so fixing it is a next action,
not a mystery.

CGST/SGST vs IGST is fully mechanical and safe to automate: both the
seller's and the buyer's GSTIN encode a state code in their first two
characters (already parsed by vendor_onboarding.validate_gstin's details
dict). Same state -> CGST+SGST split. Different state -> IGST. No
judgement call involved, unlike HSN classification.

Known scope limit, stated rather than papered over: this only invoices
customers who have a valid GSTIN on file (a B2B assumption, reasonable
for hospitals/clinics/distributors). A customer with no GSTIN — an
unregistered individual practitioner, say — can't get an invoice from
this module yet; that needs separate B2C place-of-supply handling this
doesn't attempt.

Billing off Fulfillment (not the Sales Order) so quantities reflect what
was actually shipped, not what was ordered — a partial shipment gets
invoiced for what shipped, matching real practice.

New sheets:
  Invoices        Invoice_ID | Fulfillment_ID | SO_ID | Customer_ID |
                   Customer_Name | Customer_GSTIN | Invoice_Date | Due_Date |
                   Status | Payment_Terms | Currency | Place_of_Supply |
                   Subtotal | CGST_Total | SGST_Total | IGST_Total |
                   Grand_Total | Notes
  Invoice_Items    Invoice_ID | Line_Item | Material_Code | Material_Desc |
                   HSN_Code | UOM | Qty | Unit_Price | Taxable_Value |
                   GST_Rate | CGST_Amount | SGST_Amount | IGST_Amount |
                   Line_Total

Status: Draft -> Issued -> Cancelled. is_ready_for_payment() is the hook
a future Cash Application module calls — same DI shape as everywhere else.

SQLite pilot: Invoices and Invoice_Items now live in erp_pilot.db
(tables `invoices`, `invoice_items`) — this module is the exclusive
owner of both.
"""

import os, io
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import fulfillment as ful
import sales_order as so
import customer_onboarding as co
import item_tax as it
import org_profile as op
import vendor_onboarding as vo

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

INV_SHEET = "Invoices"
INV_ITEMS_SHEET = "Invoice_Items"

INV_COLS = ["Invoice_ID", "Fulfillment_ID", "SO_ID", "Customer_ID", "Customer_Name",
           "Customer_GSTIN", "Invoice_Date", "Due_Date", "Status", "Payment_Terms",
           "Currency", "Place_of_Supply", "Subtotal", "CGST_Total", "SGST_Total",
           "IGST_Total", "Grand_Total", "Notes"]
INV_ITEM_COLS = ["Invoice_ID", "Line_Item", "Material_Code", "Material_Desc", "HSN_Code",
                 "UOM", "Qty", "Unit_Price", "Taxable_Value", "GST_Rate",
                 "CGST_Amount", "SGST_Amount", "IGST_Amount", "Line_Total"]


def ensure_sheets(wb=None):
    """Kept for signature compatibility — Invoices/Invoice_Items no
    longer live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


# ── Read ──────────────────────────────────────────────────────────────────────
def get_invoices(status=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM invoices ORDER BY invoice_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"invoice_id": r["invoice_id"], "fulfillment_id": r["fulfillment_id"],
               "so_id": r["so_id"], "customer_id": r["customer_id"],
               "customer_name": r["customer_name"], "customer_gstin": r["customer_gstin"],
               "invoice_date": r["invoice_date"], "due_date": r["due_date"],
               "status": r["status"], "payment_terms": r["payment_terms"],
               "currency": r["currency"], "place_of_supply": r["place_of_supply"],
               "subtotal": r["subtotal"], "cgst_total": r["cgst_total"],
               "sgst_total": r["sgst_total"], "igst_total": r["igst_total"],
               "grand_total": r["grand_total"], "notes": r["notes"]}
        if status and row["status"] != status:
            continue
        out.append(row)
    return out


def get_invoice(invoice_id, data_file=None):
    for inv in get_invoices(data_file=data_file):
        if inv["invoice_id"] == invoice_id:
            return inv
    return None


def get_invoice_items(invoice_id, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM invoice_items WHERE invoice_id = ? ORDER BY line_item", (invoice_id,)
        ).fetchall()
    finally:
        conn.close()
    return [{"invoice_id": r["invoice_id"], "line_item": r["line_item"],
             "mat_code": r["material_code"], "mat_desc": r["material_desc"],
             "hsn_code": r["hsn_code"], "uom": r["uom"], "qty": r["qty"],
             "unit_price": r["unit_price"], "taxable_value": r["taxable_value"],
             "gst_rate": r["gst_rate"], "cgst_amount": r["cgst_amount"],
             "sgst_amount": r["sgst_amount"], "igst_amount": r["igst_amount"],
             "line_total": r["line_total"]} for r in rows]


def fulfillment_already_invoiced(fulfillment_id, data_file=None):
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT 1 FROM invoices WHERE fulfillment_id = ? AND status != 'Cancelled' LIMIT 1",
            (fulfillment_id,),
        ).fetchone()
    finally:
        conn.close()
    return row is not None


# ── GST split — mechanical, no judgement involved ────────────────────────────
def _gst_split(taxable_value, gst_rate, buyer_state, seller_state):
    if buyer_state == seller_state:
        cgst = round(taxable_value * gst_rate / 2 / 100, 2)
        return cgst, cgst, 0.0
    return 0.0, 0.0, round(taxable_value * gst_rate / 100, 2)


# ── Create ────────────────────────────────────────────────────────────────────
def _next_invoice_id(conn):
    rows = conn.execute("SELECT invoice_id FROM invoices WHERE invoice_id LIKE 'INV-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["invoice_id"].split("-")[1]))
        except Exception: pass
    return f"INV-{mx+1:05d}"


def create_invoice(fulfillment_id, due_days=30, notes="", data_file=None):
    fpath = data_file or DATA_FILE

    if not op.is_configured(fpath):
        raise ValueError("Your organization profile has no valid GSTIN on file — "
                          "set that up before creating invoices.")
    if not ful.is_ready_for_billing(fulfillment_id, fpath):
        raise ValueError(f"{fulfillment_id} is not Delivered — only a delivered "
                          "fulfillment can be invoiced.")
    if fulfillment_already_invoiced(fulfillment_id, fpath):
        raise ValueError(f"{fulfillment_id} has already been invoiced.")

    f = ful.get_fulfillment(fulfillment_id, fpath)
    f_items = [i for i in ful.get_fulfillment_items(fulfillment_id, fpath) if (i["qty_shipped"] or 0) > 0]
    if not f_items:
        raise ValueError(f"{fulfillment_id} has no shipped quantity on any line — nothing to invoice.")

    order = so.get_order(f["so_id"], fpath)
    order_items = {i["mat_code"]: i["unit_price"] for i in so.get_order_items(f["so_id"], fpath)}

    customer = co.get_customer(f["customer_id"], fpath)
    if not customer or not customer.get("GSTIN"):
        raise ValueError(f"{f['customer_id']} has no GSTIN on file — this module only "
                          "invoices GST-registered (B2B) customers; unregistered/B2C "
                          "invoicing isn't built yet.")
    buyer_ok, buyer_msg, buyer_details = vo.validate_gstin(customer["GSTIN"])
    if not buyer_ok:
        raise ValueError(f"{f['customer_id']}'s GSTIN on file is invalid ({buyer_msg}) — fix it "
                          "in Customer Onboarding before invoicing.")

    org = op.get_org_profile(fpath)
    seller_ok, seller_msg, seller_details = vo.validate_gstin(org["GSTIN"])
    if not seller_ok:
        raise ValueError(f"Your organization's GSTIN is invalid ({seller_msg}) — fix it "
                          "in the Org Profile before invoicing.")

    # Gate on every line having HSN + GST rate before writing anything
    missing = []
    line_data = []
    for item in f_items:
        tax = it.get_item_tax_info(item["mat_code"], fpath)
        if not tax["hsn_code"] or tax["gst_rate"] is None:
            missing.append(item["mat_code"])
            continue
        unit_price = order_items.get(item["mat_code"])
        if unit_price is None:
            raise ValueError(f"No price found for {item['mat_code']} on {f['so_id']} — "
                              "data inconsistency, can't invoice.")
        line_data.append({"mat_code": item["mat_code"], "mat_desc": item["mat_desc"],
                          "uom": item["uom"], "qty": item["qty_shipped"],
                          "unit_price": unit_price, "hsn_code": tax["hsn_code"],
                          "gst_rate": float(tax["gst_rate"])})
    if missing:
        raise ValueError(f"These item(s) have no HSN code / GST rate configured yet — "
                         f"set them in Item Tax before invoicing: {', '.join(missing)}")

    buyer_state = buyer_details["state_code"]
    seller_state = seller_details["state_code"]
    place_of_supply = buyer_details["state_name"]

    subtotal = cgst_t = sgst_t = igst_t = 0.0
    computed_lines = []
    for ld in line_data:
        taxable = round(ld["qty"] * ld["unit_price"], 2)
        cgst, sgst, igst = _gst_split(taxable, ld["gst_rate"], buyer_state, seller_state)
        line_total = round(taxable + cgst + sgst + igst, 2)
        subtotal += taxable; cgst_t += cgst; sgst_t += sgst; igst_t += igst
        computed_lines.append({**ld, "taxable_value": taxable, "cgst": cgst, "sgst": sgst,
                               "igst": igst, "line_total": line_total})
    grand_total = round(subtotal + cgst_t + sgst_t + igst_t, 2)

    db.init_schema()
    conn = db.get_connection()
    try:
        invoice_id = _next_invoice_id(conn)
        conn.execute(
            "INSERT INTO invoices (invoice_id, fulfillment_id, so_id, customer_id, "
            "customer_name, customer_gstin, invoice_date, due_date, status, payment_terms, "
            "currency, place_of_supply, subtotal, cgst_total, sgst_total, igst_total, "
            "grand_total, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (invoice_id, fulfillment_id, f["so_id"], f["customer_id"], f["customer_name"],
             customer["GSTIN"], date.today().strftime("%Y-%m-%d"),
             (date.today() + timedelta(days=due_days)).strftime("%Y-%m-%d"), "Draft",
             order["payment_terms"], "INR", place_of_supply, round(subtotal, 2),
             round(cgst_t, 2), round(sgst_t, 2), round(igst_t, 2), grand_total, notes),
        )
        for seq, ld in enumerate(computed_lines, 1):
            conn.execute(
                "INSERT INTO invoice_items (invoice_id, line_item, material_code, "
                "material_desc, hsn_code, uom, qty, unit_price, taxable_value, gst_rate, "
                "cgst_amount, sgst_amount, igst_amount, line_total) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (invoice_id, seq, ld["mat_code"], ld["mat_desc"], ld["hsn_code"], ld["uom"],
                 ld["qty"], ld["unit_price"], ld["taxable_value"], ld["gst_rate"],
                 ld["cgst"], ld["sgst"], ld["igst"], ld["line_total"]),
            )
        conn.commit()
    finally:
        conn.close()
    return {"invoice_id": invoice_id, "grand_total": grand_total,
            "place_of_supply": place_of_supply,
            "tax_type": "CGST+SGST" if buyer_state == seller_state else "IGST"}


# ── Lifecycle ─────────────────────────────────────────────────────────────────
def mark_issued(invoice_id, data_file=None):
    inv = get_invoice(invoice_id, data_file)
    if inv is None:
        raise ValueError(f"{invoice_id} not found.")
    if inv["status"] != "Draft":
        raise ValueError(f"{invoice_id} is '{inv['status']}' — only a Draft invoice can be issued.")
    _set_status(invoice_id, "Issued", data_file)


def cancel_invoice(invoice_id, reason="", data_file=None):
    inv = get_invoice(invoice_id, data_file)
    if inv is None:
        raise ValueError(f"{invoice_id} not found.")
    conn = db.get_connection()
    try:
        existing = inv["notes"] or ""
        new_notes = f"{existing} | Cancelled: {reason}".strip(" |")
        conn.execute(
            "UPDATE invoices SET status='Cancelled', notes=? WHERE invoice_id=?",
            (new_notes, invoice_id),
        )
        conn.commit()
    finally:
        conn.close()


def _set_status(invoice_id, status, data_file=None):
    conn = db.get_connection()
    try:
        conn.execute("UPDATE invoices SET status = ? WHERE invoice_id = ?", (status, invoice_id))
        conn.commit()
    finally:
        conn.close()


def is_ready_for_payment(invoice_id, data_file=None):
    """The hook a future Cash Application module calls."""
    inv = get_invoice(invoice_id, data_file)
    return inv is not None and inv["status"] == "Issued"


# ── Document generation ───────────────────────────────────────────────────────
def generate_invoice_document(invoice_id, data_file=None):
    fpath = data_file or DATA_FILE
    inv = get_invoice(invoice_id, fpath)
    if inv is None:
        raise ValueError(f"{invoice_id} not found.")
    items = get_invoice_items(invoice_id, fpath)
    org = op.get_org_profile(fpath)
    customer = co.get_customer(inv["customer_id"], fpath) or {}

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    maroon = "7C2D12"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Invoice"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "TAX INVOICE")
    ws.merge_cells("A1:G1")

    label("A3", "From:")
    value("A4", org.get("Legal_Name", ""))
    value("A5", f"GSTIN: {org.get('GSTIN','')}")
    value("A6", org.get("Address", ""))

    label("D3", "Bill To:")
    value("D4", inv["customer_name"])
    value("D5", f"GSTIN: {inv['customer_gstin']}")
    value("D6", customer.get("Address", ""))

    label("A8", "Invoice No:");   value("B8", inv["invoice_id"])
    label("A9", "Date:");         value("B9", inv["invoice_date"])
    label("A10", "Due Date:");    value("B10", inv["due_date"])
    label("D8", "Place of Supply:"); value("E8", inv["place_of_supply"])
    label("D9", "Payment Terms:");   value("E9", inv["payment_terms"])
    label("D10", "Fulfillment:");    value("E10", inv["fulfillment_id"])

    hdr_row = 13
    hdrs = ["#", "HSN", "Description", "UOM", "Qty", "Rate", "Taxable Value",
            "GST%", "CGST", "SGST", "IGST", "Total"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=maroon)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = hdr_row + 1
    for i, item in enumerate(items, 1):
        vals = [i, item["hsn_code"], item["mat_desc"], item["uom"], item["qty"],
                item["unit_price"], item["taxable_value"], item["gst_rate"],
                item["cgst_amount"] or 0, item["sgst_amount"] or 0, item["igst_amount"] or 0,
                item["line_total"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 1
    for lbl, val in [("Subtotal", inv["subtotal"]), ("CGST", inv["cgst_total"]),
                      ("SGST", inv["sgst_total"]), ("IGST", inv["igst_total"]),
                      ("Grand Total", inv["grand_total"])]:
        ws.cell(r, 11, lbl).font = Font(name="Arial", size=9, bold=(lbl == "Grand Total"))
        ws.cell(r, 12, val).font = Font(name="Arial", size=9, bold=(lbl == "Grand Total"))
        r += 1

    r += 1
    label(f"A{r}", "Payment to:")
    value(f"A{r+1}", f"{org.get('Bank_Name','')} — A/C {org.get('Bank_Account_No','')} — "
                     f"IFSC {org.get('IFSC','')}")
    r += 3
    ws.cell(r, 1, "This is a computer-generated invoice.").font = Font(
        name="Arial", size=8, italic=True, color="94A3B8")

    widths = [4, 10, 28, 7, 6, 10, 12, 6, 9, 9, 9, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 28

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{invoice_id}_{inv['customer_id']}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    invoices = get_invoices(data_file=data_file)
    by_status = {}
    for inv in invoices:
        by_status[inv["status"]] = by_status.get(inv["status"], 0) + 1
    total_value = sum(inv["grand_total"] or 0 for inv in invoices if inv["status"] != "Cancelled")
    return {"total": len(invoices), "by_status": by_status, "total_value": total_value}


if __name__ == "__main__":
    print("Billing stats:", stats())
    print("Org configured:", op.is_configured())
    print("Items missing tax info:", len(it.list_items_missing_tax_info()))
