"""
rfx.py — RFx (Request-for-X) management, vendor-centric.

Picks up where pr_consolidation.py leaves off: any PR line with no
identified vendor already lands in the RFP sheet as an open RFP line.
This module adds the next stage of Source-to-Contract:

  Open RFP lines -> invite vendor(s) -> generate RFQ document
                 -> collect quotes -> select winner per line
                 -> generate POs (consolidated per vendor)

Design note — vendor-centric, not line-centric:
  Real e-sourcing tools (and your own pr_consolidation.py PO logic) treat
  a vendor as the unit of a document/PO, not a single line item. If a
  vendor is being asked to quote on 3 items, they get ONE RFQ document
  with 3 lines, and if they win all 3, they get ONE PO with 3 line items
  — not three separate RFQs/POs. This module is built around that:
    - invite_and_generate_rfq() takes a vendor + a *list* of RFP lines
    - select_winner() marks a winner per line but does NOT create a PO
    - generate_pos() is the batch step: it looks at every line with a
      selected-but-not-yet-issued winner, groups them by vendor, and
      creates one consolidated PO per vendor.

New sheets created on first use:
  RFx_Quotes       Quote_ID | RFP_Number | Vendor_ID | Vendor_Name |
                   Quoted_Price | Lead_Time_Days | MOQ | Quote_Date |
                   Status | Notes
  RFx_Invitations  Invitation_ID | RFP_Number | Vendor_ID | Vendor_Name |
                   Invited_Date | Filename

RFP sheet Status values used by this module: Open -> Selected (winner
chosen, PO not yet issued) -> Awarded (PO created).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, io, math, hashlib
from datetime import datetime, date, timedelta
from collections import defaultdict

import pr_consolidation as pc  # reuse PO-writing / PR-writeback / style helpers
import db
import vendor_onboarding as vo
import po_export

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

RFX_QUOTES_SHEET = "RFx_Quotes"
RFXQ_COLS = ["Quote_ID", "RFP_Number", "Vendor_ID", "Vendor_Name", "Quoted_Price",
             "Lead_Time_Days", "MOQ", "Quote_Date", "Status", "Notes"]

RFX_INVITATIONS_SHEET = "RFx_Invitations"
RFXI_COLS = ["Invitation_ID", "RFP_Number", "Vendor_ID", "Vendor_Name",
             "Invited_Date", "Filename"]

RFP_FIELDS = ["rfp_number", "rfp_date", "mat_code", "mat_desc", "uom", "total_qty",
              "req_by_date", "deliv_loc", "deliv_geo", "source_prs", "source_lines",
              "req_depts", "project_ids", "specs", "closing_date", "status"]


# ── Sheet bootstrap ────────────────────────────────────────────────────────────
def ensure_sheets(wb=None):
    """Kept for signature compatibility — RFx_Quotes/RFx_Invitations no
    longer live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


# ── RFP readers ─────────────────────────────────────────────────────────────────
def _row_to_rfp(vals):
    return dict(zip(RFP_FIELDS, vals))


def get_rfps_by_status(statuses, data_file=None):
    """RFP now lives in SQLite (pr_consolidation.py's pilot table
    `rfp`) — a real indexed WHERE on status, replacing the Excel
    iter_rows() scan (itself a prior fix for a 277x-slower .cell(r,c)
    pattern; SQLite removes the scan-the-sheet cost entirely, not just
    that anti-pattern)."""
    conn = db.get_connection()
    try:
        placeholders = ",".join("?" for _ in statuses)
        rows = conn.execute(
            f"SELECT * FROM rfp WHERE COALESCE(status, 'Open') IN ({placeholders})",
            list(statuses),
        ).fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        d = dict(r)
        out.append({
            "rfp_number": d["rfp_number"], "rfp_date": d["rfp_date"],
            "mat_code": d["material_code"], "mat_desc": d["material_desc"],
            "uom": d["uom"], "total_qty": d["total_qty"],
            "req_by_date": d["required_by_date"], "deliv_loc": d["delivery_location"],
            "deliv_geo": d["delivery_geolocation"], "source_prs": d["source_pr_numbers"],
            "source_lines": d["source_pr_lines"], "req_depts": d["requester_depts"],
            "project_ids": d["project_ids"], "specs": d["specifications"],
            "closing_date": d["closing_date"], "status": d["status"] or "Open",
        })
    return out


def get_open_rfps(data_file=None):
    return get_rfps_by_status(["Open"], data_file)


def get_pending_award_rfps(data_file=None):
    """RFPs with a winner selected but no PO issued yet."""
    return get_rfps_by_status(["Selected"], data_file)


# ── Vendor master / affinity ──────────────────────────────────────────────────
def _read_vendor_master():
    """Active vendors, from vendor_onboarding.py — the canonical
    Vendor_Master owner — instead of this module's own separate Excel
    read. This used to be a SECOND, independent "read Vendor_Master and
    filter Active" implementation alongside pr_consolidation.py's (which
    didn't filter at all) — exactly the kind of duplication that let the
    two disagree. There is now exactly one Vendor_Master reader in the
    whole codebase."""
    return [{"id": v["Vendor_ID"], "name": v["Vendor_Name"] or "",
             "geo": v["Geolocation"] or "", "city": v["City"] or "",
             "email": v["Contact_Email"] or ""}
            for v in vo.list_vendors(include_inactive=False)]


def _item_category_lookup():
    """{item_code: (category, subcategory)} across the WHOLE catalog
    (active_only=False) — affinity scoring should draw on historical
    tag data regardless of whether an item happens to be active today."""
    return {i["code"]: (i["category"], i["subcategory"])
            for i in po_export.load_item_master(active_only=False)}


def _category_vendor_affinity():
    known_vendors = {v["id"] for v in _read_vendor_master()}
    aff = defaultdict(lambda: defaultdict(int))
    for i in po_export.load_item_master(active_only=False):
        category, tags = i["category"], i["tags"]
        if not category or not tags:
            continue
        vendor_code = tags.split()[0]
        if vendor_code in known_vendors:
            aff[category][vendor_code] += 1
    return aff


def _parse_ll(s):
    try:
        p = str(s or "").split(",")
        if len(p) == 2:
            return float(p[0].strip()), float(p[1].strip())
    except Exception:
        pass
    return None


def _haversine_km(a, b):
    if not a or not b:
        return None
    lat1, lng1 = a; lat2, lng2 = b
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1); dl = math.radians(lng2 - lng1)
    h = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(h))


def list_vendors(data_file=None):
    """Public wrapper — active vendors from Vendor_Master, for UI pickers."""
    return _read_vendor_master()


def _rank_vendors_for_line(rfp_row, vendors, cat_lookup, affinity, top_n=5):
    """Pure computation, no file I/O — given already-loaded vendor/category/
    affinity data, ranks vendors for one RFP line. Both suggest_vendors()
    and suggest_vendors_for_lines() share this, so the ranking logic only
    exists once."""
    category, _ = cat_lookup.get(rfp_row["mat_code"], ("", ""))
    cat_scores = affinity.get(category, {})
    deliv_ll = _parse_ll(rfp_row.get("deliv_geo"))

    ranked = []
    for v in vendors:
        vendor_ll = _parse_ll(v["geo"])
        dist_km = _haversine_km(deliv_ll, vendor_ll)
        aff_count = cat_scores.get(v["id"], 0)
        score = -aff_count * 1000 + (dist_km if dist_km is not None else 99999)
        reasons = []
        if aff_count:
            reasons.append(f"supplies {category} ({aff_count} SKU history)")
        if dist_km is not None:
            reasons.append(f"{dist_km:,.0f} km from delivery point")
        ranked.append({**v, "distance_km": dist_km, "affinity": aff_count,
                        "score": score, "reason": "; ".join(reasons) or "no history — cold outreach"})
    ranked.sort(key=lambda x: x["score"])
    return ranked[:top_n]


def suggest_vendors(rfp_row, data_file=None, top_n=5):
    """Rank active vendors for a single open RFP line (affinity + proximity)."""
    vendors = _read_vendor_master()
    cat_lookup = _item_category_lookup()
    affinity = _category_vendor_affinity()
    return _rank_vendors_for_line(rfp_row, vendors, cat_lookup, affinity, top_n)


def suggest_vendors_for_lines(rfp_rows, data_file=None, top_n=6):
    """
    Aggregate vendor suggestion across MULTIPLE RFP lines at once — used when
    batch-inviting a vendor to quote on several items in one RFQ document.
    Averages each vendor's per-line score across all lines they were ranked on.

    FIX: loads vendor/category/affinity data ONCE regardless of how many
    lines are passed, instead of calling suggest_vendors() per line (which
    each opened the workbook fresh). Confirmed this was the dominant cost
    in a full RFx Management page render — the multiselect that feeds this
    defaults to ALL open RFP lines, and a BOM-driven batch can easily be
    30+ lines, meaning 30+ file opens for one page render before this fix.
    """
    if not rfp_rows:
        return []
    vendors = _read_vendor_master()
    cat_lookup = _item_category_lookup()
    affinity = _category_vendor_affinity()

    totals = defaultdict(float); counts = defaultdict(int); meta = {}
    for row in rfp_rows:
        for v in _rank_vendors_for_line(row, vendors, cat_lookup, affinity, top_n=20):
            totals[v["id"]] += v["score"]
            counts[v["id"]] += 1
            meta[v["id"]] = v
    ranked = []
    n_lines = len(rfp_rows)
    for vid, total in totals.items():
        avg_score = total / counts[vid]
        v = dict(meta[vid])
        v["score"] = avg_score
        v["lines_matched"] = counts[vid]
        v["reason"] = f"{v['reason']}  ·  relevant to {counts[vid]}/{n_lines} selected line(s)"
        ranked.append(v)
    ranked.sort(key=lambda x: x["score"])
    return ranked[:top_n]


# ── Step 1: invite vendor(s) + generate RFQ document ────────────────────────────
def _next_invitation_id(conn):
    rows = conn.execute(
        "SELECT invitation_id FROM rfx_invitations WHERE invitation_id LIKE 'INV-%'"
    ).fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["invitation_id"].split("-")[1]))
        except Exception: pass
    return f"INV-{mx+1:05d}"


def _record_invitations(vendor_id, vendor_name, rfp_rows, filename, data_file=None):
    db.init_schema()
    conn = db.get_connection()
    ids = []
    try:
        today = date.today().strftime("%Y-%m-%d")
        for row in rfp_rows:
            iid = _next_invitation_id(conn)
            conn.execute(
                "INSERT INTO rfx_invitations (invitation_id, rfp_number, vendor_id, "
                "vendor_name, invited_date, filename) VALUES (?,?,?,?,?,?)",
                (iid, row["rfp_number"], vendor_id, vendor_name, today, filename),
            )
            ids.append(iid)
        conn.commit()
    finally:
        conn.close()
    return ids


def get_invitations(vendor_id=None, rfp_number=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM rfx_invitations ORDER BY invitation_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"invitation_id": r["invitation_id"], "rfp_number": r["rfp_number"],
               "vendor_id": r["vendor_id"], "vendor_name": r["vendor_name"],
               "invited_date": r["invited_date"], "filename": r["filename"]}
        if vendor_id and row["vendor_id"] != vendor_id:
            continue
        if rfp_number and row["rfp_number"] != rfp_number:
            continue
        out.append(row)
    return out


def _rfq_workbook_bytes(vendor_id, vendor_name, vendor_email, rfp_rows, batch_id):
    """Builds the actual downloadable RFQ workbook a vendor would fill in and return."""
    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = "1E3A5F"; amber = "FFF8E8"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RFQ"

    def title(cell_ref, text, size=13, bold=True, color="0F172A"):
        c = ws[cell_ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=bold, color=color)

    def label(cell_ref, text):
        c = ws[cell_ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(cell_ref, text):
        c = ws[cell_ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "REQUEST FOR QUOTATION", 15)
    ws.merge_cells("A1:F1")
    closing = (date.today() + timedelta(days=14)).strftime("%Y-%m-%d")

    label("A3", "RFQ Batch ID:");   value("B3", batch_id)
    label("A4", "Issued:");         value("B4", date.today().strftime("%Y-%m-%d"))
    label("A5", "Response by:");    value("B5", closing)
    label("D3", "Vendor:");         value("E3", vendor_name or vendor_id)
    label("D4", "Vendor ID:");      value("E4", vendor_id)
    label("D5", "Vendor Email:");   value("E5", vendor_email or "")

    ws["A7"] = ("Please quote against the line items below. Fill in the highlighted "
                "columns (Unit Price, Lead Time, MOQ) and return this file.")
    ws["A7"].font = Font(name="Arial", size=9, italic=True, color="64748B")
    ws.merge_cells("A7:K7")

    hdrs = ["#", "RFP Ref", "Material Code", "Description", "UOM", "Qty Required",
            "Required By", "Delivery Location", "Unit Price", "Lead Time (days)",
            "MOQ", "Remarks"]
    header_row = 9
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=navy)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(header_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_cols = {9, 10, 11, 12}  # vendor-editable columns highlighted
    r = header_row + 1
    for i, row in enumerate(rfp_rows, 1):
        vals = [i, row["rfp_number"], row["mat_code"], row["mat_desc"], row["uom"],
                row["total_qty"], row["req_by_date"], row["deliv_loc"],
                None, None, None, ""]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if ci in fill_cols:
                c.fill = PatternFill("solid", fgColor=amber)
        r += 1

    widths = [4, 16, 13, 34, 8, 12, 12, 20, 12, 14, 8, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[header_row].height = 30

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def invite_and_generate_rfq(vendor_id, vendor_name, vendor_email, rfp_rows, data_file=None):
    """
    Invite a vendor to quote on a batch of RFP lines and produce the actual
    downloadable RFQ workbook (one document, all lines). Also logs the
    invitation for each line so the 'Enter Quotes' step knows what this
    vendor was asked to quote.
    Returns (filename, file_bytes, invitation_ids).
    """
    if not rfp_rows:
        raise ValueError("No RFP lines selected to invite this vendor on.")
    batch_id = f"RFQ-{vendor_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    filename = f"{batch_id}.xlsx"
    file_bytes = _rfq_workbook_bytes(vendor_id, vendor_name, vendor_email, rfp_rows, batch_id)
    inv_ids = _record_invitations(vendor_id, vendor_name, rfp_rows, filename, data_file)
    return filename, file_bytes, inv_ids


# ── Step 2: quotes (single-line and vendor-batch variants) ──────────────────────
def _next_quote_id(conn):
    rows = conn.execute("SELECT quote_id FROM rfx_quotes WHERE quote_id LIKE 'Q-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["quote_id"].split("-")[1]))
        except Exception: pass
    return f"Q-{mx+1:05d}"


def get_quotes(rfp_number=None, vendor_id=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM rfx_quotes ORDER BY quote_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"quote_id": r["quote_id"], "rfp_number": r["rfp_number"],
               "vendor_id": r["vendor_id"], "vendor_name": r["vendor_name"],
               "price": r["quoted_price"], "lead_time": r["lead_time_days"],
               "moq": r["moq"], "quote_date": r["quote_date"], "status": r["status"],
               "notes": r["notes"]}
        if rfp_number and row["rfp_number"] != rfp_number:
            continue
        if vendor_id and row["vendor_id"] != vendor_id:
            continue
        out.append(row)
    return out


def get_item_master_defaults(material_code, data_file=None):
    """List price + lead time from Item Master — the fallback used for a
    direct/skip-RFx assignment, where there's no negotiated quote to draw on.
    Point lookup by known code — active_only=False, since a code already
    referenced on an RFP line should resolve regardless of current Active
    status (see po_export.get_item_by_code's docstring)."""
    item = po_export.get_item_by_code(material_code, active_only=False)
    if not item:
        return {"price": None, "lead_time": None}
    return {"price": item["price"] or None, "lead_time": item["lead_time"]}


def quick_assign(rfp_rows, vendor_id, vendor_name, line_prices, data_file=None):
    """
    The 'RFx is recommended but not mandatory' path: skip vendor invitation
    and quote collection entirely, and go straight to marking a preferred
    vendor as the winner for each line, at a given (typically list) price.
    line_prices: {rfp_number: price}.
    Deliberately reuses record_quotes_batch() + select_winner() rather than
    a separate PO-creation path — the resulting lines land in exactly the
    same 'pending award' state a normal RFx cycle produces, so Tab 4's
    Issue POs (with the optional contract lock) handles both uniformly.
    The quote is clearly labelled as a direct assignment, never mistaken
    for a competitively-sourced price.
    """
    fpath = data_file or DATA_FILE
    line_quotes = []
    for row in rfp_rows:
        defaults = get_item_master_defaults(row["mat_code"], fpath)
        line_quotes.append({
            "rfp_number": row["rfp_number"],
            "price": line_prices.get(row["rfp_number"], defaults["price"] or 0),
            "lead_time": defaults["lead_time"] or 7, "moq": 1,
            "notes": "Direct assignment — RFx skipped, no competitive quote",
        })
    quote_ids = record_quotes_batch(vendor_id, vendor_name, line_quotes, data_file=fpath)
    for row, qid in zip(rfp_rows, quote_ids):
        select_winner(row["rfp_number"], qid, data_file=fpath)
    return quote_ids


def record_quotes_batch(vendor_id, vendor_name, line_quotes, data_file=None):
    """
    line_quotes: list of dicts {rfp_number, price, lead_time, moq, notes}
    One vendor responding to several RFP lines from a single RFQ document —
    written in one SQLite transaction instead of N.
    """
    db.init_schema()
    conn = db.get_connection()
    ids = []
    try:
        today = date.today().strftime("%Y-%m-%d")
        for lq in line_quotes:
            qid = _next_quote_id(conn)
            conn.execute(
                "INSERT INTO rfx_quotes (quote_id, rfp_number, vendor_id, vendor_name, "
                "quoted_price, lead_time_days, moq, quote_date, status, notes) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (qid, lq["rfp_number"], vendor_id, vendor_name, lq["price"],
                 lq["lead_time"], lq.get("moq", 1), today, "Submitted", lq.get("notes", "")),
            )
            ids.append(qid)
        conn.commit()
    finally:
        conn.close()
    return ids


def record_quote(rfp_number, vendor_id, vendor_name, price, lead_time_days,
                  moq=1, notes="", data_file=None):
    """Single-line quote entry (e.g. manually keying in one phone-quote)."""
    return record_quotes_batch(
        vendor_id, vendor_name,
        [{"rfp_number": rfp_number, "price": price, "lead_time": lead_time_days,
          "moq": moq, "notes": notes}],
        data_file)[0]


def _simulate_price_and_terms(rfp_row, vendor_id, base_price, seed_salt="v1"):
    h = int(hashlib.sha1(f"{rfp_row['rfp_number']}-{vendor_id}-{seed_salt}".encode()).hexdigest(), 16)
    price_variance = 0.85 + (h % 31) / 100.0   # 0.85x .. 1.15x
    lead_time = 3 + (h // 31) % 13             # 3..15 days
    moq = max(1, (h // 997) % 10)
    return round(base_price * price_variance, 2), lead_time, moq


def simulate_quotes(rfp_row, vendor_ids, data_file=None, seed_salt="v1"):
    """Demo helper: multiple vendors quoting on ONE line (for side-by-side comparison)."""
    fpath = data_file or DATA_FILE
    item = po_export.get_item_by_code(rfp_row["mat_code"], active_only=False)
    base_price = item["price"] if item and item["price"] else None
    vendors = {v["id"]: v for v in _read_vendor_master()}
    base_price = base_price or 100.0

    line_quotes = []
    vendor_ids_ordered = []
    for vid in vendor_ids:
        price, lead_time, moq = _simulate_price_and_terms(rfp_row, vid, base_price, seed_salt)
        line_quotes.append({"rfp_number": rfp_row["rfp_number"], "price": price,
                             "lead_time": lead_time, "moq": moq,
                             "notes": "Simulated response (demo)"})
        vendor_ids_ordered.append(vid)
    # write one-by-one under each vendor's own name (batch call would attribute all to one vendor)
    ids = []
    for vid, lq in zip(vendor_ids_ordered, line_quotes):
        vname = vendors.get(vid, {}).get("name", vid)
        ids += record_quotes_batch(vid, vname, [lq], data_file=fpath)
    return ids


def simulate_quotes_batch(rfp_rows, vendor_id, data_file=None, seed_salt="v1"):
    """Demo helper: ONE vendor responding to MULTIPLE lines from their RFQ document."""
    fpath = data_file or DATA_FILE
    prices = {i["code"]: (i["price"] or None) for i in po_export.load_item_master(active_only=False)}
    vendors = {v["id"]: v for v in _read_vendor_master()}
    vendor_name = vendors.get(vendor_id, {}).get("name", vendor_id)

    line_quotes = []
    for row in rfp_rows:
        base_price = prices.get(row["mat_code"]) or 100.0
        price, lead_time, moq = _simulate_price_and_terms(row, vendor_id, base_price, seed_salt)
        line_quotes.append({"rfp_number": row["rfp_number"], "price": price,
                             "lead_time": lead_time, "moq": moq,
                             "notes": "Simulated response (demo)"})
    return record_quotes_batch(vendor_id, vendor_name, line_quotes, data_file=fpath)


# ── Step 3: select a winner per line (no PO yet) ────────────────────────────────
def select_winner(rfp_number, quote_id, data_file=None):
    """
    Marks the winning quote for one RFP line and flips the RFP's status to
    'Selected' — a decision, not yet a PO. Run generate_pos() afterwards
    (any time, for any number of accumulated decisions) to actually issue POs.
    RFP and RFx_Quotes both live in SQLite now.
    """
    conn = db.get_connection()
    found = False
    try:
        rows = conn.execute(
            "SELECT quote_id FROM rfx_quotes WHERE rfp_number = ?", (rfp_number,)
        ).fetchall()
        for r in rows:
            qid = r["quote_id"]
            is_winner = qid == quote_id
            conn.execute(
                "UPDATE rfx_quotes SET status = ? WHERE quote_id = ?",
                ("Selected" if is_winner else "Rejected", qid),
            )
            found = found or is_winner
        conn.execute("UPDATE rfp SET status='Selected' WHERE rfp_number=?", (rfp_number,))
        conn.commit()
    finally:
        conn.close()
    return found


# ── Step 4: batch-generate POs, consolidated per vendor ─────────────────────────
def preview_pending_pos(data_file=None):
    """Groups pending (Selected) RFP lines by winning vendor — a preview before issuing."""
    fpath = data_file or DATA_FILE
    pending = get_pending_award_rfps(fpath)
    groups = defaultdict(list)
    for rfp in pending:
        quotes = get_quotes(rfp_number=rfp["rfp_number"], data_file=fpath)
        winner = next((q for q in quotes if q["status"] == "Selected"), None)
        if winner:
            groups[(winner["vendor_id"], winner["vendor_name"])].append((rfp, winner))
    out = []
    for (vid, vname), pairs in groups.items():
        total = sum(w["price"] * (r["total_qty"] or 0) for r, w in pairs)
        out.append({"vendor_id": vid, "vendor_name": vname, "lines": len(pairs),
                     "total_value": total, "pairs": pairs})
    out.sort(key=lambda g: -g["total_value"])
    return out


def generate_pos(data_file=None):
    """
    Issues consolidated POs: one PO per vendor, covering every pending
    (winner-selected-but-not-yet-issued) RFP line awarded to them.

    PO_Header/PO_Items/RFP/PR_Items now live in SQLite. PO writing and
    the PR_Items writeback go through pr_consolidation.py's shared
    functions (next_po_number/insert_po/mark_pr_items_ordered) — the
    SAME functions run() (PR consolidation) uses for the same tables,
    so there is exactly one implementation of "issue a PO" and one of
    "mark these PR lines as ordered," not two independent ones that
    could drift apart (which is exactly how PO numbering collided
    before this project's earlier bug fixes). RFx_Quotes stays Excel
    (out of migration scope).
    """
    fpath = data_file or DATA_FILE
    groups = preview_pending_pos(fpath)
    if not groups:
        return []

    vendors = {v["id"]: v for v in _read_vendor_master()}
    pr_headers = pc.get_pr_headers(fpath)
    results = []
    year = datetime.now().year

    for g in groups:
        vendor_id = g["vendor_id"]
        org = {}
        first_pair = g["pairs"][0][0]
        for pr_number, hdr in pr_headers.items():
            if pr_number and pr_number in str(first_pair["source_prs"]):
                org = {"po_type": hdr["po_type"], "legal_entity": hdr["legal_entity"],
                       "purch_entity": hdr["purch_entity"], "purch_group": hdr["purch_group"],
                       "currency": hdr["currency"], "plant_code": hdr["plant_code"]}
                break
        org.setdefault("po_type", "NB"); org.setdefault("legal_entity", "LE-001")
        org.setdefault("purch_entity", "PE-001"); org.setdefault("purch_group", "PG-001")
        org.setdefault("currency", "INR"); org.setdefault("plant_code", "PLANT-01")

        sup_geo = vendors.get(vendor_id, {}).get("geo", "")
        po_num = pc.next_po_number(year, fpath)

        po_lines = []
        rfp_awarded = []
        quote_awarded = []
        for rfp, winner in g["pairs"]:
            source_prs = [p.strip() for p in str(rfp["source_prs"]).split(",") if p.strip()]
            source_lines = [l.strip() for l in str(rfp["source_lines"]).split(",") if l.strip()]
            po_lines.append({"mat_code": rfp["mat_code"], "mat_desc": rfp["mat_desc"],
                   "uom": rfp["uom"], "qty": rfp["total_qty"], "unit_price": winner["price"],
                   "deliv_date": rfp["req_by_date"],
                   "deliv_loc": rfp["deliv_loc"], "deliv_geo": rfp["deliv_geo"],
                   "source_pr": ", ".join(source_prs), "source_pr_line": ", ".join(source_lines),
                   "req_id": "", "req_dept": rfp["req_depts"], "project_id": rfp["project_ids"]})
            rfp_awarded.append(rfp["rfp_number"])
            quote_awarded.append((rfp["rfp_number"], winner["quote_id"]))

        header = {"po_type": org["po_type"], "legal_entity": org["legal_entity"],
                  "purch_entity": org["purch_entity"], "purch_group": org["purch_group"],
                  "currency": org["currency"], "plant_code": org["plant_code"],
                  "supplier_id": vendor_id, "supplier_name": g["vendor_name"], "supplier_geo": sup_geo}
        written = pc.insert_po(po_num, header, po_lines, fpath)

        pr_writeback_pairs = []
        for it in written:
            prs = str(it["source_pr"]).split(", ")
            lns = str(it["source_pr_line"]).split(", ")
            for pr_r, ln_r in zip(prs, lns):
                pr_r, ln_r = pr_r.strip(), ln_r.strip()
                if pr_r and ln_r:
                    pr_writeback_pairs.append((pr_r, int(float(ln_r)), po_num, it["po_item"], vendor_id))
        pc.mark_pr_items_ordered(pr_writeback_pairs, fpath)
        # RFx-awarded POs skip 'Proposed' entirely — there's no separate
        # "send" screen for this path the way Consolidate has (see
        # mark_po_created()'s docstring), so this runs immediately rather
        # than waiting on a later user action. The caller (erp_ui.py's
        # RFx Management page) is responsible for actually generating and
        # offering the AV export file right after this returns — this
        # function only handles the status transition, not the export
        # itself (that's UI-layer, matching how Direct PO Entry works).
        pc.mark_po_created(po_num, fpath)

        conn = db.get_connection()
        try:
            for rfp_number in rfp_awarded:
                conn.execute("UPDATE rfp SET status='Awarded' WHERE rfp_number=?", (rfp_number,))
            for rfp_number, quote_id in quote_awarded:
                conn.execute(
                    "UPDATE rfx_quotes SET status='Awarded' WHERE rfp_number=? AND quote_id=?",
                    (rfp_number, quote_id),
                )
            conn.commit()
        finally:
            conn.close()

        results.append({"po_number": po_num, "vendor_id": vendor_id,
                         "vendor_name": g["vendor_name"], "lines": g["lines"],
                         "total_value": g["total_value"]})

    return results



def stats(data_file=None):
    fpath = data_file or DATA_FILE
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT status FROM rfp").fetchall()
    finally:
        conn.close()
    counts = defaultdict(int); total = 0
    for row in rows:
        total += 1
        counts[str(row["status"] or "Open")] += 1

    conn2 = db.get_connection()
    try:
        n_quotes = conn2.execute("SELECT COUNT(*) FROM rfx_quotes").fetchone()[0]
        n_invites = conn2.execute("SELECT COUNT(*) FROM rfx_invitations").fetchone()[0]
    finally:
        conn2.close()
    return {"total_rfps": total, "open": counts.get("Open", 0),
            "pending_po": counts.get("Selected", 0), "awarded": counts.get("Awarded", 0),
            "quotes": n_quotes, "invitations": n_invites}


if __name__ == "__main__":
    print("Stats:", stats())
    rfps = get_open_rfps()
    print(f"\n{len(rfps)} open RFP line(s)")
    if rfps:
        vs = suggest_vendors_for_lines(rfps, top_n=3)
        for v in vs:
            print(f"  -> {v['id']:10s} avg_score={v['score']:8.1f}  {v['reason']}")
