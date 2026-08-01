"""
migrate_mfg_remainder.py — one-time port of the remaining MFG tables
(GR_Header, GR_Items, Quality_Inspections, Production_Confirmations,
Inventory_Transactions) from data.xlsx into erp_pilot.db.

Idempotent: clears each table before inserting.

All five sheets use the "no title row" convention (header row 1, data
row 2+) — same as every other pilot-created sheet.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_gr(wb, conn):
    n_hdr = 0
    if "GR_Header" in wb.sheetnames:
        ws = wb["GR_Header"]
        hdrs = _headers(ws)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            gid = str(row[hdrs.get("GR_ID", 0)] or "").strip()
            if not gid:
                continue
            def g(col):
                i = hdrs.get(col)
                return row[i] if i is not None else None
            rows.append((gid, g("PO_Number"), g("Vendor_ID"), g("Vendor_Name"), g("GR_Date"),
                        g("Status"), g("Delivery_Location"), g("Received_By"), g("Notes")))
        conn.execute("DELETE FROM gr_header")
        conn.executemany(
            "INSERT INTO gr_header (gr_id, po_number, vendor_id, vendor_name, gr_date, "
            "status, delivery_location, received_by, notes) VALUES (?,?,?,?,?,?,?,?,?)",
            rows,
        )
        n_hdr = len(rows)

    n_items = 0
    if "GR_Items" in wb.sheetnames:
        ws2 = wb["GR_Items"]
        hdrs2 = _headers(ws2)
        rows2 = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            gid = str(row[hdrs2.get("GR_ID", 0)] or "").strip()
            if not gid:
                continue
            def g2(col):
                i = hdrs2.get(col)
                return row[i] if i is not None else None
            rows2.append((gid, g2("Line_Item"), g2("Material_Code"), g2("Material_Desc"),
                         g2("UOM"), g2("PO_Qty"), g2("Qty_Received"), g2("Unit_Price"), g2("Notes")))
        conn.execute("DELETE FROM gr_items")
        conn.executemany(
            "INSERT INTO gr_items (gr_id, line_item, material_code, material_desc, uom, "
            "po_qty, qty_received, unit_price, notes) VALUES (?,?,?,?,?,?,?,?,?)",
            rows2,
        )
        n_items = len(rows2)
    return n_hdr, n_items


def migrate_qi(wb, conn):
    if "Quality_Inspections" not in wb.sheetnames:
        return 0
    ws = wb["Quality_Inspections"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        qid = str(row[hdrs.get("QI_ID", 0)] or "").strip()
        if not qid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((qid, g("GR_ID"), g("PO_Number"), g("Line_Item"), g("Material_Code"),
                    g("Material_Desc"), g("Qty_Received"), g("Qty_Passed"), g("Qty_Failed"),
                    g("Inspected_By"), g("Inspection_Date"), g("Notes")))
    conn.execute("DELETE FROM quality_inspections")
    conn.executemany(
        "INSERT INTO quality_inspections (qi_id, gr_id, po_number, line_item, material_code, "
        "material_desc, qty_received, qty_passed, qty_failed, inspected_by, inspection_date, "
        "notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_production(wb, conn):
    if "Production_Confirmations" not in wb.sheetnames:
        return 0
    ws = wb["Production_Confirmations"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = str(row[hdrs.get("Confirmation_ID", 0)] or "").strip()
        if not cid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((cid, g("Parent_Item_Code"), g("Parent_Item_Desc"), g("Quantity_Built"),
                    g("Location_ID"), g("Confirmation_Date"), g("Confirmed_By"), g("Notes")))
    conn.execute("DELETE FROM production_confirmations")
    conn.executemany(
        "INSERT INTO production_confirmations (confirmation_id, parent_item_code, "
        "parent_item_desc, quantity_built, location_id, confirmation_date, confirmed_by, "
        "notes) VALUES (?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_inventory(wb, conn):
    if "Inventory_Transactions" not in wb.sheetnames:
        return 0
    ws = wb["Inventory_Transactions"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = str(row[hdrs.get("Txn_ID", 0)] or "").strip()
        if not tid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((tid, g("Txn_Date"), g("Material_Code"), g("Material_Desc"),
                    g("Location_ID"), g("Location_Name"), g("Quantity"), g("Txn_Type"),
                    g("Reference_Type"), g("Reference_ID"), g("Notes")))
    conn.execute("DELETE FROM inventory_transactions")
    conn.executemany(
        "INSERT INTO inventory_transactions (txn_id, txn_date, material_code, material_desc, "
        "location_id, location_name, quantity, txn_type, reference_type, reference_id, notes) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_grh, n_gri = migrate_gr(wb, conn)
        n_qi = migrate_qi(wb, conn)
        n_pc = migrate_production(wb, conn)
        n_txn = migrate_inventory(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"gr_header": n_grh, "gr_items": n_gri, "quality_inspections": n_qi,
            "production_confirmations": n_pc, "inventory_transactions": n_txn}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate MFG remainder tables from Excel to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to source data.xlsx")
    parser.add_argument("--db", default=None, help="Path to target SQLite db (default: erp_pilot.db)")
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
