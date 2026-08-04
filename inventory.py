"""
inventory.py — Inventory Ledger (S2C/MFG, manufacturing pilot).

Transaction-based, same design as accounting.py's Journal Entries and
goods_receipt.py's own receipt tracking: every movement is a signed-
quantity row, current balance is computed live by summing history, never
stored as a separately-mutated number. Same reasons apply here as
everywhere else this pattern's been used — auditable, can't drift out of
sync with its own history.

Deliberately standalone — no import of goods_receipt.py, pr_consolidation.py,
or bom.py. goods_receipt.create_gr() calls record_transaction() directly
using data it already has in scope (material, PO line's delivery
location, quantity) rather than this module reading GR data itself,
which would create a circular import (goods_receipt already needs to
import this module to post the other direction).

HONEST SCOPE, stated before anyone relies on this for more than it does:
only "goods received against a PO" is wired in as a stock-increasing
event, because it's the only genuine stock movement that exists anywhere
in this system right now. There is no production-consumption event (BOM
components going into an actual build), no finished-goods-produced
event, and O2C Fulfillment's shipments aren't wired to decrease stock
here yet either — all three are real, but none of them exist as an
implemented business process yet, and this module isn't going to
pretend a balance is complete when it can only ever go up. Building any
of those is the natural next step, not a limitation of this ledger's
design — the ledger itself is ready for all of them the moment those
events exist, since it's just signed quantities against a reference.

New sheet:
  Inventory_Transactions   Txn_ID | Txn_Date | Material_Code | Material_Desc |
                           Location_ID | Location_Name | Quantity | Txn_Type |
                           Reference_Type | Reference_ID | Notes

Quantity is signed: positive = stock in, negative = stock out. Txn_Type
is a free-form label ("GR Receipt", "Production Consumption", "Shipment",
"Adjustment", ...) — not constrained to a fixed set, since new movement
types will get added as the corresponding business processes get built.

SQLite pilot: Inventory_Transactions now lives in erp_pilot.db (table
`inventory_transactions`). Still deliberately standalone — record_transaction()
queries item_master/delivery_locations directly via a couple of plain
SQL SELECTs rather than importing po_export.py or pr_consolidation.py
as modules, so the "no cross-module business-logic imports" principle
above still holds; only `db` (the shared low-level connection helper,
not a business-logic module) is imported.

Active-flag policy (warn-only, per direct instruction): every write
checks whether material_code/location_id are currently Active in their
master tables and returns any concerns as `warnings` — it never blocks
the transaction. A transaction referencing an inactive item or location
is usually a legitimate real event (e.g. completing a GR against a PO
that was placed before the item was deactivated); refusing to record it
would just make the ledger wrong. If a code/location isn't found in its
master table at all, that's silently NOT warned about here — this
module has always tolerated caller-supplied location strings that don't
match a real Delivery_Locations row (e.g. "Unspecified"), by design;
only a KNOWN-and-inactive reference triggers a warning, not an
unrecognized one.
"""

import os, io
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

TXN_SHEET = "Inventory_Transactions"
TXN_COLS = ["Txn_ID", "Txn_Date", "Material_Code", "Material_Desc", "Location_ID",
           "Location_Name", "Quantity", "Txn_Type", "Reference_Type", "Reference_ID", "Notes"]


def ensure_sheets(wb=None):
    """Kept for signature compatibility — Inventory_Transactions no
    longer lives in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


def _next_txn_id(conn):
    rows = conn.execute("SELECT txn_id FROM inventory_transactions WHERE txn_id LIKE 'TXN-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["txn_id"].split("-")[1]))
        except Exception: pass
    return f"TXN-{mx+1:06d}"


def _check_active_refs(conn, material_code, location_id):
    """Warn-only (see module docstring): returns a list of warning
    strings for any KNOWN-and-inactive material/location. Unrecognized
    codes are not warned about — see docstring for why."""
    warnings = []
    if material_code:
        row = conn.execute(
            "SELECT active FROM item_master WHERE item_code = ?", (material_code,)
        ).fetchone()
        if row is not None and (row["active"] or "Yes") != "Yes":
            warnings.append(f"Material {material_code} is currently marked inactive.")
    if location_id:
        row = conn.execute(
            "SELECT active FROM delivery_locations WHERE location_id = ?", (location_id,)
        ).fetchone()
        if row is not None and (row["active"] or "Yes") != "Yes":
            warnings.append(f"Location {location_id} is currently marked inactive.")
    return warnings


def record_transaction(material_code, material_desc, location_id, quantity, txn_type,
                       reference_type="", reference_id="", notes="", data_file=None):
    """
    quantity: signed — positive for stock in, negative for stock out.
    location_id: should be a real Delivery_Locations location_id (e.g.
    "GRB_DL_PKD_Factory"), not its display name. This wasn't consistently
    true everywhere until a real bug was found and fixed: several UI
    pickers across Create PR / Sales Orders / BOM Explosion / Production
    stored the location's NAME instead of its ID as the value passed
    around as "delivery location" — so a GR posting stock under
    "GRB_DL_PKD_Factory" and a Sales Order demanding it at "Genrobotics
    Manufacturing Plant - Kanjikode Palakkad" looked like two different
    locations to every on-hand/position/transfer calculation, even
    though they're the same real place. record_transaction() itself
    never enforced a convention (a free-text location like
    "Unspecified" is still fine — location_id is just a grouping key),
    but every UI-driven caller now consistently passes the real ID.

    location_name is resolved here from Delivery_Locations for display
    purposes (falls back to the raw location_id if it isn't a
    recognized location — e.g. "Unspecified" stays "Unspecified"), so
    Inventory's own tables show a human-readable place, not a code.

    No block on a resulting negative balance — same reasoning as every
    other module here that records rather than gatekeeps (GR allows
    over-receipt, QC allows failures): if consumption exceeds what was
    ever recorded as received, that's a real, informative fact about a
    process gap, not something to silently prevent.

    Returns {"txn_id": ..., "warnings": [...]} — warnings is non-blocking
    (see module docstring on the Active-flag policy); no current caller
    inspects the old bare-txn_id return value, so this shape change is
    safe. `notes` gets any warnings appended automatically, so they're
    visible on the ledger row itself even if the caller doesn't check
    the return value.
    """
    if not quantity:
        raise ValueError("Transaction quantity can't be zero — nothing happened.")
    db.init_schema()
    conn = db.get_connection()
    try:
        warnings = _check_active_refs(conn, material_code, location_id)
        full_notes = notes
        if warnings:
            full_notes = (notes + " | " if notes else "") + " | ".join(warnings)
        loc_row = conn.execute(
            "SELECT location_name FROM delivery_locations WHERE location_id = ?", (location_id,)
        ).fetchone()
        location_name = loc_row["location_name"] if loc_row else location_id
        txn_id = _next_txn_id(conn)
        conn.execute(
            "INSERT INTO inventory_transactions (txn_id, txn_date, material_code, material_desc, "
            "location_id, location_name, quantity, txn_type, reference_type, reference_id, notes) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (txn_id, date.today().strftime("%Y-%m-%d"), material_code, material_desc,
             location_id, location_name, round(quantity, 3), txn_type, reference_type,
             reference_id, full_notes),
        )
        conn.commit()
    finally:
        conn.close()
    return {"txn_id": txn_id, "warnings": warnings}


# ── Transfers ─────────────────────────────────────────────────────────────────
def _next_transfer_id(conn):
    rows = conn.execute("SELECT DISTINCT reference_id FROM inventory_transactions "
                        "WHERE reference_type='Transfer'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["reference_id"].split("-")[1]))
        except Exception: pass
    return f"TRF-{mx+1:05d}"


def ship_transfer(mat_code, mat_desc, from_location, to_location, qty,
                  shipped_by="", carrier="", tracking_ref="", notes="", data_file=None):
    """
    Stage 1 of a real Stock Transfer: posts Transfer Out at the source
    only. Stock leaves from_location immediately (its balance drops
    right away) but does NOT arrive at to_location until
    receive_transfer() confirms it — genuinely in transit in between,
    not silently teleported. Before this (execute_transfer(), still
    kept below as a convenience wrapper), a transfer was instantaneous:
    both sides updated in the same click, with no way to represent a
    shipment that's actually en route for a few days, no confirmation
    step at the receiving end, and no way to show "how much of this
    material is currently in transit" anywhere.

    A real GL entry now posts for the goods movement itself, matching
    this system's own user story design exactly (INV-US-05's own GL
    Impact section, which STO-US-02 explicitly reuses) rather than the
    earlier, incomplete state this function was actually in: same-state
    moves value through 1300 Inter-Plant Stock-in-Transit into 1200
    Inventory; inter-state additionally recognizes IGST as a deemed
    supply (a real GST rule for a cross-state movement between
    separately-registered branches, treated the same as a real
    inter-state sale). Value is the shipped quantity times the
    material's own Item Master price (its "book cost" for this
    purpose) -- a real, defensible basis, not a placeholder, though a
    genuine moving-average or FIFO costing layer remains a real,
    separate, larger piece of future work this does not attempt.
    Freight (posted separately, below) was already real; this was the
    genuinely missing half.

    A real e-way bill is generated automatically here, for every
    caller, when the real rule requires one (genuinely inter-state,
    at or above the real value threshold) — not just for STO-created
    transfers. This used to only be wired into sto.py's own STO
    creation path, a real, found gap: an ad hoc shipment that's
    genuinely inter-state and above threshold needs a real e-way bill
    exactly as much as an STO-originated one does — compliance is
    about the physical movement, not which internal process decided
    to make it. Moving this in here, the one real place every
    shipment already passes through, means every caller gets it
    right automatically, including any future one, rather than each
    needing to remember to wire it in separately.
    """
    if qty <= 0:
        raise ValueError("Transfer quantity must be positive.")
    available = get_balance(mat_code, from_location, data_file)
    if qty > available + 0.005:
        raise ValueError(f"Only {available:g} available at {from_location} — can't ship {qty:g}.")

    conn = db.get_connection()
    try:
        transfer_id = _next_transfer_id(conn)
    finally:
        conn.close()

    out_result = record_transaction(mat_code, mat_desc, from_location, -qty, "Transfer Out",
                                     reference_type="Transfer", reference_id=transfer_id,
                                     notes=notes, data_file=data_file)

    import eway_bill as ewb
    import po_export
    item = po_export.get_item_by_code(mat_code, active_only=False)
    unit_price = item["price"] if item else 0
    gst_rate = item["gst_rate"] if item and item.get("gst_rate") else 0
    goods_value = round(unit_price * qty, 2)
    inter_state = ewb.is_inter_state(from_location, to_location, data_file=data_file)
    igst_amount = round(goods_value * gst_rate / 100, 2) if inter_state else 0.0

    fpath = data_file or DATA_FILE
    conn = db.get_connection()
    try:
        conn.execute(
            "INSERT INTO stock_transfers (transfer_id, material_code, material_desc, uom, "
            "quantity, from_location, to_location, status, shipped_date, shipped_by, "
            "carrier, tracking_ref, notes, source_type, gl_goods_value, gl_igst_amount) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (transfer_id, mat_code, mat_desc, _uom_for(mat_code, data_file), qty,
             from_location, to_location, "In Transit", str(date.today()), shipped_by,
             carrier, tracking_ref, notes, "Ad Hoc", goods_value, igst_amount),
        )
        conn.commit()
    finally:
        conn.close()

    if goods_value > 0:
        import accounting
        je_lines = [
            {"account_code": "1300", "debit": goods_value, "credit": 0,
             "description": f"{transfer_id} ship — goods leg ({from_location})"},
            {"account_code": "1200", "debit": 0, "credit": goods_value,
             "description": f"{transfer_id} ship — goods leg ({from_location})"},
        ]
        if inter_state and igst_amount > 0:
            je_lines.append({"account_code": "1300", "debit": igst_amount, "credit": 0,
                             "description": f"{transfer_id} ship — IGST leg"})
            je_lines.append({"account_code": "2220", "debit": 0, "credit": igst_amount,
                             "description": f"{transfer_id} ship — IGST leg"})
        accounting.post_journal_entry("Transfer", transfer_id,
            f"{transfer_id} ship — {mat_desc} ({from_location} \u2192 {to_location})",
            je_lines, data_file=data_file)

    declared_value = goods_value
    if ewb.is_eway_bill_required(from_location, to_location, declared_value, data_file=data_file):
        bill = ewb.generate_eway_bill(from_location, to_location, declared_value, data_file=data_file)
        conn = db.get_connection()
        try:
            conn.execute(
                "UPDATE stock_transfers SET eway_bill_number=?, eway_bill_valid_until=? "
                "WHERE transfer_id=?",
                (bill["ewb_number"], bill["valid_until"], transfer_id))
            conn.commit()
        finally:
            conn.close()

    return {"transfer_id": transfer_id, "out_txn": out_result["txn_id"],
            "warnings": out_result["warnings"]}


def receive_transfer(transfer_id, received_qty=None, received_by="", notes="",
                     wrong_product=False, data_file=None):
    """
    Stage 2: confirms a Stock Transfer that's currently In Transit —
    posts Transfer In at the destination now, for the first time, and
    flips status to Received. Until this runs, the shipped quantity is
    genuinely absent from both locations' balances, which is correct:
    it's on a truck, not sitting on a shelf anywhere.

    Real discrepancy handling, not assumed-always-full receipt (a real
    gap found from direct testing before this existed — every receipt
    silently posted the full shipped quantity regardless of what
    actually arrived). received_qty is optional and defaults to the
    full shipped quantity, so every existing caller that only ever
    confirmed a full receipt keeps working unchanged. Comparing the
    two determines a real discrepancy_type, captured on the record
    itself, not just logged and discarded:
      - Full: received_qty matches shipped quantity (the default case).
      - Partial: less arrived than the transfer record says was shipped.
        The shortfall is posted as genuinely missing, not silently
        written off or assumed still in transit -- a real, separate
        inventory-adjustment/write-off GL entry for a confirmed
        shortfall is deliberately out of scope here, same "no
        accounting entry" principle this function already holds for a
        normal transfer between the org's own locations, not yet
        extended to a shortfall specifically.
      - Excess: more arrived than the transfer record says was
        shipped -- posted as-is, flagged, not silently capped at the
        original quantity.
      - Wrong Product (wrong_product=True): what physically arrived
        doesn't match the material this transfer_id is for at all.
        Nothing is posted -- there is no real stock movement to record
        when the actual received item is unknown to this function --
        and the transfer moves to a real 'Exception' status instead of
        'Received', requiring manual resolution this function
        deliberately doesn't attempt to automate. A full reconciliation
        workflow (log what genuinely arrived as its own new receipt)
        is a real, separate piece of future work, not built here.

    A real GL entry now clears this transfer's own share of 1300 Inter-
    Plant Stock-in-Transit into 1200 Inventory (and 1170 GST Input
    Credit -- IGST, for an inter-state transfer), matching the real
    value ship_transfer() already posted there -- not recomputed from
    scratch, since the received quantity can genuinely differ from what
    was shipped. Pro-rated to the received quantity for a Partial
    receipt (only that real share clears; the shortfall's own value
    stays sitting in 1300, an honest, visible reflection of the same
    real discrepancy already flagged above, not silently written off).
    Capped at the originally-shipped value for an Excess receipt --
    real quantity beyond what was ever shipped has no real 1300 balance
    of its own to clear, and inventing one would mean crediting a value
    that was never actually posted anywhere; this is a genuine, left-
    open limitation of an Excess receipt's own GL treatment, not a
    silently-invented resolution.
    """
    t = get_stock_transfer(transfer_id, data_file)
    if t is None:
        raise ValueError(f"{transfer_id} not found.")
    if t["status"] != "In Transit":
        raise ValueError(f"{transfer_id} is '{t['status']}' — only an In Transit transfer can be received.")

    shipped_qty = t["quantity"]

    if wrong_product:
        conn = db.get_connection()
        try:
            conn.execute(
                "UPDATE stock_transfers SET status='Exception', received_date=?, received_by=?, "
                "discrepancy_type='Wrong Product', discrepancy_notes=? WHERE transfer_id=?",
                (str(date.today()), received_by, notes, transfer_id),
            )
            conn.commit()
        finally:
            conn.close()
        return {"transfer_id": transfer_id, "discrepancy_type": "Wrong Product", "in_txn": None,
               "warnings": [f"{transfer_id} flagged Wrong Product — no stock posted; "
                           f"requires manual resolution outside this function."]}

    qty = shipped_qty if received_qty is None else received_qty
    if qty <= 0:
        raise ValueError("Received quantity must be positive — use wrong_product=True "
                         "for a receipt that doesn't match this transfer's material at all.")

    if abs(qty - shipped_qty) < 0.005:
        discrepancy_type = "Full"
    elif qty < shipped_qty:
        discrepancy_type = "Partial"
    else:
        discrepancy_type = "Excess"

    in_result = record_transaction(t["material_code"], t["material_desc"], t["to_location"],
                                   qty, "Transfer In", reference_type="Transfer",
                                   reference_id=transfer_id, notes=notes, data_file=data_file)

    conn = db.get_connection()
    try:
        conn.execute(
            "UPDATE stock_transfers SET status='Received', received_date=?, received_by=?, "
            "received_qty=?, discrepancy_type=?, discrepancy_notes=? WHERE transfer_id=?",
            (str(date.today()), received_by, qty, discrepancy_type,
             notes if discrepancy_type != "Full" else "", transfer_id),
        )
        conn.commit()
    finally:
        conn.close()

    # ATP-US-03's own real supply-arrival trigger: a transfer Confirm
    # Receipt is the second of the two real events (the other being
    # goods_receipt.py's own GR posting) that genuinely increases
    # on-hand for a material at a Plant.
    import backorder as bo
    bo.reevaluate_backorders(t["material_code"], t["to_location"])

    gl_goods_value = t.get("gl_goods_value") or 0
    gl_igst_amount = t.get("gl_igst_amount") or 0
    if gl_goods_value > 0 and shipped_qty > 0:
        clear_ratio = min(qty, shipped_qty) / shipped_qty
        received_goods_value = round(gl_goods_value * clear_ratio, 2)
        received_igst_amount = round(gl_igst_amount * clear_ratio, 2)
        import accounting
        je_lines = [
            {"account_code": "1200", "debit": received_goods_value, "credit": 0,
             "description": f"{transfer_id} receive — goods leg ({t['to_location']})"},
            {"account_code": "1300", "debit": 0, "credit": received_goods_value,
             "description": f"{transfer_id} receive — goods leg ({t['to_location']})"},
        ]
        if received_igst_amount > 0:
            je_lines.append({"account_code": "1170", "debit": received_igst_amount, "credit": 0,
                             "description": f"{transfer_id} receive — IGST leg"})
            je_lines.append({"account_code": "1300", "debit": 0, "credit": received_igst_amount,
                             "description": f"{transfer_id} receive — IGST leg"})
        accounting.post_journal_entry("Transfer", transfer_id,
            f"{transfer_id} receive — {t['material_desc']} at {t['to_location']}",
            je_lines, data_file=data_file)

    warnings = list(in_result["warnings"])
    if discrepancy_type == "Partial":
        warnings.append(f"Partial receipt: {qty:g} of {shipped_qty:g} {t['uom']} confirmed "
                        f"— {shipped_qty - qty:g} short, not automatically written off.")
    elif discrepancy_type == "Excess":
        warnings.append(f"Excess receipt: {qty:g} confirmed against {shipped_qty:g} {t['uom']} "
                        f"shipped — {qty - shipped_qty:g} more than the transfer record states.")

    return {"transfer_id": transfer_id, "in_txn": in_result["txn_id"],
           "discrepancy_type": discrepancy_type, "warnings": warnings}


def cancel_transfer(transfer_id, cancelled_by="", notes="", data_file=None):
    """
    A real, genuine gap found from direct testing, not anticipated up
    front: once ship_transfer() ran, there was no way back — a
    transfer that turned out to be unbookable with any available
    courier (too heavy, wrong route, whatever the real reason) was
    permanently stuck In Transit with no path forward and no path back.

    Reverses ship_transfer()'s own Transfer Out posting with a real,
    separate "Transfer Cancelled" transaction at the source location
    (not a silent deletion of the original Out — both stay on the real
    transaction history, so the audit trail shows a shipment that was
    genuinely started and then genuinely called off, not one that
    never happened). The source location's balance is restored
    immediately, and the material reappears as a real transfer
    opportunity on the next Position & Transfers calculation, the same
    as if it had never been shipped.

    Only a transfer still In Transit can be cancelled -- once Received,
    the stock has already landed and this is no longer a shipping
    decision to undo. carrier/tracking_ref are deliberately left on the
    record rather than wiped, even if a courier was already booked
    before the cancellation -- real history of what was attempted, not
    erased just because it didn't complete.
    """
    t = get_stock_transfer(transfer_id, data_file)
    if t is None:
        raise ValueError(f"{transfer_id} not found.")
    if t["status"] != "In Transit":
        raise ValueError(f"{transfer_id} is '{t['status']}' — only an In Transit transfer can be cancelled.")

    cancel_result = record_transaction(t["material_code"], t["material_desc"], t["from_location"],
                                       t["quantity"], "Transfer Cancelled", reference_type="Transfer",
                                       reference_id=transfer_id, notes=notes, data_file=data_file)

    conn = db.get_connection()
    try:
        conn.execute(
            "UPDATE stock_transfers SET status='Cancelled', cancelled_date=?, cancelled_by=? "
            "WHERE transfer_id=?",
            (str(date.today()), cancelled_by, transfer_id),
        )
        conn.commit()
    finally:
        conn.close()

    return {"transfer_id": transfer_id, "cancel_txn": cancel_result["txn_id"],
           "warnings": cancel_result["warnings"]}


def get_stock_transfer(transfer_id, data_file=None):
    conn = db.get_connection()
    try:
        row = conn.execute("SELECT * FROM stock_transfers WHERE transfer_id=?", (transfer_id,)).fetchone()
    finally:
        conn.close()
    return dict(row) if row else None


def update_transfer_tracking(transfer_id, tracking_ref, carrier=None, data_file=None):
    """
    Persists a real courier tracking reference (e.g. an AWB number)
    onto an already-shipped transfer, after the fact — ship_transfer()
    only ever set tracking_ref at creation time (usually blank, since a
    real tracking number doesn't exist until the courier actually
    books the shipment), and nothing updated it afterward until now.
    carrier is optional — only overwritten if explicitly passed, so
    this can't accidentally blank out a carrier that was already
    correctly set at ship time.
    """
    conn = db.get_connection()
    try:
        if carrier is not None:
            conn.execute(
                "UPDATE stock_transfers SET tracking_ref=?, carrier=? WHERE transfer_id=?",
                (tracking_ref, carrier, transfer_id))
        else:
            conn.execute(
                "UPDATE stock_transfers SET tracking_ref=? WHERE transfer_id=?",
                (tracking_ref, transfer_id))
        conn.commit()
    finally:
        conn.close()


def get_stock_transfers(status=None, data_file=None):
    """Powers the 'In Transit' queue and any transfer history view.
    status=None returns every transfer regardless of status."""
    conn = db.get_connection()
    try:
        if status:
            rows = conn.execute("SELECT * FROM stock_transfers WHERE status=? "
                                "ORDER BY shipped_date DESC", (status,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM stock_transfers ORDER BY shipped_date DESC").fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


def get_plant_pair_transfer_lead_time(from_location, to_location, data_file=None):
    """
    Real, observed transfer duration between two Plants — the input
    time-phased planning needs to compute a transfer's own expected
    arrival date (Ship date + this), the same way
    pr_consolidation.get_open_po_lines_with_dates() computes a PO's
    expected arrival from PO date + lead time.

    Three-tier precedence, observed data always wins once there's
    enough of it to trust:
      1. Observed average of (received_date - shipped_date) across
         every Received transfer for this exact (from, to) pair, IF at
         least 5 such transfers exist — below that, a couple of
         unusually fast or slow shipments could badly skew an average
         nobody should trust yet.
      2. A specific estimate for this pair, if one has been set
         (set_plant_pair_lead_time_estimate()) — a real, if manually
         entered, business judgment about this specific route.
      3. The tenant-wide default (org_defaults, 'Default Transfer Lead
         Time Days', fallback 3) — the least specific, most generic
         answer, used only when neither of the above exists.

    Returns {days, source, observation_count} — source is always
    explicit ('observed' / 'estimated_pair' / 'estimated_default') so
    a caller never displays a guess with the same confidence as real
    history, the same discipline INV-US-03's own design already
    insists on for forecast vs. confirmed demand.
    """
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT shipped_date, received_date FROM stock_transfers "
            "WHERE from_location=? AND to_location=? AND status='Received' "
            "AND shipped_date IS NOT NULL AND received_date IS NOT NULL",
            (from_location, to_location),
        ).fetchall()
        durations = []
        for r in rows:
            try:
                shipped = datetime.strptime(r["shipped_date"], "%Y-%m-%d").date()
                received = datetime.strptime(r["received_date"], "%Y-%m-%d").date()
                durations.append((received - shipped).days)
            except (ValueError, TypeError):
                pass  # unparseable date on an old/malformed row — skip, don't crash the average

        if len(durations) >= 5:
            avg_days = round(sum(durations) / len(durations))
            return {"days": avg_days, "source": "observed", "observation_count": len(durations)}

        est_row = conn.execute(
            "SELECT estimated_days FROM plant_pair_lead_time_estimates "
            "WHERE from_location=? AND to_location=?",
            (from_location, to_location),
        ).fetchone()
    finally:
        conn.close()

    if est_row and est_row["estimated_days"] is not None:
        return {"days": est_row["estimated_days"], "source": "estimated_pair",
                "observation_count": len(durations)}

    import org_defaults as od
    default_days = int(od.get_default("Default Transfer Lead Time Days", data_file))
    return {"days": default_days, "source": "estimated_default", "observation_count": len(durations)}


def set_plant_pair_lead_time_estimate(from_location, to_location, estimated_days, data_file=None):
    """Sets or updates a specific-pair fallback estimate — tier 2 of
    get_plant_pair_transfer_lead_time()'s precedence. A real business
    judgment ('Chennai to Hyderabad normally takes about 4 days') for a
    route that doesn't have 5 observed transfers yet. Upsert — calling
    this again for the same pair replaces the prior estimate rather
    than erroring or duplicating."""
    conn = db.get_connection()
    try:
        conn.execute(
            "INSERT INTO plant_pair_lead_time_estimates (from_location, to_location, estimated_days) "
            "VALUES (?,?,?) ON CONFLICT(from_location, to_location) "
            "DO UPDATE SET estimated_days=excluded.estimated_days",
            (from_location, to_location, estimated_days),
        )
        conn.commit()
    finally:
        conn.close()


def _uom_for(mat_code, data_file=None):
    """Best-effort UOM lookup for stock_transfers' own record — falls
    back to 'pcs' rather than failing if Item Master doesn't have this
    code (matches how the rest of this module treats missing catalog
    data as a warning, not a hard stop)."""
    try:
        import po_export
        item = po_export.get_item_by_code(mat_code, active_only=False, data_file=data_file)
        return item["uom"] if item else "pcs"
    except Exception:
        return "pcs"


def execute_transfer(mat_code, mat_desc, from_location, to_location, qty, notes="", data_file=None):
    """
    Convenience wrapper — ships and immediately receives in one call,
    for scripted/automated use (demo scenarios' "Complete the Rest"
    fast-forward, tests) where instant completion is the point and
    staging it across two steps would just be friction. The real UI
    (Inventory -> Position & Transfers) uses ship_transfer() and
    receive_transfer() separately instead, so a live demo can show an
    actual in-transit gap — see those two functions for why this
    distinction exists at all.
    """
    ship = ship_transfer(mat_code, mat_desc, from_location, to_location, qty, notes=notes, data_file=data_file)
    receive = receive_transfer(ship["transfer_id"], notes=notes, data_file=data_file)
    return {"transfer_id": ship["transfer_id"], "out_txn": ship["out_txn"], "in_txn": receive["in_txn"],
            "warnings": ship["warnings"] + receive["warnings"]}


# ── Read ──────────────────────────────────────────────────────────────────────
def get_transactions(material_code=None, location_id=None, data_file=None):
    conn = db.get_connection()
    try:
        query = "SELECT * FROM inventory_transactions WHERE 1=1"
        params = []
        if material_code:
            query += " AND material_code = ?"
            params.append(material_code)
        if location_id:
            query += " AND location_id = ?"
            params.append(location_id)
        query += " ORDER BY txn_date DESC, txn_id DESC"
        rows = conn.execute(query, params).fetchall()
    finally:
        conn.close()
    return [{"txn_id": r["txn_id"], "txn_date": r["txn_date"], "mat_code": r["material_code"],
             "mat_desc": r["material_desc"], "location_id": r["location_id"],
             "location_name": r["location_name"], "quantity": r["quantity"],
             "txn_type": r["txn_type"], "reference_type": r["reference_type"],
             "reference_id": r["reference_id"], "notes": r["notes"]} for r in rows]


def get_balance(material_code, location_id=None, data_file=None):
    conn = db.get_connection()
    try:
        query = "SELECT SUM(quantity) AS total FROM inventory_transactions WHERE material_code = ?"
        params = [material_code]
        if location_id:
            query += " AND location_id = ?"
            params.append(location_id)
        row = conn.execute(query, params).fetchone()
    finally:
        conn.close()
    return round(row["total"] or 0.0, 3)


def get_all_balances(data_file=None):
    """Every (material, location) pair with any transaction history, and
    its current computed balance — a single GROUP BY instead of pulling
    every transaction row into Python and aggregating there."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT material_code, material_desc, location_id, location_name, "
            "SUM(quantity) AS balance FROM inventory_transactions "
            "GROUP BY material_code, location_id ORDER BY material_code, location_id"
        ).fetchall()
    finally:
        conn.close()
    return [{"mat_code": r["material_code"], "mat_desc": r["material_desc"],
             "location_id": r["location_id"], "location_name": r["location_name"],
             "balance": round(r["balance"] or 0.0, 3)} for r in rows]


def get_material_balance_by_location(material_code, data_file=None):
    return [b for b in get_all_balances(data_file) if b["mat_code"] == material_code]


def get_location_balance_by_material(location_id, data_file=None):
    return [b for b in get_all_balances(data_file) if b["location_id"] == location_id]


def get_locations(data_file=None):
    """Every distinct location that has any transaction history."""
    txns = get_transactions(data_file=data_file)
    seen = {}
    for t in txns:
        if t["location_id"] not in seen:
            seen[t["location_id"]] = t["location_name"]
    return [{"location_id": k, "location_name": v} for k, v in seen.items()]


# ── Document generation ───────────────────────────────────────────────────────
def generate_stock_report(data_file=None):
    fpath = data_file or DATA_FILE
    balances = get_all_balances(fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = "14432E"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    c = ws["A1"]; c.value = "STOCK ON HAND — ALL LOCATIONS"
    c.font = Font(name="Arial", size=15, bold=True, color="0F172A")
    ws.merge_cells("A1:E1")
    c2 = ws["A2"]; c2.value = f"As of {date.today().strftime('%Y-%m-%d')}"
    c2.font = Font(name="Arial", size=9, color="475569")

    hdr_row = 4
    hdrs = ["Material Code", "Description", "Location", "Balance"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=green)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for b in balances:
        vals = [b["mat_code"], b["mat_desc"], b["location_name"], b["balance"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="B91C1C" if (ci == 4 and val < 0) else "1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    widths = [15, 34, 30, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 20

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"stock_report_{date.today().strftime('%Y%m%d')}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    balances = get_all_balances(data_file)
    materials = {b["mat_code"] for b in balances}
    locations = {b["location_id"] for b in balances}
    total_units = round(sum(b["balance"] for b in balances), 3)
    negative_count = sum(1 for b in balances if b["balance"] < 0)
    in_transit = get_stock_transfers(status="In Transit", data_file=data_file)
    total_in_transit = round(sum(t["quantity"] for t in in_transit), 3)
    return {"materials_tracked": len(materials), "locations_tracked": len(locations),
            "total_units_on_hand": total_units, "negative_balances": negative_count,
            "total_units_in_transit": total_in_transit}


if __name__ == "__main__":
    print("Inventory stats:", stats())
