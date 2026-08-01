"""
migrate_o2c_fulfillment.py — one-time port of Fulfillments and
Fulfillment_Items from data.xlsx into erp_pilot.db.

Fourth O2C slice — depends on Sales_Orders already being migrated.
Idempotent.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_fulfillments(wb, conn):
    if "Fulfillments" not in wb.sheetnames:
        return 0
    ws = wb["Fulfillments"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fid = str(row[hdrs.get("Fulfillment_ID", 0)] or "").strip()
        if not fid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((fid, g("SO_ID"), g("Customer_ID"), g("Customer_Name"), g("Status"),
                    g("Created_Date"), g("Shipped_Date"), g("Delivered_Date"), g("Carrier"),
                    g("Tracking_Ref"), g("Delivery_Location"), g("POD_Reference"), g("Notes")))
    conn.execute("DELETE FROM fulfillments")
    conn.executemany(
        "INSERT INTO fulfillments (fulfillment_id, so_id, customer_id, customer_name, status, "
        "created_date, shipped_date, delivered_date, carrier, tracking_ref, delivery_location, "
        "pod_reference, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_fulfillment_items(wb, conn):
    if "Fulfillment_Items" not in wb.sheetnames:
        return 0
    ws = wb["Fulfillment_Items"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fid = str(row[hdrs.get("Fulfillment_ID", 0)] or "").strip()
        if not fid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((fid, g("Line_Item"), g("Material_Code"), g("Material_Desc"),
                    g("UOM"), g("Qty_Ordered"), g("Qty_Shipped")))
    conn.execute("DELETE FROM fulfillment_items")
    conn.executemany(
        "INSERT INTO fulfillment_items (fulfillment_id, line_item, material_code, material_desc, "
        "uom, qty_ordered, qty_shipped) VALUES (?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_f = migrate_fulfillments(wb, conn)
        n_fi = migrate_fulfillment_items(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"fulfillments": n_f, "fulfillment_items": n_fi}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Fulfillments/Fulfillment_Items to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--db", default=None)
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
