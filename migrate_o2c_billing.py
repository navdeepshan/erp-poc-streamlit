"""
migrate_o2c_billing.py — one-time port of Invoices and Invoice_Items
from data.xlsx into erp_pilot.db.

Fifth O2C slice — depends on Fulfillments, Sales_Orders, Customer_Master
already being migrated. Idempotent.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_invoices(wb, conn):
    if "Invoices" not in wb.sheetnames:
        return 0
    ws = wb["Invoices"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        iid = str(row[hdrs.get("Invoice_ID", 0)] or "").strip()
        if not iid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((iid, g("Fulfillment_ID"), g("SO_ID"), g("Customer_ID"), g("Customer_Name"),
                    g("Customer_GSTIN"), g("Invoice_Date"), g("Due_Date"), g("Status"),
                    g("Payment_Terms"), g("Currency"), g("Place_of_Supply"), g("Subtotal"),
                    g("CGST_Total"), g("SGST_Total"), g("IGST_Total"), g("Grand_Total"), g("Notes")))
    conn.execute("DELETE FROM invoices")
    conn.executemany(
        "INSERT INTO invoices (invoice_id, fulfillment_id, so_id, customer_id, customer_name, "
        "customer_gstin, invoice_date, due_date, status, payment_terms, currency, "
        "place_of_supply, subtotal, cgst_total, sgst_total, igst_total, grand_total, notes) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_invoice_items(wb, conn):
    if "Invoice_Items" not in wb.sheetnames:
        return 0
    ws = wb["Invoice_Items"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        iid = str(row[hdrs.get("Invoice_ID", 0)] or "").strip()
        if not iid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((iid, g("Line_Item"), g("Material_Code"), g("Material_Desc"), g("HSN_Code"),
                    g("UOM"), g("Qty"), g("Unit_Price"), g("Taxable_Value"), g("GST_Rate"),
                    g("CGST_Amount"), g("SGST_Amount"), g("IGST_Amount"), g("Line_Total")))
    conn.execute("DELETE FROM invoice_items")
    conn.executemany(
        "INSERT INTO invoice_items (invoice_id, line_item, material_code, material_desc, "
        "hsn_code, uom, qty, unit_price, taxable_value, gst_rate, cgst_amount, sgst_amount, "
        "igst_amount, line_total) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_inv = migrate_invoices(wb, conn)
        n_ii = migrate_invoice_items(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"invoices": n_inv, "invoice_items": n_ii}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Invoices/Invoice_Items to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--db", default=None)
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
