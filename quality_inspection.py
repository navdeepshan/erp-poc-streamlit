"""
quality_inspection.py — Quality Inspection (S2C, manufacturing pilot).

Deliberately record-only. Per direct instruction: goods receipt and
invoice processing must never wait on inspection turnaround, and a
production payment-proposal system outside this PoC's scope is what
actually gates payment — this module's only job is to give that system
something real to check. Nothing here blocks GR, blocks an invoice, or
blocks anything else. get_po_quality_status() is the hook a payment
proposal system would call; this module doesn't implement one.

Anchored to GR, not an invoice — deliberately. Quality inspection
examines what physically arrived, which is exactly what a GR represents;
it has no dependency on whether that shipment has been billed yet.
Skipping vendor-invoice/three-way-match (per direct instruction) doesn't
block this module at all, which is exactly why QC was sequenced to not
need it.

Five statuses, not three — "not done" and "failed" carry very different
weight for whoever reads this downstream, and a batch is rarely a clean
pass/fail:
  Not Done      no inspection recorded for this line yet
  In Progress   some but not all of the received qty has been inspected
  Passed        entire received qty inspected, all passed
  Failed        entire received qty inspected, all failed
  Partial Pass  entire received qty inspected, a genuine mix

Upsert semantics on record_inspection() — re-inspecting a line updates
its one record rather than creating a growing audit trail. Fine for a
PoC; a real QMS would want inspection history, not just current state.

New sheet:
  Quality_Inspections   QI_ID | GR_ID | PO_Number | Line_Item |
                        Material_Code | Material_Desc | Qty_Received |
                        Qty_Passed | Qty_Failed | Inspected_By |
                        Inspection_Date | Notes

SQLite pilot: Quality_Inspections now lives in erp_pilot.db (table
`quality_inspections`). This module's own upsert semantics are now a
real UNIQUE(gr_id, material_code) constraint + ON CONFLICT, same
pattern as bom.add_bom_line(). Also fixed along the way: this module
used to read GR_Items directly (a second, independent implementation
of "read GR_Items" alongside goods_receipt.py's own) — it now delegates
to goods_receipt.get_gr_items_index(), so GR_Items has exactly one
reader in the whole codebase.

Quality Hold (QHD-US-01 extension, 2026-08-09): a real, corrective
change — until now, "Failed" was recorded here but never actually
excluded from available on-hand, meaning a Sales Order could genuinely
reserve or ship material that had already failed inbound inspection.
Fixed the same way ship_transfer()'s own Transfer Out already handles
a real physical state change: record_inspection() now posts a real
"Quality Hold" transaction against inventory.py's own ledger for
whatever quantity is newly failed (a signed delta, so a correction to
an existing inspection can move the held quantity up or down before
any disposition exists). Because inventory.get_balance() is what every
"available" calculation in this codebase already reads from (ATP's
reservation check, ship_transfer()'s own eligibility check, every
bom.py position function), this one change makes all of them correctly
exclude quality-held stock with no separate exclusion term threaded
into any of them — the same reasoning that already makes In Transit
work without its own bucket. A real quality_holds table tracks
disposition state (Held -> Return to Vendor or Scrap) since that's a
business decision with its own real GL consequence, not a physical
ledger movement — see dispose_quality_hold() and rtv.py (RTV-US-01)
for the physical-return half of this story.
"""

import os, io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import goods_receipt as gr
import inventory as inv

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

QI_SHEET = "Quality_Inspections"
QI_COLS = ["QI_ID", "GR_ID", "PO_Number", "Line_Item", "Material_Code", "Material_Desc",
          "Qty_Received", "Qty_Passed", "Qty_Failed", "Inspected_By", "Inspection_Date", "Notes"]

STATUS_NOT_DONE = "Not Done"
STATUS_IN_PROGRESS = "In Progress"
STATUS_PASSED = "Passed"
STATUS_FAILED = "Failed"
STATUS_PARTIAL = "Partial Pass"


def ensure_sheets(wb=None):
    """Kept for signature compatibility — Quality_Inspections no longer
    lives in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


def _compute_status(qty_received, qty_passed, qty_failed):
    qty_received = qty_received or 0
    qty_passed = qty_passed or 0
    qty_failed = qty_failed or 0
    decided = qty_passed + qty_failed
    if decided <= 0:
        return STATUS_NOT_DONE
    if decided < qty_received - 0.001:
        return STATUS_IN_PROGRESS
    if qty_failed <= 0.001:
        return STATUS_PASSED
    if qty_passed <= 0.001:
        return STATUS_FAILED
    return STATUS_PARTIAL


# ── Read ──────────────────────────────────────────────────────────────────────
def get_inspections(gr_id=None, po_number=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM quality_inspections ORDER BY qi_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"qi_id": r["qi_id"], "gr_id": r["gr_id"], "po_number": r["po_number"],
               "line_item": r["line_item"], "mat_code": r["material_code"],
               "mat_desc": r["material_desc"], "qty_received": r["qty_received"],
               "qty_passed": r["qty_passed"], "qty_failed": r["qty_failed"],
               "inspected_by": r["inspected_by"], "inspection_date": r["inspection_date"],
               "notes": r["notes"]}
        row["status"] = _compute_status(row["qty_received"], row["qty_passed"], row["qty_failed"])
        if gr_id and row["gr_id"] != gr_id:
            continue
        if po_number and row["po_number"] != po_number:
            continue
        out.append(row)
    return out


def _load_gr_items_index(data_file):
    """Delegates to goods_receipt.get_gr_items_index() — GR_Items now
    has exactly one reader in the codebase (see module docstring)."""
    return gr.get_gr_items_index(data_file)


def _load_inspections_index(data_file):
    """Reads Quality_Inspections once, returns {(gr_id, po_item): inspection}
    — keyed by PO line number, not material_code (see record_inspection()'s
    docstring for why: two GR lines can now legitimately share a material
    code, since PR consolidation can produce two PO lines for the same
    item at two different delivery locations)."""
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM quality_inspections").fetchall()
    finally:
        conn.close()
    index = {}
    for r in rows:
        insp = {"qi_id": r["qi_id"], "gr_id": r["gr_id"], "po_number": r["po_number"],
                "line_item": r["line_item"], "po_item": r["po_item"], "mat_code": r["material_code"],
                "mat_desc": r["material_desc"], "qty_received": r["qty_received"],
                "qty_passed": r["qty_passed"], "qty_failed": r["qty_failed"],
                "inspected_by": r["inspected_by"], "inspection_date": r["inspection_date"],
                "notes": r["notes"]}
        insp["status"] = _compute_status(insp["qty_received"], insp["qty_passed"], insp["qty_failed"])
        index[(insp["gr_id"], insp["po_item"])] = insp
    return index


def _gr_quality_status_from_index(gr_id, gr_items_index, inspections_index):
    """Pure computation, no I/O — given pre-loaded indices from the two
    functions above. get_gr_quality_status() below is the convenience
    wrapper for a single GR; anything looping over multiple GRs should
    load the indices once and call this directly instead."""
    gr_items = gr_items_index.get(gr_id, [])
    out = []
    for item in gr_items:
        insp = inspections_index.get((gr_id, item["po_item"]))
        if insp:
            out.append(insp)
        else:
            out.append({"qi_id": None, "gr_id": gr_id, "po_number": None,
                        "line_item": item["line_item"], "po_item": item["po_item"],
                        "mat_code": item["mat_code"],
                        "mat_desc": item["mat_desc"], "qty_received": item["qty_received"],
                        "qty_passed": 0, "qty_failed": 0, "inspected_by": None,
                        "inspection_date": None, "notes": None, "status": STATUS_NOT_DONE})
    return out


def get_gr_quality_status(gr_id, data_file=None):
    """Every line of a GR, inspected or not — mirrors
    goods_receipt.get_po_receipt_status()'s shape so the two read the
    same way in the UI. Single-GR convenience wrapper; if you need this
    for multiple GRs, load the indices once (see _load_gr_items_index/
    _load_inspections_index) and call _gr_quality_status_from_index()
    directly — this version reopens the workbook every call."""
    fpath = data_file or DATA_FILE
    gr_items_index = _load_gr_items_index(fpath)
    inspections_index = _load_inspections_index(fpath)
    return _gr_quality_status_from_index(gr_id, gr_items_index, inspections_index)


def get_grs_needing_inspection(data_file=None):
    """GRs with at least one line still Not Done or In Progress — the
    natural candidate list for the 'record inspection' UI. Loads GR_Items
    and Quality_Inspections once regardless of how many GRs exist,
    instead of get_gr_quality_status() reopening the workbook per GR."""
    fpath = data_file or DATA_FILE
    gr_items_index = _load_gr_items_index(fpath)
    inspections_index = _load_inspections_index(fpath)
    out = []
    for g in gr.get_grs(status="Posted", data_file=fpath):
        lines = _gr_quality_status_from_index(g["gr_id"], gr_items_index, inspections_index)
        pending = [l for l in lines if l["status"] in (STATUS_NOT_DONE, STATUS_IN_PROGRESS)]
        if pending:
            out.append({"gr_id": g["gr_id"], "po_number": g["po_number"],
                        "vendor_name": g["vendor_name"], "lines": len(lines),
                        "pending_lines": len(pending)})
    return out


def get_po_quality_status(po_number, data_file=None):
    """
    The hook a payment-proposal system (out of scope here) would call.
    Aggregates QC across every GR received against this PO, and
    deliberately keeps 'not yet inspected' and 'failed' as separate
    flags rather than one combined 'ok/not ok' — those mean very
    different things to whoever reads this before approving payment.
    """
    fpath = data_file or DATA_FILE
    gr_items_index = _load_gr_items_index(fpath)
    inspections_index = _load_inspections_index(fpath)
    grs = gr.get_grs(po_number=po_number, status="Posted", data_file=fpath)
    all_lines = []
    for g in grs:
        all_lines.extend(_gr_quality_status_from_index(g["gr_id"], gr_items_index, inspections_index))
    not_done = [l for l in all_lines if l["status"] == STATUS_NOT_DONE]
    in_progress = [l for l in all_lines if l["status"] == STATUS_IN_PROGRESS]
    failed = [l for l in all_lines if l["status"] == STATUS_FAILED]
    partial = [l for l in all_lines if l["status"] == STATUS_PARTIAL]
    passed = [l for l in all_lines if l["status"] == STATUS_PASSED]
    return {
        "po_number": po_number, "total_lines": len(all_lines),
        "not_yet_inspected": len(not_done) + len(in_progress),
        "has_failures": len(failed) + len(partial) > 0,
        "failed_lines": len(failed), "partial_lines": len(partial),
        "passed_lines": len(passed),
        "clean": len(not_done) == 0 and len(in_progress) == 0
                 and len(failed) == 0 and len(partial) == 0 and len(all_lines) > 0,
    }


# ── Write ─────────────────────────────────────────────────────────────────────
def _next_qi_id(conn):
    rows = conn.execute("SELECT qi_id FROM quality_inspections WHERE qi_id LIKE 'QI-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["qi_id"].split("-")[1]))
        except Exception: pass
    return f"QI-{mx+1:05d}"


def record_inspection(gr_id, po_item, qty_passed, qty_failed, inspected_by="", notes="", data_file=None):
    """
    po_item identifies which GR line this inspects — the PO line number,
    NOT material code. Changed from material-code-keying after the same
    bug found in goods_receipt.py: two GR lines can now legitimately
    share a material code (a PO consolidated from two different
    delivery locations), so material_code alone can no longer identify
    a single line. mat_code is still accepted implicitly via the GR
    line looked up by po_item — callers that need the material code for
    display can read it off the returned/looked-up line, not pass it in.
    """
    fpath = data_file or DATA_FILE
    if not gr.is_ready_for_quality_check(gr_id, fpath):
        raise ValueError(f"{gr_id} not found or cancelled — nothing to inspect.")
    gr_items = {i["po_item"]: i for i in gr.get_gr_items(gr_id, fpath)}
    if po_item not in gr_items:
        raise ValueError(f"PO line {po_item} is not on {gr_id}.")
    item = gr_items[po_item]
    mat_code = item["mat_code"]
    qty_received = item["qty_received"] or 0
    qty_passed = qty_passed or 0
    qty_failed = qty_failed or 0
    if qty_passed < 0 or qty_failed < 0:
        raise ValueError("Passed/failed quantities can't be negative.")
    if qty_passed + qty_failed > qty_received + 0.001:
        raise ValueError(f"Passed + failed ({qty_passed + qty_failed:g}) exceeds the "
                          f"received quantity ({qty_received:g}) for {mat_code}.")

    g = gr.get_gr(gr_id, fpath)
    db.init_schema()
    conn = db.get_connection()
    try:
        existing = conn.execute(
            "SELECT qi_id, qty_failed FROM quality_inspections WHERE gr_id = ? AND po_item = ?",
            (gr_id, po_item),
        ).fetchone()
        qi_id = existing["qi_id"] if existing else _next_qi_id(conn)
        old_qty_failed = (existing["qty_failed"] or 0) if existing else 0

        if existing:
            disposed = conn.execute(
                "SELECT status FROM quality_holds WHERE qi_id=? AND status != 'Held'",
                (qi_id,)).fetchone()
            if disposed:
                raise ValueError(f"{qi_id} already has a disposition on file "
                                  f"({disposed['status']}) — re-inspecting a disposed line "
                                  f"isn't allowed. A correction needs an explicit reversal, "
                                  f"never a silent edit.")

        conn.execute(
            "INSERT INTO quality_inspections (qi_id, gr_id, po_number, line_item, po_item, "
            "material_code, material_desc, qty_received, qty_passed, qty_failed, inspected_by, "
            "inspection_date, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(gr_id, po_item) DO UPDATE SET "
            "qty_passed=excluded.qty_passed, qty_failed=excluded.qty_failed, "
            "inspected_by=excluded.inspected_by, inspection_date=excluded.inspection_date, "
            "notes=excluded.notes",
            (qi_id, gr_id, g["po_number"], item["line_item"], po_item, mat_code, item["mat_desc"],
             qty_received, qty_passed, qty_failed, inspected_by,
             date.today().strftime("%Y-%m-%d"), notes),
        )
        conn.commit()
    finally:
        conn.close()

    # Quality Hold: a Fail is a real physical state (quarantined stock),
    # posted as a real inventory movement rather than a side-table exclusion
    # term — see module docstring for why this is the whole fix. delta
    # handles a genuine correction (still allowed pre-disposition, blocked
    # above once disposed) moving the held quantity up or down, never
    # assumes a re-inspection only ever increases it.
    delta = round(qty_failed - old_qty_failed, 3)
    if delta != 0:
        inv.record_transaction(mat_code, item["mat_desc"], g["delivery_location"], -delta,
                               "Quality Hold", reference_type="QualityInspection",
                               reference_id=qi_id, notes=f"QC fail on {gr_id} line {po_item}",
                               data_file=fpath)
        conn = db.get_connection()
        try:
            hold = conn.execute(
                "SELECT hold_id FROM quality_holds WHERE qi_id=? AND status='Held'",
                (qi_id,)).fetchone()
            if qty_failed > 0.001:
                if hold:
                    conn.execute("UPDATE quality_holds SET qty=? WHERE hold_id=?",
                                (qty_failed, hold["hold_id"]))
                else:
                    hold_id = _next_hold_id(conn)
                    conn.execute(
                        "INSERT INTO quality_holds (hold_id, qi_id, gr_id, po_number, po_item, "
                        "material_code, material_desc, location_id, qty, status, created_date) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (hold_id, qi_id, gr_id, g["po_number"], po_item, mat_code,
                         item["mat_desc"], g["delivery_location"], qty_failed, "Held",
                         date.today().strftime("%Y-%m-%d")))
            elif hold:
                # A correction eliminated the failure entirely -- the transaction
                # above already restored the quantity to available; nothing left
                # to hold, so the now-empty Held row is removed, not left at zero.
                conn.execute("DELETE FROM quality_holds WHERE hold_id=?", (hold["hold_id"],))
            conn.commit()
        finally:
            conn.close()

    return {"qi_id": qi_id, "status": _compute_status(qty_received, qty_passed, qty_failed)}


# ── Quality Hold ────────────────────────────────────────────────────────────────
def _next_hold_id(conn):
    rows = conn.execute("SELECT hold_id FROM quality_holds WHERE hold_id LIKE 'QH-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["hold_id"].split("-")[1]))
        except Exception: pass
    return f"QH-{mx+1:05d}"


def _row_to_hold(r):
    """Normalizes to this codebase's established mat_code/mat_desc
    convention (get_inspections(), get_fulfillment_items(), etc. all do
    the same) — raw column names stay material_code/material_desc only
    at the SQL layer, never in a dict a caller reads from."""
    d = dict(r)
    d["mat_code"] = d.pop("material_code")
    d["mat_desc"] = d.pop("material_desc")
    return d


def get_quality_holds(status=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM quality_holds ORDER BY hold_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = _row_to_hold(r)
        if status and row["status"] != status:
            continue
        out.append(row)
    return out


def get_quality_hold(hold_id, data_file=None):
    conn = db.get_connection()
    try:
        row = conn.execute("SELECT * FROM quality_holds WHERE hold_id=?", (hold_id,)).fetchone()
    finally:
        conn.close()
    return _row_to_hold(row) if row else None


def dispose_quality_hold(hold_id, disposition, disposed_by="", notes="", data_file=None):
    """
    All-or-nothing disposition per hold — matches this project's own
    documented QHD-US-01/RTV-US-01 design (a single Return to Vendor or
    Scrap decision per line, not a partial-quantity split). Once
    disposed, record_inspection() refuses to re-inspect this same line
    (see its own docstring) — a genuine correction after this point is
    an explicit, separately-authorized reversal, never a silent edit.

    Scrap posts a real GL write-off immediately, at the GR's own
    received unit cost (the same valuation basis GR itself posted at):
    Dr Scrap Expense (5200) / Cr Inventory Clearing (1200) — the
    quantity already left available on-hand at hold time, so this is
    purely a balance-sheet consequence, not a further ledger movement.

    Return to Vendor posts no GL entry here at all — only moves the
    hold into a state rtv.py's ship_return_to_vendor() can act on,
    exactly mirroring GR/IR Clearing's own "don't assume resolution
    prematurely" pattern: the real GL consequence is RTV-US-01's own,
    posted only once the physical return shipment actually happens.
    """
    fpath = data_file or DATA_FILE
    hold = get_quality_hold(hold_id, fpath)
    if hold is None:
        raise ValueError(f"{hold_id} not found.")
    if hold["status"] != "Held":
        raise ValueError(f"{hold_id} is '{hold['status']}' — only a Held quantity can be disposed.")
    if disposition not in ("Return to Vendor", "Scrap"):
        raise ValueError(f"Unknown disposition '{disposition}' — must be 'Return to Vendor' or 'Scrap'.")

    je_id = None
    if disposition == "Scrap":
        gr_items = {i["po_item"]: i for i in gr.get_gr_items(hold["gr_id"], fpath)}
        unit_price = (gr_items.get(hold["po_item"]) or {}).get("unit_price") or 0
        write_off_value = round(unit_price * hold["qty"], 2)
        if write_off_value > 0:
            import accounting as acct
            je_id = acct.post_journal_entry(
                "QualityHold", hold_id,
                f"{hold_id} — Scrap write-off, {hold['mat_desc']} ({hold['qty']:g} units)",
                [{"account_code": "5200", "debit": write_off_value, "credit": 0,
                  "description": f"{hold['mat_desc']} scrapped from {hold['gr_id']}"},
                 {"account_code": "1200", "debit": 0, "credit": write_off_value,
                  "description": f"Inventory Clearing write-off — {hold_id}"}],
                data_file=fpath)

    new_status = "Scrapped" if disposition == "Scrap" else "Pending RTV Shipment"
    conn = db.get_connection()
    try:
        conn.execute(
            "UPDATE quality_holds SET status=?, disposed_date=?, disposed_by=?, "
            "disposition_notes=? WHERE hold_id=?",
            (new_status, date.today().strftime("%Y-%m-%d"), disposed_by, notes, hold_id))
        conn.commit()
    finally:
        conn.close()

    return {"hold_id": hold_id, "status": new_status, "je_id": je_id}


# ── Document generation ───────────────────────────────────────────────────────
def generate_inspection_report(gr_id, data_file=None):
    fpath = data_file or DATA_FILE
    g = gr.get_gr(gr_id, fpath)
    if g is None:
        raise ValueError(f"{gr_id} not found.")
    lines = get_gr_quality_status(gr_id, fpath)

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    maroon = "7C2D12"
    status_colors = {STATUS_PASSED: "DCFCE7", STATUS_FAILED: "FEE2E2",
                     STATUS_PARTIAL: "FEF3C7", STATUS_IN_PROGRESS: "DBEAFE", STATUS_NOT_DONE: "F1F5F9"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspection Report"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "QUALITY INSPECTION REPORT")
    ws.merge_cells("A1:G1")
    label("A3", "GR No:");   value("B3", g["gr_id"])
    label("A4", "PO No:");   value("B4", g["po_number"])
    label("D3", "Vendor:");  value("E3", g["vendor_name"])
    label("D4", "Date:");    value("E4", date.today().strftime("%Y-%m-%d"))

    hdr_row = 7
    hdrs = ["#", "Material Code", "Description", "Received", "Passed", "Failed", "Status"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=maroon)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for i, ln in enumerate(lines, 1):
        vals = [i, ln["mat_code"], ln["mat_desc"], ln["qty_received"],
                ln["qty_passed"], ln["qty_failed"], ln["status"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color="1A1A2E")
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if ci == 7:
                c.fill = PatternFill("solid", fgColor=status_colors.get(ln["status"], "FFFFFF"))
        r += 1

    r += 3
    ws.cell(r, 1, "Inspected by:").font = Font(name="Arial", size=10, bold=True)
    r += 3
    ws.cell(r, 1, "Signature: ____________________").font = Font(name="Arial", size=9)
    r += 1
    ws.cell(r, 1, "Name / Date: ____________________").font = Font(name="Arial", size=9)

    widths = [4, 15, 32, 10, 9, 9, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"{gr_id}_QC.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    fpath = data_file or DATA_FILE
    all_grs = gr.get_grs(status="Posted", data_file=fpath)
    by_status = {}
    for g in all_grs:
        for ln in get_gr_quality_status(g["gr_id"], fpath):
            by_status[ln["status"]] = by_status.get(ln["status"], 0) + 1
    return {"by_status": by_status,
            "grs_pending": len(get_grs_needing_inspection(fpath))}


if __name__ == "__main__":
    print("Quality Inspection stats:", stats())
