"""
bom.py — Bill of Materials + explosion-to-PR (S2C, manufacturing pilot).

Deliberately a flat parent-child table, not a header/version structure —
multi-level BOMs come naturally from this shape: a component is treated
as purchasable the moment it has no BOM of its own, and as "build from
sub-components" the moment it does. No separate flag needed; make-vs-buy
falls straight out of whether a BOM row exists for that item. That also
means the same flat table handles arbitrarily deep nesting (assembly of
assemblies) without any special-casing in the explosion logic itself.

Trigger is deliberately manual, per direct instruction — pick a finished
good and a quantity, explode, propose PRs. This is NOT continuous MRP
(no time-phasing, no safety stock, no reorder points) — that's a
different, much bigger system nobody asked for yet.

Netting is now genuinely gross-to-net: propose_pr_lines() nets the
exploded gross requirement against BOTH open PO exposure (reusing
goods_receipt.py's own outstanding-quantity logic) AND on-hand inventory
at the delivery location (reusing inventory.py's ledger). This used to
only net against open POs, because inventory.py didn't exist yet — that
gap is closed now that it does, not left as a permanent limitation.

PRs generated here use the exact same PR_Header/PR_Items schema and
PR-numbering convention as Create PR, and populate Preferred_Vendor from
Item Master's Tags/Keywords the same way the item picker does — so a
BOM-generated PR line flows through the exact same PO-vs-RFP split
already tested, no special handling needed downstream.

SQLite pilot (BOM -> PR -> Consolidation -> PO): BOM_Items now lives
entirely in erp_pilot.db (table `bom_items`), not in data.xlsx. This is
the first migrated slice — confirmed BOM_Items is exclusively owned by
this module, so no other file needed to change. Every public function
below keeps its original signature and return shape (including the
`data_file` parameter, which continues to mean "the Excel path" for the
Item Master / PO / inventory reads still done here) — callers of this
module did not need to change.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, io
from datetime import date, datetime, timedelta

import pr_consolidation as pc
import po_export
import goods_receipt as gr
import inventory as inv
import db

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")


def ensure_sheets(wb):
    """Kept for signature compatibility with any external caller, but
    BOM_Items no longer lives in the Excel workbook — this now just
    makes sure the SQLite schema exists. `wb` is accepted and ignored."""
    db.init_schema()
    return wb


def _vendor_from_tags(tags):
    """Same rule Create PR's item picker uses — an ALL-CAPS first word
    of Tags/Keywords (the convention every real vendor tag follows,
    e.g. WOODPECKER, MELAG), not just any first word. Reimplemented
    locally rather than imported from erp_ui.py, since a backend
    module importing a UI module would be the wrong direction of
    dependency. Real bug fixed here 2026-07-31, not hypothetical:
    a lowercase first word ("handpiece high speed turbine clinical")
    was resolving to vendor="handpiece" — not a real vendor id at all
    — silently affecting 9 items across the current Item Master."""
    t = (tags or "").strip()
    first = t.split()[0] if t else ""
    return first if first.isupper() else ""


# ── Read ──────────────────────────────────────────────────────────────────────
def _load_bom_index(data_file=None):
    """Reads bom_items from SQLite once, returns {parent_code: [child
    rows]} — same shape as before the migration. `data_file` is accepted
    and deliberately IGNORED here: it used to mean "the Excel path to
    read BOM_Items from," but BOM_Items now lives in erp_pilot.db, which
    has exactly one live copy (same pattern as one live data.xlsx), so
    there's nothing for this parameter to select between. Keeping the
    parameter (as a no-op) rather than dropping it means every existing
    call site — many of which pass an Excel path here — keeps working
    without modification, and avoids the exact "which kind of path"
    mistake this migration flagged as a risk: this function never treats
    its argument as a database path, so an Excel path landing here can
    never be mis-opened as one.

    Single query, not one open per lookup — this was the real fix for
    explode_bom()'s ~1.9s of file-open overhead (51 opens x ~0.025s each)
    that iter_rows() alone didn't solve; SQLite removes that cost
    entirely since the whole index still comes from one query."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT parent_code, parent_desc, component_code, component_desc, "
            "qty_per, uom, notes FROM bom_items"
        ).fetchall()
    finally:
        conn.close()
    index = {}
    for row in rows:
        rowd = dict(row)
        index.setdefault(rowd["parent_code"], []).append(rowd)
    return index


def get_bom(parent_code, data_file=None):
    """Direct (single-level) children only. Loads the whole BOM once per
    call — fine for a single lookup (the UI's 'view BOM' tab), but
    explode_bom() below uses _load_bom_index() directly instead, to
    avoid paying this cost on every one of ~50 recursive calls."""
    return _load_bom_index(data_file).get(parent_code, [])


def get_bom_matched_open_sales_orders(data_file=None):
    """
    Confirmed Sales Orders whose ordered item counts as demand for
    inventory-optimization purposes — the O2C side of the chain this was
    built for. Imports sales_order.py directly (no circular risk:
    sales_order.py imports customer_onboarding, quotation,
    pr_consolidation, none of which import bom.py).

    What counts as "demand" depends on org_defaults' Demand Detection
    Mode (default "Manufactured Items Only", preserving the original
    behavior for anyone who hasn't set this explicitly):
      - "Manufactured Items Only" — item must be a BOM parent (a
        finished good this org actually builds). The original behavior,
        right for a manufacturer like Genrobotics.
      - "All Items" — every confirmed Sales Order line counts, BOM or
        not. Right for a pure-trading/distribution business (e.g. IDS
        Denmed) that sells finished goods it never manufactures — under
        the old BOM-only logic, none of their Sales Orders would ever
        register as demand, and Inventory Position & Transfers would
        show nothing no matter how real their stock imbalance was. This
        is the one function get_inventory_position() and
        get_transfer_opportunities() both depend on for their demand
        signal, so fixing it here is the whole fix — nothing downstream
        needed to change.

    Deliberately just SURFACES these — doesn't auto-explode or
    auto-propose anything. Per direct instruction, auto-*proposing* a PR
    is the right level of autonomy; auto-*committing* one is not. A human
    still picks one from this list and clicks explode/propose, same as
    picking a finished good manually.
    """
    fpath = data_file or DATA_FILE
    import sales_order as so
    import org_defaults as od
    all_items_mode = od.get_default("Demand Detection Mode", fpath) == "All Items"
    parents = None if all_items_mode else {fg["code"] for fg in get_finished_goods(fpath)}
    out = []
    for order in so.get_orders(status="Confirmed", data_file=fpath):
        for item in so.get_order_items(order["so_id"], fpath):
            if all_items_mode or item["mat_code"] in parents:
                out.append({"so_id": order["so_id"], "customer_name": order["customer_name"],
                           "mat_code": item["mat_code"], "mat_desc": item["mat_desc"],
                           "qty": item["qty"], "delivery_location": order["delivery_location"],
                           "delivery_geo": order["delivery_geo"],
                           "requested_delivery_date": order["requested_delivery_date"]})
    return out


def get_inventory_position(data_file=None):
    """
    The persistent version of what net_requirements() computes on demand.
    Aggregates gross demand at each location, nets against on-hand stock,
    open PO exposure, and in-transit Stock Transfers already heading
    there, and separately reports what's on hand at every OTHER
    location — the raw material get_transfer_opportunities() below
    turns into a suggestion.

    Demand source is mode-aware (2026-07-30) — goes through
    get_active_demand_events() rather than hand-exploding confirmed
    Sales Orders directly, so this function (and everything downstream
    of it: get_transfer_opportunities(), get_procurement_
    recommendations(), the Position & Transfers tab) reflects whichever
    'Time-Phased Planning Mode' is currently active, not always Sales
    Order Based specifically. Fixed after finding the gap directly: a
    tenant on Reorder Qty Based mode with zero Sales Orders would
    otherwise see this function return an empty list — and everything
    built on top of it along with it — even though
    project_position() was correctly finding a real projected
    stock-out from the very same rate-based demand. Sums every event
    within the horizon regardless of its individual date, matching
    this function's own point-in-time, not time-phased, nature — see
    project_position() for the date-aware view of the same data.

    In-transit netting added 2026-07-30 — a real bug, not a refinement:
    without it, a destination's shortfall never shrank the moment stock
    was shipped toward it, only once it was actually received (which
    can be days later). get_transfer_opportunities() would then keep
    suggesting, and a Ship click would keep creating, the exact same
    transfer over and over — found by shipping the same route three
    times in a row during testing and getting three real, separate
    9-unit-total transfers for a 3-unit need, not a display glitch.
    In-transit quantity is netted the same way on-hand and open PO
    exposure already were; it does not, on its own, change what a
    Plant is willing to give away as a source (that ceiling is still
    project_position()'s own horizon-aware minimum, computed
    separately in get_transfer_opportunities()).
    """
    fpath = data_file or DATA_FILE
    demand_events = get_active_demand_events(fpath)
    if not demand_events:
        return []

    # Aggregate gross demand per (material, location)
    demand = {key: sum(qty for _d, qty, _ref in events) for key, events in demand_events.items()}

    catalog = {i["code"]: i for i in po_export.load_item_master(fpath)}
    open_exposure = _batch_open_po_exposure(fpath)
    all_balances = inv.get_all_balances(fpath)
    balance_by_key = {(b["mat_code"], b["location_id"]): b["balance"] for b in all_balances}

    in_transit_by_key = {}
    for t in inv.get_stock_transfers(status="In Transit", data_file=fpath):
        key = (t["material_code"], t["to_location"])
        in_transit_by_key[key] = in_transit_by_key.get(key, 0) + (t["quantity"] or 0)

    out = []
    for (mat_code, loc), gross in demand.items():
        on_hand = max(0.0, balance_by_key.get((mat_code, loc), 0.0))
        open_po = open_exposure.get(mat_code, 0)
        in_transit = in_transit_by_key.get((mat_code, loc), 0.0)
        net = max(0, gross - on_hand - open_po - in_transit)
        elsewhere = [{"location_id": b["location_id"], "balance": b["balance"]}
                    for b in all_balances if b["mat_code"] == mat_code and b["location_id"] != loc
                    and b["balance"] > 0]
        out.append({"mat_code": mat_code, "mat_desc": catalog.get(mat_code, {}).get("desc", mat_code),
                    "location_id": loc, "gross_demand": gross, "on_hand": on_hand,
                    "open_po": open_po, "in_transit": in_transit, "net_position": net,
                    "stock_elsewhere": elsewhere})
    return sorted(out, key=lambda p: (-p["net_position"], p["mat_code"]))


# ── Time-phased planning (INV-US-03, Phase 3) ────────────────────────────────
def set_planning_params(material_code, location, min_qty, max_qty, reorder_cadence_days,
                        data_file=None):
    """
    Configures a (material, location) for Reorder Qty Based demand —
    the manual, zero-history-required mode: a Planner declares min/max
    and how often they typically reorder, and this becomes an implied
    consumption rate ((max - min) / cadence), converted into synthetic
    recurring demand events by _reorder_qty_demand_events() below.
    Reference/config data, same weight-class as plant_pair_lead_time_
    estimates — not wiped by seed_manager's reset, an upsert like that
    table's own setter.
    """
    conn = db.get_connection()
    try:
        conn.execute(
            "INSERT INTO material_location_planning_params "
            "(material_code, location, min_qty, max_qty, reorder_cadence_days) "
            "VALUES (?,?,?,?,?) ON CONFLICT(material_code, location) "
            "DO UPDATE SET min_qty=excluded.min_qty, max_qty=excluded.max_qty, "
            "reorder_cadence_days=excluded.reorder_cadence_days",
            (material_code, location, min_qty, max_qty, reorder_cadence_days),
        )
        conn.commit()
    finally:
        conn.close()


def get_planning_params(data_file=None):
    """Every configured (material, location) planning parameter set —
    what the Reorder Qty Based mode's settings screen lists, and what
    _reorder_qty_demand_events() reads to generate synthetic demand."""
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM material_location_planning_params").fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


def _reorder_qty_demand_events(horizon_days=None, data_file=None):
    """
    Reorder Qty Based mode's demand producer — same output shape as
    _batch_demand_events() ({(mat_code, location): [(date, qty, ref),
    ...]}), so project_position() can't tell the difference and needs
    no special-casing. The events themselves are assumed, not
    confirmed — every reference is tagged 'REORDER-QTY-ASSUMED'
    precisely so nothing downstream can mistake one for a real order,
    and the dispatcher below (get_active_demand_events) only ever calls
    this function when Reorder Qty Based is the tenant's single active
    mode — an assumed event never sits on the same trajectory as a
    confirmed one.

    For each configured (material, location): implied rate = (max -
    min) / cadence. Generates one recurring demand event of (max - min)
    units every cadence days, starting one cadence out from today,
    through the horizon. A location with no cadence configured
    (cadence <= 0) is skipped entirely rather than dividing by zero or
    guessing one.
    """
    fpath = data_file or DATA_FILE
    horizon = int(horizon_days or 90)
    horizon_cutoff = date.today() + timedelta(days=horizon)

    events = {}
    for p in get_planning_params(fpath):
        cadence = p["reorder_cadence_days"]
        min_q, max_q = p["min_qty"] or 0, p["max_qty"] or 0
        if not cadence or cadence <= 0 or max_q <= min_q:
            continue
        qty_per_cycle = max_q - min_q
        key = (p["material_code"], p["location"])
        d = date.today() + timedelta(days=cadence)
        cycle = 1
        while d <= horizon_cutoff:
            events.setdefault(key, []).append(
                (d.isoformat(), qty_per_cycle, f"REORDER-QTY-ASSUMED-cycle{cycle}"))
            d += timedelta(days=cadence)
            cycle += 1
    return events


def get_active_demand_events(data_file=None):
    """
    Dispatcher — returns _batch_demand_events()'s shape from whichever
    producer the tenant's 'Time-Phased Planning Mode' (org_defaults)
    currently selects. Every caller that needs demand events
    (project_position(), project_all_positions(),
    get_transfer_opportunities(), get_procurement_recommendations())
    goes through this instead of calling _batch_demand_events()
    directly, so switching modes changes the whole engine's input in
    one place.

    'Optimize Existing PRs' deliberately returns no demand events at
    all here — an Open PR's implied need doesn't inject as trajectory
    demand (it would roughly cancel against that same PR's own later
    PO-arrival supply event, telling the projection nothing real). That
    mode's actual value — lead-time sufficiency checks, cross-location
    duplicate warnings, and feeding PR-implied need into the transfer
    side — lives in get_pr_optimization_analysis() instead, a
    genuinely different kind of output, not a demand-event producer.
    """
    import org_defaults as od
    fpath = data_file or DATA_FILE
    mode = od.get_default("Time-Phased Planning Mode", fpath)
    if mode == "Reorder Qty Based":
        return _reorder_qty_demand_events(data_file=fpath)
    if mode == "Optimize Existing PRs":
        return {}
    return _batch_demand_events(fpath)  # "Sales Order Based" — the default


def _batch_demand_events(data_file=None):
    """
    Every confirmed order's demand, exploded to component level but —
    unlike get_inventory_position() — keeping each order's own
    requested_delivery_date attached per event, not aggregated away.
    get_inventory_position() nets everything as if it's all needed
    'now'; a time-phased projection needs to know exactly which date
    each unit of demand actually falls due on.

    Returns {(mat_code, location): [(date, qty, so_id), ...]}, dates as
    'YYYY-MM-DD' strings, unsorted (the caller sorts once it's merged
    with supply events).
    """
    fpath = data_file or DATA_FILE
    events = {}
    for o in get_bom_matched_open_sales_orders(fpath):
        if not o.get("requested_delivery_date"):
            continue  # no date to place this demand on — excluded, never guessed
        exploded = explode_bom(o["mat_code"], o["qty"], fpath)
        loc = o["delivery_location"] or "Unspecified"
        for mat_code, qty in exploded.items():
            events.setdefault((mat_code, loc), []).append(
                (o["requested_delivery_date"], qty, o["so_id"]))
    return events


def _batch_supply_events(data_file=None):
    """
    Every expected inbound with a real, computable arrival date — open
    PO lines (Phase 1) and in-transit Stock Transfers (Phase 2), each
    tagged to the (material, destination location) they'll land at.

    Deliberately excludes Open PRs entirely — a PR has no committed
    vendor, price or reliable lead time yet, so there's no honest date
    to place it on. Counting it here would understate real risk. Its
    only role in this story is suppressing a duplicate recommendation
    later, not feeding the trajectory itself (see INV-US-03's own
    business rules).

    Returns {(mat_code, location): [(date, qty, kind, ref), ...]},
    same shape as _batch_demand_events() for a clean merge.
    """
    import pr_consolidation as pc
    fpath = data_file or DATA_FILE
    events = {}

    for line in pc.get_open_po_lines_with_dates(data_file=fpath):
        if not line["expected_arrival_date"]:
            continue  # no resolvable lead time — excluded, never guessed
        key = (line["material_code"], line["delivery_location"])
        events.setdefault(key, []).append(
            (line["expected_arrival_date"], line["outstanding_qty"], line["po_number"]))

    for t in inv.get_stock_transfers(status="In Transit", data_file=fpath):
        if not t["shipped_date"]:
            continue
        lt = inv.get_plant_pair_transfer_lead_time(t["from_location"], t["to_location"], fpath)
        arrival = (datetime.strptime(t["shipped_date"], "%Y-%m-%d").date() +
                  timedelta(days=lt["days"])).isoformat()
        key = (t["material_code"], t["to_location"])
        events.setdefault(key, []).append((arrival, t["quantity"], t["transfer_id"]))

    return events


def project_position(mat_code, location, horizon_days=None, data_file=None,
                     demand_events=None, supply_events=None):
    """
    The actual time-phased engine: a day-by-day (really event-by-event
    — see below) projected position for one material at one location,
    starting from real current on-hand and walking forward through
    every known future event in date order.

    Event-driven rather than a literal loop over every day in the
    horizon — inventory position is a step function, it only changes
    on the specific dates something actually happens, so this is both
    cheaper and a more honest model of reality than a continuous daily
    walk would be.

    Same-day tie-break: an arrival is applied before a demand event on
    the same date (optimistic but reasonable — stock landing this
    morning can cover a shipment going out this afternoon). Encoded by
    sorting supply events ahead of demand events within the same date,
    not by any special-casing in the walk itself.

    demand_events/supply_events accept pre-fetched batch results (see
    project_all_positions() below) so a tenant-wide run doesn't re-run
    the same expensive queries once per material/location pair —
    passing neither runs this one pair standalone, fetching fresh.

    Returns {mat_code, location, starting_balance, trajectory
    (list of {date, balance, event} in order), stockout_date (first
    date balance goes negative, or None), min_projected_balance
    (the lowest point in the trajectory — what get_transfer_
    opportunities() will eventually use as 'safely available to give
    away', once wired up)}.
    """
    fpath = data_file or DATA_FILE
    horizon = int(horizon_days or 90)
    horizon_cutoff = (date.today() + timedelta(days=horizon)).isoformat()

    if demand_events is None:
        demand_events = get_active_demand_events(fpath)
    if supply_events is None:
        supply_events = _batch_supply_events(fpath)

    starting_balance = inv.get_balance(mat_code, location, fpath)

    events = []
    for d, qty, ref in supply_events.get((mat_code, location), []):
        if d <= horizon_cutoff:
            events.append((d, 0, qty, "supply", ref))   # tie-break rank 0 — before demand
    for d, qty, ref in demand_events.get((mat_code, location), []):
        if d <= horizon_cutoff:
            events.append((d, 1, -qty, "demand", ref))  # tie-break rank 1 — after supply
    events.sort(key=lambda e: (e[0], e[1]))

    running = starting_balance
    trajectory = [{"date": date.today().isoformat(), "balance": running, "event": None}]
    stockout_date = None
    min_balance = running
    for d, _rank, delta, kind, ref in events:
        running += delta
        trajectory.append({"date": d, "balance": running, "event": f"{kind}:{ref}"})
        min_balance = min(min_balance, running)
        if running < 0 and stockout_date is None:
            stockout_date = d

    return {"mat_code": mat_code, "location": location, "starting_balance": starting_balance,
            "trajectory": trajectory, "stockout_date": stockout_date,
            "min_projected_balance": min_balance}


def project_all_positions(horizon_days=None, data_file=None):
    """
    Batch entry point — projects every (material, location) pair that
    has real demand inside the horizon, in one pass, fetching every
    demand and supply event exactly once rather than re-querying per
    pair. This is the NFR-driving concern: a naive per-material version
    of this would blow the tenant-wide 60-second target immediately at
    any real data volume, the same reason
    bom._batch_open_po_exposure() already batches instead of querying
    per material.

    Only projects pairs with at least one demand event — never every
    material/location in the whole Item Master, matching this story's
    own refusal to manufacture a signal the data doesn't support.

    Returns a list of project_position()'s own return shape, one per
    projected (material, location) pair.
    """
    fpath = data_file or DATA_FILE
    demand_events = get_active_demand_events(fpath)
    supply_events = _batch_supply_events(fpath)
    return [project_position(mat_code, loc, horizon_days, fpath, demand_events, supply_events)
            for (mat_code, loc) in demand_events.keys()]


def get_transfer_opportunities(data_file=None):
    """
    Materials with a real shortage at one location while sitting on hand
    at another — derived directly from get_inventory_position(), not a
    separate guess. suggested_qty is capped at the destination's own
    shortage, what the source can actually spare, AND — when the same
    material is short at more than one destination at once — against
    what's already been allocated to other destinations earlier in this
    same pass.

    Fixed 2026-07-29: an earlier version priced every destination's
    suggestion off the source's full, undiminished balance, independent
    of every other destination also drawing on it — so two destinations
    competing for the same limited source stock could each show a full
    suggestion, and the two together could add up to more than the
    source actually has (each individual execute_transfer() call still
    safely refused if the maths didn't work out by the time someone
    clicked it, but the suggestions themselves didn't agree with each
    other in advance).

    Allocation order: largest shortage served first, a real business-
    logic choice the data doesn't dictate on its own — see
    CONTEXT_HANDOFF_v2.md for the reasoning. A location that itself has
    unmet demand for a material is never treated as a source for
    another destination's shortage in the same material, even if it
    happens to be sitting on some on-hand balance — it needs that stock
    itself.

    Upgraded 2026-07-29 to use each source's horizon-aware minimum
    projected balance (project_position()'s own min_projected_balance,
    floored at 0) instead of today's raw on-hand snapshot as the
    ceiling on what a source can safely give away. A source could look
    like it has surplus today and still be walking into its own
    shortage in three weeks — the old snapshot-only version couldn't
    see that; suggesting a transfer today at a quantity that leaves the
    source short of its own confirmed demand later would just move the
    shortfall around, not resolve anything. Structurally this changes
    nothing about the allocation logic above (still largest-shortage-
    first, still a shared, decrementing pool per source) — it only
    changes what number that pool starts from.

    Only ever proposes a starting point — the UI lets a person edit the
    suggested quantity before executing, and inventory.execute_transfer()
    is what actually enforces real availability at the moment of
    execution, not this function.
    """
    fpath = data_file or DATA_FILE
    position = [p for p in get_inventory_position(fpath) if p["net_position"] > 0]
    if not position:
        return []

    by_material = {}
    for p in position:
        by_material.setdefault(p["mat_code"], []).append(p)

    # Batch-fetched once for the whole call — project_position() takes
    # pre-fetched events precisely so this doesn't re-query per source,
    # the same discipline project_all_positions() already applies.
    demand_events = get_active_demand_events(fpath)
    supply_events = _batch_supply_events(fpath)

    out = []
    for mat_code, dests in by_material.items():
        dest_location_ids = {p["location_id"] for p in dests}
        # Every location this material has stock at, across every destination's own
        # stock_elsewhere list, excluding any location that's itself short of this
        # material — a location with unmet demand doesn't supply another.
        remaining_at_source = {}
        original_available = {}
        current_balance = {}
        for p in dests:
            for s in p["stock_elsewhere"]:
                if s["location_id"] in dest_location_ids:
                    continue
                if s["location_id"] in remaining_at_source:
                    continue
                proj = project_position(mat_code, s["location_id"], data_file=fpath,
                                        demand_events=demand_events, supply_events=supply_events)
                safe_qty = max(0.0, proj["min_projected_balance"])
                remaining_at_source[s["location_id"]] = safe_qty
                original_available[s["location_id"]] = safe_qty
                current_balance[s["location_id"]] = s["balance"]
        if not remaining_at_source:
            continue

        for p in sorted(dests, key=lambda p: -p["net_position"]):  # largest shortage first
            need = p["net_position"]
            candidate_sources = sorted(
                (s for s in p["stock_elsewhere"] if s["location_id"] in remaining_at_source),
                key=lambda s: -remaining_at_source[s["location_id"]])
            for src in candidate_sources:
                if need <= 0:
                    break
                avail = remaining_at_source[src["location_id"]]
                suggested = min(need, avail)
                if suggested > 0:
                    out.append({"mat_code": mat_code, "mat_desc": p["mat_desc"],
                               "from_location": src["location_id"], "to_location": p["location_id"],
                               "suggested_qty": suggested, "shortage_at_destination": p["net_position"],
                               "available_at_source": original_available[src["location_id"]],
                               "current_balance_at_source": current_balance[src["location_id"]]})
                    remaining_at_source[src["location_id"]] -= suggested
                    need -= suggested
    return sorted(out, key=lambda t: -t["suggested_qty"])


def get_procurement_recommendations(data_file=None):
    """
    Phase 4 (INV-US-03) — the final outcome for every shortfall that
    transfers (get_transfer_opportunities(), above) can't fully
    resolve. Implements the story's own three-outcome resolution
    order, tried in this sequence, mutually exclusive per gap:

      1. Transfer (already handled by get_transfer_opportunities() —
         this function only ever looks at what's LEFT after transfers,
         never recomputes or duplicates that decision).
      2. Timed PR — proposed only if there's still enough time for
         normal procurement to arrive before the projected stock-out
         date (days remaining >= the material's own procurement lead
         time), AND no Open PR already covers this exact gap.
      3. At Risk — the remaining gap can't be closed by normal means
         in time. Deliberately produces NO PR proposal at all, ever,
         for this gap: proposing one that everyone can already tell
         will arrive too late would misrepresent a real problem as
         'being handled' when it isn't. This is a hard rule, not a
         style choice — see INV-US-03's own business rules.

    Deduplication reuses the real Open PR data already in the system
    (pr_consolidation.get_pr_items(open_only=True)) rather than a new
    check — a gap already covered by an existing PR or PO surfaces that
    instead of proposing a second one. Real bug fixed 2026-07-31, found
    from a direct report, not discovered independently: this used to
    check only PR lines still in status "Open", so the moment
    consolidation moved a covering line to "PO Proposed" or "PO
    Created" — genuinely MORE committed procurement, not less — the
    coverage became invisible and every one of those materials
    incorrectly fell through to a fresh "Action Needed" recommendation
    for a gap that was already real, active procurement in progress.
    Confirmed directly: consolidating the Genrobotics customer wave's
    12 PRs flipped 22 rows from "Already Covered" to "Action Needed"
    in one step, all real double-counting, none of it an actual new
    gap. Fixed by checking every PR line not yet fully closed out
    (Open, RFP, PO Proposed, PO Created all count), not just "Open".

    Returns a list of {mat_code, mat_desc, location, remaining_gap,
    stockout_date, days_until_stockout, pipeline_lead_time_days,
    lead_time_source, outcome} where outcome is one of 'At Risk',
    'Already Covered by Existing PR/PO', or 'Action Needed' — the last
    of these also carries recommended_qty and required_by_date, ready
    to hand straight to PR-US-01.
    """
    import pr_consolidation as pc
    fpath = data_file or DATA_FILE

    position = [p for p in get_inventory_position(fpath) if p["net_position"] > 0]
    if not position:
        return []

    transfers = get_transfer_opportunities(fpath)
    covered_by_transfer = {}
    for t in transfers:
        key = (t["mat_code"], t["to_location"])
        covered_by_transfer[key] = covered_by_transfer.get(key, 0) + t["suggested_qty"]

    # Every currently-active PR line, batch-fetched once — the basis for
    # deduplication below, matching PR-US-01's own documented duplicate-
    # demand check rather than adding separate logic for it. "Active"
    # means not yet fully closed out: Open, RFP, PO Proposed, and PO
    # Created all represent real, in-progress procurement for that
    # (material, location) — only checking "Open" was the real bug
    # described above.
    open_pr_by_key = {}
    for line in pc.get_pr_items(open_only=False, data_file=fpath):
        if line["status"] not in ("Open", "RFP", "PO Proposed", "PO Created"):
            continue
        key = (line["mat_code"], line["deliv_loc"])
        open_pr_by_key.setdefault(key, []).append(line)

    demand_events = get_active_demand_events(fpath)
    supply_events = _batch_supply_events(fpath)
    today = date.today()

    out = []
    for p in position:
        key = (p["mat_code"], p["location_id"])
        remaining_gap = p["net_position"] - covered_by_transfer.get(key, 0)
        if remaining_gap <= 0:
            continue  # fully resolved by transfer alone — no further outcome needed

        proj = project_position(p["mat_code"], p["location_id"], data_file=fpath,
                                demand_events=demand_events, supply_events=supply_events)
        if not proj["stockout_date"]:
            continue  # defensive — shouldn't normally happen when net_position > 0

        stockout = datetime.strptime(proj["stockout_date"], "%Y-%m-%d").date()
        days_until_stockout = (stockout - today).days

        lt = pc.get_procurement_lead_time(p["mat_code"], fpath)
        pipeline_days = lt["lead_time_days"]

        row = {"mat_code": p["mat_code"], "mat_desc": p["mat_desc"], "location": p["location_id"],
               "remaining_gap": remaining_gap, "stockout_date": proj["stockout_date"],
               "days_until_stockout": days_until_stockout,
               "pipeline_lead_time_days": pipeline_days, "lead_time_source": lt["source"]}

        if pipeline_days is None or days_until_stockout < pipeline_days:
            row["outcome"] = "At Risk"
        elif key in open_pr_by_key:
            row["outcome"] = "Already Covered by Existing PR/PO"
            row["covering_pr"] = open_pr_by_key[key][0]["pr_number"]
        else:
            row["outcome"] = "Action Needed"
            row["recommended_qty"] = remaining_gap
            row["required_by_date"] = proj["stockout_date"]
        out.append(row)

    order = {"At Risk": 0, "Action Needed": 1, "Already Covered by Existing PR/PO": 2}
    return sorted(out, key=lambda r: (order[r["outcome"]], r["days_until_stockout"]))


def get_pr_optimization_analysis(data_file=None):
    """
    'Optimize Existing PRs' mode's real output — deliberately not a
    demand-event producer (see get_active_demand_events()'s own
    docstring for why: an Open PR's implied need would roughly cancel
    against that same PR's later PO-arrival supply event, telling a
    trajectory nothing real). Two independent checks over every
    currently-Open PR line instead:

    1. Lead-time sufficiency — if this PR were consolidated into a PO
       today, would it actually arrive by its own stated required
       date? Compares (today + get_procurement_lead_time()) against
       required_date. This is the same At-Risk-style comparison
       get_procurement_recommendations() makes against a computed
       stock-out date, just anchored to a human's own stated need
       instead.
    2. Cross-location duplicate warning — every (material, location)
       combination with more than one Open PR line. Deliberately
       location-scoped, not material-only: the same material needed at
       two different locations at the same time is two genuine,
       independent needs, not a duplicate. Warn-only, never a block —
       two PRs from different requesters can both be legitimate, and a
       central planner reconciling PRs raised independently across
       locations is exactly the situation where a false positive is
       more likely than in PR-US-01's own single-requester duplicate
       check, so an unoverridable block would be actively harmful here.

    Returns {sufficiency: [...], duplicates: [...]} — sufficiency rows
    carry pr_number, pr_line, mat_code, location, required_date,
    implied_arrival_date, shortfall_days, status ('Sufficient' /
    'Insufficient Lead Time'); duplicate rows carry mat_code, location,
    and the list of competing (pr_number, pr_line, qty, required_date)
    entries.
    """
    import pr_consolidation as pc
    fpath = data_file or DATA_FILE
    today = date.today()

    open_lines = pc.get_pr_items(open_only=True, data_file=fpath)

    sufficiency = []
    by_key = {}
    for line in open_lines:
        key = (line["mat_code"], line["deliv_loc"])
        by_key.setdefault(key, []).append(line)

        if not line["req_date"]:
            continue  # no stated need date — nothing to check sufficiency against
        lt = pc.get_procurement_lead_time(line["mat_code"], fpath)
        row = {"pr_number": line["pr_number"], "pr_line": line["pr_line"],
               "mat_code": line["mat_code"], "mat_desc": line["mat_desc"],
               "location": line["deliv_loc"], "required_date": line["req_date"],
               "pipeline_lead_time_days": lt["lead_time_days"], "lead_time_source": lt["source"]}
        if lt["lead_time_days"] is None:
            row["status"] = "Unknown Lead Time"
            row["implied_arrival_date"] = None
            row["shortfall_days"] = None
        else:
            implied_arrival = today + timedelta(days=int(lt["lead_time_days"]))
            required = datetime.strptime(line["req_date"], "%Y-%m-%d").date()
            row["implied_arrival_date"] = implied_arrival.isoformat()
            if implied_arrival > required:
                row["status"] = "Insufficient Lead Time"
                row["shortfall_days"] = (implied_arrival - required).days
            else:
                row["status"] = "Sufficient"
                row["shortfall_days"] = 0
        sufficiency.append(row)

    duplicates = [{"mat_code": k[0], "location": k[1],
                   "competing_lines": [{"pr_number": l["pr_number"], "pr_line": l["pr_line"],
                                        "qty": l["qty"], "required_date": l["req_date"]}
                                       for l in v]}
                 for k, v in by_key.items() if len(v) > 1]

    return {"sufficiency": sorted(sufficiency,
                                  key=lambda r: (r["status"] != "Insufficient Lead Time",
                                                 -(r.get("shortfall_days") or 0))),
            "duplicates": duplicates}


def get_material_flow_detail(mat_code, from_location, data_file=None):
    """
    Drill-down data for one transfer-opportunity row: which vendor(s)
    supplied mat_code at from_location (via GR history), alongside the
    current on-hand balance there. Feeds the Sankey-style flow view —
    Supplier(s) -> from_location -> {stays here, suggested transfer}.

    Deliberately uses on_hand (not a separately-summed "total ever
    received") as the single flow quantity entering from_location: since
    suggested transfers are always <= on_hand by construction
    (get_transfer_opportunities caps at available_at_source), the
    diagram balances exactly without needing to reconcile historical
    receipts against any consumption that may have happened since.
    """
    fpath = data_file or DATA_FILE
    txns = inv.get_transactions(material_code=mat_code, location_id=from_location, data_file=fpath)
    gr_ids = {t["reference_id"] for t in txns if t["txn_type"] == "GR Receipt" and t["reference_id"]}
    vendor_counts = {}
    for gid in gr_ids:
        g = gr.get_gr(gid, fpath)
        if g and g.get("vendor_name"):
            vendor_counts[g["vendor_name"]] = vendor_counts.get(g["vendor_name"], 0) + 1
    vendors = sorted(vendor_counts.items(), key=lambda x: -x[1])
    on_hand = inv.get_balance(mat_code, from_location, fpath)
    return {"vendors": vendors, "on_hand": on_hand}


def check_items_coverage(mat_codes, data_file=None):
    """
    Batch lookup for the 'about to order something you may already have'
    alert — one pass over inventory/PO data for the whole list, not one
    query per item (that was the exact class of bug that caused BOM's
    explosion to hang before it got fixed; this doesn't repeat it).
    Returns {mat_code: {"total_on_hand": X, "by_location": [...], "open_po": Y}}.
    """
    fpath = data_file or DATA_FILE
    codes = set(mat_codes)
    all_balances = inv.get_all_balances(fpath)
    open_exposure = _batch_open_po_exposure(fpath)
    out = {c: {"total_on_hand": 0.0, "by_location": [], "open_po": open_exposure.get(c, 0)} for c in codes}
    for b in all_balances:
        if b["mat_code"] in codes and b["balance"] > 0:
            out[b["mat_code"]]["total_on_hand"] += b["balance"]
            out[b["mat_code"]]["by_location"].append({"location_id": b["location_id"], "balance": b["balance"]})
    return out


def get_finished_goods(data_file=None, active_only=True):
    """Items that appear as a parent anywhere — the natural 'what can I
    explode' picker list. `data_file` is accepted and ignored, same
    reasoning as _load_bom_index(). Explicit alphabetical ordering by
    parent_code, for predictable UI dropdown behavior (not guaranteed by
    a plain SQL SELECT without ORDER BY).

    active_only=True (the default — every picker in mfg_ui.py that lists
    finished goods to build/explode uses this) excludes parents whose
    Item Master record is currently inactive, same "listing/selection
    filters Active, completeness doesn't" policy used everywhere else
    in this migration. stats() below explicitly passes active_only=False,
    since a structural fact about the BOM (how many finished goods it
    defines) shouldn't change just because an item was deactivated."""
    conn = db.get_connection()
    try:
        if active_only:
            rows = conn.execute(
                "SELECT DISTINCT b.parent_code, b.parent_desc FROM bom_items b "
                "JOIN item_master m ON m.item_code = b.parent_code "
                "WHERE COALESCE(m.active, 'Yes') = 'Yes' "
                "ORDER BY b.parent_code"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT DISTINCT parent_code, parent_desc FROM bom_items "
                "ORDER BY parent_code"
            ).fetchall()
    finally:
        conn.close()
    return [{"code": r["parent_code"], "desc": r["parent_desc"]} for r in rows]


def _would_create_cycle(parent_code, component_code, data_file):
    """True if adding parent_code -> component_code would close a loop
    — i.e. parent_code already appears somewhere in component_code's
    own descendant tree."""
    if parent_code == component_code:
        return True
    index = _load_bom_index(data_file)
    stack = [component_code]
    seen = set()
    while stack:
        node = stack.pop()
        if node in seen:
            continue
        seen.add(node)
        for child in index.get(node, []):
            if child["component_code"] == parent_code:
                return True
            stack.append(child["component_code"])
    return False


def explode_bom(parent_code, quantity, data_file=None, _visited=None, _index=None):
    """
    Recursively explodes parent_code x quantity down to purchasable
    leaves (items with no BOM of their own), aggregating quantities
    where the same leaf is reached through more than one path. Returns
    {material_code: total_qty}. If parent_code itself has no BOM, it's
    treated as its own leaf and returned as-is — makes the function safe
    to call on anything, not just confirmed assemblies.

    `data_file` is threaded through purely for signature/call-site
    compatibility (callers still pass an Excel path here) — it's never
    used to locate BOM data anymore, since _load_bom_index() ignores it.
    """
    fpath = data_file or DATA_FILE
    _visited = _visited or frozenset()
    if _index is None:
        _index = _load_bom_index(fpath)
    if parent_code in _visited:
        raise ValueError(f"Circular BOM detected involving {parent_code}.")
    children = _index.get(parent_code, [])
    if not children:
        return {parent_code: quantity}
    next_visited = _visited | {parent_code}
    requirements = {}
    for child in children:
        child_qty = (child["qty_per"] or 0) * quantity
        sub = explode_bom(child["component_code"], child_qty, fpath, next_visited, _index)
        for code, qty in sub.items():
            requirements[code] = requirements.get(code, 0) + qty
    return requirements


def explode_bom_detailed(parent_code, quantity, data_file=None):
    """explode_bom() enriched with description/UOM/cost/tags from Item
    Master, for display and for PR line construction."""
    fpath = data_file or DATA_FILE
    gross = explode_bom(parent_code, quantity, fpath)
    catalog = {i["code"]: i for i in po_export.load_item_master(fpath)}
    out = []
    for code, qty in gross.items():
        item = catalog.get(code, {})
        out.append({"mat_code": code, "mat_desc": item.get("desc", code),
                    "uom": item.get("uom", ""), "unit_price": item.get("price", 0),
                    "tags": item.get("tags", ""), "gross_qty": qty})
    return out


def _batch_open_po_exposure(data_file):
    """
    Computes {material_code: total_outstanding_qty} across every PO in
    one pass — {ordered - received} per material, floored at 0.

    PO_Items, GR_Header, and GR_Items all now live in SQLite
    (pr_consolidation.py and goods_receipt.py respectively) — this was
    found still reading them from Excel via openpyxl during a final
    sweep after the whole application's migration was otherwise
    complete. Since neither table has been written to Excel since their
    respective migrations, this function had been silently computing
    open-PO exposure as if every PO were empty (real Excel sheets
    minus everything ever consolidated into them = stale zero data) —
    net_requirements()'s open-PO netting was therefore silently a
    no-op for as long as those two migrations have existed. Not caught
    by earlier testing because bom.py's own test suite ran before
    PO_Items/GR_Header/GR_Items moved out from under it; nothing since
    re-exercised net_requirements() against a real open PO.
    """
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT material_code, quantity FROM po_items"
        ).fetchall()
        ordered_by_material = {}
        for row in rows:
            mat = row["material_code"]
            ordered_by_material[mat] = ordered_by_material.get(mat, 0) + (row["quantity"] or 0)

        active_grs = {r["gr_id"] for r in conn.execute(
            "SELECT gr_id FROM gr_header WHERE status != 'Cancelled'"
        ).fetchall()}

        received_by_material = {}
        if active_grs:
            placeholders = ",".join("?" for _ in active_grs)
            gr_rows = conn.execute(
                f"SELECT gr_id, material_code, qty_received FROM gr_items "
                f"WHERE gr_id IN ({placeholders})",
                list(active_grs),
            ).fetchall()
            for row in gr_rows:
                mat = row["material_code"]
                received_by_material[mat] = received_by_material.get(mat, 0) + (row["qty_received"] or 0)
    finally:
        conn.close()

    return {mat: max(0, ordered - received_by_material.get(mat, 0))
            for mat, ordered in ordered_by_material.items()}


def net_requirements(gross_requirements, location_id=None, data_file=None):
    """
    True gross-to-net now: nets against BOTH on-hand inventory at
    location_id (if given) AND open PO exposure. This used to only net
    against open POs — inventory.py didn't exist yet when this was first
    built, and the module docstring said so honestly rather than
    pretending otherwise. Now that a real inventory ledger exists, this
    closes that gap instead of leaving it as a permanent limitation.

    location_id: which location's on-hand stock to net against. If not
    given, on-hand inventory is skipped (same behavior as before this
    function existed) — only makes sense to net against a specific
    location's stock, not a cross-location total, since that's what's
    actually available for a build at that location.

    A negative on-hand balance (which inventory.py allows — see its own
    docstring on why) is clamped to 0 here rather than allowed to
    increase the net requirement — a data/process issue elsewhere
    shouldn't make this function ask for MORE than the gross requirement.
    """
    fpath = data_file or DATA_FILE
    if isinstance(gross_requirements, dict):
        items = [{"mat_code": k, "gross_qty": v} for k, v in gross_requirements.items()]
    else:
        items = gross_requirements
    open_exposure = _batch_open_po_exposure(fpath)
    out = []
    for item in items:
        on_hand = max(0.0, inv.get_balance(item["mat_code"], location_id, fpath)) if location_id else 0.0
        open_qty = open_exposure.get(item["mat_code"], 0)
        net_qty = max(0, item["gross_qty"] - on_hand - open_qty)
        if net_qty > 0:
            out.append({**item, "on_hand_qty": on_hand, "open_po_qty": open_qty, "net_qty": net_qty})
    return out


# Old name kept as an alias — propose_pr_lines() and any external caller
# written against the previous signature still works, just without
# inventory netting (location_id defaults to None).
def net_against_open_pos(gross_requirements, data_file=None):
    return net_requirements(gross_requirements, location_id=None, data_file=data_file)


# ── Write ─────────────────────────────────────────────────────────────────────
def add_bom_line(parent_code, parent_desc, component_code, component_desc,
                  qty_per, uom, notes="", data_file=None):
    """SQLite upsert: same parent+component pair updates qty_per/notes
    instead of duplicating, same behavior as before the migration — now
    expressed as a real UNIQUE(parent_code, component_code) constraint
    plus ON CONFLICT, instead of a manual row-scan. `data_file` is
    accepted and ignored (see _load_bom_index() docstring)."""
    if qty_per is None or qty_per <= 0:
        raise ValueError("Qty per unit must be greater than zero.")
    if _would_create_cycle(parent_code, component_code, data_file):
        raise ValueError(f"Adding {component_code} under {parent_code} would create a "
                          "circular BOM — refusing.")
    db.init_schema()
    conn = db.get_connection()
    try:
        conn.execute(
            "INSERT INTO bom_items "
            "(parent_code, parent_desc, component_code, component_desc, qty_per, uom, notes) "
            "VALUES (?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(parent_code, component_code) DO UPDATE SET "
            "qty_per=excluded.qty_per, notes=excluded.notes",
            (parent_code, parent_desc, component_code, component_desc, qty_per, uom, notes),
        )
        conn.commit()
    finally:
        conn.close()


def propose_pr_lines(parent_code, quantity, requester_id="", requester_name="",
                      requester_dept="", project_id="", delivery_location="",
                      delivery_geo="", required_date=None, net_against_pos=True,
                      data_file=None):
    """
    Explodes the BOM, optionally nets against on-hand inventory at
    delivery_location AND open PO exposure, and writes a real PR
    (PR_Header + PR_Items) — the same schema and PR-numbering Create PR
    uses. Preferred_Vendor is populated from Item Master's Tags/Keywords
    exactly like the item picker does, so these lines flow through the
    normal Consolidate PO-vs-RFP split unchanged.

    PR_Header/PR_Items now live in SQLite (pr_consolidation.py's
    pilot) — this delegates to pr_consolidation.next_pr_number() and
    .create_pr() instead of writing Excel cells directly, so BOM-
    generated PRs use the exact same numbering sequence and storage as
    Create PR's, with no separate implementation to keep in sync.
    """
    fpath = data_file or DATA_FILE
    detailed = explode_bom_detailed(parent_code, quantity, fpath)
    if not detailed:
        raise ValueError(f"{parent_code} has no BOM — nothing to explode.")

    lines = net_requirements(detailed, delivery_location, fpath) if net_against_pos else \
            [{**d, "net_qty": d["gross_qty"]} for d in detailed]
    if not lines:
        raise ValueError("Nothing to requisition — every component is already covered "
                          "by open POs.")

    pr_number = pc.next_pr_number()
    pr_lines = [{
        "vendor": _vendor_from_tags(line.get("tags", "")),
        "mat_code": line["mat_code"], "mat_desc": line["mat_desc"], "uom": line["uom"],
        "qty": line["net_qty"], "req_date": str(required_date) if required_date else "",
        "deliv_loc": delivery_location, "deliv_geo": delivery_geo,
    } for line in lines]
    pc.create_pr(pr_number, requester_id=requester_id, requester_name=requester_name,
                 requester_dept=requester_dept, project_id=project_id, lines=pr_lines)

    return {"pr_number": pr_number, "lines": len(lines),
            "with_vendor": sum(1 for l in lines if _vendor_from_tags(l.get("tags", "")))}


# ── Document generation ───────────────────────────────────────────────────────
def generate_bom_document(parent_code, data_file=None):
    fpath = data_file or DATA_FILE
    index = _load_bom_index(fpath)
    children = index.get(parent_code, [])
    if not children:
        raise ValueError(f"{parent_code} has no BOM defined.")
    parent_desc = children[0]["parent_desc"]

    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = "1E3A5F"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    title("A1", "BILL OF MATERIALS")
    ws.merge_cells("A1:F1")
    ws["A3"] = "Parent:"; ws["A3"].font = Font(name="Arial", size=9, bold=True)
    ws["B3"] = f"{parent_code} — {parent_desc}"; ws["B3"].font = Font(name="Arial", size=10)

    hdr_row = 6
    hdrs = ["#", "Component Code", "Description", "Qty Per Unit", "UOM", "Level"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=navy)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    def write_level(parent, indent, row):
        kids = index.get(parent, [])
        for child in kids:
            vals = [None, child["component_code"], ("  " * indent) + child["component_desc"],
                    child["qty_per"], child["uom"], indent + 1]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row, ci, val)
                c.font = Font(name="Arial", size=9, color="1A1A2E")
                c.border = bdr
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row += 1
            row = write_level(child["component_code"], indent + 1, row)
        return row

    r = hdr_row + 1
    final_r = write_level(parent_code, 0, r)
    for i, row_num in enumerate(range(r, final_r), 1):
        ws.cell(row_num, 1, i)

    widths = [4, 15, 40, 12, 8, 7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"BOM_{parent_code}.xlsx"
    return filename, buf.read()


def stats(data_file=None):
    fgs = get_finished_goods(data_file, active_only=False)
    conn = db.get_connection()
    try:
        line_count = conn.execute("SELECT COUNT(*) FROM bom_items").fetchone()[0]
    finally:
        conn.close()
    return {"finished_goods": len(fgs), "bom_lines": line_count}


if __name__ == "__main__":
    print("BOM stats:", stats())
