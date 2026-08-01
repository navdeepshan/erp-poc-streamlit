"""
migrate_o2c_accounting.py — one-time port of Chart_of_Accounts,
Journal_Entries, and Journal_Entry_Lines from data.xlsx into
erp_pilot.db.

Sixth O2C slice — depends on Invoices/Fulfillments already being
migrated. If Chart_of_Accounts has no pre-existing data (the seed
accounts were never persisted to Excel — e.g. get_accounts() was
never called), the same SEED_ACCOUNTS list accounting.py has always
used is inserted here instead, so a fresh SQLite db still starts with
the accounts every posting function needs. Idempotent.
"""

import argparse
import os
import openpyxl

import db

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

# Same seed list accounting.py has always used — kept here too so a
# fresh SQLite db (no prior Chart_of_Accounts data in Excel) still
# starts with the accounts every posting function needs.
SEED_ACCOUNTS = [
    ("1000", "Cash and Bank", "Asset", "Cash and bank balances"),
    ("1100", "Accounts Receivable", "Asset", "Amounts owed by customers"),
    ("1150", "GST Input Credit - CGST", "Asset",
     "Central GST paid on purchases, recoverable against GST Output"),
    ("1160", "GST Input Credit - SGST", "Asset",
     "State GST paid on purchases, recoverable against GST Output"),
    ("1170", "GST Input Credit - IGST", "Asset",
     "Integrated GST paid on inter-state purchases, recoverable against GST Output"),
    ("1200", "Inventory Clearing", "Asset",
     "Clearing account for inventory value — debited when Goods Receipt posts "
     "(stock arriving), credited when Fulfillment posts COGS (stock leaving). "
     "Production-confirmed stock (manufactured, not purchased) still isn't "
     "modeled here — only GR-driven receipts and O2C-driven shipments post to it."),
    ("2000", "Accounts Payable", "Liability", "Amounts owed to vendors for goods received"),
    ("2100", "GR/IR Clearing", "Liability",
     "Goods Received / Invoice Received clearing — credited (ex-GST) when a GR posts, "
     "debited when the matching Vendor Invoice posts. A nonzero balance is a real, "
     "meaningful signal: goods received but not yet invoiced. Zero once every GR "
     "has a matching invoice."),
    ("2300", "Customer Advances", "Liability",
     "Payments received but not yet applied to a specific invoice"),
    ("4000", "Sales Revenue", "Revenue", "Revenue from sale of goods"),
    ("5000", "Cost of Goods Sold", "Expense", "Cost of goods sold, from Item Master cost basis"),
    ("2200", "GST Output - CGST", "Liability", "Central GST collected on sales, payable to government"),
    ("2210", "GST Output - SGST", "Liability", "State GST collected on sales, payable to government"),
    ("2220", "GST Output - IGST", "Liability", "Integrated GST collected on inter-state sales, payable to government"),
]


def _headers(ws):
    return {str(v or "").strip(): i for i, v in
            enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}


def migrate_coa(wb, conn):
    rows = []
    if "Chart_of_Accounts" in wb.sheetnames:
        ws = wb["Chart_of_Accounts"]
        hdrs = _headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = str(row[hdrs.get("Account_Code", 0)] or "").strip()
            if not code:
                continue
            def g(col):
                i = hdrs.get(col)
                return row[i] if i is not None else None
            rows.append((code, g("Account_Name"), g("Account_Type"), g("Description")))
    if not rows:
        rows = list(SEED_ACCOUNTS)
    conn.execute("DELETE FROM chart_of_accounts")
    conn.executemany(
        "INSERT INTO chart_of_accounts (account_code, account_name, account_type, description) "
        "VALUES (?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_je(wb, conn):
    if "Journal_Entries" not in wb.sheetnames:
        return 0
    ws = wb["Journal_Entries"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        jid = str(row[hdrs.get("JE_ID", 0)] or "").strip()
        if not jid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((jid, g("Entry_Date"), g("Source_Type"), g("Source_ID"), g("Description"),
                    g("Total_Debit"), g("Total_Credit")))
    conn.execute("DELETE FROM journal_entries")
    conn.executemany(
        "INSERT INTO journal_entries (je_id, entry_date, source_type, source_id, description, "
        "total_debit, total_credit) VALUES (?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate_je_lines(wb, conn):
    if "Journal_Entry_Lines" not in wb.sheetnames:
        return 0
    ws = wb["Journal_Entry_Lines"]
    hdrs = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        jid = str(row[hdrs.get("JE_ID", 0)] or "").strip()
        if not jid:
            continue
        def g(col):
            i = hdrs.get(col)
            return row[i] if i is not None else None
        rows.append((jid, g("Line_Item"), g("Account_Code"), g("Account_Name"),
                    g("Debit"), g("Credit"), g("Description")))
    conn.execute("DELETE FROM journal_entry_lines")
    conn.executemany(
        "INSERT INTO journal_entry_lines (je_id, line_item, account_code, account_name, debit, "
        "credit, description) VALUES (?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def migrate(xlsx_path=None, db_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    db.init_schema(db_path)

    # read_only=False deliberately — read_only mode trusts the sheet's
    # <dimension> XML tag to know where to stop iterating, and that tag can
    # go stale (common after rows are pasted/appended by tools other than
    # native Excel save). Confirmed directly: a file with a stale dimension
    # tag silently truncated a 3000-row sheet to 816 rows under read_only=True,
    # with no error at all. Migrations are one-time, not a hot path, so the
    # memory/speed cost of full-load mode is a non-issue here — correctness
    # matters far more than speed for a one-time import.
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    conn = db.get_connection(db_path)
    try:
        n_coa = migrate_coa(wb, conn)
        n_je = migrate_je(wb, conn)
        n_jel = migrate_je_lines(wb, conn)
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"chart_of_accounts": n_coa, "journal_entries": n_je, "journal_entry_lines": n_jel}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrate Chart_of_Accounts/Journal_Entries to SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--db", default=None)
    args = parser.parse_args()

    counts = migrate(args.xlsx, args.db)
    for name, n in counts.items():
        print(f"Migrated {n} {name} row(s) from {args.xlsx} into SQLite.")
