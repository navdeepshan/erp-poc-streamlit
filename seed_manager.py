"""
seed_manager.py — full data reset + reseed from an uploaded Excel file,
for switching between industry-profile demo scenarios (e.g. the Dental
pilot vs. the Genrobotics pilot this PoC currently runs).

Two phases, deliberately kept separate:

  validate_seed_file()  — a pure, read-only parse of the uploaded file.
                          Touches the database not at all. Reports every
                          structural problem found: missing sheets,
                          missing columns, blank required fields,
                          non-numeric values where a number is expected,
                          malformed GSTIN/PAN, duplicate IDs within a
                          sheet, and broken cross-sheet references
                          (a Purchase Bundle line naming an Item Code
                          that isn't in Item Master, etc.).

  reset_and_reseed()    — only ever called after validate_seed_file()
                          has already returned zero errors. Wipes every
                          one of the 45 tables in the schema, then loads
                          the 13 seed/master-data tables from the
                          validated file. The other 32 (every
                          transactional table — PRs, POs, GRs, Sales
                          Orders, Invoices, Journal Entries, etc.) are
                          left empty on purpose: none of that history
                          means anything once the underlying business
                          (items, vendors, customers) has changed out
                          from under it.

Deliberately conservative on malformed data: if ANY row in ANY sheet
fails validation, the whole file is refused — nothing partial ever
gets loaded. A demo's master data is the foundation everything else is
built on; a silently-skipped bad row here would surface much later, in
a much more confusing place (a GR that won't post, an invoice with the
wrong GST split), for no reason a person testing that flow would ever
connect back to a data problem from a completely different session.

Sheet -> table mapping (12 sheets -> 13 tables, Purchase Bundles is
header + lines):
  Org_Profile      -> org_profile            (1 row expected)
  Org_Defaults     -> org_defaults
  Legal_Entities   -> legal_entities
  Delivery_Locations -> delivery_locations
  Item Master      -> item_master
  Purchase Bundles -> purchase_bundles + purchase_bundle_items
  BOM_Items        -> bom_items
  Vendor_Types     -> vendor_types
  Vendor_Master    -> vendor_master
  Customer_Types   -> customer_types
  Customer_Master  -> customer_master
  Chart_of_Accounts -> chart_of_accounts

Every other table in the schema is transactional and gets wiped, not
reseeded — see TRANSACTIONAL_TABLES below for the explicit list (kept
explicit, not "everything not in the seed list", so a newly-added table
that someone forgets to categorize fails loudly instead of silently
landing in the wrong bucket).
"""

import io
import os
import zipfile
from datetime import date

import openpyxl

import db
import vendor_onboarding as vo  # reuse GSTIN/PAN validators — same rules, not a second copy

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")
DB_FILE = db.DB_FILE

ACCOUNT_TYPES = {"Asset", "Liability", "Revenue", "Expense"}

TRANSACTIONAL_TABLES = [
    "pr_items", "pr_header", "po_items", "po_header", "rfp",
    "vendor_documents", "contracts", "contract_items",
    "vrq_requests", "vrq_responses", "rfx_quotes", "rfx_invitations", "rfx_clarifications",
    "gr_header", "gr_items", "quality_inspections", "quality_holds", "rtv_shipments",
    "vendor_ratings", "rmas", "credit_memos", "lots",
    "production_confirmations", "inventory_transactions", "stock_transfers",
    "customer_documents", "quotes", "quote_items",
    "sales_orders", "sales_order_items", "fulfillments", "fulfillment_items",
    "invoices", "invoice_items", "e_invoices", "journal_entries", "journal_entry_lines",
    "payments", "payment_applications",
    "vendor_invoices", "vendor_invoice_payments",
    "sto_header", "sto_lines", "reservations", "backorders",
]

SEED_TABLES = [
    "org_profile", "org_defaults", "legal_entities", "delivery_locations",
    "item_master", "purchase_bundles", "purchase_bundle_items", "bom_items",
    "vendor_types", "vendor_master", "customer_types", "customer_master",
    "chart_of_accounts",
]


# ── Helpers ───────────────────────────────────────────────────────────────────
def _rows(ws, header_row=1):
    """Header dict (name -> col index) + data rows (as dicts), from the
    given header row onward. Blank leading rows (a title row above the
    real header, as Item Master has) are handled by the caller passing
    the right header_row, not guessed here."""
    all_rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    if not all_rows:
        return {}, []
    headers = {str(h).strip(): i for i, h in enumerate(all_rows[0]) if h}
    data = []
    for r in all_rows[1:]:
        if any(c is not None and str(c).strip() != "" for c in r):
            data.append(r)
    return headers, data


def _get(row, headers, col):
    i = headers.get(col)
    if i is None or i >= len(row):
        return None
    v = row[i]
    return v.strip() if isinstance(v, str) else v


def _num(v):
    try:
        return float(v) if v is not None and str(v).strip() != "" else None
    except (ValueError, TypeError):
        return None


# ── Validation ────────────────────────────────────────────────────────────────
def validate_seed_file(file_path):
    """
    Read-only. Returns {"valid": bool, "errors": [...], "warnings": [...],
    "summary": {sheet_name: row_count}, "parsed": {...}} — "parsed" holds
    everything reset_and_reseed() needs, so a valid file only gets
    opened and parsed once, not twice.
    """
    errors = []
    warnings = []
    summary = {}
    parsed = {}

    try:
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    except Exception as e:
        return {"valid": False, "errors": [f"Couldn't open the file: {e}"],
                "warnings": [], "summary": {}, "parsed": {}}

    required_sheets = ["Org_Profile", "Org_Defaults", "Legal_Entities",
                       "Delivery_Locations", "Item Master", "Purchase Bundles",
                       "BOM_Items", "Vendor_Types", "Vendor_Master",
                       "Customer_Types", "Customer_Master", "Chart_of_Accounts"]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        return {"valid": False, "errors": [f"Missing sheet(s): {', '.join(missing)}"],
                "warnings": [], "summary": {}, "parsed": {}}

    # ---- Org_Profile (exactly 1 row expected) ----
    headers, rows = _rows(wb["Org_Profile"])
    req_cols = ["Org_ID", "Legal_Name", "GSTIN", "PAN", "Address", "City", "State",
               "Country", "Bank_Account_No", "IFSC", "Bank_Name", "Contact_Email", "Contact_Phone"]
    missing_cols = [c for c in req_cols if c not in headers]
    if missing_cols:
        errors.append(f"Org_Profile: missing column(s) {missing_cols}")
    elif not rows:
        errors.append("Org_Profile: no data row found")
    else:
        r = rows[0]
        org_id = _get(r, headers, "Org_ID")
        legal_name = _get(r, headers, "Legal_Name")
        if not org_id:
            errors.append("Org_Profile: Org_ID is blank")
        if not legal_name:
            errors.append("Org_Profile: Legal_Name is blank")
        gstin = _get(r, headers, "GSTIN")
        if gstin:
            ok, msg, _ = vo.validate_gstin(gstin)
            if not ok:
                errors.append(f"Org_Profile: GSTIN '{gstin}' invalid — {msg}")
        pan = _get(r, headers, "PAN")
        if pan:
            ok, msg = vo.validate_pan_format(pan)
            if not ok:
                errors.append(f"Org_Profile: PAN '{pan}' invalid — {msg}")
        parsed["org_profile"] = {c: _get(r, headers, c) for c in req_cols}
        # Pincode is optional at the file level (2026-08-10, gst_einvoice.py's
        # own SellerDtls.Pin) so an older seed file without the column still
        # validates clean — _get() already returns None for a missing header.
        parsed["org_profile"]["Pincode"] = _get(r, headers, "Pincode")
    summary["Org_Profile"] = len(rows)

    # ---- Org_Defaults ----
    headers, rows = _rows(wb["Org_Defaults"])
    if "Org_Element" not in headers or "Default_Value" not in headers:
        errors.append("Org_Defaults: missing column(s) Org_Element / Default_Value")
    else:
        seen = set()
        out = []
        for i, r in enumerate(rows, 2):
            el = _get(r, headers, "Org_Element")
            val = _get(r, headers, "Default_Value")
            if not el:
                errors.append(f"Org_Defaults row {i}: Org_Element is blank")
                continue
            if el in seen:
                errors.append(f"Org_Defaults: duplicate Org_Element '{el}'")
            seen.add(el)
            if val is None or str(val).strip() == "":
                errors.append(f"Org_Defaults row {i} ('{el}'): Default_Value is blank")
            out.append((el, val))
        parsed["org_defaults"] = out
    summary["Org_Defaults"] = len(rows)

    # ---- Legal_Entities ----
    headers, rows = _rows(wb["Legal_Entities"])
    req_cols = ["LE_ID", "LE_Name", "GSTIN", "PAN", "Address", "City", "State",
               "Country", "Bank_Account_No", "IFSC", "Bank_Name", "Contact_Email", "Contact_Phone"]
    missing_cols = [c for c in req_cols if c not in headers]
    if missing_cols:
        errors.append(f"Legal_Entities: missing column(s) {missing_cols}")
    else:
        seen = set()
        out = []
        for i, r in enumerate(rows, 2):
            le_id = _get(r, headers, "LE_ID")
            if not le_id:
                errors.append(f"Legal_Entities row {i}: LE_ID is blank")
                continue
            if le_id in seen:
                errors.append(f"Legal_Entities: duplicate LE_ID '{le_id}'")
            seen.add(le_id)
            if not _get(r, headers, "LE_Name"):
                errors.append(f"Legal_Entities row {i} ('{le_id}'): LE_Name is blank")
            gstin = _get(r, headers, "GSTIN")
            if gstin:
                ok, msg, _ = vo.validate_gstin(gstin)
                if not ok:
                    errors.append(f"Legal_Entities '{le_id}': GSTIN invalid — {msg}")
            pan = _get(r, headers, "PAN")
            if pan:
                ok, msg = vo.validate_pan_format(pan)
                if not ok:
                    errors.append(f"Legal_Entities '{le_id}': PAN invalid — {msg}")
            out.append({c: _get(r, headers, c) for c in req_cols})
        parsed["legal_entities"] = out
    summary["Legal_Entities"] = len(rows)

    # ---- Delivery_Locations ----
    headers, rows = _rows(wb["Delivery_Locations"])
    req_cols = ["Location_ID", "Location_Name", "Geolocation", "City", "State",
               "Country", "Address", "Active"]
    missing_cols = [c for c in req_cols if c not in headers]
    if missing_cols:
        errors.append(f"Delivery_Locations: missing column(s) {missing_cols}")
    else:
        seen = set()
        out = []
        for i, r in enumerate(rows, 2):
            lid = _get(r, headers, "Location_ID")
            if not lid:
                errors.append(f"Delivery_Locations row {i}: Location_ID is blank")
                continue
            if lid in seen:
                errors.append(f"Delivery_Locations: duplicate Location_ID '{lid}'")
            seen.add(lid)
            out.append({c: _get(r, headers, c) for c in req_cols})
        parsed["delivery_locations"] = out
    summary["Delivery_Locations"] = len(rows)

    # ---- Item Master (title row + real header row 2, matches the live sheet) ----
    ws = wb["Item Master"]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_row_idx = 2 if (first_row and first_row[0] and "Item Code" not in str(first_row)) else 1
    headers, rows = _rows(ws, header_row=header_row_idx)
    req_cols = ["Item Code", "Item Description", "Category", "Sub-Category",
               "Unit of Measure", "Unit Price", "Lead Time (days)", "In Stock",
               "Tags / Keywords", "Active", "HSN_Code", "GST_Rate"]
    # Weight_KG is deliberately optional, not in req_cols — added 2026-07-31
    # for real shipping-detail collection, but a seed file predating this
    # addition should still validate and load cleanly rather than fail
    # outright over a column that didn't exist yet when it was built.
    has_weight_col = "Weight_KG" in headers
    # Tracking_Type / Shelf_Life_Tracked are the identical kind of
    # deliberately-optional addition Weight_KG already is (TRC-US-01,
    # 2026-08-09) -- a seed file predating this should still load
    # cleanly, every item defaulting to untracked, not fail outright.
    has_tracking_col = "Tracking_Type" in headers
    has_shelf_life_col = "Shelf_Life_Tracked" in headers
    read_cols = (req_cols + (["Weight_KG"] if has_weight_col else [])
                + (["Tracking_Type"] if has_tracking_col else [])
                + (["Shelf_Life_Tracked"] if has_shelf_life_col else []))
    missing_cols = [c for c in req_cols if c not in headers]
    item_codes = set()
    if missing_cols:
        errors.append(f"Item Master: missing column(s) {missing_cols}")
    else:
        out = []
        for i, r in enumerate(rows, header_row_idx + 1):
            code = _get(r, headers, "Item Code")
            if not code:
                errors.append(f"Item Master row {i}: Item Code is blank")
                continue
            if code in item_codes:
                errors.append(f"Item Master: duplicate Item Code '{code}'")
            item_codes.add(code)
            if not _get(r, headers, "Item Description"):
                errors.append(f"Item Master '{code}': Item Description is blank")
            price = _get(r, headers, "Unit Price")
            if price is not None and _num(price) is None:
                errors.append(f"Item Master '{code}': Unit Price '{price}' isn't numeric")
            gst = _get(r, headers, "GST_Rate")
            if gst is not None and str(gst).strip() != "" and _num(gst) is None:
                errors.append(f"Item Master '{code}': GST_Rate '{gst}' isn't numeric")
            row_dict = {c: _get(r, headers, c) for c in read_cols}
            if not has_weight_col:
                row_dict["Weight_KG"] = None
            # A blank cell (column exists but this row's value is empty)
            # means the same thing as the column not existing at all --
            # untracked -- not a null tracking type to guess at later.
            if not row_dict.get("Tracking_Type"):
                row_dict["Tracking_Type"] = "None"
            if not row_dict.get("Shelf_Life_Tracked"):
                row_dict["Shelf_Life_Tracked"] = "No"
            out.append(row_dict)
        parsed["item_master"] = out
    summary["Item Master"] = len(rows)

    # ---- Vendor_Types (single column) ----
    headers, rows = _rows(wb["Vendor_Types"])
    if "Vendor_Type" not in headers:
        errors.append("Vendor_Types: missing column Vendor_Type")
        vendor_types = set()
    else:
        seen = set()
        for i, r in enumerate(rows, 2):
            vt = _get(r, headers, "Vendor_Type")
            if not vt:
                errors.append(f"Vendor_Types row {i}: blank value")
                continue
            if vt in seen:
                errors.append(f"Vendor_Types: duplicate value '{vt}'")
            seen.add(vt)
        vendor_types = seen
        parsed["vendor_types"] = sorted(seen)
    summary["Vendor_Types"] = len(rows)

    # ---- Customer_Types (single column) ----
    headers, rows = _rows(wb["Customer_Types"])
    if "Customer_Type" not in headers:
        errors.append("Customer_Types: missing column Customer_Type")
        customer_types = set()
    else:
        seen = set()
        for i, r in enumerate(rows, 2):
            ct = _get(r, headers, "Customer_Type")
            if not ct:
                errors.append(f"Customer_Types row {i}: blank value")
                continue
            if ct in seen:
                errors.append(f"Customer_Types: duplicate value '{ct}'")
            seen.add(ct)
        customer_types = seen
        parsed["customer_types"] = sorted(seen)
    summary["Customer_Types"] = len(rows)

    # ---- Vendor_Master ----
    headers, rows = _rows(wb["Vendor_Master"])
    req_cols = ["Vendor_ID", "Vendor_Name", "Geolocation", "City", "Country", "Address",
               "Contact_Name", "Contact_Email", "Active", "GSTIN", "PAN",
               "Bank_Account_No", "IFSC", "Bank_Name", "Bank_Branch",
               "Onboarding_Status", "KYC_Flag", "Onboarded_Date"]
    missing_cols = [c for c in req_cols if c not in headers]
    has_vtype_col = "Vendor_Type" in headers
    if missing_cols:
        errors.append(f"Vendor_Master: missing column(s) {missing_cols}")
    else:
        seen = set()
        out = []
        for i, r in enumerate(rows, 2):
            vid = _get(r, headers, "Vendor_ID")
            if not vid:
                errors.append(f"Vendor_Master row {i}: Vendor_ID is blank")
                continue
            if vid in seen:
                errors.append(f"Vendor_Master: duplicate Vendor_ID '{vid}'")
            seen.add(vid)
            if not _get(r, headers, "Vendor_Name"):
                errors.append(f"Vendor_Master '{vid}': Vendor_Name is blank")
            gstin = _get(r, headers, "GSTIN")
            if gstin:
                ok, msg, _ = vo.validate_gstin(gstin)
                if not ok:
                    errors.append(f"Vendor_Master '{vid}': GSTIN invalid — {msg}")
            pan = _get(r, headers, "PAN")
            if pan:
                ok, msg = vo.validate_pan_format(pan)
                if not ok:
                    errors.append(f"Vendor_Master '{vid}': PAN invalid — {msg}")
            vtype = _get(r, headers, "Vendor_Type") if has_vtype_col else None
            if vtype and vtype not in vendor_types:
                errors.append(f"Vendor_Master '{vid}': Vendor_Type '{vtype}' isn't in the Vendor_Types sheet")
            row_dict = {c: _get(r, headers, c) for c in req_cols}
            row_dict["Vendor_Type"] = vtype
            out.append(row_dict)
        parsed["vendor_master"] = out
    summary["Vendor_Master"] = len(rows)

    # ---- Customer_Master ----
    headers, rows = _rows(wb["Customer_Master"])
    req_cols = ["Customer_ID", "Customer_Name", "Customer_Type", "Geolocation", "City",
               "Country", "Address", "Contact_Name", "Contact_Email", "Contact_Phone",
               "GSTIN", "PAN", "Credit_Limit", "Credit_Status", "Payment_Terms",
               "Onboarding_Status", "KYC_Flag", "Onboarded_Date", "Active"]
    missing_cols = [c for c in req_cols if c not in headers]
    if missing_cols:
        errors.append(f"Customer_Master: missing column(s) {missing_cols}")
    else:
        seen = set()
        out = []
        for i, r in enumerate(rows, 2):
            cid = _get(r, headers, "Customer_ID")
            if not cid:
                errors.append(f"Customer_Master row {i}: Customer_ID is blank")
                continue
            if cid in seen:
                errors.append(f"Customer_Master: duplicate Customer_ID '{cid}'")
            seen.add(cid)
            if not _get(r, headers, "Customer_Name"):
                errors.append(f"Customer_Master '{cid}': Customer_Name is blank")
            gstin = _get(r, headers, "GSTIN")
            if gstin:
                ok, msg, _ = vo.validate_gstin(gstin)
                if not ok:
                    errors.append(f"Customer_Master '{cid}': GSTIN invalid — {msg}")
            pan = _get(r, headers, "PAN")
            if pan:
                ok, msg = vo.validate_pan_format(pan)
                if not ok:
                    errors.append(f"Customer_Master '{cid}': PAN invalid — {msg}")
            ctype = _get(r, headers, "Customer_Type")
            if ctype and ctype not in customer_types:
                errors.append(f"Customer_Master '{cid}': Customer_Type '{ctype}' isn't in the Customer_Types sheet")
            limit = _get(r, headers, "Credit_Limit")
            if limit is not None and str(limit).strip() != "" and _num(limit) is None:
                errors.append(f"Customer_Master '{cid}': Credit_Limit '{limit}' isn't numeric")
            row_out = {c: _get(r, headers, c) for c in req_cols}
            # Pincode is optional at the file level, same reasoning as
            # Org_Profile above — BuyerDtls.Pin for gst_einvoice.py.
            row_out["Pincode"] = _get(r, headers, "Pincode")
            out.append(row_out)
        parsed["customer_master"] = out
    summary["Customer_Master"] = len(rows)

    # ---- BOM_Items (positional, matches migrate_bom_to_sqlite.py exactly) ----
    ws = wb["BOM_Items"]
    bom_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue
        if len(row) < 7:
            errors.append(f"BOM_Items row {i}: expected 7 columns, found {len(row)}")
            continue
        parent_code, parent_desc, comp_code, comp_desc, qty_per, uom, notes = row[:7]
        if parent_code not in item_codes:
            errors.append(f"BOM_Items row {i}: Parent code '{parent_code}' isn't in Item Master")
        if comp_code not in item_codes:
            errors.append(f"BOM_Items row {i}: Component code '{comp_code}' isn't in Item Master")
        if _num(qty_per) is None:
            errors.append(f"BOM_Items row {i}: Qty_Per '{qty_per}' isn't numeric")
        bom_rows.append((parent_code, parent_desc, comp_code, comp_desc, qty_per, uom, notes))
    parsed["bom_items"] = bom_rows
    summary["BOM_Items"] = len(bom_rows)

    # ---- Purchase Bundles ----
    headers, rows = _rows(wb["Purchase Bundles"])
    req_cols = ["Bundle_ID", "Bundle_Name", "Description", "Department",
               "Material_Code", "Material_Desc", "UOM", "Default_Qty", "Notes"]
    missing_cols = [c for c in req_cols if c not in headers]
    if missing_cols:
        errors.append(f"Purchase Bundles: missing column(s) {missing_cols}")
    elif rows:
        out = []
        for i, r in enumerate(rows, 2):
            mat_code = _get(r, headers, "Material_Code")
            if mat_code not in item_codes:
                errors.append(f"Purchase Bundles row {i}: Material_Code '{mat_code}' isn't in Item Master")
            qty = _get(r, headers, "Default_Qty")
            if _num(qty) is None:
                errors.append(f"Purchase Bundles row {i}: Default_Qty '{qty}' isn't numeric")
            out.append({c: _get(r, headers, c) for c in req_cols})
        parsed["purchase_bundles"] = out
    summary["Purchase Bundles"] = len(rows)

    # ---- Chart_of_Accounts ----
    headers, rows = _rows(wb["Chart_of_Accounts"])
    req_cols = ["Account_Code", "Account_Name", "Account_Type", "Description"]
    missing_cols = [c for c in req_cols if c not in headers]
    if missing_cols:
        errors.append(f"Chart_of_Accounts: missing column(s) {missing_cols}")
    else:
        seen = set()
        out = []
        for i, r in enumerate(rows, 2):
            code = _get(r, headers, "Account_Code")
            if not code:
                errors.append(f"Chart_of_Accounts row {i}: Account_Code is blank")
                continue
            if code in seen:
                errors.append(f"Chart_of_Accounts: duplicate Account_Code '{code}'")
            seen.add(code)
            atype = _get(r, headers, "Account_Type")
            if atype not in ACCOUNT_TYPES:
                errors.append(f"Chart_of_Accounts '{code}': Account_Type '{atype}' "
                              f"must be one of {sorted(ACCOUNT_TYPES)}")
            out.append({c: _get(r, headers, c) for c in req_cols})
        parsed["chart_of_accounts"] = out
    summary["Chart_of_Accounts"] = len(rows)

    wb.close()
    return {"valid": len(errors) == 0, "errors": errors, "warnings": warnings,
            "summary": summary, "parsed": parsed}


# ── Backup ────────────────────────────────────────────────────────────────────
def backup_current_state():
    """Zips the current erp_pilot.db (+ data.xlsx if present) into bytes
    — a downloadable snapshot of whatever's about to be wiped. Not
    mandatory (the caller decides whether to offer/require it), but
    cheap insurance against "wait, I didn't mean to lose that.\""""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if os.path.exists(DB_FILE):
            zf.write(DB_FILE, arcname="erp_pilot.db")
        if os.path.exists(DATA_FILE):
            zf.write(DATA_FILE, arcname="data.xlsx")
    buf.seek(0)
    return buf.read()


# ── Reset + reseed ────────────────────────────────────────────────────────────
def reset_and_reseed(file_path):
    """
    Re-validates the file itself (doesn't just trust a UI checkbox that
    validation "already passed" — a defensive re-check, since this is
    the one function in this module that's actually destructive).
    Raises ValueError with the full error list if validation fails;
    wipes and reloads only if it passes clean.

    Returns {"tables_wiped": int, "seed_counts": {table: count}}.
    """
    result = validate_seed_file(file_path)
    if not result["valid"]:
        raise ValueError("Seed file failed validation:\n" + "\n".join(result["errors"]))
    parsed = result["parsed"]

    db.init_schema()
    conn = db.get_connection()
    seed_counts = {}
    try:
        for t in TRANSACTIONAL_TABLES + SEED_TABLES:
            conn.execute(f"DELETE FROM {t}")

        op = parsed["org_profile"]
        conn.execute(
            "INSERT INTO org_profile (org_id, legal_name, gstin, pan, address, city, "
            "state, country, pincode, bank_account_no, ifsc, bank_name, contact_email, "
            "contact_phone) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (op["Org_ID"], op["Legal_Name"], op["GSTIN"], op["PAN"], op["Address"],
             op["City"], op["State"], op["Country"], op.get("Pincode"), op["Bank_Account_No"],
             op["IFSC"], op["Bank_Name"], op["Contact_Email"], op["Contact_Phone"]))
        seed_counts["org_profile"] = 1

        for el, val in parsed["org_defaults"]:
            conn.execute("INSERT INTO org_defaults (org_element, default_value) VALUES (?,?)", (el, val))
        seed_counts["org_defaults"] = len(parsed["org_defaults"])

        for le in parsed["legal_entities"]:
            conn.execute(
                "INSERT INTO legal_entities (le_id, le_name, gstin, pan, address, city, "
                "state, country, bank_account_no, ifsc, bank_name, contact_email, contact_phone) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (le["LE_ID"], le["LE_Name"], le["GSTIN"], le["PAN"], le["Address"], le["City"],
                 le["State"], le["Country"], le["Bank_Account_No"], le["IFSC"], le["Bank_Name"],
                 le["Contact_Email"], le["Contact_Phone"]))
        seed_counts["legal_entities"] = len(parsed["legal_entities"])

        for dl in parsed["delivery_locations"]:
            conn.execute(
                "INSERT INTO delivery_locations (location_id, location_name, geolocation, "
                "city, state, country, address, active) VALUES (?,?,?,?,?,?,?,?)",
                (dl["Location_ID"], dl["Location_Name"], dl["Geolocation"], dl["City"],
                 dl["State"], dl["Country"], dl["Address"], dl["Active"]))
        seed_counts["delivery_locations"] = len(parsed["delivery_locations"])

        for it in parsed["item_master"]:
            conn.execute(
                "INSERT INTO item_master (item_code, item_desc, category, subcategory, uom, "
                "unit_price, lead_time_days, in_stock, tags, active, hsn_code, gst_rate, "
                "weight_kg, tracking_type, shelf_life_tracked) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (it["Item Code"], it["Item Description"], it["Category"], it["Sub-Category"],
                 it["Unit of Measure"], _num(it["Unit Price"]), it["Lead Time (days)"],
                 it["In Stock"], it["Tags / Keywords"], it["Active"], it["HSN_Code"],
                 _num(it["GST_Rate"]), _num(it.get("Weight_KG")),
                 it.get("Tracking_Type", "None"), it.get("Shelf_Life_Tracked", "No")))
        seed_counts["item_master"] = len(parsed["item_master"])

        for vt in parsed["vendor_types"]:
            conn.execute("INSERT INTO vendor_types (vendor_type) VALUES (?)", (vt,))
        seed_counts["vendor_types"] = len(parsed["vendor_types"])

        for ct in parsed["customer_types"]:
            conn.execute("INSERT INTO customer_types (customer_type) VALUES (?)", (ct,))
        seed_counts["customer_types"] = len(parsed["customer_types"])

        for v in parsed["vendor_master"]:
            conn.execute(
                "INSERT INTO vendor_master (vendor_id, vendor_name, geolocation, city, country, "
                "address, contact_name, contact_email, active, gstin, pan, bank_account_no, ifsc, "
                "bank_name, bank_branch, onboarding_status, kyc_flag, onboarded_date, vendor_type) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (v["Vendor_ID"], v["Vendor_Name"], v["Geolocation"], v["City"], v["Country"],
                 v["Address"], v["Contact_Name"], v["Contact_Email"], v["Active"], v["GSTIN"],
                 v["PAN"], v["Bank_Account_No"], v["IFSC"], v["Bank_Name"], v["Bank_Branch"],
                 v["Onboarding_Status"], v["KYC_Flag"], v["Onboarded_Date"], v["Vendor_Type"]))
        seed_counts["vendor_master"] = len(parsed["vendor_master"])

        for c in parsed["customer_master"]:
            conn.execute(
                "INSERT INTO customer_master (customer_id, customer_name, customer_type, "
                "geolocation, city, country, address, contact_name, contact_email, contact_phone, "
                "gstin, pan, credit_limit, credit_status, payment_terms, onboarding_status, "
                "kyc_flag, onboarded_date, active, pincode) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (c["Customer_ID"], c["Customer_Name"], c["Customer_Type"], c["Geolocation"],
                 c["City"], c["Country"], c["Address"], c["Contact_Name"], c["Contact_Email"],
                 c["Contact_Phone"], c["GSTIN"], c["PAN"], _num(c["Credit_Limit"]),
                 c["Credit_Status"], c["Payment_Terms"], c["Onboarding_Status"], c["KYC_Flag"],
                 c["Onboarded_Date"], c["Active"], c.get("Pincode")))
        seed_counts["customer_master"] = len(parsed["customer_master"])

        for b in parsed["bom_items"]:
            conn.execute(
                "INSERT INTO bom_items (parent_code, parent_desc, component_code, component_desc, "
                "qty_per, uom, notes) VALUES (?,?,?,?,?,?,?)", b)
        seed_counts["bom_items"] = len(parsed["bom_items"])

        bundle_ids = set()
        pb_lines = 0
        for pb in parsed.get("purchase_bundles", []):
            bid = pb["Bundle_ID"]
            if bid not in bundle_ids:
                conn.execute(
                    "INSERT INTO purchase_bundles (bundle_id, bundle_name, description, "
                    "department, active) VALUES (?,?,?,?,?)",
                    (bid, pb["Bundle_Name"], pb["Description"], pb["Department"], "Yes"))
                bundle_ids.add(bid)
            conn.execute(
                "INSERT INTO purchase_bundle_items (bundle_id, material_code, material_desc, "
                "uom, default_qty, notes) VALUES (?,?,?,?,?,?)",
                (bid, pb["Material_Code"], pb["Material_Desc"], pb["UOM"],
                 _num(pb["Default_Qty"]), pb["Notes"]))
            pb_lines += 1
        seed_counts["purchase_bundles"] = len(bundle_ids)
        seed_counts["purchase_bundle_items"] = pb_lines

        for a in parsed["chart_of_accounts"]:
            conn.execute(
                "INSERT INTO chart_of_accounts (account_code, account_name, account_type, "
                "description) VALUES (?,?,?,?)",
                (a["Account_Code"], a["Account_Name"], a["Account_Type"], a["Description"]))
        seed_counts["chart_of_accounts"] = len(parsed["chart_of_accounts"])

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    # init_schema() itself runs before any seed data is even inserted
    # above — its own backfills (pr_date, po_date, and now
    # default_delivery_location) would otherwise run against an empty
    # or stale table and have nothing real to work with. init_schema()
    # is explicitly idempotent and cheap to call again, so this just
    # re-runs it now that real seed data actually exists to backfill
    # against.
    db.init_schema()

    return {"tables_wiped": len(TRANSACTIONAL_TABLES) + len(SEED_TABLES),
            "seed_counts": seed_counts}
