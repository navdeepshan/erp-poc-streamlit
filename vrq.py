"""
vrq.py — Vendor Registration Questionnaire (VRQ).

Mirrors how large Indian corporates actually run vendor intake (checked
against Adani Group's published process: EOI -> shortlist -> a detailed
Vendor Registration Questionnaire covering operations/compliance/
certifications -> review -> approve -> portal access). This module builds
that middle step, and plugs directly into the vendor_onboarding.py gate
already built:

  create_vrq_request()  -> generate_vrq_document()  -> vendor fills it in
       -> parse_uploaded_vrq() / record_responses()  -> validate_responses()
       -> promote_to_vendor()  -> vendor_onboarding.approve_vendor()

promote_to_vendor() hands off to vendor_onboarding.upsert_vendor(), so VRQ
answers get the exact same GSTIN/PAN checksum validation as the direct
intake form — one validation path, two ways in.

New sheets:
  VRQ_Requests    VRQ_ID | Vendor_Name | Contact_Email | Sent_Date |
                  Status | Filename | Vendor_ID
  VRQ_Responses   VRQ_ID | Question_Key | Section | Question | Answer
                  (long/EAV format — easy to extend the questionnaire
                  later without a schema migration)

The generated document embeds each question's machine key in a narrow
reference column, so parsing a filled-in copy back is exact-match, not
fuzzy text matching.

SQLite pilot: VRQ_Requests and VRQ_Responses now live in erp_pilot.db
(tables `vrq_requests`, `vrq_responses`), not in data.xlsx. vrq.py is
the exclusive owner of both — no other module reads or writes them —
so this was a contained, single-file migration.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, io
from datetime import date, timedelta

import db
import pr_consolidation as pc
import vendor_onboarding as vo

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")

VRQ_REQUESTS_SHEET = "VRQ_Requests"
VRQ_RESPONSES_SHEET = "VRQ_Responses"
VRQ_REQ_COLS = ["VRQ_ID", "Vendor_Name", "Contact_Email", "Sent_Date",
               "Status", "Filename", "Vendor_ID"]
VRQ_RESP_COLS = ["VRQ_ID", "Question_Key", "Section", "Question", "Answer"]

# ── The questionnaire itself ─────────────────────────────────────────────────
# Expanded to match the real scope of large-enterprise VRQs (cross-checked
# against Adani Group's published ~10-section process, plus the ESG/insurance
# sections that have become standard in enterprise SRM questionnaires since).
# Sections/questions are data, not code — extend this list to change the form;
# generate_vrq_document(), parse_uploaded_vrq(), and validate_responses() all
# read it dynamically, so no other code needs to change.
VRQ_SECTIONS = [
    {"section": "Company Information", "questions": [
        {"key": "legal_name", "text": "Legal Company Name", "required": True},
        {"key": "entity_type", "text": "Entity Type (Proprietorship / Partnership / LLP / Pvt Ltd / Public Ltd)", "required": True},
        {"key": "cin", "text": "Corporate Identification Number (CIN), if applicable", "required": False},
        {"key": "year_established", "text": "Year Established", "required": False},
        {"key": "group_holding_company", "text": "Parent / Holding Company, if any", "required": False},
        {"key": "registered_address", "text": "Registered Address", "required": True},
        {"key": "corporate_office_address", "text": "Corporate Office Address, if different", "required": False},
        {"key": "manufacturing_locations", "text": "Manufacturing / Warehouse Location(s)", "required": False},
        {"key": "website", "text": "Company Website", "required": False},
        {"key": "employee_count", "text": "Approx. Number of Employees", "required": False},
    ]},
    {"section": "Tax & Statutory Registration", "questions": [
        {"key": "gstin", "text": "GSTIN", "required": True},
        {"key": "pan", "text": "PAN", "required": True},
        {"key": "tan", "text": "TAN (Tax Deduction Account Number)", "required": False},
        {"key": "udyam_number", "text": "Udyam (MSME) Registration Number, if applicable", "required": False},
        {"key": "iec_code", "text": "Import Export Code (IEC), if applicable", "required": False},
        {"key": "pf_registration", "text": "Provident Fund (PF) Registration Number", "required": False},
        {"key": "esi_registration", "text": "ESI Registration Number", "required": False},
    ]},
    {"section": "Banking Details", "questions": [
        {"key": "bank_account_no", "text": "Bank Account Number", "required": True},
        {"key": "ifsc", "text": "IFSC Code", "required": True},
        {"key": "bank_name", "text": "Bank Name", "required": False},
        {"key": "bank_branch", "text": "Bank Branch", "required": False},
        {"key": "account_type", "text": "Account Type (Current / Savings)", "required": False},
    ]},
    {"section": "Products, Services & Capacity", "questions": [
        {"key": "categories_supplied", "text": "Product/Service Categories Supplied", "required": True},
        {"key": "brands_represented", "text": "Brands Represented / Authorized Dealership For", "required": False},
        {"key": "manufacturer_or_trader", "text": "Manufacturer / Trader / Both", "required": False},
        {"key": "annual_capacity", "text": "Approx. Annual Supply Capacity", "required": False},
        {"key": "annual_turnover_supply", "text": "Approx. Annual Turnover from Supply of These Categories", "required": False},
        {"key": "key_raw_materials_sourcing", "text": "Key Raw Materials / Sourcing (if manufacturer)", "required": False},
    ]},
    {"section": "Quality & Certifications", "questions": [
        {"key": "iso_9001", "text": "ISO 9001 Certified? (Yes/No + Cert No.)", "required": False},
        {"key": "iso_13485", "text": "ISO 13485 (Medical Devices) Certified? (Yes/No + Cert No.)", "required": False},
        {"key": "iso_14001", "text": "ISO 14001 (Environmental Management) Certified? (Yes/No + Cert No.)", "required": False},
        {"key": "iso_45001", "text": "ISO 45001 (Occupational Health & Safety) Certified? (Yes/No + Cert No.)", "required": False},
        {"key": "ce_marking", "text": "CE Marking on Relevant Products? (Yes/No)", "required": False},
        {"key": "cdsco_license", "text": "CDSCO / Drug License Number, if applicable", "required": False},
        {"key": "quality_control_process", "text": "Brief Description of Quality Control Process", "required": False},
    ]},
    {"section": "Health, Safety & Environment (HSE)", "questions": [
        {"key": "hse_policy", "text": "Documented HSE Policy in Place? (Yes/No)", "required": False},
        {"key": "safety_incidents_3yr", "text": "Any Major Safety/Environmental Incidents in the Last 3 Years?", "required": False},
        {"key": "environmental_clearance", "text": "Environmental Clearance Certificate, if manufacturing", "required": False},
    ]},
    {"section": "Compliance & Ethics Declarations", "questions": [
        {"key": "anti_bribery", "text": "Do you have an Anti-Bribery & Anti-Corruption Policy? (Yes/No)", "required": True},
        {"key": "code_of_conduct", "text": "Do you agree to our Supplier Code of Conduct? (Yes/No)", "required": True},
        {"key": "conflict_of_interest", "text": "Any Conflict of Interest to Declare with Our Employees?", "required": True},
        {"key": "child_forced_labor_policy", "text": "Policy Against Child/Forced Labor in Place? (Yes/No)", "required": False},
        {"key": "data_privacy_agreement", "text": "Willing to Sign a Data Privacy/Confidentiality Agreement? (Yes/No)", "required": False},
        {"key": "sanctions_declaration", "text": "Are You or Your Directors/Owners on Any Government Sanctions or Denied-Party List? (Yes/No)", "required": True},
        {"key": "litigation_declaration", "text": "Any Pending Litigation, Insolvency, or Bankruptcy Proceedings? (Yes/No + details)", "required": False},
    ]},
    {"section": "Financial Information", "questions": [
        {"key": "annual_turnover_total", "text": "Total Annual Turnover, Last Financial Year", "required": False},
        {"key": "net_worth", "text": "Approx. Net Worth", "required": False},
        {"key": "credit_rating", "text": "Credit Rating, if Available (Agency + Rating)", "required": False},
    ]},
    {"section": "Insurance", "questions": [
        {"key": "product_liability_insurance", "text": "Product Liability Insurance? (Yes/No + Coverage Amount)", "required": False},
        {"key": "general_liability_insurance", "text": "General Liability Insurance? (Yes/No)", "required": False},
    ]},
    {"section": "ESG & Sustainability", "questions": [
        {"key": "msme_women_owned", "text": "Women-Owned / Minority-Owned / MSME Enterprise? (specify)", "required": False},
        {"key": "sustainability_initiatives", "text": "Any Sustainability/Environmental Initiatives?", "required": False},
        {"key": "csr_policy", "text": "CSR Policy in Place? (Yes/No)", "required": False},
    ]},
    {"section": "References & Track Record", "questions": [
        {"key": "years_in_industry", "text": "Years of Experience in This Industry", "required": False},
        {"key": "reference_1", "text": "Reference Client 1 (Name, Contact)", "required": False},
        {"key": "reference_2", "text": "Reference Client 2 (Name, Contact)", "required": False},
        {"key": "major_past_projects", "text": "Major Past Supply Contracts/Projects, if any", "required": False},
    ]},
    {"section": "Contact Information", "questions": [
        {"key": "contact_name", "text": "Primary Contact Name", "required": True},
        {"key": "contact_designation", "text": "Designation / Title", "required": False},
        {"key": "contact_email", "text": "Primary Contact Email", "required": True},
        {"key": "contact_phone", "text": "Primary Contact Phone", "required": False},
        {"key": "escalation_contact_name", "text": "Escalation Contact Name, if different", "required": False},
        {"key": "escalation_contact_email", "text": "Escalation Contact Email", "required": False},
    ]},
]

_ALL_QUESTIONS = {q["key"]: {**q, "section": s["section"]}
                  for s in VRQ_SECTIONS for q in s["questions"]}


# ── Sheet bootstrap ────────────────────────────────────────────────────────────
def ensure_sheets(wb=None):
    """Kept for signature compatibility — VRQ_Requests/VRQ_Responses no
    longer live in the Excel workbook. `wb` is accepted and ignored."""
    db.init_schema()


# ── Requests ──────────────────────────────────────────────────────────────────
def _next_vrq_id(conn):
    rows = conn.execute("SELECT vrq_id FROM vrq_requests WHERE vrq_id LIKE 'VRQ-%'").fetchall()
    mx = 0
    for r in rows:
        try: mx = max(mx, int(r["vrq_id"].split("-")[1]))
        except Exception: pass
    return f"VRQ-{mx+1:05d}"


def create_vrq_request(vendor_name, contact_email="", data_file=None):
    db.init_schema()
    conn = db.get_connection()
    try:
        vrq_id = _next_vrq_id(conn)
        conn.execute(
            "INSERT INTO vrq_requests (vrq_id, vendor_name, contact_email, sent_date, "
            "status, filename, vendor_id) VALUES (?,?,?,?,?,?,?)",
            (vrq_id, vendor_name, contact_email, date.today().strftime("%Y-%m-%d"),
             "Sent", "", ""),
        )
        conn.commit()
    finally:
        conn.close()
    return vrq_id


def get_vrq_requests(status=None, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute("SELECT * FROM vrq_requests ORDER BY vrq_id").fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        row = {"vrq_id": r["vrq_id"], "vendor_name": r["vendor_name"],
               "contact_email": r["contact_email"], "sent_date": r["sent_date"],
               "status": r["status"], "filename": r["filename"], "vendor_id": r["vendor_id"]}
        if status and row["status"] != status:
            continue
        out.append(row)
    return out


def get_vrq(vrq_id, data_file=None):
    for r in get_vrq_requests(data_file=data_file):
        if r["vrq_id"] == vrq_id:
            return r
    return None


def _set_vrq_status(vrq_id, status, data_file=None, vendor_id=None, filename=None):
    conn = db.get_connection()
    try:
        conn.execute("UPDATE vrq_requests SET status = ? WHERE vrq_id = ?", (status, vrq_id))
        if vendor_id is not None:
            conn.execute("UPDATE vrq_requests SET vendor_id = ? WHERE vrq_id = ?", (vendor_id, vrq_id))
        if filename is not None:
            conn.execute("UPDATE vrq_requests SET filename = ? WHERE vrq_id = ?", (filename, vrq_id))
        conn.commit()
    finally:
        conn.close()


# ── Document generation ──────────────────────────────────────────────────────
def generate_vrq_document(vrq_id, vendor_name, contact_email=""):
    """Builds the actual fillable VRQ workbook. Column A carries each
    question's machine key for exact-match parsing later — narrow &
    greyed-out so it doesn't distract a human filling the form."""
    thin = Side(style="thin", color="CBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = "1E3A5F"; purple = "4C1D95"; amber = "FFF8E8"; grey = "94A3B8"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VRQ"

    def title(ref, text, size=15):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=size, bold=True, color="0F172A")

    def label(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=9, bold=True, color="475569")

    def value(ref, text):
        c = ws[ref]; c.value = text
        c.font = Font(name="Arial", size=10, color="1A1A2E")

    title("A1", "VENDOR REGISTRATION QUESTIONNAIRE")
    ws.merge_cells("A1:E1")
    closing = (date.today() + timedelta(days=14)).strftime("%Y-%m-%d")
    label("A3", "VRQ ID:");       value("B3", vrq_id)
    label("A4", "Issued:");       value("B4", date.today().strftime("%Y-%m-%d"))
    label("A5", "Response by:");  value("B5", closing)
    label("D3", "Vendor:");       value("E3", vendor_name)
    label("D4", "Contact Email:"); value("E4", contact_email)

    ws["A7"] = ("Please complete every required question in the 'Your Answer' column "
                "and return this file. Do not edit column A.")
    ws["A7"].font = Font(name="Arial", size=9, italic=True, color="64748B")
    ws.merge_cells("A7:E7")

    hdr_row = 9
    hdrs = ["Key", "Section", "Question", "Required", "Your Answer"]
    hf = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor=navy)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = hf; c.fill = hb; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = hdr_row + 1
    for section in VRQ_SECTIONS:
        sc = ws.cell(r, 1, "")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        sec_cell = ws.cell(r, 1, section["section"])
        sec_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        sec_cell.fill = PatternFill("solid", fgColor=purple)
        sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        r += 1
        for q in section["questions"]:
            ws.cell(r, 1, q["key"]).font = Font(name="Arial", size=6, color=grey)
            ws.cell(r, 2, section["section"]).font = Font(name="Arial", size=9, color="64748B")
            qc = ws.cell(r, 3, q["text"])
            qc.font = Font(name="Arial", size=9, color="1A1A2E")
            qc.alignment = Alignment(wrap_text=True, vertical="center")
            reqc = ws.cell(r, 4, "Yes" if q["required"] else "No")
            reqc.font = Font(name="Arial", size=9, color=("B91C1C" if q["required"] else "94A3B8"))
            ansc = ws.cell(r, 5, None)
            ansc.fill = PatternFill("solid", fgColor=amber)
            for c in range(1, 6):
                ws.cell(r, c).border = bdr
            r += 1

    widths = [3, 20, 46, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def send_vrq(vendor_name, contact_email="", data_file=None):
    """Convenience: create the request AND generate the document in one call."""
    vrq_id = create_vrq_request(vendor_name, contact_email, data_file)
    filename = f"{vrq_id}.xlsx"
    file_bytes = generate_vrq_document(vrq_id, vendor_name, contact_email)
    _set_vrq_status(vrq_id, "Sent", data_file, filename=filename)
    return vrq_id, filename, file_bytes


# ── Parsing a filled-in copy ──────────────────────────────────────────────────
def parse_uploaded_vrq(file_bytes):
    """Pure parsing — reads column A (key) + column E (answer) from an
    uploaded workbook built by generate_vrq_document(). Returns {key: answer}.
    Raises ValueError if it doesn't look like a VRQ file at all."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Could not read this file as an Excel workbook — {e}")

    answers = {}
    found_any_key = False
    for row in ws.iter_rows(min_row=10, max_col=5, values_only=True):
        key = str(row[0] or "").strip()
        if not key or key not in _ALL_QUESTIONS:
            continue
        found_any_key = True
        ans = row[4]
        if ans is not None and str(ans).strip():
            answers[key] = str(ans).strip()
    if not found_any_key:
        raise ValueError("This doesn't look like a VRQ file generated by this app "
                          "(no recognizable question keys found in column A).")
    return answers


# ── Responses ────────────────────────────────────────────────────────────────
def record_responses(vrq_id, answers, data_file=None):
    """answers: {question_key: answer_text}. Overwrites any previous
    responses for this VRQ_ID (so a corrected re-upload just works) —
    one transaction: delete existing rows for this vrq_id, then insert
    fresh ones."""
    db.init_schema()
    conn = db.get_connection()
    try:
        conn.execute("DELETE FROM vrq_responses WHERE vrq_id = ?", (vrq_id,))
        for key, ans in answers.items():
            meta = _ALL_QUESTIONS.get(key, {"section": "", "text": key})
            conn.execute(
                "INSERT INTO vrq_responses (vrq_id, question_key, section, question, answer) "
                "VALUES (?,?,?,?,?)",
                (vrq_id, key, meta["section"], meta.get("text", key), ans),
            )
        conn.commit()
    finally:
        conn.close()
    _set_vrq_status(vrq_id, "Received", data_file)
    return len(answers)


def get_responses(vrq_id, data_file=None):
    conn = db.get_connection()
    try:
        rows = conn.execute(
            "SELECT question_key, answer FROM vrq_responses WHERE vrq_id = ?", (vrq_id,)
        ).fetchall()
    finally:
        conn.close()
    return {r["question_key"]: r["answer"] for r in rows}


def simulate_response(vrq_id, data_file=None):
    """Demo helper: plausible answers for every question, including a
    genuinely checksum-valid dummy GSTIN (via vendor_onboarding.demo_gstin)
    so downstream validation has something real to pass/fail against."""
    vrq_row = get_vrq(vrq_id, data_file)
    name = vrq_row["vendor_name"] if vrq_row else "Demo Vendor Pvt Ltd"
    answers = {
        # Company Information
        "legal_name": name, "entity_type": "Private Limited",
        "cin": "U33110KL2014PTC012345", "year_established": "2014",
        "group_holding_company": "None", "registered_address": "MG Road, Kochi, Kerala",
        "corporate_office_address": "Same as registered address",
        "manufacturing_locations": "N/A (trading entity)",
        "website": "www.example-vendor.in", "employee_count": "45",
        # Tax & Statutory Registration
        "gstin": vo.demo_gstin("32", "AABCU9603R", "1"), "pan": "AABCU9603R",
        "tan": "COCK12345D", "udyam_number": "UDYAM-KL-03-1234567",
        "iec_code": "N/A", "pf_registration": "KR/KCH/1234567",
        "esi_registration": "42000112330000999",
        # Banking Details
        "bank_account_no": "50100123456789", "ifsc": "HDFC0001234",
        "bank_name": "HDFC Bank", "bank_branch": "MG Road, Kochi", "account_type": "Current",
        # Products, Services & Capacity
        "categories_supplied": "Endodontics, Restorative & Aesthetic",
        "brands_represented": "Mani, 3M ESPE (authorized distributor)",
        "manufacturer_or_trader": "Trader", "annual_capacity": "50,000 units/year",
        "annual_turnover_supply": "\u20b91.2 crore", "key_raw_materials_sourcing": "N/A",
        # Quality & Certifications
        "iso_9001": "Yes - ISO9001:2015 Cert #IN-2024-8891",
        "iso_13485": "No", "iso_14001": "No", "iso_45001": "No",
        "cdsco_license": "KL-DL-2024-77213", "ce_marking": "Yes",
        "quality_control_process": "Incoming batch inspection + expiry tracking on all consumables",
        # HSE
        "hse_policy": "Yes", "safety_incidents_3yr": "None",
        "environmental_clearance": "N/A (trading entity)",
        # Compliance & Ethics
        "anti_bribery": "Yes", "code_of_conduct": "Yes",
        "conflict_of_interest": "None declared", "child_forced_labor_policy": "Yes",
        "data_privacy_agreement": "Yes", "sanctions_declaration": "No",
        "litigation_declaration": "None",
        # Financial
        "annual_turnover_total": "\u20b91.5 crore", "net_worth": "\u20b945 lakh",
        "credit_rating": "Not rated",
        # Insurance
        "product_liability_insurance": "Yes - \u20b950 lakh coverage",
        "general_liability_insurance": "Yes",
        # ESG
        "msme_women_owned": "MSME (Micro)", "sustainability_initiatives": "Paperless invoicing",
        "csr_policy": "No",
        # References
        "years_in_industry": "12", "reference_1": "City Dental Hospital, Kochi",
        "reference_2": "Apex Dental Clinic, Trivandrum",
        "major_past_projects": "Regular supplier to 8 dental clinics across Kerala",
        # Contact
        "contact_name": "Ravi Menon", "contact_designation": "Sales Manager",
        "contact_email": vrq_row["contact_email"] if vrq_row else "ravi@example-vendor.in",
        "contact_phone": "+91 98765 43210",
        "escalation_contact_name": "Suresh Kumar (Director)",
        "escalation_contact_email": "suresh@example-vendor.in",
    }
    record_responses(vrq_id, answers, data_file)
    return answers


# ── Validation & promotion ───────────────────────────────────────────────────
def validate_responses(vrq_id, data_file=None):
    """Checks required-field completeness + runs the same GSTIN/PAN checksum
    validators used in the direct intake form. Returns {key: {ok, message}}."""
    answers = get_responses(vrq_id, data_file)
    checks = {}
    for key, meta in _ALL_QUESTIONS.items():
        if meta["required"] and not answers.get(key):
            checks[key] = {"ok": False, "message": f"Required — not answered"}
    if answers.get("gstin"):
        ok, msg, _ = vo.validate_gstin(answers["gstin"])
        checks["gstin"] = {"ok": ok, "message": msg}
    if answers.get("pan"):
        ok, msg = vo.validate_pan_format(answers["pan"])
        checks["pan"] = {"ok": ok, "message": msg}
    return checks


def promote_to_vendor(vrq_id, vendor_id, data_file=None):
    """Maps VRQ answers onto vendor_onboarding.upsert_vendor()'s field names
    and runs it through the exact same GSTIN/PAN validation + approval-gate
    pipeline as the direct intake form."""
    answers = get_responses(vrq_id, data_file)
    if not answers:
        raise ValueError(f"No responses recorded for {vrq_id} yet.")
    fields = {
        "Vendor_Name": answers.get("legal_name", ""),
        "Address": answers.get("registered_address", ""),
        "GSTIN": answers.get("gstin", ""), "PAN": answers.get("pan", ""),
        "Bank_Account_No": answers.get("bank_account_no", ""),
        "IFSC": answers.get("ifsc", ""),
        "Bank_Name": answers.get("bank_name", ""),
        "Bank_Branch": answers.get("bank_branch", ""),
        "Contact_Name": answers.get("contact_name", ""),
        "Contact_Email": answers.get("contact_email", ""),
    }
    result = vo.upsert_vendor(vendor_id, fields, data_file)
    _set_vrq_status(vrq_id, "Promoted", data_file, vendor_id=vendor_id)
    return result


def reject_vrq(vrq_id, data_file=None):
    _set_vrq_status(vrq_id, "Rejected", data_file)


def stats(data_file=None):
    reqs = get_vrq_requests(data_file=data_file)
    by_status = {}
    for r in reqs:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    return {"total": len(reqs), "by_status": by_status}


if __name__ == "__main__":
    print("VRQ stats:", stats())
    vrq_id, fname, fbytes = send_vrq("Test Dental Traders", "test@example.in")
    print(f"Created {vrq_id}, doc {fname} ({len(fbytes)} bytes)")
    answers = simulate_response(vrq_id)
    print(f"Simulated {len(answers)} answers")
    checks = validate_responses(vrq_id)
    print("Validation issues:", {k: v for k, v in checks.items() if not v["ok"]} or "none")
